from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── スコア正規化 ──
def normalize_score(v):
    if v is None:
        return ""
    if isinstance(v, int):
        return v
    s = str(v).strip()
    if "/" in s:
        return int(s.split("/")[0])
    try:
        return int(s)
    except:
        return s

# ── ソース読み込み ──
SRC = r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx"
src_wb = load_workbook(SRC)

SHEETS = ["ZECZEC", "Indiegogo", "FlyingV"]
FILLS  = {
    "ZECZEC":    "FFF2CC",
    "Indiegogo": "FCE4D6",
    "FlyingV":   "E2EFDA",
}

# ── 新ワークブック ──
wb = Workbook()
ws = wb.active
ws.title = "統合リスト"

# ── スタイル定義 ──
header_font   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
normal_font   = Font(name="Arial", size=10)
link_font_z   = Font(name="Arial", size=10, color="0563C1", underline="single")
thin_side     = Side(style="thin", color="AAAAAA")
border        = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

header_fill   = PatternFill(fill_type="solid", fgColor="2F4F4F")  # ダークグリーン

PLATFORM_FILL = {k: PatternFill(fill_type="solid", fgColor=v) for k, v in FILLS.items()}

# ── ヘッダー ──
HEADERS = [
    "カテゴリ",
    "商品名",
    "メーカー名",
    "CFプラットフォーム",
    "CFページURL",
    "メーカーHP",
    "コンタクトメール",
    "スコア",
    "コメント",
]
COL_WIDTHS = [18, 55, 20, 14, 55, 35, 28, 8, 55]

for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.border    = border
    cell.alignment = center_align
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 22

# ── データ書き込み ──
current_row = 2
total = 0

for sheet_name in SHEETS:
    src_ws = src_wb[sheet_name]
    fill   = PLATFORM_FILL[sheet_name]

    for r in range(2, src_ws.max_row + 1):
        # 空行スキップ（商品名が空）
        if not src_ws.cell(r, 2).value:
            continue

        cat      = src_ws.cell(r, 1).value
        name     = src_ws.cell(r, 2).value
        maker    = src_ws.cell(r, 3).value
        cf_url   = src_ws.cell(r, 4).value
        hp       = src_ws.cell(r, 5).value
        mail     = src_ws.cell(r, 6).value
        score    = normalize_score(src_ws.cell(r, 7).value)
        comment  = src_ws.cell(r, 8).value

        row_data = [cat, name, maker, sheet_name, cf_url, hp, mail, score, comment]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font      = normal_font
            cell.fill      = fill
            cell.border    = border
            # 列5(CFページURL)と列6(メーカーHP)はリンク化
            if col_idx in (5, 6) and val and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
                cell.alignment = left_align
            elif col_idx in (1, 4, 8):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        ws.row_dimensions[current_row].height = 30
        current_row += 1
        total += 1

# ── ウィンドウ枠の固定（ヘッダー行を固定）──
ws.freeze_panes = "A2"

# ── オートフィルター ──
ws.auto_filter.ref = f"A1:I{current_row - 1}"

# ── 保存 ──
OUT = r"C:\Users\hiro\HP0524\海外CF_統合リスト.xlsx"
wb.save(OUT)
print(f"✓ 保存完了: {OUT}")
print(f"  合計 {total} 件を統合")
print(f"  ZECZEC / Indiegogo / FlyingV を1シートに集約")
