from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote

wb = load_workbook(r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸705件.xlsx")

link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
alt_fill = PatternFill("solid", start_color="EBF3FB")

def build_mailto(email, maker_name, product_name):
    subject = "Potential Distribution Partnership for Japan"

    body = f"""Dear {maker_name} Team,

My name is Hiroyuki Inoguchi from
Sumai pluS Co., Ltd. in Japan

I hope this message finds you all well.

Our company focuses on promoting products that enrich people's daily lives. We are currently seeking unique international brands to introduce to the Japanese market and support their growth.

I recently had the opportunity to review your product ({product_name}) and was very impressed by its potential in the Japanese market.

I am confident that your product will strongly appeal to Japanese customers, and I would like to explore the possibility of building a partnership with your company to help ensure its success.

Furthermore, our company has a proven track record of promoting overseas products through our proprietary sales network and Japanese distribution channels.

Would it be possible to schedule a brief online meeting sometime next week to discuss this matter?

I look forward to hearing from you.


Sincerely,

Hiroyuki Inoguchi
Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )

================================================

Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )
Hiroyuki Inoguchi

E-mail:yutorin.ino@gmail.com

Address: 49-5 Kitazakuno, Higashi-Itsushiro, Ichinomiya City, Aichi Prefecture, Japan

================================================="""

    encoded_to      = quote(email, safe="")
    encoded_subject = quote(subject, safe="")
    encoded_body    = quote(body, safe="")
    return f"https://mail.google.com/mail/?view=cm&fs=1&to={encoded_to}&su={encoded_subject}&body={encoded_body}"

for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=2):
        row_num = row[0].row
        bg = alt_fill if row_num % 2 == 0 else None

        if len(row) < 8:
            continue

        product_name = str(row[2].value or "").strip()
        maker_name   = str(row[3].value or "").strip()
        email_cell   = row[7]
        email_val    = str(email_cell.value or "").strip()

        if not email_val or not product_name or not maker_name:
            continue

        if "@" in email_val and not email_val.startswith("http") and "問い合わせ" not in email_val:
            mailto_link = build_mailto(email_val, maker_name, product_name)
            email_cell.hyperlink = mailto_link
            email_cell.value = email_val
            email_cell.font = link_font
            email_cell.alignment = Alignment(vertical="center", wrap_text=True)
            email_cell.border = border
            if bg:
                email_cell.fill = bg

output_path = r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸705件.xlsx"
wb.save(output_path)
print("完了")
print(f"保存先: {output_path}")
