from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 第5弾 新規追加製品リスト（11件）
# ZECZEC 7件 / FlyingV 1件 / Indiegogo 3件
# 日本CF（Makuake / Campfire / Greenfunding）未掲載確認済み
# ============================================================
#
# 列構成: カテゴリ(1) | 商品名(2) | メーカー名(3) | CFページURL(4)
#         メーカーHP(5) | コンタクトメール(6) | スコア(7) | コメント(8)

NEW_PRODUCTS = {
    "ZECZEC": [
        (
            "映像・プロジェクター",
            "Yaber T1 Pro 小影霸｜500lm AI智慧對焦 × 180°雲台 × 120吋掌上便攜智慧投影機",
            "Yaber",
            "https://www.zeczec.com/projects/yaber-smart-projector",
            "https://www.yaber.com", "要調査", 4,
            "500lm AI自動フォーカス 180°首振り 120インチ 欧州EISA賞受賞コンパクト投影機"
        ),
        (
            "映像・プロジェクター",
            "Tairona R-01｜真4K掌上劇院 277g超軽量 DLP × 内蔵AndroidTV × モバイルバッテリー駆動",
            "Tairona",
            "https://www.zeczec.com/projects/tairona-r-01-l-4k",
            "要調査", "要調査", 4,
            "DLP真4K 277g超軽量 Android TV内蔵 モバイルバッテリー給電対応"
        ),
        (
            "スマートリング・ウォッチ",
            "IS愛思 SW-30 PLUS｜智慧通話運動門禁手錶 全台首發CUID門禁卡機能",
            "IS愛思",
            "https://www.zeczec.com/projects/is-sw-30-plus",
            "要調査", "要調査", 4,
            "全台初 CUID非接触ドアアクセス 2.02型OLED 血氧・心拍・睡眠・通話"
        ),
        (
            "スマートリング・ウォッチ",
            "小米有品 KFI AI智慧手錶GT6｜ChatGPT搭載 × 100+運動 × AMOLED × Bluetooth通話",
            "KFI (小米有品)",
            "https://www.zeczec.com/projects/KFI",
            "要調査", "要調査", 4,
            "ChatGPT内蔵 100+スポーツ AMOLED 7日間バッテリー Bluetooth通話対応"
        ),
        (
            "ヘルスケア・フィットネス",
            "筋Good棒 智慧觸控筋膜槍｜螢幕智能提示 × 五種律動情境 居家放鬆教練",
            "筋Good棒",
            "https://www.zeczec.com/projects/good",
            "要調査", "要調査", 4,
            "タッチスクリーン搭載スマートマッサージガン 5リズム智能ガイド 台湾製造"
        ),
        (
            "ヘルスケア・フィットネス",
            "QImoji 刮罐師呼吸式拔罐按摩儀｜一吸一放 釋放身體痠緊繃 台灣製造SGS認證",
            "QImoji / AURAI",
            "https://www.zeczec.com/projects/aurai-qimoji",
            "要調査", "要調査", 3,
            "伝統カッピング×現代技術 呼吸式+マイクロ振動 台湾製SGS認証取得"
        ),
        (
            "ヘルスケア・フィットネス",
            "RheoFit 全自動按摩滾輪｜100%解放雙手 AI輔助筋膜放鬆 改變世界的運動放鬆方式",
            "RheoFit",
            "https://www.zeczec.com/projects/rheofit",
            "https://rheofit.com", "要調査", 4,
            "100%ハンズフリー自動マッサージローラー AI筋肉回復 筋膜リリース革命"
        ),
    ],

    "FlyingV": [
        (
            "スマートホーム・IoT・ロボット",
            "Acer Gadget 碁智鎖 Pro｜旗艦級智慧門鎖 指紋×IC卡×密碼×APP 多重驗證",
            "Acer Gadget",
            "https://www.flyingv.cc/projects/36332",
            "https://www.acer.com",
            "要調査", 4,
            "Acer製フラグシップスマートロック 指紋/ICカード/暗証番号/アプリ多重認証"
        ),
    ],

    "Indiegogo": [
        (
            "PC周辺・入力機器",
            "EDGE 4K OLED Ultralight Portable Magnetic Monitor｜mmWave Zero-Latency Wireless",
            "InnLead",
            "https://www.indiegogo.com/projects/edge-4k-oled-ultralight-portable-magnetic-monitor",
            "https://innlead.shop",
            "要調査", 4,
            "15.6型4K OLEDモバイルモニター マグネット着脱 mmWave100フィートワイヤレス"
        ),
        (
            "スマートリング・ウォッチ",
            "DTECTIO Wearable ECG Monitor｜AI-Empowered 24/7 Heart Health Smartwatch",
            "DTECTIO",
            "https://www.indiegogo.com/projects/dtectio-wearable-ecg-monitor-ai-empowered",
            "要調査", "要調査", 4,
            "AI搭載24時間連続心電図スマートウォッチ 処方箋不要の心臓モニタリング"
        ),
        (
            "PC周辺・入力機器",
            "NANO: Self Powered 4K Portable Monitor｜4K/2.5K 165Hz Touch by Armorylabs",
            "Armorylabs",
            "https://www.indiegogo.com/en/projects/armorylabs/nano-self-powered-4k-2-5k-165hz-portable-monitor",
            "要調査", "要調査", 4,
            "自己電源供給 4K/2.5K 165Hzタッチポータブルモニター ゲーミング対応"
        ),
    ],
}

# ── スタイル定義 ──
normal_font  = Font(name="Arial", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

sheet_fills = {
    "ZECZEC":    PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "Indiegogo": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "FlyingV":   PatternFill(fill_type="solid", fgColor="E2EFDA"),
}

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total_added = 0

print("=== 第5弾 新規製品追加 ===")
for sheet_name, products in NEW_PRODUCTS.items():
    ws = wb[sheet_name]
    fill = sheet_fills[sheet_name]
    start_row = ws.max_row + 1

    print(f"\n[{sheet_name}] {len(products)}件追加 (行{start_row}〜)")
    for i, (cat, name, maker, cf_url, hp, mail, score, comment) in enumerate(products):
        r = start_row + i
        values = [cat, name, maker, cf_url, hp, mail, score, comment]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = normal_font
            c.fill      = fill
            c.border    = border
            c.alignment = center_align if col == 1 else left_align
        ws.row_dimensions[r].height = 30
        total_added += 1
        print(f"  + {name[:45]}")

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"\n✓ 保存完了: 合計 {total_added}件追加")

# ── 追加後の件数確認 ──
print("\n=== 各シート件数（追加後） ===")
wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
grand_total = 0
for sn in wb2.sheetnames:
    if sn == "削除済みリスト":
        continue
    ws2 = wb2[sn]
    cnt = sum(1 for r in range(2, ws2.max_row + 1) if ws2.cell(r, 2).value)
    print(f"  [{sn}] {cnt}件")
    grand_total += cnt
print(f"  合計: {grand_total}件")
