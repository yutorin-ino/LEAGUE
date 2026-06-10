from openpyxl import load_workbook
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 第2弾 日本CFサイト確認済み製品リスト
# 調査で確認されたMakuake / Campfire / Greenfunding掲載製品
# ============================================================

JAPAN_CF_ROWS2 = {
    "ZECZEC": {
        4:  ("Allite 氮化鎵快充 史上最小65W雙孔",                 "Greenfunding"),
        9:  ("LilNob 3C1A 四孔充電器 最高100W",                   "Makuake / Greenfunding"),
        10: ("TITANSHIELD 固態行動電源 防爆五合一",                "Makuake"),
        19: ("TRACPOINT 実体+仮想 二合一マウス×プレゼンペン",      "Campfire"),
        21: ("AIFA i-Ctrl Pro Plus 家電遠端遙控×智慧防霉",         "Makuake"),
        23: ("Slimca 超薄録音カード 随録随傳",                     "Makuake"),
        24: ("TileRec 隠形録音片 全球最小智慧録音設備",             "Makuake / Campfire / Greenfunding"),
        33: ("FLAT 壁掛畫空淨機 外銷突破20國",                    "Makuake (Sauberair)"),
    },
    "Indiegogo": {
        2: ("RayNeo X2 World's 1st AI-Powered True AR Glasses",   "Greenfunding"),
        3: ("Looktech AI: The Smart Glasses That Truly Know You",  "Makuake"),
    },
    "FlyingV": {
        9: ("智米 空氣清淨機 400m³/h CADR VOC監測",               "Greenfunding"),
    },
}

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")

# ─── プレビュー表示 ───
total_delete = 0
for sn in wb.sheetnames:
    rows = JAPAN_CF_ROWS2.get(sn, {})
    if rows:
        print(f"\n=== [{sn}] 削除予定 {len(rows)}件 ===")
        for r in sorted(rows):
            name, site = rows[r]
            print(f"  行{r:2d}: {name}  ← {site}")
        total_delete += len(rows)

print(f"\n合計削除予定: {total_delete}件")
print("─" * 50)

# ─── 実際の削除（行番号が大きい順に削除してズレを防ぐ） ───
deleted = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = JAPAN_CF_ROWS2.get(sn, {})
    for r in sorted(rows.keys(), reverse=True):
        ws.delete_rows(r)
        deleted += 1

# ─── 末尾の空行クリーンアップ ───
for sn in wb.sheetnames:
    if sn == "削除済みリスト":
        continue
    ws = wb[sn]
    while ws.max_row > 1 and ws.cell(ws.max_row, 2).value is None:
        ws.delete_rows(ws.max_row)

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"\n削除完了: {deleted}件")

# ─── 削除後の件数確認 ───
wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total = 0
for sn in wb2.sheetnames:
    if sn == "削除済みリスト":
        continue
    ws = wb2[sn]
    cnt = ws.max_row - 1
    print(f"  [{sn}] 残り {cnt}件")
    total += cnt
print(f"  合計残り: {total}件")
