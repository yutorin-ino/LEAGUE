from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# 追加調査済みメーカーHP辞書
MAKER_DATA = {
    # 充電・モバイル（新規）
    "GrantClassic":      ("https://grantclassic.com", ""),
    "TITANSHIELD":       ("https://grantclassic.com", ""),
    "DEZCTOP":           ("https://dezctop.com", "support@dezctop.com"),
    "MicroNovelty":      ("https://www.micronovelty.com", ""),
    "Tmobi":             ("https://www.magink.app", ""),
    "MagInk":            ("https://www.magink.app", ""),
    # ウェアラブル（新規）
    "Aipower":           ("https://www.myaipower.com", ""),
    "Wearbuds":          ("https://www.myaipower.com", ""),
    "TikTik":            ("https://tiktikglobal.com", ""),
    "Aether X":          ("https://tiktikglobal.com", ""),
    "LARMI":             ("https://larmi.tw", ""),
    # スマートフォン・PC
    "Minimal Company":   ("https://minimalcompany.com", ""),
    "Minimal Phone":     ("https://minimalcompany.com", ""),
    "Vision Keyboard":   ("https://visionkeyboard.com", ""),
    "Absolute Cycling":  ("https://absolutecycling.com", ""),
    # ロボット・IoT
    "LOOI":              ("https://looirobot.com", ""),
    "Tangible Future":   ("https://looirobot.com", ""),
    "Blotch":            ("https://blotch.app", ""),
    # セキュリティ
    "小島快鎖":           ("https://www.3elock.com", ""),
    "3E":                ("https://www.3elock.com", ""),
    # オーディオ（新規）
    "XROUND":            ("https://xroundaudio.com", ""),
    "AERO":              ("https://xroundaudio.com", ""),
    # 空気清浄（新規）
    "MyComfyTech":       ("https://www.mycomfytech.com", ""),
    "SOLEUS AIR":        ("https://soleusair.com", ""),
    "SoleusAir":         ("https://soleusair.com", ""),
    # スマートホーム・家具（新規）
    "THINKTHING":        ("https://www.thinkthingstudio.com", ""),
    "eterno":            ("https://eternopet.com", ""),
    # 睡眠・ヘルス
    "OuOu":              ("https://ouou.love", ""),
    "Osleepy":           ("https://ouou.love", ""),
}

link_font   = Font(name="Arial", size=10, color="0563C1", underline="single")
normal_font = Font(name="Arial", size=10)

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
updated = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        product = str(ws.cell(row=row, column=2).value or "")
        maker   = str(ws.cell(row=row, column=3).value or "")
        hp_cell   = ws.cell(row=row, column=5)
        mail_cell = ws.cell(row=row, column=6)

        if str(hp_cell.value or "").startswith("http"):
            continue  # 既に入力済み

        combined = product + " " + maker
        for kw, (hp, mail) in MAKER_DATA.items():
            if kw.lower() in combined.lower():
                hp_cell.value = hp
                hp_cell.hyperlink = hp
                hp_cell.font = link_font
                hp_cell.alignment = Alignment(vertical="top", wrap_text=True)
                if mail and str(mail_cell.value or "") in ("要調査", ""):
                    mail_cell.value = mail
                    mail_cell.font = normal_font
                updated += 1
                break

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"更新件数: {updated}件")

wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total = remain = 0
for sn in wb2.sheetnames:
    ws = wb2[sn]
    s_total = ws.max_row - 1
    s_remain = sum(1 for r in range(2, ws.max_row + 1)
                   if not str(ws.cell(row=r, column=5).value or "").startswith("http"))
    print(f"  [{sn}] 全{s_total}件 / HP未入力: {s_remain}件")
    total += s_total; remain += s_remain
print(f"  合計: {total}件 / HP未入力: {remain}件 / 入力済: {total-remain}件")
