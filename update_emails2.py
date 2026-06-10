from openpyxl import load_workbook
from openpyxl.styles import Font
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 追加調査済みメールアドレス辞書（第2弾）
EMAIL_DATA = {
    # PC周辺
    "JCH":               "service.tw@j5create.com",
    "j5create":          "service.tw@j5create.com",
    "Vision Keyboard":   "service@visionkeyboard.com",
    # AR/VR
    "MAD Gaze":          "info@madgaze.com",
    "Dream Glass":       "contact@dreamworldvision.com",
    "DPVR":              "support@dpvr.com",
    # 空気清浄
    "POIEMA":            "custcare@poiema.com.tw",
    "Smartmi":           "support@smartmiglobal.com",
    "智米":               "support@smartmiglobal.com",
    "SOLEUS AIR":        "contact@soleusairwest.com",
    "SoleusAir":         "contact@soleusairwest.com",
    # E-ink・ディスプレイ
    "Bloomin8":          "hello@bloomin8.com",
    "DASUNG":            "contact@dasung.com",
    # ウェアラブル
    "VIITA":             "hello@viita-watches.com",
    "Sequent":           "care@sequent.ch",
}

normal_font = Font(name="Arial", size=10)

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
updated = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        hp_cell   = ws.cell(row=row, column=5)
        mail_cell = ws.cell(row=row, column=6)

        if not str(hp_cell.value or "").startswith("http"):
            continue
        if str(mail_cell.value or "") not in ("要調査", "", "None"):
            continue

        product = str(ws.cell(row=row, column=2).value or "")
        maker   = str(ws.cell(row=row, column=3).value or "")
        combined = product + " " + maker

        for kw, mail in EMAIL_DATA.items():
            if kw.lower() in combined.lower():
                mail_cell.value = mail
                mail_cell.font = normal_font
                updated += 1
                break

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"メール更新件数: {updated}件")

wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total_hp = total_mail = 0
for sn in wb2.sheetnames:
    ws = wb2[sn]
    s_hp   = sum(1 for r in range(2, ws.max_row + 1)
                 if str(ws.cell(row=r, column=5).value or "").startswith("http"))
    s_mail = sum(1 for r in range(2, ws.max_row + 1)
                 if str(ws.cell(row=r, column=6).value or "") not in ("要調査", "", "None")
                 and str(ws.cell(row=r, column=5).value or "").startswith("http"))
    print(f"  [{sn}] HP入力済: {s_hp}件 / メール入力済: {s_mail}件 / メール未入力: {s_hp-s_mail}件")
    total_hp += s_hp; total_mail += s_mail
print(f"  合計: HP入力済{total_hp}件 / メール入力済{total_mail}件({round(total_mail/total_hp*100)}%) / 未入力{total_hp-total_mail}件")

# 残り未入力リスト
print("\n--- メール未入力リスト（HP有り）---")
wb3 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
for sn in wb3.sheetnames:
    ws = wb3[sn]
    for r in range(2, ws.max_row + 1):
        hp = str(ws.cell(row=r, column=5).value or '')
        mail = str(ws.cell(row=r, column=6).value or '')
        if hp.startswith('http') and mail in ('要調査', '', 'None'):
            maker = ws.cell(row=r, column=3).value
            print(f"  [{sn}] {maker} | {hp}")
