from openpyxl import load_workbook
from openpyxl.styles import Font
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 全シートの列4（CFページURL）と列5（メーカーHP）を
# クリック可能なハイパーリンクに変換するスクリプト
# ============================================================

HYPERLINK_FONT = Font(
    name="Arial", size=10,
    color="0563C1",   # 標準的なリンク青色
    underline="single"
)

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")

total_links = 0

for sn in wb.sheetnames:
    ws = wb[sn]
    sheet_links = 0
    for row in range(2, ws.max_row + 1):
        for col in [4, 5]:   # 列4=CFページURL / 列5=メーカーHP
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = HYPERLINK_FONT
                sheet_links += 1
                total_links += 1
    if sheet_links > 0:
        print(f"  [{sn}] {sheet_links}件のリンクを設定")

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"\n✓ 保存完了: 合計 {total_links}件のハイパーリンクを設定しました")
