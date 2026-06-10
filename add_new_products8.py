from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 第8弾 新規追加製品リスト（8件）
# ZECZEC 5件 / Indiegogo 3件
# 日本CF（Makuake / Campfire / Greenfunding）未掲載確認済み
# ============================================================
# 精査除外済み:
#   Wanbo Mozart 1 Pro       → Amazon Japan 一般販売済み
#   Smini 空気清浄機          → Makuake + CAMPFIRE掲載済み
#   COFO Chair Pro 2         → Makuake掲載済み（2億円超達成）
#   Tineco S9 Artist Pro     → Makuake掲載済み
#   Rux Robot (Letianpai)    → CAMPFIRE + Greenfunding掲載済み
#   TITANSHIELD 固態行動電源  → Makuake掲載済み
#   PLAUD NotePin            → Amazon Japan / 日本公式サイトで展開済み
#   Halliday AI Glasses      → Makuake掲載済み
#   TriCreate 折疳三螢幕      → CAMPFIRE + Greenfunding掲載済み（R7除外済）
#   ASUS AirVision M1        → Greenfunding掲載済み
#
# 列構成: カテゴリ(1) | 商品名(2) | メーカー名(3) | CFページURL(4)
#         メーカーHP(5) | コンタクトメール(6) | スコア(7) | コメント(8)

NEW_PRODUCTS = {
    "ZECZEC": [
        (
            "スマートホーム・IoT・ロボット",
            "Meetech《AI雙球循環扇》｜等離子淨化+20m超遠送風 720°全域循環 語音觸控 親子必備",
            "Meetech (遇科技)",
            "https://www.zeczec.com/projects/meetech",
            "要調査", "要調査", 4,
            "全台首創720°二球同時循環 等離子イオン浄化 20m超遠距離送風 AI音声制御 省エネ72%減"
        ),
        (
            "オーディオ・録音",
            "Wonder 快點翻錄機｜業界創新「翻譯 × 錄音」二合一 離線可用 隱私零外洩 出貨中",
            "Wonder",
            "https://www.zeczec.com/projects/wonder-pro",
            "要調査", "要調査", 5,
            "翻訳×録音 業界初二合一 オフライン対応 全データ端末内保存 登録不要 2026年5月出荷 7百万NTD超達成"
        ),
        (
            "スマートホーム・IoT・ロボット",
            "惟家智能星幻開關｜模組化設計 Zigbee × Matter 重新定義你的空間！",
            "JustPlus (惟家)",
            "https://www.zeczec.com/projects/justplus",
            "https://shop.justplus.com.tw", "要調査", 3,
            "モジュール交換可能なスマートスイッチ Zigbee/Matter両対応 中性線不要版あり Apple HomeKit/Google Home連携"
        ),
        (
            "ヘルスケア・フィットネス",
            "3ZeBra｜溫揉足 3D溫感腳底按摩機【擬人指壓×完整包覆】溫熱×氣壓×滾球三効合一",
            "3ZeBra",
            "https://www.zeczec.com/projects/foot-massage-machine",
            "https://www.3zebra.com.tw", "要調査", 3,
            "3D温感足底マッサージ機 温熱(32-40℃)×気圧×ローリングボール3効合一 タイマー付リモコン 洗えるインナー"
        ),
        (
            "PC周辺・入力機器",
            "【鍵境】AZOMA AZ330/AZ660 無線機械式鍵盤｜三模連線×自訂全彩LCD螢幕×ホットスワップ",
            "靈動數碼 (lingdong)",
            "https://www.zeczec.com/projects/AZ330-AZ660",
            "https://www.lingdong.com.tw", "要調査", 4,
            "カスタマイズ可能な全彩LCD搭載 ホットスワップ 三モード(有線/2.4G/BT) 200時間バッテリー RGB"
        ),
    ],

    "Indiegogo": [
        (
            "ガジェット・充電",
            "Panak: Solar-Powered MagSafe Wallet & Power Bank｜10000mAh × Solar × 45W × 20W WL",
            "Panak",
            "https://www.indiegogo.com/projects/panak-solar-powered-magsafe-wallet-power-bank",
            "要調査", "要調査", 4,
            "MagSafe対応 ソーラー充電内蔵 10000mAh 45W Type-C急速充電 20Wワイヤレス 5枚カード収納 Indiegogo3660%達成"
        ),
        (
            "スマートホーム・IoT・ロボット",
            "Productivity Redefined: AI-Powered Desk Lamp by Blinder｜ChatGPT×Camera×355° Adj",
            "Blinder Team",
            "https://www.indiegogo.com/projects/productivity-redefined-ai-powered-desk-lamp",
            "要調査", "要調査", 3,
            "ChatGPT搭載AIデスクライト 内蔵カメラ 355°調整 太陽光スペクトル×RGB6モード 生産性向上 1193%達成"
        ),
        (
            "PC周辺・入力機器",
            "XNote: AI-Powered Smart Writing Set｜Smart Pen × Notebook × Handwriting→Digital × AI Chat",
            "Kodion Innovation",
            "https://www.indiegogo.com/projects/xnote-ai-powered-smart-writing-set",
            "https://xnote.ai", "要調査", 4,
            "AIスマートペン+専用ノート 手書きをリアルタイムデジタル化 AI検索/チャット/タスク自動検出 エンドツーエンド暗号化"
        ),
    ],
}

# ── スタイル定義 ──
normal_font  = Font(name="Arial", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

sheet_fills = {
    "ZECZEC":    PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "Indiegogo": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "FlyingV":   PatternFill(fill_type="solid", fgColor="E2EFDA"),
}

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total_added = 0

print("=== 第8弾 新規製品追加 ===")
for sheet_name, products in NEW_PRODUCTS.items():
    ws = wb[sheet_name]
    fill = sheet_fills[sheet_name]
    start_row = ws.max_row + 1

    print(f"\n[{sheet_name}] {len(products)}件追加 (行{start_row}〜)")
    for i, (cat, name, maker, cf_url, hp, mail, score, comment) in enumerate(products):
        r = start_row + i
        values = [cat, name, maker, cf_url, hp, mail, score, comment]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = normal_font
            c.fill      = fill
            c.border    = border
            c.alignment = center_align if col == 1 else left_align
        ws.row_dimensions[r].height = 30
        total_added += 1
        print(f"  + {name[:50]}")

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"\n✓ 保存完了: 合計 {total_added}件追加")

# ── 追加後の件数確認 ──
print("\n=== 各シート件数（追加後） ===")
wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
grand_total = 0
for sn in wb2.sheetnames:
    if sn == "削除済みリスト":
        continue
    ws2 = wb2[sn]
    cnt = sum(1 for r in range(2, ws2.max_row + 1) if ws2.cell(r, 2).value)
    print(f"  [{sn}] {cnt}件")
    grand_total += cnt
print(f"  合計: {grand_total}件")
