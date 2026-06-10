from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 第7弾 新規追加製品リスト（10件）
# ZECZEC 6件 / FlyingV 1件 / Indiegogo 3件
# 日本CF（Makuake / Campfire / Greenfunding）未掲載確認済み
# ============================================================
# 精査除外済み:
#   HOVERAir X1 Pro/Smart → Makuake + CAMPFIRE掲載済み
#   ASUS AirVision M1     → Greenfunding掲載済み
#   TriCreate 折疳三螢幕   → CAMPFIRE + Greenfunding掲載済み
#   LuluPet AI貓砂盆       → Makuake掲載済み
#   Ropet 洛比 (lopeto)    → Makuake掲載済み
#   Zikko 七合一充電座      → CAMPFIRE + Makuake掲載済み（2022年）
#   Kolude Keyhub / K2    → Makuake + CAMPFIRE掲載済み
#   GOOVIS 各モデル        → CAMPFIRE + Makuake掲載済み
#   Mokacam Alpha3 Flip   → Makuake掲載済み
#   NarTick スマートカレンダー→ Greenfunding掲載済み
#   Wearbuds Watch        → Makuake + Greenfunding掲載済み
#   Looktech AI眼鏡        → Makuake掲載済み
#   DONGYING A3投影機      → CAMPFIRE掲載済み
#   INMO GO3              → Kickstarter（Indiegogo対象外）
#
# 列構成: カテゴリ(1) | 商品名(2) | メーカー名(3) | CFページURL(4)
#         メーカーHP(5) | コンタクトメール(6) | スコア(7) | コメント(8)

NEW_PRODUCTS = {
    "ZECZEC": [
        (
            "スマートリング・ウォッチ",
            "Wonder Ring｜全台首發 AI 智慧戒指 微感科技 × 智慧觸控 × 獨家繁體APP 健康追蹤",
            "Wonder Ring",
            "https://www.zeczec.com/projects/wonder-ring",
            "要調査", "要調査", 4,
            "全台首発 AIタッチコントロール+健康追跡スマートリング 繁体字専用APP 微感科技採用 運動/睡眠/心拍管理"
        ),
        (
            "スマートグラス・AR",
            "SmartEcho AI智慧眼鏡｜全台首款AR同步翻譯眼鏡 1.6mm超薄波導鏡片 32語言 44g",
            "SmartEcho",
            "https://www.zeczec.com/projects/smartecho-glasses",
            "https://www.smartecho.com.tw", "要調査", 4,
            "全台初AR同期翻訳メガネ 32言語対応 1.6mm超薄型ウェーブガイドレンズ 44g ARプロンプター/翻訳/文字起こし"
        ),
        (
            "オーディオ・録音",
            "OpenRock《Link20》AI「翻譯×錄音×降噪」磁吸麥克風 開放式無線耳機｜週週出貨",
            "OpenRock (1MORE)",
            "https://www.zeczec.com/projects/link20-ai",
            "要調査", "要調査", 4,
            "AI翻訳×録音×ノイキャン統合型 磁吸着マイク付きオープンイヤーイヤホン リアルタイム文字起こし 全方向収録"
        ),
        (
            "スマートホーム・IoT・ロボット",
            "iKKO Mind ONE｜專為 AI 而生的手機！即時翻譯、會議錄音整理，商務出差更輕鬆",
            "iKKO",
            "https://www.zeczec.com/projects/ikko-mind-one-ai",
            "要調査", "要調査", 3,
            "AI専用スマートフォン リアルタイム翻訳・会議録音・議事録自動生成 ビジネス出張向けポケットAIアシスタント"
        ),
        (
            "ヘルスケア・フィットネス",
            "PETANGEL MOVE｜寵物運動記錄器 歩數×嗅聞時間×主動運動 首圈裝載 全家毛孩適用",
            "PETANGEL",
            "https://www.zeczec.com/projects/petangel-move",
            "https://www.petangel.com.tw", "要調査", 3,
            "愛犬の運動データ自動記録 歩数・嗅ぎ時間・運動強度を可視化 首輪装着型 台湾発有機ペットケアブランドPETANGEL製"
        ),
        (
            "ヘルスケア・フィットネス",
            "護心AI攜帶型心率儀 iCloud小藍｜全球最簡單操作的心率健康管理裝置",
            "iCloud AI",
            "https://www.zeczec.com/projects/icloud-ai-20",
            "要調査", "要調査", 3,
            "世界最簡単操作のAI携帯型心拍管理デバイス クリップ装着 スマホアプリ連携リアルタイム心拍健康管理"
        ),
    ],

    "FlyingV": [
        (
            "オーディオ・録音",
            "PHILIPS 小飛筆 Pro｜最聰明的AI人生副手 即時轉寫×同步翻譯×通話錄音 終生免費",
            "PHILIPS",
            "https://www.flyingv.cc/projects/36816",
            "https://www.philips.com", "要調査", 4,
            "8言語認識×24言語翻訳 4マイク配列 LINE/通話録音対応 終生無料文字起こし 16GB内蔵64時間録音 サブスク不要"
        ),
    ],

    "Indiegogo": [
        (
            "E-Ink・ディスプレイ",
            "DASUNG Link: The World First E-ink Phone Monitor｜6.7\" Touch Eye-Friendly",
            "DASUNG",
            "https://www.indiegogo.com/projects/dasung-link-the-world-first-e-ink-phone-monitor",
            "要調査", "要調査", 4,
            "世界初E-Inkスマホ外付けモニター 6.7型タッチ対応 300PPI 目に優しいフロントライト付 直射日光下でも視認可能"
        ),
        (
            "スマートリング・ウォッチ",
            "NovaSmart: ChatGPT-4o powered Stylish AI Smart Ring｜20+ Metrics No Subscription",
            "Nova Ring",
            "https://www.indiegogo.com/projects/novasmart-chatgpt-4o-powered-stylish-ai-smart-ring",
            "https://www.novaring.com", "要調査", 3,
            "ChatGPT-4o統合 20種類以上健康指標モニタリング サブスク不要 5ATM防水 4-7日バッテリー チタン/サファイア素材"
        ),
        (
            "PC周辺・入力機器",
            "Keykrush ORCA: Split Ergonomic Keyboard｜Wireless × Hot-Swap × RGB × Programmable",
            "Victor Sui",
            "https://www.indiegogo.com/en/projects/victorsui/keykrush-orca-split-ergonomic-keyboard",
            "要調査", "要調査", 3,
            "セパレート型エルゴノミクスキーボード ワイヤレス+ホットスワップ対応 RGB 40%レイアウト フル カスタマイズ可能"
        ),
    ],
}

# ── スタイル定義 ──
normal_font  = Font(name="Arial", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

sheet_fills = {
    "ZECZEC":    PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "Indiegogo": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "FlyingV":   PatternFill(fill_type="solid", fgColor="E2EFDA"),
}

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total_added = 0

print("=== 第7弾 新規製品追加 ===")
for sheet_name, products in NEW_PRODUCTS.items():
    ws = wb[sheet_name]
    fill = sheet_fills[sheet_name]
    start_row = ws.max_row + 1

    print(f"\n[{sheet_name}] {len(products)}件追加 (行{start_row}〜)")
    for i, (cat, name, maker, cf_url, hp, mail, score, comment) in enumerate(products):
        r = start_row + i
        values = [cat, name, maker, cf_url, hp, mail, score, comment]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = normal_font
            c.fill      = fill
            c.border    = border
            c.alignment = center_align if col == 1 else left_align
        ws.row_dimensions[r].height = 30
        total_added += 1
        print(f"  + {name[:50]}")

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"\n✓ 保存完了: 合計 {total_added}件追加")

# ── 追加後の件数確認 ──
print("\n=== 各シート件数（追加後） ===")
wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
grand_total = 0
for sn in wb2.sheetnames:
    if sn == "削除済みリスト":
        continue
    ws2 = wb2[sn]
    cnt = sum(1 for r in range(2, ws2.max_row + 1) if ws2.cell(r, 2).value)
    print(f"  [{sn}] {cnt}件")
    grand_total += cnt
print(f"  合計: {grand_total}件")
