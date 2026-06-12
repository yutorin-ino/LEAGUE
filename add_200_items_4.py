from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote

def build_gmail(email, maker_name, product_name):
    subject = "Potential Distribution Partnership for Japan"
    body = f"""Dear {maker_name} Team,

My name is Hiroyuki Inoguchi from
Sumai pluS Co., Ltd. in Japan

I hope this message finds you all well.

Our company focuses on promoting products that enrich people's daily lives. We are currently seeking unique international brands to introduce to the Japanese market and support their growth.

I recently had the opportunity to review your product ({product_name}) and was very impressed by its potential in the Japanese market.

I am confident that your product will strongly appeal to Japanese customers, and I would like to explore the possibility of building a partnership with your company to help ensure its success.

Furthermore, our company has a proven track record of promoting overseas products through our proprietary sales network and Japanese distribution channels.

Would it be possible to schedule a brief online meeting sometime next week to discuss this matter?

I look forward to hearing from you.


Sincerely,

Hiroyuki Inoguchi
Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )

================================================

Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )
Hiroyuki Inoguchi

E-mail:yutorin.ino@gmail.com

Address: 49-5 Kitazakuno, Higashi-Itsushiro, Ichinomiya City, Aichi Prefecture, Japan

================================================="""
    sender = quote("yutorin.ino@gmail.com", safe='')
    return f"https://mail.google.com/mail/?authuser={sender}&view=cm&fs=1&to={quote(email,safe='')}&su={quote(subject,safe='')}&body={quote(body,safe='')}"

new_products = [
    # ── キッチン・調理器具 ──────────────────────────────
    {"no":1306,"category":"キッチン・調理器具","name":"Zwilling Enfinigy Cool Touch Toaster 4-Slot","maker":"Zwilling","ec_site":"Zwilling公式サイト","prod_url":"https://www.zwilling.com/us/kitchen-electrics/toasters/","maker_hp":"https://www.zwilling.com","contact":"cs@zwilling.com"},
    {"no":1307,"category":"キッチン・調理器具","name":"Nutribullet Blender Combo 1200W 12-Piece","maker":"NutriBullet","ec_site":"NutriBullet公式サイト","prod_url":"https://www.nutribullet.com/collections/blenders","maker_hp":"https://www.nutribullet.com","contact":"support@nutribullet.com"},
    {"no":1308,"category":"キッチン・調理器具","name":"Breville Barista Express Impress Espresso Machine","maker":"Breville","ec_site":"Breville公式サイト","prod_url":"https://www.breville.com/us/en/products/espresso/bes876","maker_hp":"https://www.breville.com","contact":"support@breville.com"},
    {"no":1309,"category":"キッチン・調理器具","name":"Sage Oracle Touch Fully Automatic Espresso","maker":"Sage Appliances","ec_site":"Sage公式サイト","prod_url":"https://www.sageappliances.com/uk/en/products/espresso/bes990","maker_hp":"https://www.sageappliances.com","contact":"support@sageappliances.com"},
    {"no":1310,"category":"キッチン・調理器具","name":"Anova Precision Oven Full-Size Combi Steam","maker":"Anova Culinary","ec_site":"Anova公式サイト","prod_url":"https://anovaculinary.com/products/anova-precision-oven","maker_hp":"https://anovaculinary.com","contact":"support@anovaculinary.com"},
    {"no":1311,"category":"キッチン・調理器具","name":"Yedi Houseware Infinity 10-in-1 Instant Pot","maker":"Yedi Houseware","ec_site":"Yedi公式サイト","prod_url":"https://yedihouseware.com/collections/instant-pots","maker_hp":"https://yedihouseware.com","contact":"support@yedihouseware.com"},
    {"no":1312,"category":"キッチン・調理器具","name":"Made In Blue Carbon Steel Wok 12-Inch","maker":"Made In Cookware","ec_site":"Made In公式サイト","prod_url":"https://madeincookware.com/collections/woks","maker_hp":"https://madeincookware.com","contact":"hello@madeincookware.com"},
    {"no":1313,"category":"キッチン・調理器具","name":"Earlywood Handmade Hardwood Cooking Utensil Set","maker":"Earlywood","ec_site":"Earlywood公式サイト","prod_url":"https://www.earlywooddesigns.com/collections/sets","maker_hp":"https://www.earlywooddesigns.com","contact":"info@earlywooddesigns.com"},
    {"no":1314,"category":"キッチン・調理器具","name":"Fable Dinnerware Modern Stoneware Starter Set","maker":"Fable","ec_site":"Fable公式サイト","prod_url":"https://www.fablehome.co/collections/dinnerware","maker_hp":"https://www.fablehome.co","contact":"hello@fablehome.co"},
    {"no":1315,"category":"キッチン・調理器具","name":"Hydros Bottle 24oz Filtered Water Bottle","maker":"Hydros","ec_site":"Hydros公式サイト","prod_url":"https://www.hydros.com/collections/filtered-bottles","maker_hp":"https://www.hydros.com","contact":"hello@hydros.com"},
    {"no":1316,"category":"キッチン・調理器具","name":"Brita Hub Instant Powerful Water Filtration","maker":"Brita","ec_site":"Brita公式サイト","prod_url":"https://www.brita.com/products/hub","maker_hp":"https://www.brita.com","contact":"consumer@brita.com"},
    {"no":1317,"category":"キッチン・調理器具","name":"Clever Coffee Dripper Large Immersion Brewer","maker":"Abid","ec_site":"Clever公式サイト","prod_url":"https://cleverdrip.com/products/clever-dripper-large","maker_hp":"https://cleverdrip.com","contact":"info@cleverdrip.com"},
    {"no":1318,"category":"キッチン・調理器具","name":"Berkey Travel Water Purifier System 1.5 Gallon","maker":"Berkey Filters","ec_site":"Berkey公式サイト","prod_url":"https://www.berkeyfilters.com/products/travel-berkey","maker_hp":"https://www.berkeyfilters.com","contact":"support@berkeyfilters.com"},
    {"no":1319,"category":"キッチン・調理器具","name":"Zojirushi Induction Heating Rice Cooker 5.5-Cup","maker":"Zojirushi","ec_site":"Zojirushi公式サイト","prod_url":"https://www.zojirushi.com/app/category/rice-cookers","maker_hp":"https://www.zojirushi.com","contact":"info@zojirushi.com"},
    {"no":1320,"category":"キッチン・調理器具","name":"Crock-Pot 7-Quart Manual Slow Cooker","maker":"Crock-Pot","ec_site":"Crock-Pot公式サイト","prod_url":"https://www.crockpot.com/collections/slow-cookers","maker_hp":"https://www.crockpot.com","contact":"consumer@crockpot.com"},
    {"no":1321,"category":"キッチン・調理器具","name":"Keurig K-Supreme Plus SMART Coffee Maker","maker":"Keurig","ec_site":"Keurig公式サイト","prod_url":"https://www.keurig.com/brewers/k-supreme-plus-smart","maker_hp":"https://www.keurig.com","contact":"support@keurig.com"},
    {"no":1322,"category":"キッチン・調理器具","name":"Nespresso Vertuo Pop+ Capsule Coffee Machine","maker":"Nespresso","ec_site":"Nespresso公式サイト","prod_url":"https://www.nespresso.com/us/en/vertuo-pop-plus","maker_hp":"https://www.nespresso.com","contact":"us.info@nespresso.com"},
    {"no":1323,"category":"キッチン・調理器具","name":"Wolf Gourmet Elite High-Performance Blender","maker":"Wolf Gourmet","ec_site":"Wolf Gourmet公式サイト","prod_url":"https://www.wolfgourmet.com/products/blenders","maker_hp":"https://www.wolfgourmet.com","contact":"support@wolfgourmet.com"},
    {"no":1324,"category":"キッチン・調理器具","name":"Alessi 9090 Espresso Maker Stainless","maker":"Alessi","ec_site":"Alessi公式サイト","prod_url":"https://www.alessi.com/us/en/collections/coffee-tea","maker_hp":"https://www.alessi.com","contact":"info@alessi.com"},
    {"no":1325,"category":"キッチン・調理器具","name":"Peugeot Paris u'Select Pepper Mill 9-Inch","maker":"Peugeot Saveurs","ec_site":"Peugeot公式サイト","prod_url":"https://www.peugeot-saveurs.com/en/mills/pepper-mills","maker_hp":"https://www.peugeot-saveurs.com","contact":"info@peugeot-saveurs.com"},
    # ── スマートホーム・インテリア・照明 ───────────────────
    {"no":1326,"category":"スマートホーム・インテリア・照明","name":"Hoobs 4 Smart Home Hub Bridge","maker":"HOOBS","ec_site":"HOOBS公式サイト","prod_url":"https://hoobs.com/products/hoobs-4","maker_hp":"https://hoobs.com","contact":"support@hoobs.com"},
    {"no":1327,"category":"スマートホーム・インテリア・照明","name":"Aqara Hub M3 Smart Home Controller","maker":"Aqara","ec_site":"Aqara公式サイト","prod_url":"https://www.aqara.com/en/product/hub-m3","maker_hp":"https://www.aqara.com","contact":"support@aqara.com"},
    {"no":1328,"category":"スマートホーム・インテリア・照明","name":"Midea U-Shaped Window Air Conditioner 8000BTU","maker":"Midea","ec_site":"Midea公式サイト","prod_url":"https://www.midea.com/us/air-conditioners/window","maker_hp":"https://www.midea.com","contact":"support@midea.com"},
    {"no":1329,"category":"スマートホーム・インテリア・照明","name":"Dyson Purifier Hot+Cool HP09 Air Purifier Fan","maker":"Dyson","ec_site":"Dyson公式サイト","prod_url":"https://www.dyson.com/en/air-treatment/purifiers/dyson-purifier-hot-cool","maker_hp":"https://www.dyson.com","contact":"support@dyson.com"},
    {"no":1330,"category":"スマートホーム・インテリア・照明","name":"Molekule Air Pro PECO Air Purifier","maker":"Molekule","ec_site":"Molekule公式サイト","prod_url":"https://molekule.com/products/air-pro","maker_hp":"https://molekule.com","contact":"support@molekule.com"},
    {"no":1331,"category":"スマートホーム・インテリア・照明","name":"IQAir HealthPro Plus Air Purifier HyperHEPA","maker":"IQAir","ec_site":"IQAir公式サイト","prod_url":"https://www.iqair.com/us/air-purifiers/healthpro-plus","maker_hp":"https://www.iqair.com","contact":"info@iqair.com"},
    {"no":1332,"category":"スマートホーム・インテリア・照明","name":"Blueair Blue Pure 411i Max Smart Air Purifier","maker":"Blueair","ec_site":"Blueair公式サイト","prod_url":"https://www.blueair.com/us/air-purifiers/blue-pure-411i-max","maker_hp":"https://www.blueair.com","contact":"support@blueair.com"},
    {"no":1333,"category":"スマートホーム・インテリア・照明","name":"Levoit Core 600S Smart Air Purifier 635sqft","maker":"Levoit","ec_site":"Levoit公式サイト","prod_url":"https://levoit.com/products/core-600s-smart-true-hepa-air-purifier","maker_hp":"https://levoit.com","contact":"support@levoit.com"},
    {"no":1334,"category":"スマートホーム・インテリア・照明","name":"Sengled Smart LED Edison Bulb Multicolor","maker":"Sengled","ec_site":"Sengled公式サイト","prod_url":"https://us.sengled.com/collections/smart-bulbs","maker_hp":"https://us.sengled.com","contact":"support@sengled.com"},
    {"no":1335,"category":"スマートホーム・インテリア・照明","name":"Signify WiZ Smart LED Spotlight Starter Kit","maker":"Signify","ec_site":"WiZ公式サイト","prod_url":"https://www.wizconnected.com/en-US/consumer/products/","maker_hp":"https://www.wizconnected.com","contact":"support@wizconnected.com"},
    {"no":1336,"category":"スマートホーム・インテリア・照明","name":"Elgato Key Light Air Professional LED Panel","maker":"Elgato","ec_site":"Elgato公式サイト","prod_url":"https://www.elgato.com/en/key-light-air","maker_hp":"https://www.elgato.com","contact":"support@elgato.com"},
    {"no":1337,"category":"スマートホーム・インテリア・照明","name":"Govee TV Backlight T2 with Camera","maker":"Govee","ec_site":"Govee公式サイト","prod_url":"https://us.govee.com/products/govee-tv-backlight-t2","maker_hp":"https://us.govee.com","contact":"support@govee.com"},
    {"no":1338,"category":"スマートホーム・インテリア・照明","name":"Coway Airmega 400S Smart Air Purifier","maker":"Coway","ec_site":"Coway公式サイト","prod_url":"https://www.cowaymega.com/products/airmega-400s","maker_hp":"https://www.cowaymega.com","contact":"support@coway.com"},
    {"no":1339,"category":"スマートホーム・インテリア・照明","name":"La Crosse Technology Color Wireless Weather Station","maker":"La Crosse Technology","ec_site":"La Crosse公式サイト","prod_url":"https://www.lacrossetechnology.com/collections/weather-stations","maker_hp":"https://www.lacrossetechnology.com","contact":"support@lacrossetechnology.com"},
    {"no":1340,"category":"スマートホーム・インテリア・照明","name":"Logitech POP Smart Button Starter Kit","maker":"Logitech","ec_site":"Logitech公式サイト","prod_url":"https://www.logitech.com/en-us/products/pop-smart-buttons.html","maker_hp":"https://www.logitech.com","contact":"support@logitech.com"},
    {"no":1341,"category":"スマートホーム・インテリア・照明","name":"Roost Smart Water Sensor & Leak Detector","maker":"Roost","ec_site":"Roost公式サイト","prod_url":"https://www.getroost.com/products/roost-water-sensor","maker_hp":"https://www.getroost.com","contact":"support@getroost.com"},
    {"no":1342,"category":"スマートホーム・インテリア・照明","name":"Notion Smart Home Sensor Monitoring Kit","maker":"Notion","ec_site":"Notion公式サイト","prod_url":"https://getnotion.com/products/sensor-kit","maker_hp":"https://getnotion.com","contact":"support@getnotion.com"},
    {"no":1343,"category":"スマートホーム・インテリア・照明","name":"Elgato Facecam Pro 4K60 Webcam","maker":"Elgato","ec_site":"Elgato公式サイト","prod_url":"https://www.elgato.com/en/facecam-pro","maker_hp":"https://www.elgato.com","contact":"support@elgato.com"},
    {"no":1344,"category":"スマートホーム・インテリア・照明","name":"Hatch Restore 2 Smart Alarm Clock Sound Machine","maker":"Hatch","ec_site":"Hatch公式サイト","prod_url":"https://www.hatch.co/restore","maker_hp":"https://www.hatch.co","contact":"support@hatch.co"},
    {"no":1345,"category":"スマートホーム・インテリア・照明","name":"Casper Glow Light Smart Dimmable Nightlight","maker":"Casper","ec_site":"Casper公式サイト","prod_url":"https://casper.com/products/glow-light","maker_hp":"https://casper.com","contact":"support@casper.com"},
    # ── ウェアラブル・ヘルス・フィットネス ──────────────────
    {"no":1346,"category":"ウェアラブル・ヘルス・フィットネス","name":"Fitbit Charge 6 Advanced Fitness Tracker","maker":"Fitbit","ec_site":"Fitbit公式サイト","prod_url":"https://www.fitbit.com/global/us/products/trackers/charge6","maker_hp":"https://www.fitbit.com","contact":"support@fitbit.com"},
    {"no":1347,"category":"ウェアラブル・ヘルス・フィットネス","name":"Dexcom G7 Continuous Glucose Monitor","maker":"Dexcom","ec_site":"Dexcom公式サイト","prod_url":"https://www.dexcom.com/en-us/dexcom-g7","maker_hp":"https://www.dexcom.com","contact":"support@dexcom.com"},
    {"no":1348,"category":"ウェアラブル・ヘルス・フィットネス","name":"Levels Continuous Glucose Monitoring Program","maker":"Levels Health","ec_site":"Levels公式サイト","prod_url":"https://www.levelshealth.com","maker_hp":"https://www.levelshealth.com","contact":"support@levelshealth.com"},
    {"no":1349,"category":"ウェアラブル・ヘルス・フィットネス","name":"Apollo Neuro Wearable Stress Relief Band","maker":"Apollo Neuroscience","ec_site":"Apollo公式サイト","prod_url":"https://apolloneuro.com/products/apollo-wearable","maker_hp":"https://apolloneuro.com","contact":"hello@apolloneuro.com"},
    {"no":1350,"category":"ウェアラブル・ヘルス・フィットネス","name":"Whoop 5.0 Next-Gen Performance Tracker","maker":"WHOOP","ec_site":"WHOOP公式サイト","prod_url":"https://www.whoop.com/us/en/","maker_hp":"https://www.whoop.com","contact":"support@whoop.com"},
    {"no":1351,"category":"ウェアラブル・ヘルス・フィットネス","name":"Lumen Metabolism Tracker Breath Device","maker":"Lumen","ec_site":"Lumen公式サイト","prod_url":"https://www.lumen.me/products/lumen","maker_hp":"https://www.lumen.me","contact":"support@lumen.me"},
    {"no":1352,"category":"ウェアラブル・ヘルス・フィットネス","name":"Polar Pacer Pro Advanced Running GPS Watch","maker":"Polar Electro","ec_site":"Polar公式サイト","prod_url":"https://www.polar.com/en/pacer-pro","maker_hp":"https://www.polar.com","contact":"support@polar.com"},
    {"no":1353,"category":"ウェアラブル・ヘルス・フィットネス","name":"Garmin Forerunner 965 Premium GPS Running Watch","maker":"Garmin","ec_site":"Garmin公式サイト","prod_url":"https://www.garmin.com/en-US/p/forerunner-965","maker_hp":"https://www.garmin.com","contact":"support@garmin.com"},
    {"no":1354,"category":"ウェアラブル・ヘルス・フィットネス","name":"HigherDOSE Infrared PEMF Mat Full Body","maker":"HigherDOSE","ec_site":"HigherDOSE公式サイト","prod_url":"https://higherdose.com/products/infrared-pemf-mat","maker_hp":"https://higherdose.com","contact":"hi@higherdose.com"},
    {"no":1355,"category":"ウェアラブル・ヘルス・フィットネス","name":"Biostrap Pulse Overnight Recovery Band","maker":"Biostrap","ec_site":"Biostrap公式サイト","prod_url":"https://biostrap.com/products/biostrap-pulse","maker_hp":"https://biostrap.com","contact":"support@biostrap.com"},
    {"no":1356,"category":"ウェアラブル・ヘルス・フィットネス","name":"Tone & Strengthen Band Smart EMS Abdominal Belt","maker":"Slendertone","ec_site":"Slendertone公式サイト","prod_url":"https://www.slendertone.com/products","maker_hp":"https://www.slendertone.com","contact":"support@slendertone.com"},
    {"no":1357,"category":"ウェアラブル・ヘルス・フィットネス","name":"NovaBay Avenova Eye Spray Lid & Lash Cleanser","maker":"NovaBay Pharmaceuticals","ec_site":"Avenova公式サイト","prod_url":"https://avenova.com/products/avenova-spray","maker_hp":"https://avenova.com","contact":"info@avenova.com"},
    {"no":1358,"category":"ウェアラブル・ヘルス・フィットネス","name":"Withings Body Comp Smart Scale Wi-Fi","maker":"Withings","ec_site":"Withings公式サイト","prod_url":"https://www.withings.com/us/en/body-comp","maker_hp":"https://www.withings.com","contact":"support@withings.com"},
    {"no":1359,"category":"ウェアラブル・ヘルス・フィットネス","name":"Renpho Smart Body Fat Scale Bluetooth","maker":"Renpho","ec_site":"Renpho公式サイト","prod_url":"https://renpho.com/collections/smart-scale","maker_hp":"https://renpho.com","contact":"support@renpho.com"},
    {"no":1360,"category":"ウェアラブル・ヘルス・フィットネス","name":"Omron Evolv Wireless Upper Arm Blood Pressure","maker":"Omron Healthcare","ec_site":"Omron公式サイト","prod_url":"https://omronhealthcare.com/products/evolv-wireless-upper-arm-blood-pressure-monitor","maker_hp":"https://omronhealthcare.com","contact":"consumer@omron.com"},
    # ── アウトドア・スポーツ・旅行 ────────────────────────
    {"no":1361,"category":"アウトドア・スポーツ・旅行","name":"Tentsile Stingray 3-Person Tree Tent","maker":"Tentsile","ec_site":"Tentsile公式サイト","prod_url":"https://www.tentsile.com/collections/tree-tents","maker_hp":"https://www.tentsile.com","contact":"info@tentsile.com"},
    {"no":1362,"category":"アウトドア・スポーツ・旅行","name":"ENO DoubleNest Hammock Lightweight 2-Person","maker":"Eagles Nest Outfitters","ec_site":"ENO公式サイト","prod_url":"https://eaglesnestoutfittersinc.com/collections/hammocks","maker_hp":"https://eaglesnestoutfittersinc.com","contact":"info@eaglesnestoutfittersinc.com"},
    {"no":1363,"category":"アウトドア・スポーツ・旅行","name":"Big Agnes Tiger Wall UL3 Ultralight Tent","maker":"Big Agnes","ec_site":"Big Agnes公式サイト","prod_url":"https://www.bigagnes.com/collections/tents","maker_hp":"https://www.bigagnes.com","contact":"info@bigagnes.com"},
    {"no":1364,"category":"アウトドア・スポーツ・旅行","name":"Jetboil Flash Cooking System 1L Camp Stove","maker":"Jetboil","ec_site":"Jetboil公式サイト","prod_url":"https://jetboil.johnsondiversey.com/products/flash","maker_hp":"https://jetboil.com","contact":"info@jetboil.com"},
    {"no":1365,"category":"アウトドア・スポーツ・旅行","name":"Snow Peak GigaPower Stove Auto Ignition","maker":"Snow Peak","ec_site":"Snow Peak公式サイト","prod_url":"https://snowpeak.com/collections/stoves","maker_hp":"https://snowpeak.com","contact":"info@snowpeak.com"},
    {"no":1366,"category":"アウトドア・スポーツ・旅行","name":"Hydro Flask 32oz Wide Mouth Trail Series","maker":"Hydro Flask","ec_site":"Hydro Flask公式サイト","prod_url":"https://www.hydroflask.com/collections/wide-mouth","maker_hp":"https://www.hydroflask.com","contact":"info@hydroflask.com"},
    {"no":1367,"category":"アウトドア・スポーツ・旅行","name":"Yeti Rambler 30oz Travel Mug Magslider Lid","maker":"YETI","ec_site":"YETI公式サイト","prod_url":"https://www.yeti.com/en_US/drinkware/rambler-30-oz-travel-mug","maker_hp":"https://www.yeti.com","contact":"info@yeti.com"},
    {"no":1368,"category":"アウトドア・スポーツ・旅行","name":"Klymit Static V2 Lightweight Sleeping Pad","maker":"Klymit","ec_site":"Klymit公式サイト","prod_url":"https://www.klymit.com/collections/sleeping-pads","maker_hp":"https://www.klymit.com","contact":"support@klymit.com"},
    {"no":1369,"category":"アウトドア・スポーツ・旅行","name":"Cascade Mountain Tech Carbon Fiber Trekking Poles","maker":"Cascade Mountain Tech","ec_site":"CMT公式サイト","prod_url":"https://cascademountaintech.com/collections/trekking-poles","maker_hp":"https://cascademountaintech.com","contact":"support@cascademountaintech.com"},
    {"no":1370,"category":"アウトドア・スポーツ・旅行","name":"Outdoor Research Helium AscentShell Jacket","maker":"Outdoor Research","ec_site":"OR公式サイト","prod_url":"https://www.outdoorresearch.com/collections/mens-rain-jackets","maker_hp":"https://www.outdoorresearch.com","contact":"info@outdoorresearch.com"},
    {"no":1371,"category":"アウトドア・スポーツ・旅行","name":"Arc'teryx Atom Hoody Insulated Jacket","maker":"Arc'teryx","ec_site":"Arc'teryx公式サイト","prod_url":"https://arcteryx.com/us/en/shop/mens/atom-hoody","maker_hp":"https://arcteryx.com","contact":"customer.service@arcteryx.com"},
    {"no":1372,"category":"アウトドア・スポーツ・旅行","name":"Patagonia Nano Puff Hoody Lightweight","maker":"Patagonia","ec_site":"Patagonia公式サイト","prod_url":"https://www.patagonia.com/shop/mens-insulated-jackets","maker_hp":"https://www.patagonia.com","contact":"customerservice@patagonia.com"},
    {"no":1373,"category":"アウトドア・スポーツ・旅行","name":"Osprey Talon 33 Hiking Backpack","maker":"Osprey Packs","ec_site":"Osprey公式サイト","prod_url":"https://www.osprey.com/us/en/category/mens-hiking-packs/","maker_hp":"https://www.osprey.com","contact":"custserv@ospreypacks.com"},
    {"no":1374,"category":"アウトドア・スポーツ・旅行","name":"Deuter Futura Air Trek 55+10 Trekking Pack","maker":"Deuter","ec_site":"Deuter公式サイト","prod_url":"https://www.deuter.com/us-en/backpacks/trekking","maker_hp":"https://www.deuter.com","contact":"info@deuter.com"},
    {"no":1375,"category":"アウトドア・スポーツ・旅行","name":"Black Diamond Spot 350 Headlamp Waterproof","maker":"Black Diamond Equipment","ec_site":"Black Diamond公式サイト","prod_url":"https://www.blackdiamondequipment.com/en_US/headlamps","maker_hp":"https://www.blackdiamondequipment.com","contact":"info@bdel.com"},
    # ── ペット用品 ─────────────────────────────────────
    {"no":1376,"category":"ペット用品","name":"Outward Hound Fun Feeder Slo Bowl Anti-Bloat","maker":"Outward Hound","ec_site":"Outward Hound公式サイト","prod_url":"https://www.outwardhound.com/collections/feeding","maker_hp":"https://www.outwardhound.com","contact":"support@outwardhound.com"},
    {"no":1377,"category":"ペット用品","name":"West Paw Zogoflex Qwizl Treat Dispensing Toy","maker":"West Paw","ec_site":"West Paw公式サイト","prod_url":"https://www.westpaw.com/collections/dog-chews","maker_hp":"https://www.westpaw.com","contact":"info@westpaw.com"},
    {"no":1378,"category":"ペット用品","name":"Ruffwear Web Master Dog Harness","maker":"Ruffwear","ec_site":"Ruffwear公式サイト","prod_url":"https://ruffwear.com/collections/dog-harnesses","maker_hp":"https://ruffwear.com","contact":"support@ruffwear.com"},
    {"no":1379,"category":"ペット用品","name":"Kong Wobbler Treat Dispensing Dog Toy","maker":"Kong Company","ec_site":"Kong公式サイト","prod_url":"https://www.kongcompany.com/products/for-dogs/chew-toys/treat-dispensing","maker_hp":"https://www.kongcompany.com","contact":"info@kongcompany.com"},
    {"no":1380,"category":"ペット用品","name":"Chewy Ramp Foldable Pet Steps 3-Step","maker":"PetSafe","ec_site":"PetSafe公式サイト","prod_url":"https://www.petsafe.net/steps-and-ramps","maker_hp":"https://www.petsafe.net","contact":"csc@petsafe.net"},
    {"no":1381,"category":"ペット用品","name":"Bissell BarkBath Portable Dog Bath System","maker":"Bissell","ec_site":"Bissell公式サイト","prod_url":"https://www.bissell.com/collections/barkbath","maker_hp":"https://www.bissell.com","contact":"consumercare@bissell.com"},
    {"no":1382,"category":"ペット用品","name":"CleanPaws Smart Paw Washer Electric Dog","maker":"Dexas","ec_site":"Dexas公式サイト","prod_url":"https://www.dexas.com/products/mudbuster","maker_hp":"https://www.dexas.com","contact":"info@dexas.com"},
    {"no":1383,"category":"ペット用品","name":"Motorola Halo+ Smart Video Baby Monitor","maker":"Motorola","ec_site":"Motorola公式サイト","prod_url":"https://www.motorola.com/us/baby-monitors","maker_hp":"https://www.motorola.com","contact":"support@motorola.com"},
    {"no":1384,"category":"ペット用品","name":"Sure Petcare Animo Dog Behavior Monitor","maker":"Sure Petcare","ec_site":"Sure Petcare公式サイト","prod_url":"https://www.surepetcare.com/en-us/animo","maker_hp":"https://www.surepetcare.com","contact":"support@surepetcare.com"},
    {"no":1385,"category":"ペット用品","name":"Vetericyn Plus All Animal Wound & Skin Care","maker":"Vetericyn","ec_site":"Vetericyn公式サイト","prod_url":"https://vetericyn.com/products/plus-wound-skin-care-spray","maker_hp":"https://vetericyn.com","contact":"info@vetericyn.com"},
    {"no":1386,"category":"ペット用品","name":"GoPro Fetch Dog Harness Mount","maker":"GoPro","ec_site":"GoPro公式サイト","prod_url":"https://gopro.com/en/us/update/fetch","maker_hp":"https://gopro.com","contact":"support@gopro.com"},
    {"no":1387,"category":"ペット用品","name":"Pura Smart Fragrance Diffuser Pet Safe","maker":"Pura","ec_site":"Pura公式サイト","prod_url":"https://www.trypura.com/products/smart-diffuser","maker_hp":"https://www.trypura.com","contact":"support@trypura.com"},
    {"no":1388,"category":"ペット用品","name":"Torus Pet Water Bowl Filtered 68oz","maker":"Torus Pet","ec_site":"Torus公式サイト","prod_url":"https://www.toruspet.com/products/torus-water-bowl","maker_hp":"https://www.toruspet.com","contact":"info@toruspet.com"},
    {"no":1389,"category":"ペット用品","name":"Mendota Pet Slip Lead Rope Training Leash","maker":"Mendota Products","ec_site":"Mendota公式サイト","prod_url":"https://mendotapet.com/collections/slip-leads","maker_hp":"https://mendotapet.com","contact":"info@mendotapet.com"},
    {"no":1390,"category":"ペット用品","name":"Embark Breed & Health Dog DNA Test Kit","maker":"Embark Veterinary","ec_site":"Embark公式サイト","prod_url":"https://embarkvet.com/products/dog-dna-test","maker_hp":"https://embarkvet.com","contact":"support@embarkvet.com"},
    # ── テクノロジー・ガジェット ────────────────────────
    {"no":1391,"category":"テクノロジー・ガジェット","name":"Thelio Mega Desktop Linux Computer System76","maker":"System76","ec_site":"System76公式サイト","prod_url":"https://system76.com/desktops/thelio-mega","maker_hp":"https://system76.com","contact":"info@system76.com"},
    {"no":1392,"category":"テクノロジー・ガジェット","name":"Loupedeck Live S Streamlined Photo Editing Console","maker":"Loupedeck","ec_site":"Loupedeck公式サイト","prod_url":"https://loupedeck.com/products/loupedeck-live-s","maker_hp":"https://loupedeck.com","contact":"support@loupedeck.com"},
    {"no":1393,"category":"テクノロジー・ガジェット","name":"Valnet MiniPC M680 Mini Desktop Computer","maker":"MinisForum","ec_site":"MinisForum公式サイト","prod_url":"https://www.minisforum.com/products/um780","maker_hp":"https://www.minisforum.com","contact":"support@minisforum.com"},
    {"no":1394,"category":"テクノロジー・ガジェット","name":"Elgato Wave Neo USB Condenser Microphone","maker":"Elgato","ec_site":"Elgato公式サイト","prod_url":"https://www.elgato.com/en/wave-neo","maker_hp":"https://www.elgato.com","contact":"support@elgato.com"},
    {"no":1395,"category":"テクノロジー・ガジェット","name":"Blue Yeti X USB Condenser Microphone","maker":"Blue Microphones","ec_site":"Blue公式サイト","prod_url":"https://www.bluemic.com/en-us/products/yeti-x/","maker_hp":"https://www.bluemic.com","contact":"support@bluemic.com"},
    {"no":1396,"category":"テクノロジー・ガジェット","name":"Peak Design Mobile Tripod Compact Phone Mount","maker":"Peak Design","ec_site":"Peak Design公式サイト","prod_url":"https://www.peakdesign.com/collections/mobile","maker_hp":"https://www.peakdesign.com","contact":"help@peakdesign.com"},
    {"no":1397,"category":"テクノロジー・ガジェット","name":"Moment V3 Anamorphic 1.33x Phone Lens","maker":"Moment","ec_site":"Moment公式サイト","prod_url":"https://www.shopmoment.com/collections/lenses","maker_hp":"https://www.shopmoment.com","contact":"support@shopmoment.com"},
    {"no":1398,"category":"テクノロジー・ガジェット","name":"Insta360 GO 3S Action Camera Tiny Body","maker":"Insta360","ec_site":"Insta360公式サイト","prod_url":"https://www.insta360.com/product/insta360-go3s","maker_hp":"https://www.insta360.com","contact":"support@insta360.com"},
    {"no":1399,"category":"テクノロジー・ガジェット","name":"DJI Action 5 Pro Action Camera","maker":"DJI","ec_site":"DJI公式サイト","prod_url":"https://www.dji.com/action-5-pro","maker_hp":"https://www.dji.com","contact":"support@dji.com"},
    {"no":1400,"category":"テクノロジー・ガジェット","name":"Rode Wireless PRO Dual Channel Wireless Mic","maker":"Rode Microphones","ec_site":"Rode公式サイト","prod_url":"https://rode.com/en/microphones/wireless/wireless-pro","maker_hp":"https://rode.com","contact":"support@rode.com"},
    {"no":1401,"category":"テクノロジー・ガジェット","name":"Ulanzi MT-44 Tabletop Tripod Phone Stand","maker":"Ulanzi","ec_site":"Ulanzi公式サイト","prod_url":"https://www.ulanzi.com/collections/phone-tripods","maker_hp":"https://www.ulanzi.com","contact":"support@ulanzi.com"},
    {"no":1402,"category":"テクノロジー・ガジェット","name":"Anker 778 Thunderbolt 4 Docking Station 12-in-1","maker":"Anker","ec_site":"Anker公式サイト","prod_url":"https://www.anker.com/products/anker-778-thunderbolt-docking-station","maker_hp":"https://www.anker.com","contact":"support@anker.com"},
    {"no":1403,"category":"テクノロジー・ガジェット","name":"CalDigit TS4 Thunderbolt 4 Dock 18-Port","maker":"CalDigit","ec_site":"CalDigit公式サイト","prod_url":"https://www.caldigit.com/thunderbolt-4-element-hub/","maker_hp":"https://www.caldigit.com","contact":"support@caldigit.com"},
    {"no":1404,"category":"テクノロジー・ガジェット","name":"Jackery Solar Generator 2000 Pro","maker":"Jackery","ec_site":"Jackery公式サイト","prod_url":"https://www.jackery.com/products/solar-generator-2000-pro","maker_hp":"https://www.jackery.com","contact":"support@jackery.com"},
    {"no":1405,"category":"テクノロジー・ガジェット","name":"EcoFlow Delta 2 Max Portable Power Station","maker":"EcoFlow","ec_site":"EcoFlow公式サイト","prod_url":"https://www.ecoflow.com/products/delta2-max-portable-power-station","maker_hp":"https://www.ecoflow.com","contact":"service@ecoflow.com"},
    # ── 美容・スキンケア ───────────────────────────────
    {"no":1406,"category":"美容・スキンケア","name":"Therabody RecoveryAir JetBoots Compression","maker":"Therabody","ec_site":"Therabody公式サイト","prod_url":"https://www.therabody.com/us/en-us/recoveryair-jetboots.html","maker_hp":"https://www.therabody.com","contact":"support@therabody.com"},
    {"no":1407,"category":"美容・スキンケア","name":"Gua Sha Stainless Steel Face Scraping Tool","maker":"Mount Lai","ec_site":"Mount Lai公式サイト","prod_url":"https://mountlai.com/collections/tools","maker_hp":"https://mountlai.com","contact":"hello@mountlai.com"},
    {"no":1408,"category":"美容・スキンケア","name":"Tula Skincare Glow & Get It Cooling & Brightening Eye Balm","maker":"Tula Skincare","ec_site":"Tula公式サイト","prod_url":"https://www.tula.com/collections/eye-treatment","maker_hp":"https://www.tula.com","contact":"help@tula.com"},
    {"no":1409,"category":"美容・スキンケア","name":"Elemis Pro-Collagen Marine Cream 50ml","maker":"Elemis","ec_site":"Elemis公式サイト","prod_url":"https://www.elemis.com/pro-collagen-marine-cream","maker_hp":"https://www.elemis.com","contact":"customerservice@elemis.com"},
    {"no":1410,"category":"美容・スキンケア","name":"Summer Fridays Cloud Dew Oil-Free Moisturizer","maker":"Summer Fridays","ec_site":"Summer Fridays公式サイト","prod_url":"https://www.summerfridays.com/collections/moisturizers","maker_hp":"https://www.summerfridays.com","contact":"support@summerfridays.com"},
    {"no":1411,"category":"美容・スキンケア","name":"Kiehl's Ultra Facial Cream SPF 30 125ml","maker":"Kiehl's","ec_site":"Kiehl's公式サイト","prod_url":"https://www.kiehls.com/skincare/moisturizers","maker_hp":"https://www.kiehls.com","contact":"kiehls@kiehls.com"},
    {"no":1412,"category":"美容・スキンケア","name":"Tatcha The Dewy Skin Cream Rich Moisturizer","maker":"Tatcha","ec_site":"Tatcha公式サイト","prod_url":"https://www.tatcha.com/collections/moisturizers","maker_hp":"https://www.tatcha.com","contact":"support@tatcha.com"},
    {"no":1413,"category":"美容・スキンケア","name":"Sunday Riley Good Genes All-In-One AHA Lactic Acid","maker":"Sunday Riley","ec_site":"Sunday Riley公式サイト","prod_url":"https://sundayriley.com/collections/serums","maker_hp":"https://sundayriley.com","contact":"support@sundayriley.com"},
    {"no":1414,"category":"美容・スキンケア","name":"Dr. Dennis Gross Alpha Beta Universal Daily Peel","maker":"Dr. Dennis Gross","ec_site":"DDG公式サイト","prod_url":"https://www.drdennisgross.com/collections/peels","maker_hp":"https://www.drdennisgross.com","contact":"support@drdennisgross.com"},
    {"no":1415,"category":"美容・スキンケア","name":"Herbivore Botanicals Emerald Cannabis Sativa Hemp Oil","maker":"Herbivore Botanicals","ec_site":"Herbivore公式サイト","prod_url":"https://www.herbivorebotanicals.com/collections/serums","maker_hp":"https://www.herbivorebotanicals.com","contact":"hello@herbivorebotanicals.com"},
    {"no":1416,"category":"美容・スキンケア","name":"Biossance Squalane + Vitamin C Rose Oil","maker":"Biossance","ec_site":"Biossance公式サイト","prod_url":"https://biossance.com/collections/oils","maker_hp":"https://biossance.com","contact":"support@biossance.com"},
    {"no":1417,"category":"美容・スキンケア","name":"Votary Super Seed Facial Oil Rosehip & Retinoid","maker":"Votary","ec_site":"Votary公式サイト","prod_url":"https://votary.com/collections/face-oils","maker_hp":"https://votary.com","contact":"hello@votary.com"},
    {"no":1418,"category":"美容・スキンケア","name":"Alpyn Beauty PlantGenius Melt Moisturizer","maker":"Alpyn Beauty","ec_site":"Alpyn Beauty公式サイト","prod_url":"https://alpynbeauty.com/collections/moisturizers","maker_hp":"https://alpynbeauty.com","contact":"hello@alpynbeauty.com"},
    {"no":1419,"category":"美容・スキンケア","name":"Versed Dew Point Moisturizing Gel-Cream","maker":"Versed Skin","ec_site":"Versed公式サイト","prod_url":"https://www.versedskin.com/collections/moisturizers","maker_hp":"https://www.versedskin.com","contact":"hello@versedskin.com"},
    {"no":1420,"category":"美容・スキンケア","name":"Glow Recipe Watermelon Glow Niacinamide Dew Drops","maker":"Glow Recipe","ec_site":"Glow Recipe公式サイト","prod_url":"https://www.glowrecipe.com/collections/serums","maker_hp":"https://www.glowrecipe.com","contact":"support@glowrecipe.com"},
    # ── 子供・教育 ─────────────────────────────────────
    {"no":1421,"category":"子供・教育","name":"KiwiCo Atlas Crate Subscription Geography Box","maker":"KiwiCo","ec_site":"KiwiCo公式サイト","prod_url":"https://www.kiwico.com/atlas","maker_hp":"https://www.kiwico.com","contact":"support@kiwico.com"},
    {"no":1422,"category":"子供・教育","name":"Snap Circuits Extreme 750 Electronics Kit","maker":"Elenco Electronics","ec_site":"Elenco公式サイト","prod_url":"https://www.elenco.com/products/snap-circuits/","maker_hp":"https://www.elenco.com","contact":"info@elenco.com"},
    {"no":1423,"category":"子供・教育","name":"Learning Resources Coding Critters Ranger & Rex","maker":"Learning Resources","ec_site":"LR公式サイト","prod_url":"https://www.learningresources.com/coding-critters","maker_hp":"https://www.learningresources.com","contact":"info@learningresources.com"},
    {"no":1424,"category":"子供・教育","name":"Popular Science Hydraulic Robot Arm Kit","maker":"Popular Science","ec_site":"Amazon","prod_url":"https://www.amazon.com/s?k=popular+science+hydraulic+robot+arm","maker_hp":"https://www.popsci.com","contact":"letters@popsci.com"},
    {"no":1425,"category":"子供・教育","name":"Botley 2.0 Coding Robot Activity Set","maker":"Learning Resources","ec_site":"LR公式サイト","prod_url":"https://www.learningresources.com/botley-2","maker_hp":"https://www.learningresources.com","contact":"info@learningresources.com"},
    {"no":1426,"category":"子供・教育","name":"Geomag Mechanics Gravity Motor Set 169pcs","maker":"Geomag","ec_site":"Geomag公式サイト","prod_url":"https://geomagworld.com/en/ranges/mechanics","maker_hp":"https://geomagworld.com","contact":"info@geomagworld.com"},
    {"no":1427,"category":"子供・教育","name":"Magformers Stem Challenger 144-Piece Set","maker":"Magformers","ec_site":"Magformers公式サイト","prod_url":"https://www.magformers.com/en/stem-challenger","maker_hp":"https://www.magformers.com","contact":"info@magformers.com"},
    {"no":1428,"category":"子供・教育","name":"Strawbees Maker Kit Creative STEM Building","maker":"Strawbees","ec_site":"Strawbees公式サイト","prod_url":"https://www.strawbees.com/products/maker-kit","maker_hp":"https://www.strawbees.com","contact":"hello@strawbees.com"},
    {"no":1429,"category":"子供・教育","name":"Discovery Mindblown Coding Robot & Activity Set","maker":"Discovery Mindblown","ec_site":"Discovery公式サイト","prod_url":"https://www.discoverymindblown.com/collections/coding-robots","maker_hp":"https://www.discoverymindblown.com","contact":"support@discoverymindblown.com"},
    {"no":1430,"category":"子供・教育","name":"Thames & Kosmos Electrical Circuit Board Game","maker":"Thames & Kosmos","ec_site":"Thames公式サイト","prod_url":"https://www.thamesandkosmos.com/products/electricity","maker_hp":"https://www.thamesandkosmos.com","contact":"science@thamesandkosmos.com"},
    {"no":1431,"category":"子供・教育","name":"Kano PC Kit Build-Your-Own Laptop Touchscreen","maker":"Kano","ec_site":"Kano公式サイト","prod_url":"https://kano.me/store/en-us/collections/pc","maker_hp":"https://kano.me","contact":"support@kano.me"},
    {"no":1432,"category":"子供・教育","name":"Artie 3000 The Coding Robot Draw STEM","maker":"Educational Insights","ec_site":"EI公式サイト","prod_url":"https://www.educationalinsights.com/artie-3000","maker_hp":"https://www.educationalinsights.com","contact":"service@educationalinsights.com"},
    {"no":1433,"category":"子供・教育","name":"Lego Technic Bugatti Bolide Model Car 905pc","maker":"Lego","ec_site":"Lego公式サイト","prod_url":"https://www.lego.com/en-us/themes/technic","maker_hp":"https://www.lego.com","contact":"service@lego.com"},
    {"no":1434,"category":"子供・教育","name":"Beebot Programmable Floor Robot for Kids","maker":"TTS Group","ec_site":"TTS公式サイト","prod_url":"https://www.tts-group.co.uk/bee-bot/","maker_hp":"https://www.tts-group.co.uk","contact":"enquiries@tts-group.co.uk"},
    {"no":1435,"category":"子供・教育","name":"Osmo Genius Numbers Learning Game Kit iPad","maker":"Osmo","ec_site":"Osmo公式サイト","prod_url":"https://www.playosmo.com/en/numbers/","maker_hp":"https://www.playosmo.com","contact":"support@playosmo.com"},
    # ── ファッション・アクセサリー ──────────────────────
    {"no":1436,"category":"ファッション・アクセサリー","name":"Baggallini Anti-Theft Crossbody Travel Bag","maker":"Baggallini","ec_site":"Baggallini公式サイト","prod_url":"https://www.baggallini.com/collections/crossbody-bags","maker_hp":"https://www.baggallini.com","contact":"customerservice@baggallini.com"},
    {"no":1437,"category":"ファッション・アクセサリー","name":"Osprey Daylite Travel Daypack 24+6L","maker":"Osprey Packs","ec_site":"Osprey公式サイト","prod_url":"https://www.osprey.com/us/en/product/daylite-travel/","maker_hp":"https://www.osprey.com","contact":"custserv@ospreypacks.com"},
    {"no":1438,"category":"ファッション・アクセサリー","name":"Pacsafe Vibe 25L Anti-Theft Backpack","maker":"Pacsafe","ec_site":"Pacsafe公式サイト","prod_url":"https://www.pacsafe.com/collections/anti-theft-backpacks","maker_hp":"https://www.pacsafe.com","contact":"help@pacsafe.com"},
    {"no":1439,"category":"ファッション・アクセサリー","name":"Knomo Beauchamp 14 Inch Roll-Top Backpack","maker":"Knomo London","ec_site":"Knomo公式サイト","prod_url":"https://www.knomo.com/collections/backpacks","maker_hp":"https://www.knomo.com","contact":"support@knomo.com"},
    {"no":1440,"category":"ファッション・アクセサリー","name":"Db Journey Hugger 30L Backpack Snow Pro","maker":"Db","ec_site":"Db公式サイト","prod_url":"https://www.db-journey.com/collections/bags","maker_hp":"https://www.db-journey.com","contact":"support@db-journey.com"},
    {"no":1441,"category":"ファッション・アクセサリー","name":"Tortuga Travel Backpack 40L Laptop Carry-On","maker":"Tortuga Backpacks","ec_site":"Tortuga公式サイト","prod_url":"https://www.tortugabackpacks.com/collections/travel-backpacks","maker_hp":"https://www.tortugabackpacks.com","contact":"hello@tortugabackpacks.com"},
    {"no":1442,"category":"ファッション・アクセサリー","name":"SLNT Signal Blocking Faraday Phone Sleeve","maker":"SLNT","ec_site":"SLNT公式サイト","prod_url":"https://slnt.com/collections/phone-cases","maker_hp":"https://slnt.com","contact":"support@slnt.com"},
    {"no":1443,"category":"ファッション・アクセサリー","name":"Chrome Industries Vega 2.0 Messenger Bag","maker":"Chrome Industries","ec_site":"Chrome公式サイト","prod_url":"https://www.chromeindustries.com/collections/bags","maker_hp":"https://www.chromeindustries.com","contact":"support@chromeindustries.com"},
    {"no":1444,"category":"ファッション・アクセサリー","name":"Rains Rolltop Backpack Mini Waterproof","maker":"Rains","ec_site":"Rains公式サイト","prod_url":"https://www.rains.com/collections/backpacks","maker_hp":"https://www.rains.com","contact":"info@rains.com"},
    {"no":1445,"category":"ファッション・アクセサリー","name":"Bellroy Transit Backpack Plus 38L","maker":"Bellroy","ec_site":"Bellroy公式サイト","prod_url":"https://bellroy.com/products/transit-backpack-plus","maker_hp":"https://bellroy.com","contact":"help@bellroy.com"},
    {"no":1446,"category":"ファッション・アクセサリー","name":"Dagne Dover Landon Carryall Medium Tote","maker":"Dagne Dover","ec_site":"Dagne Dover公式サイト","prod_url":"https://www.dagnedover.com/collections/landon-carryall","maker_hp":"https://www.dagnedover.com","contact":"hello@dagnedover.com"},
    {"no":1447,"category":"ファッション・アクセサリー","name":"Calpak Luka Duffel Bag 50L Carry-On Size","maker":"CALPAK","ec_site":"CALPAK公式サイト","prod_url":"https://www.calpaktravel.com/collections/duffel-bags","maker_hp":"https://www.calpaktravel.com","contact":"hello@calpaktravel.com"},
    {"no":1448,"category":"ファッション・アクセサリー","name":"Hex Camera DSLR Convertible Messenger Bag","maker":"Hex Brand","ec_site":"Hex公式サイト","prod_url":"https://hexbrand.com/collections/camera-bags","maker_hp":"https://hexbrand.com","contact":"hello@hexbrand.com"},
    {"no":1449,"category":"ファッション・アクセサリー","name":"Tom Bihn Aeronaut 45 One-Bag Travel Pack","maker":"Tom Bihn","ec_site":"Tom Bihn公式サイト","prod_url":"https://www.tombihn.com/products/aeronaut-45","maker_hp":"https://www.tombihn.com","contact":"tbcs@tombihn.com"},
    {"no":1450,"category":"ファッション・アクセサリー","name":"Minaal Carry-On 3.0 Backpack 35L","maker":"Minaal","ec_site":"Minaal公式サイト","prod_url":"https://www.minaal.com/products/carry-on-3","maker_hp":"https://www.minaal.com","contact":"hello@minaal.com"},
    # ── クリーニング・収納・整理 ───────────────────────
    {"no":1451,"category":"クリーニング・収納・整理","name":"Eufy Clean X9 Pro Self-Cleaning Robot with Mop","maker":"Eufy","ec_site":"Eufy公式サイト","prod_url":"https://www.eufy.com/products/t2150111","maker_hp":"https://www.eufy.com","contact":"support@eufylife.com"},
    {"no":1452,"category":"クリーニング・収納・整理","name":"Shark AI Ultra 2-in-1 Robot Vacuum with Matrix Clean","maker":"Shark","ec_site":"Shark公式サイト","prod_url":"https://www.sharkclean.com/collections/robot-vacuums","maker_hp":"https://www.sharkclean.com","contact":"customerservice@sharkclean.com"},
    {"no":1453,"category":"クリーニング・収納・整理","name":"Karcher WD 4 V-20/6 Wet and Dry Vacuum","maker":"Kärcher","ec_site":"Kärcher公式サイト","prod_url":"https://www.karcher.com/us/products/wet-and-dry-vacuum-cleaners.html","maker_hp":"https://www.karcher.com","contact":"info@karcher.com"},
    {"no":1454,"category":"クリーニング・収納・整理","name":"BLACK+DECKER 20V MAX Dustbuster AdvancedClean","maker":"BLACK+DECKER","ec_site":"B+D公式サイト","prod_url":"https://www.blackanddecker.com/collections/dustbusters","maker_hp":"https://www.blackanddecker.com","contact":"consumer@sbdinc.com"},
    {"no":1455,"category":"クリーニング・収納・整理","name":"Miele Triflex HX2 Bagless Cordless Stick Vacuum","maker":"Miele","ec_site":"Miele公式サイト","prod_url":"https://www.mieleusa.com/domestic/cordless-stick-vacuums","maker_hp":"https://www.mieleusa.com","contact":"info@mieleusa.com"},
    {"no":1456,"category":"クリーニング・収納・整理","name":"Anova Culinary Precision Vacuum Sealer Pro","maker":"Anova Culinary","ec_site":"Anova公式サイト","prod_url":"https://anovaculinary.com/products/anova-precision-vacuum-sealer-pro","maker_hp":"https://anovaculinary.com","contact":"support@anovaculinary.com"},
    {"no":1457,"category":"クリーニング・収納・整理","name":"Simplehuman Rectangular Stainless Step Trash Can 50L","maker":"simplehuman","ec_site":"simplehuman公式サイト","prod_url":"https://www.simplehuman.com/collections/rectangular-step-cans","maker_hp":"https://www.simplehuman.com","contact":"support@simplehuman.com"},
    {"no":1458,"category":"クリーニング・収納・整理","name":"Joseph Joseph DrawerStore Expandable Cutlery Organiser","maker":"Joseph Joseph","ec_site":"Joseph Joseph公式サイト","prod_url":"https://www.josephjoseph.com/collections/drawer-organisers","maker_hp":"https://www.josephjoseph.com","contact":"help@josephjoseph.com"},
    {"no":1459,"category":"クリーニング・収納・整理","name":"OXO Good Grips 20-Piece Smart Seal Container Set","maker":"OXO","ec_site":"OXO公式サイト","prod_url":"https://www.oxo.com/categories/organization/food-storage","maker_hp":"https://www.oxo.com","contact":"info@oxo.com"},
    {"no":1460,"category":"クリーニング・収納・整理","name":"Brabantia Bo Touch Bin 2x30L Dual Compartment","maker":"Brabantia","ec_site":"Brabantia公式サイト","prod_url":"https://www.brabantia.com/en/bo-touch-bin","maker_hp":"https://www.brabantia.com","contact":"consumer@brabantia.com"},
    {"no":1461,"category":"クリーニング・収納・整理","name":"Oxo Tot Dishwasher Basket for Bottle Parts","maker":"OXO","ec_site":"OXO公式サイト","prod_url":"https://www.oxo.com/categories/baby-and-toddler","maker_hp":"https://www.oxo.com","contact":"info@oxo.com"},
    {"no":1462,"category":"クリーニング・収納・整理","name":"Umbra Trigg Corner Floating Shelf Organizer","maker":"Umbra","ec_site":"Umbra公式サイト","prod_url":"https://www.umbra.com/collections/shelves","maker_hp":"https://www.umbra.com","contact":"customerservice@umbra.com"},
    {"no":1463,"category":"クリーニング・収納・整理","name":"Yamazaki Home Tower Slim Organizer Over Toilet","maker":"Yamazaki Home","ec_site":"Yamazaki公式サイト","prod_url":"https://yamazaki-store.com/collections/bathroom-storage","maker_hp":"https://yamazaki-store.com","contact":"support@yamazaki-store.com"},
    {"no":1464,"category":"クリーニング・収納・整理","name":"Honey-Can-Do Rolling Garment Rack with Shelves","maker":"Honey-Can-Do","ec_site":"HCD公式サイト","prod_url":"https://www.honeycando.com/collections/closet-storage","maker_hp":"https://www.honeycando.com","contact":"customercare@honeycando.com"},
    {"no":1465,"category":"クリーニング・収納・整理","name":"iDesign Eco Stackable Bin Organizer 4-Piece","maker":"iDesign","ec_site":"iDesign公式サイト","prod_url":"https://idesignliving.com/collections/eco-collection","maker_hp":"https://idesignliving.com","contact":"support@idesignliving.com"},
    # ── キッチン続き ──────────────────────────────────
    {"no":1466,"category":"キッチン・調理器具","name":"Yeti Hopper Flip 18 Soft Cooler Bag","maker":"YETI","ec_site":"YETI公式サイト","prod_url":"https://www.yeti.com/en_US/coolers/soft-coolers/hopper-flip-18","maker_hp":"https://www.yeti.com","contact":"info@yeti.com"},
    {"no":1467,"category":"キッチン・調理器具","name":"Igloo BMX 72qt Heavy Duty Cooler Ice Chest","maker":"Igloo Products","ec_site":"Igloo公式サイト","prod_url":"https://www.igloocoolers.com/collections/bmx","maker_hp":"https://www.igloocoolers.com","contact":"support@igloocoolers.com"},
    {"no":1468,"category":"キッチン・調理器具","name":"Oster BLSTVB-R00-000 Blender with XL Personal Jar","maker":"Oster","ec_site":"Oster公式サイト","prod_url":"https://www.oster.com/collections/blenders","maker_hp":"https://www.oster.com","contact":"consumer@oster.com"},
    {"no":1469,"category":"キッチン・調理器具","name":"Cuisinart MCP-12N Multiclad Pro Stainless Set 12pc","maker":"Cuisinart","ec_site":"Cuisinart公式サイト","prod_url":"https://www.cuisinart.com/collections/cookware","maker_hp":"https://www.cuisinart.com","contact":"consumer_service@cuisinart.com"},
    {"no":1470,"category":"キッチン・調理器具","name":"Instant Vortex Plus 6-in-1 Air Fryer XL 6Qt","maker":"Instant Brands","ec_site":"Instant公式サイト","prod_url":"https://www.instantbrands.com/collections/air-fryers","maker_hp":"https://www.instantbrands.com","contact":"support@instantbrands.com"},
    # ── スマートホーム続き ─────────────────────────────
    {"no":1471,"category":"スマートホーム・インテリア・照明","name":"Wink Hub 2 Smart Home Controller","maker":"Wink","ec_site":"Wink公式サイト","prod_url":"https://www.wink.com/products/wink-hub-2","maker_hp":"https://www.wink.com","contact":"support@wink.com"},
    {"no":1472,"category":"スマートホーム・インテリア・照明","name":"SmartThings Aeotec Hub v3 Smart Home","maker":"Aeotec","ec_site":"Aeotec公式サイト","prod_url":"https://aeotec.com/products/aeotec-smart-home-hub","maker_hp":"https://aeotec.com","contact":"support@aeotec.com"},
    {"no":1473,"category":"スマートホーム・インテリア・照明","name":"Brilliant Control 4-Switch Smart Home Panel","maker":"Brilliant","ec_site":"Brilliant公式サイト","prod_url":"https://www.brilliant.tech/products/4-switch-panel","maker_hp":"https://www.brilliant.tech","contact":"hello@brilliant.tech"},
    {"no":1474,"category":"スマートホーム・インテリア・照明","name":"Dreo Cruiser Pro T1 Smart Tower Fan","maker":"Dreo","ec_site":"Dreo公式サイト","prod_url":"https://www.dreo.com/products/tower-fans","maker_hp":"https://www.dreo.com","contact":"support@dreo.com"},
    {"no":1475,"category":"スマートホーム・インテリア・照明","name":"Vornado ATOM AE 1-in-1 Air Purifier Fan","maker":"Vornado","ec_site":"Vornado公式サイト","prod_url":"https://www.vornado.com/collections/air-purifiers","maker_hp":"https://www.vornado.com","contact":"info@vornado.com"},
    # ── ウェアラブル続き ───────────────────────────────
    {"no":1476,"category":"ウェアラブル・ヘルス・フィットネス","name":"Amazfit GTR 4 Smartwatch 14-Day Battery","maker":"Amazfit","ec_site":"Amazfit公式サイト","prod_url":"https://www.amazfit.com/en/gtr-4.html","maker_hp":"https://www.amazfit.com","contact":"support@amazfit.com"},
    {"no":1477,"category":"ウェアラブル・ヘルス・フィットネス","name":"Mobvoi TicWatch Pro 5 Wear OS Smartwatch","maker":"Mobvoi","ec_site":"Mobvoi公式サイト","prod_url":"https://www.mobvoi.com/us/products/ticwatchpro5","maker_hp":"https://www.mobvoi.com","contact":"support@mobvoi.com"},
    {"no":1478,"category":"ウェアラブル・ヘルス・フィットネス","name":"Zepp Health Amazfit T-Rex Ultra Rugged Smartwatch","maker":"Zepp Health","ec_site":"Amazfit公式サイト","prod_url":"https://www.amazfit.com/en/t-rex-ultra.html","maker_hp":"https://www.amazfit.com","contact":"support@amazfit.com"},
    {"no":1479,"category":"ウェアラブル・ヘルス・フィットネス","name":"Xiaomi Smart Band 9 Pro Fitness Tracker","maker":"Xiaomi","ec_site":"Xiaomi公式サイト","prod_url":"https://www.mi.com/us/product/mi-smart-band-9-pro","maker_hp":"https://www.mi.com","contact":"global-support@xiaomi.com"},
    {"no":1480,"category":"ウェアラブル・ヘルス・フィットネス","name":"Jabra Elite 10 Gen 2 True Wireless Earbuds","maker":"Jabra","ec_site":"Jabra公式サイト","prod_url":"https://www.jabra.com/bluetooth-headsets/jabra-elite-10","maker_hp":"https://www.jabra.com","contact":"support@jabra.com"},
    # ── アウトドア続き ─────────────────────────────────
    {"no":1481,"category":"アウトドア・スポーツ・旅行","name":"Nemo Tensor Elite Ultralight Sleeping Pad Regular","maker":"NEMO Equipment","ec_site":"NEMO公式サイト","prod_url":"https://www.nemoequipment.com/collections/sleeping-pads","maker_hp":"https://www.nemoequipment.com","contact":"info@nemoequipment.com"},
    {"no":1482,"category":"アウトドア・スポーツ・旅行","name":"Western Mountaineering Ultralite 6-Degree Sleeping Bag","maker":"Western Mountaineering","ec_site":"WM公式サイト","prod_url":"https://www.westernmountaineering.com/sleeping-bags","maker_hp":"https://www.westernmountaineering.com","contact":"info@westernmountaineering.com"},
    {"no":1483,"category":"アウトドア・スポーツ・旅行","name":"Sitka Gear Kelvin Active Jacket Insulated","maker":"Sitka Gear","ec_site":"Sitka公式サイト","prod_url":"https://www.sitkagear.com/collections/insulated-jackets","maker_hp":"https://www.sitkagear.com","contact":"customerservice@sitkagear.com"},
    {"no":1484,"category":"アウトドア・スポーツ・旅行","name":"Mammut Eiger Speed SO Hooded Jacket Softshell","maker":"Mammut","ec_site":"Mammut公式サイト","prod_url":"https://www.mammut.com/us/en/c/all-jackets","maker_hp":"https://www.mammut.com","contact":"info@mammut.com"},
    {"no":1485,"category":"アウトドア・スポーツ・旅行","name":"Suunto Core Outdoor Watch Altimeter Compass","maker":"Suunto","ec_site":"Suunto公式サイト","prod_url":"https://www.suunto.com/en-us/Products/sports-watches/suunto-core/","maker_hp":"https://www.suunto.com","contact":"support@suunto.com"},
    # ── ペット続き ─────────────────────────────────────
    {"no":1486,"category":"ペット用品","name":"Diggs Revol Dog Crate Collapsible Travel","maker":"Diggs","ec_site":"Diggs公式サイト","prod_url":"https://www.diggs.pet/products/revol-dog-crate","maker_hp":"https://www.diggs.pet","contact":"hello@diggs.pet"},
    {"no":1487,"category":"ペット用品","name":"Scollar Personalized Smart Dog Collar","maker":"Scollar","ec_site":"Scollar公式サイト","prod_url":"https://www.scollar.com/products/scollar","maker_hp":"https://www.scollar.com","contact":"support@scollar.com"},
    {"no":1488,"category":"ペット用品","name":"Feliway Optimum Enhanced Calm Diffuser Kit","maker":"CEVA Animal Health","ec_site":"Feliway公式サイト","prod_url":"https://www.feliway.com/us/Products/FELIWAY-Optimum","maker_hp":"https://www.feliway.com","contact":"info@feliway.com"},
    {"no":1489,"category":"ペット用品","name":"Ruffwear Switchbak Harness Multi-Use Dog","maker":"Ruffwear","ec_site":"Ruffwear公式サイト","prod_url":"https://ruffwear.com/collections/dog-harnesses","maker_hp":"https://ruffwear.com","contact":"support@ruffwear.com"},
    {"no":1490,"category":"ペット用品","name":"Freeze Dried Raw Boost Mixers Dog Food Topper","maker":"Instinct Pet Food","ec_site":"Instinct公式サイト","prod_url":"https://www.instinctpetfood.com/collections/mixers","maker_hp":"https://www.instinctpetfood.com","contact":"help@instinctpetfood.com"},
    # ── テクノロジー続き ──────────────────────────────
    {"no":1491,"category":"テクノロジー・ガジェット","name":"Bambu Lab A1 Mini Combo 3D Printer","maker":"Bambu Lab","ec_site":"Bambu Lab公式サイト","prod_url":"https://bambulab.com/en-us/a1-mini","maker_hp":"https://bambulab.com","contact":"support@bambulab.com"},
    {"no":1492,"category":"テクノロジー・ガジェット","name":"Prusa MK4S 3D Printer Fully Assembled","maker":"Prusa Research","ec_site":"Prusa公式サイト","prod_url":"https://www.prusa3d.com/product/prusa-mk4s-3d-printer","maker_hp":"https://www.prusa3d.com","contact":"info@prusa3d.com"},
    {"no":1493,"category":"テクノロジー・ガジェット","name":"Elegoo Saturn 4 Ultra 16K Resin 3D Printer","maker":"Elegoo","ec_site":"Elegoo公式サイト","prod_url":"https://www.elegoo.com/products/saturn-4-ultra","maker_hp":"https://www.elegoo.com","contact":"support@elegoo.com"},
    {"no":1494,"category":"テクノロジー・ガジェット","name":"Anycubic Photon Mono M7 Pro 3D Resin Printer","maker":"Anycubic","ec_site":"Anycubic公式サイト","prod_url":"https://www.anycubic.com/collections/photon-mono","maker_hp":"https://www.anycubic.com","contact":"support@anycubic.com"},
    {"no":1495,"category":"テクノロジー・ガジェット","name":"xTool F2 Ultra 20W Enclosed Fiber Laser Engraver","maker":"xTool","ec_site":"xTool公式サイト","prod_url":"https://www.xtool.com/products/xtool-f2-ultra","maker_hp":"https://www.xtool.com","contact":"support@xtool.com"},
    # ── 美容続き ─────────────────────────────────────
    {"no":1496,"category":"美容・スキンケア","name":"Joanna Vargas Twilight Eye Cream Retinol Peptide","maker":"Joanna Vargas","ec_site":"Joanna Vargas公式サイト","prod_url":"https://joannavargas.com/collections/eye-treatment","maker_hp":"https://joannavargas.com","contact":"info@joannavargas.com"},
    {"no":1497,"category":"美容・スキンケア","name":"Augustinus Bader The Rich Cream 50ml","maker":"Augustinus Bader","ec_site":"AB公式サイト","prod_url":"https://augustinusbader.com/us/en/the-rich-cream","maker_hp":"https://augustinusbader.com","contact":"info@augustinusbader.com"},
    {"no":1498,"category":"美容・スキンケア","name":"Retrouvé Intensive Replenishing Facial Oil","maker":"Retrouvé","ec_site":"Retrouvé公式サイト","prod_url":"https://retrouve.com/collections/face-oils","maker_hp":"https://retrouve.com","contact":"hello@retrouve.com"},
    {"no":1499,"category":"美容・スキンケア","name":"Westman Atelier Vital Skincare Complexion Drops","maker":"Westman Atelier","ec_site":"Westman公式サイト","prod_url":"https://westmanatelier.com/collections/face","maker_hp":"https://westmanatelier.com","contact":"info@westmanatelier.com"},
    {"no":1500,"category":"美容・スキンケア","name":"Tower 28 Beauty SOS Daily Rescue Facial Spray","maker":"Tower 28 Beauty","ec_site":"Tower 28公式サイト","prod_url":"https://tower28beauty.com/collections/face","maker_hp":"https://tower28beauty.com","contact":"hello@tower28beauty.com"},
    # ── 子供続き ─────────────────────────────────────
    {"no":1501,"category":"子供・教育","name":"Brackitz Pulley STEM Building Toys 60pc","maker":"Brackitz","ec_site":"Brackitz公式サイト","prod_url":"https://brackitz.com/collections/building-toys","maker_hp":"https://brackitz.com","contact":"info@brackitz.com"},
    # ── ファッション続き ───────────────────────────────
    {"no":1502,"category":"ファッション・アクセサリー","name":"Waterfield Designs Bolt Laptop Bag Slim Sleeve","maker":"WaterField Designs","ec_site":"WaterField公式サイト","prod_url":"https://www.sfbags.com/collections/laptop-sleeves","maker_hp":"https://www.sfbags.com","contact":"info@sfbags.com"},
    # ── クリーニング続き ──────────────────────────────
    {"no":1503,"category":"クリーニング・収納・整理","name":"Electrolux WM40 Wascator Washing Machine Portable","maker":"Electrolux","ec_site":"Electrolux公式サイト","prod_url":"https://www.electrolux.com/en/washing-machines/","maker_hp":"https://www.electrolux.com","contact":"consumer@electrolux.com"},
    {"no":1504,"category":"クリーニング・収納・整理","name":"Boon Clutch Dispensing Spoon Baby Feeding Set","maker":"Boon","ec_site":"Boon公式サイト","prod_url":"https://www.boonhome.com/collections/feeding","maker_hp":"https://www.boonhome.com","contact":"support@boonhome.com"},
    {"no":1505,"category":"クリーニング・収納・整理","name":"Sistema Nest IT 5-Piece Container Set Stackable","maker":"Sistema Plastics","ec_site":"Sistema公式サイト","prod_url":"https://www.sistema.co.nz/collections/nest-it","maker_hp":"https://www.sistema.co.nz","contact":"info@sistema.co.nz"},
]

# ===================== Styles =====================
cat_fills = {
    "キッチン・調理器具":             PatternFill("solid", start_color="2E75B6"),
    "スマートホーム・インテリア・照明": PatternFill("solid", start_color="538135"),
    "ウェアラブル・ヘルス・フィットネス": PatternFill("solid", start_color="7030A0"),
    "アウトドア・スポーツ・旅行":       PatternFill("solid", start_color="C55A11"),
    "ペット用品":                      PatternFill("solid", start_color="843C0C"),
    "テクノロジー・ガジェット":         PatternFill("solid", start_color="244185"),
    "美容・スキンケア":                PatternFill("solid", start_color="C00000"),
    "子供・教育":                      PatternFill("solid", start_color="375623"),
    "ファッション・アクセサリー":       PatternFill("solid", start_color="7B3F00"),
    "クリーニング・収納・整理":         PatternFill("solid", start_color="404040"),
}
white_font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
normal_font = Font(name="Arial", size=10)
link_font   = Font(name="Arial", size=10, color="0563C1", underline="single")
thin        = Side(style="thin", color="CCCCCC")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)
alt_fill    = PatternFill("solid", start_color="EBF3FB")
center_va   = Alignment(vertical="center", wrap_text=True)

wb = load_workbook(r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸1305件.xlsx")
ws = wb.active

for p in new_products:
    row_num = p["no"] + 1
    bg = alt_fill if row_num % 2 == 0 else None
    cat = p["category"]
    cat_fill = cat_fills.get(cat)
    contact_val = str(p["contact"]).strip()
    has_email = "@" in contact_val and not contact_val.startswith("http") and "問い合わせ" not in contact_val
    values = [p["no"], cat, p["name"], p["maker"], p["ec_site"], p["prod_url"], p["maker_hp"], contact_val]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.border = border
        cell.alignment = center_va
        if col_idx == 2:
            if cat_fill:
                cell.fill = cat_fill
                cell.font = white_font
            else:
                cell.font = normal_font
        elif col_idx in (6, 7) and val and str(val).startswith("http"):
            cell.hyperlink = str(val)
            cell.font = link_font
            if bg:
                cell.fill = bg
        elif col_idx == 8:
            if has_email:
                gmail_url = build_gmail(contact_val, p["maker"], p["name"])
                cell.hyperlink = gmail_url
                cell.font = link_font
            if bg:
                cell.fill = bg
        else:
            cell.font = normal_font
            if bg and col_idx != 2:
                cell.fill = bg

output_path = r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸1505件.xlsx"
wb.save(output_path)
print(f"完了: {output_path}")
print(f"追加件数: {len(new_products)} 件 (No.1306〜1505)")
email_count = sum(1 for p in new_products if "@" in str(p["contact"]) and not str(p["contact"]).startswith("http"))
print(f"メールアドレスあり: {email_count} 件 ({email_count*100//len(new_products)}%)")
