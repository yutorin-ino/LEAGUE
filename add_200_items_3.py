from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote

def build_gmail(email, maker_name, product_name):
    subject = "Potential Distribution Partnership for Japan"
    body = f"""Dear {maker_name} Team,

My name is Hiroyuki Inoguchi from
Sumai pluS Co., Ltd. in Japan

I hope this message finds you all well.

Our company focuses on promoting products that enrich people's daily lives. We are currently seeking unique international brands to introduce to the Japanese market and support their growth.

I recently had the opportunity to review your product ({product_name}) and was very impressed by its potential in the Japanese market.

I am confident that your product will strongly appeal to Japanese customers, and I would like to explore the possibility of building a partnership with your company to help ensure its success.

Furthermore, our company has a proven track record of promoting overseas products through our proprietary sales network and Japanese distribution channels.

Would it be possible to schedule a brief online meeting sometime next week to discuss this matter?

I look forward to hearing from you.


Sincerely,

Hiroyuki Inoguchi
Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )

================================================

Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )
Hiroyuki Inoguchi

E-mail:yutorin.ino@gmail.com

Address: 49-5 Kitazakuno, Higashi-Itsushiro, Ichinomiya City, Aichi Prefecture, Japan

================================================="""
    sender = quote("yutorin.ino@gmail.com", safe='')
    return f"https://mail.google.com/mail/?authuser={sender}&view=cm&fs=1&to={quote(email,safe='')}&su={quote(subject,safe='')}&body={quote(body,safe='')}"

new_products = [
    # ── キッチン・調理器具 ──────────────────────────────
    {"no":1106,"category":"キッチン・調理器具","name":"Hestan Cue Smart Induction Burner","maker":"Hestan Smart Cooking","ec_site":"Hestan公式サイト","prod_url":"https://hestancue.com/products/induction-burner","maker_hp":"https://hestancue.com","contact":"support@hestancue.com"},
    {"no":1107,"category":"キッチン・調理器具","name":"Mellow Sous Vide Smart Cooker","maker":"Mellow","ec_site":"Mellow公式サイト","prod_url":"https://cookmellow.com/products/mellow","maker_hp":"https://cookmellow.com","contact":"hello@cookmellow.com"},
    {"no":1108,"category":"キッチン・調理器具","name":"Suvie Kitchen Robot 4-Zone Cooking","maker":"Suvie","ec_site":"Suvie公式サイト","prod_url":"https://www.suvie.com/products/suvie-kitchen-robot","maker_hp":"https://www.suvie.com","contact":"support@suvie.com"},
    {"no":1109,"category":"キッチン・調理器具","name":"Pico Brew Craft Beer Brewing Machine","maker":"PicoBrew","ec_site":"PicoBrew公式サイト","prod_url":"https://picobrew.com/products/picoc","maker_hp":"https://picobrew.com","contact":"support@picobrew.com"},
    {"no":1110,"category":"キッチン・調理器具","name":"ChefSteps Joule Turbo Sous Vide","maker":"ChefSteps","ec_site":"ChefSteps公式サイト","prod_url":"https://www.chefsteps.com/joule","maker_hp":"https://www.chefsteps.com","contact":"support@chefsteps.com"},
    {"no":1111,"category":"キッチン・調理器具","name":"Balmuda The Toaster Steam Oven","maker":"Balmuda","ec_site":"Balmuda公式サイト","prod_url":"https://www.balmuda.com/en/toaster","maker_hp":"https://www.balmuda.com","contact":"support@balmuda.com"},
    {"no":1112,"category":"キッチン・調理器具","name":"Typhur Sync Gold Wireless Thermometer","maker":"Typhur","ec_site":"Typhur公式サイト","prod_url":"https://www.typhur.com/products/sync-gold","maker_hp":"https://www.typhur.com","contact":"support@typhur.com"},
    {"no":1113,"category":"キッチン・調理器具","name":"CHEF iQ Smart Pressure Cooker","maker":"CHEF iQ","ec_site":"CHEF iQ公式サイト","prod_url":"https://chefiq.com/products/smart-cooker","maker_hp":"https://chefiq.com","contact":"support@chefiq.com"},
    {"no":1114,"category":"キッチン・調理器具","name":"Furo Smart Electric Kettle App Control","maker":"Furo","ec_site":"Amazon","prod_url":"https://www.amazon.com/s?k=Furo+smart+kettle","maker_hp":"https://furoliving.com","contact":"hello@furoliving.com"},
    {"no":1115,"category":"キッチン・調理器具","name":"Crux Artisan Series Stand Mixer 5.5Qt","maker":"Crux Kitchen","ec_site":"Crux公式サイト","prod_url":"https://cruxkitchen.com/collections/mixers","maker_hp":"https://cruxkitchen.com","contact":"info@cruxkitchen.com"},
    {"no":1116,"category":"キッチン・調理器具","name":"W&P Porter Ceramic Reusable Mug","maker":"W&P Design","ec_site":"W&P公式サイト","prod_url":"https://wandpdesign.com/collections/mugs","maker_hp":"https://wandpdesign.com","contact":"hello@wandpdesign.com"},
    {"no":1117,"category":"キッチン・調理器具","name":"Caraway Maison Non-Toxic Ceramic Cookware Set","maker":"Caraway Home","ec_site":"Caraway公式サイト","prod_url":"https://www.carawayhome.com/products/cookware-set","maker_hp":"https://www.carawayhome.com","contact":"support@carawayhome.com"},
    {"no":1118,"category":"キッチン・調理器具","name":"Our Place Always Pan 2.0 Ceramic","maker":"Our Place","ec_site":"Our Place公式サイト","prod_url":"https://fromourplace.com/products/always-pan","maker_hp":"https://fromourplace.com","contact":"support@fromourplace.com"},
    {"no":1119,"category":"キッチン・調理器具","name":"Smeg 50s Retro Style Drip Coffee Maker","maker":"Smeg","ec_site":"Smeg公式サイト","prod_url":"https://www.smeg.com/en-us/kitchen/coffee-makers/","maker_hp":"https://www.smeg.com","contact":"info@smeg.com"},
    {"no":1120,"category":"キッチン・調理器具","name":"Spinn Coffee Machine Wi-Fi Centrifugal Brewer","maker":"Spinn Coffee","ec_site":"Spinn公式サイト","prod_url":"https://www.spinn.com/products/spinn-coffee-machine","maker_hp":"https://www.spinn.com","contact":"hello@spinn.com"},
    {"no":1121,"category":"キッチン・調理器具","name":"Bartesian Premium Cocktail Machine","maker":"Bartesian","ec_site":"Bartesian公式サイト","prod_url":"https://bartesian.com/products/bartesian-premium-cocktail-machine","maker_hp":"https://bartesian.com","contact":"support@bartesian.com"},
    {"no":1122,"category":"キッチン・調理器具","name":"Dreo ChefMaker Combi Cooker","maker":"Dreo","ec_site":"Dreo公式サイト","prod_url":"https://www.dreo.com/products/chefmaker-combi-cooker","maker_hp":"https://www.dreo.com","contact":"support@dreo.com"},
    {"no":1123,"category":"キッチン・調理器具","name":"Bedsure Silicone Baking Mat Non-Stick Set","maker":"Bedsure","ec_site":"Bedsure公式サイト","prod_url":"https://bedsure.com/collections/kitchen","maker_hp":"https://bedsure.com","contact":"support@bedsure.com"},
    {"no":1124,"category":"キッチン・調理器具","name":"Dash Rapid Egg Cooker 6 Capacity","maker":"Dash","ec_site":"Dash公式サイト","prod_url":"https://www.bydash.com/products/rapid-egg-cooker","maker_hp":"https://www.bydash.com","contact":"support@bydash.com"},
    {"no":1125,"category":"キッチン・調理器具","name":"Vitamix A3500 Ascent Series Smart Blender","maker":"Vitamix","ec_site":"Vitamix公式サイト","prod_url":"https://www.vitamix.com/us/en_us/shop/a3500","maker_hp":"https://www.vitamix.com","contact":"consumer@vitamix.com"},
    # ── スマートホーム・インテリア・照明 ───────────────────
    {"no":1126,"category":"スマートホーム・インテリア・照明","name":"Brilliant All-in-One Smart Home Control 4-Switch","maker":"Brilliant","ec_site":"Brilliant公式サイト","prod_url":"https://www.brilliant.tech/products/4-switch-panel","maker_hp":"https://www.brilliant.tech","contact":"hello@brilliant.tech"},
    {"no":1127,"category":"スマートホーム・インテリア・照明","name":"Arlo Pro 5S 2K Wireless Security Camera","maker":"Arlo Technologies","ec_site":"Arlo公式サイト","prod_url":"https://www.arlo.com/en-us/cameras/pro/arlo-pro-5s-2k.html","maker_hp":"https://www.arlo.com","contact":"support@arlo.com"},
    {"no":1128,"category":"スマートホーム・インテリア・照明","name":"Eufy Security Video Doorbell E340 Dual Camera","maker":"Eufy","ec_site":"Eufy公式サイト","prod_url":"https://www.eufy.com/products/e8213cd1","maker_hp":"https://www.eufy.com","contact":"support@eufylife.com"},
    {"no":1129,"category":"スマートホーム・インテリア・照明","name":"Mysa Smart Thermostat for Electric Baseboard","maker":"Mysa","ec_site":"Mysa公式サイト","prod_url":"https://getmysa.com/products/mysa-for-electric-baseboard","maker_hp":"https://getmysa.com","contact":"support@getmysa.com"},
    {"no":1130,"category":"スマートホーム・インテリア・照明","name":"Noon Room Director Smart Dimmer System","maker":"Noon Home","ec_site":"Noon公式サイト","prod_url":"https://noonhome.com/products/room-director","maker_hp":"https://noonhome.com","contact":"hello@noonhome.com"},
    {"no":1131,"category":"スマートホーム・インテリア・照明","name":"Lutron Caseta Wireless Smart Lighting Starter Kit","maker":"Lutron Electronics","ec_site":"Lutron公式サイト","prod_url":"https://www.casetawireless.com/products/starter-kits","maker_hp":"https://www.lutron.com","contact":"cs@lutron.com"},
    {"no":1132,"category":"スマートホーム・インテリア・照明","name":"Leviton Decora Smart Wi-Fi 2nd Gen Dimmer","maker":"Leviton","ec_site":"Leviton公式サイト","prod_url":"https://www.leviton.com/en/products/smart-home/dimmers","maker_hp":"https://www.leviton.com","contact":"customerservice@leviton.com"},
    {"no":1133,"category":"スマートホーム・インテリア・照明","name":"Wemo Stage Scene Controller Smart Button","maker":"Wemo","ec_site":"Wemo公式サイト","prod_url":"https://www.wemo.com/collections/smart-plugs","maker_hp":"https://www.wemo.com","contact":"support@wemo.com"},
    {"no":1134,"category":"スマートホーム・インテリア・照明","name":"Wyze Lock Bolt Fingerprint Smart Lock","maker":"Wyze Labs","ec_site":"Wyze公式サイト","prod_url":"https://www.wyze.com/products/wyze-lock-bolt","maker_hp":"https://www.wyze.com","contact":"support@wyze.com"},
    {"no":1135,"category":"スマートホーム・インテリア・照明","name":"Schlage Encode Plus Smart WiFi Deadbolt","maker":"Schlage","ec_site":"Schlage公式サイト","prod_url":"https://www.schlage.com/en/home/smart-locks/encode-plus.html","maker_hp":"https://www.schlage.com","contact":"consumer@allegion.com"},
    {"no":1136,"category":"スマートホーム・インテリア・照明","name":"Lasso Loop Recycling Smart Home Recycler","maker":"Lasso Loop","ec_site":"Lasso公式サイト","prod_url":"https://lassoloop.com/products/lasso-loop-recycling","maker_hp":"https://lassoloop.com","contact":"hello@lassoloop.com"},
    {"no":1137,"category":"スマートホーム・インテリア・照明","name":"Moen Flo Smart Water Shutoff","maker":"Moen","ec_site":"Moen公式サイト","prod_url":"https://www.moen.com/flo","maker_hp":"https://www.moen.com","contact":"consumer.service@moen.com"},
    {"no":1138,"category":"スマートホーム・インテリア・照明","name":"Phyn Plus Smart Water Assistant","maker":"Phyn","ec_site":"Phyn公式サイト","prod_url":"https://www.phyn.com/products/phyn-plus","maker_hp":"https://www.phyn.com","contact":"support@phyn.com"},
    {"no":1139,"category":"スマートホーム・インテリア・照明","name":"Canary Pro Indoor Security Camera","maker":"Canary","ec_site":"Canary公式サイト","prod_url":"https://canary.is/products/canary-pro","maker_hp":"https://canary.is","contact":"support@canary.is"},
    {"no":1140,"category":"スマートホーム・インテリア・照明","name":"Ambient Weather WS-2902 Smart Weather Station","maker":"Ambient Weather","ec_site":"Ambient Weather公式サイト","prod_url":"https://www.ambientweather.com/amws2902.html","maker_hp":"https://www.ambientweather.com","contact":"support@ambientweather.com"},
    {"no":1141,"category":"スマートホーム・インテリア・照明","name":"iDevices Instinct Smart Light Switch Alexa Built-in","maker":"iDevices","ec_site":"iDevices公式サイト","prod_url":"https://idevicesinc.com/products/instinct","maker_hp":"https://idevicesinc.com","contact":"support@idevicesinc.com"},
    {"no":1142,"category":"スマートホーム・インテリア・照明","name":"Orbit B-Hyve Smart 12-Zone Sprinkler Controller","maker":"Orbit","ec_site":"Orbit公式サイト","prod_url":"https://bhyve.orbitonline.com/products/12-zone-smart-sprinkler-controller","maker_hp":"https://orbitonline.com","contact":"support@orbitonline.com"},
    {"no":1143,"category":"スマートホーム・インテリア・照明","name":"Ecolink Z-Wave Plus Motion Detector","maker":"Ecolink","ec_site":"Ecolink公式サイト","prod_url":"https://ecolink.com/products/z-wave-plus-motion-detector","maker_hp":"https://ecolink.com","contact":"support@ecolink.com"},
    {"no":1144,"category":"スマートホーム・インテリア・照明","name":"Adurosmart ERIA Smart Lighting Starter Kit","maker":"AduroSmart","ec_site":"AduroSmart公式サイト","prod_url":"https://adurosmart.com/collections/eria","maker_hp":"https://adurosmart.com","contact":"info@adurosmart.com"},
    {"no":1145,"category":"スマートホーム・インテリア・照明","name":"Kasa Smart EP40 Outdoor Smart Plug 2-Outlet","maker":"Kasa Smart","ec_site":"Kasa公式サイト","prod_url":"https://www.kasasmart.com/us/products/smart-plugs/kasa-smart-wi-fi-outdoor-plug-ep40","maker_hp":"https://www.kasasmart.com","contact":"support@tp-link.com"},
    # ── ウェアラブル・ヘルス・フィットネス ──────────────────
    {"no":1146,"category":"ウェアラブル・ヘルス・フィットネス","name":"Muse S Gen 2 Brain Sensing Meditation Headband","maker":"InteraXon","ec_site":"Muse公式サイト","prod_url":"https://choosemuse.com/products/muse-s-gen-2","maker_hp":"https://choosemuse.com","contact":"support@choosemuse.com"},
    {"no":1147,"category":"ウェアラブル・ヘルス・フィットネス","name":"Flowtime Biosensing Meditation Headset","maker":"Flowtime","ec_site":"Flowtime公式サイト","prod_url":"https://www.flowtime.cn/en/product","maker_hp":"https://www.flowtime.cn","contact":"support@flowtime.cn"},
    {"no":1148,"category":"ウェアラブル・ヘルス・フィットネス","name":"Biostrap EVO Wristband Health Tracker","maker":"Biostrap","ec_site":"Biostrap公式サイト","prod_url":"https://biostrap.com/products/biostrap-evo-set","maker_hp":"https://biostrap.com","contact":"support@biostrap.com"},
    {"no":1149,"category":"ウェアラブル・ヘルス・フィットネス","name":"Polar Grit X2 Pro Outdoor GPS Smartwatch","maker":"Polar Electro","ec_site":"Polar公式サイト","prod_url":"https://www.polar.com/en/grit-x2-pro","maker_hp":"https://www.polar.com","contact":"support@polar.com"},
    {"no":1150,"category":"ウェアラブル・ヘルス・フィットネス","name":"Suunto Race S Solar GPS Multisport Watch","maker":"Suunto","ec_site":"Suunto公式サイト","prod_url":"https://www.suunto.com/en-us/Products/sports-watches/suunto-race-s/","maker_hp":"https://www.suunto.com","contact":"support@suunto.com"},
    {"no":1151,"category":"ウェアラブル・ヘルス・フィットネス","name":"Movella Xsens DOT Wireless Motion Sensor","maker":"Movella","ec_site":"Movella公式サイト","prod_url":"https://www.movella.com/products/wearables/xsens-dot","maker_hp":"https://www.movella.com","contact":"info@movella.com"},
    {"no":1152,"category":"ウェアラブル・ヘルス・フィットネス","name":"Halo Rise Sleep Tracker Bedside Device","maker":"Amazon Halo","ec_site":"Amazon公式サイト","prod_url":"https://www.amazon.com/dp/B09K36WNDP","maker_hp":"https://www.amazon.com/halo","contact":"halo-feedback@amazon.com"},
    {"no":1153,"category":"ウェアラブル・ヘルス・フィットネス","name":"Eight Sleep Pod 4 Smart Mattress Cover","maker":"Eight Sleep","ec_site":"Eight Sleep公式サイト","prod_url":"https://www.eightsleep.com/eight-pod-smart-mattress/","maker_hp":"https://www.eightsleep.com","contact":"support@eightsleep.com"},
    {"no":1154,"category":"ウェアラブル・ヘルス・フィットネス","name":"Dreem 3 EEG Sleep Headband","maker":"Dreem","ec_site":"Dreem公式サイト","prod_url":"https://dreem.com/en/products/dreem-3","maker_hp":"https://dreem.com","contact":"contact@dreem.com"},
    {"no":1155,"category":"ウェアラブル・ヘルス・フィットネス","name":"Sensate 2 Vagus Nerve Stimulation Wearable","maker":"Sensate","ec_site":"Sensate公式サイト","prod_url":"https://www.getsensate.com/products/sensate-2","maker_hp":"https://www.getsensate.com","contact":"hello@getsensate.com"},
    {"no":1156,"category":"ウェアラブル・ヘルス・フィットネス","name":"Pavlok 3 Habit-Breaking Wristband","maker":"Pavlok","ec_site":"Pavlok公式サイト","prod_url":"https://pavlok.com/products/pavlok-3","maker_hp":"https://pavlok.com","contact":"support@pavlok.com"},
    {"no":1157,"category":"ウェアラブル・ヘルス・フィットネス","name":"Halo Band Activity Tracker Health Coaching","maker":"Halo Wearables","ec_site":"Halo公式サイト","prod_url":"https://halowearables.com","maker_hp":"https://halowearables.com","contact":"info@halowearables.com"},
    {"no":1158,"category":"ウェアラブル・ヘルス・フィットネス","name":"Garmin Vivosmart 5 Slim Health Tracker","maker":"Garmin","ec_site":"Garmin公式サイト","prod_url":"https://www.garmin.com/en-US/p/vivosmart-5","maker_hp":"https://www.garmin.com","contact":"support@garmin.com"},
    {"no":1159,"category":"ウェアラブル・ヘルス・フィットネス","name":"Wahoo TICKR X Heart Rate Monitor","maker":"Wahoo Fitness","ec_site":"Wahoo公式サイト","prod_url":"https://www.wahoofitness.com/devices/heart-rate-monitors/tickr-x","maker_hp":"https://www.wahoofitness.com","contact":"support@wahoofitness.com"},
    {"no":1160,"category":"ウェアラブル・ヘルス・フィットネス","name":"Myontec Mbody 3 Smart Compression Shorts","maker":"Myontec","ec_site":"Myontec公式サイト","prod_url":"https://www.myontec.com/en/products/","maker_hp":"https://www.myontec.com","contact":"info@myontec.com"},
    # ── アウトドア・スポーツ・旅行 ────────────────────────
    {"no":1161,"category":"アウトドア・スポーツ・旅行","name":"Goal Zero Yeti 500X Portable Power Station","maker":"Goal Zero","ec_site":"Goal Zero公式サイト","prod_url":"https://www.goalzero.com/products/goal-zero-yeti-500x","maker_hp":"https://www.goalzero.com","contact":"support@goalzero.com"},
    {"no":1162,"category":"アウトドア・スポーツ・旅行","name":"Nomad Base Station Hub Edition Wireless Charger","maker":"Nomad Goods","ec_site":"Nomad公式サイト","prod_url":"https://nomadgoods.com/products/base-station-hub","maker_hp":"https://nomadgoods.com","contact":"support@nomadgoods.com"},
    {"no":1163,"category":"アウトドア・スポーツ・旅行","name":"Hyperlite Mountain Gear 2400 Southwest Pack","maker":"Hyperlite Mountain Gear","ec_site":"HMG公式サイト","prod_url":"https://www.hyperlitemountaingear.com/collections/packs","maker_hp":"https://www.hyperlitemountaingear.com","contact":"info@hyperlitemountaingear.com"},
    {"no":1164,"category":"アウトドア・スポーツ・旅行","name":"Matador NanoDry Trek Towel Large","maker":"Matador","ec_site":"Matador公式サイト","prod_url":"https://matadorup.com/collections/towels","maker_hp":"https://matadorup.com","contact":"support@matadorup.com"},
    {"no":1165,"category":"アウトドア・スポーツ・旅行","name":"Helinox Chair Zero Ultralight Camp Chair","maker":"Helinox","ec_site":"Helinox公式サイト","prod_url":"https://helinox.com/collections/chairs/products/chair-zero","maker_hp":"https://helinox.com","contact":"info@helinox.com"},
    {"no":1166,"category":"アウトドア・スポーツ・旅行","name":"Primus Lite+ Backpacking Stove System","maker":"Primus","ec_site":"Primus公式サイト","prod_url":"https://www.primus.eu/en/products/stoves/backpacking-stoves","maker_hp":"https://www.primus.eu","contact":"info@primus.se"},
    {"no":1167,"category":"アウトドア・スポーツ・旅行","name":"MSR Hubba Hubba Bikepack 2-Person Tent","maker":"MSR","ec_site":"MSR公式サイト","prod_url":"https://www.msrgear.com/tents","maker_hp":"https://www.msrgear.com","contact":"info@msrgear.com"},
    {"no":1168,"category":"アウトドア・スポーツ・旅行","name":"Zpacks Duplex Flex Ultralight Two Person Tent","maker":"Zpacks","ec_site":"Zpacks公式サイト","prod_url":"https://zpacks.com/products/duplex-flex-tent","maker_hp":"https://zpacks.com","contact":"info@zpacks.com"},
    {"no":1169,"category":"アウトドア・スポーツ・旅行","name":"Kammock Roo Double Camping Hammock","maker":"Kammock","ec_site":"Kammock公式サイト","prod_url":"https://www.kammock.com/products/roo-double","maker_hp":"https://www.kammock.com","contact":"support@kammock.com"},
    {"no":1170,"category":"アウトドア・スポーツ・旅行","name":"Rumpl Original Puffy Blanket","maker":"Rumpl","ec_site":"Rumpl公式サイト","prod_url":"https://rumpl.com/collections/puffy-blankets","maker_hp":"https://rumpl.com","contact":"hello@rumpl.com"},
    {"no":1171,"category":"アウトドア・スポーツ・旅行","name":"Loki Basecamp 4-Season Dog Tent","maker":"Loki Products","ec_site":"Loki公式サイト","prod_url":"https://lokiproducts.com/products/basecamp","maker_hp":"https://lokiproducts.com","contact":"info@lokiproducts.com"},
    {"no":1172,"category":"アウトドア・スポーツ・旅行","name":"Cotopaxi Allpa 35L Travel Pack","maker":"Cotopaxi","ec_site":"Cotopaxi公式サイト","prod_url":"https://www.cotopaxi.com/collections/travel-bags","maker_hp":"https://www.cotopaxi.com","contact":"help@cotopaxi.com"},
    {"no":1173,"category":"アウトドア・スポーツ・旅行","name":"Enlightened Equipment Enigma Quilt 20F","maker":"Enlightened Equipment","ec_site":"EE公式サイト","prod_url":"https://enlightenedequipment.com/enigma/","maker_hp":"https://enlightenedequipment.com","contact":"info@enlightenedequipment.com"},
    {"no":1174,"category":"アウトドア・スポーツ・旅行","name":"Sawyer Products Squeeze Water Filter System","maker":"Sawyer Products","ec_site":"Sawyer公式サイト","prod_url":"https://sawyer.com/products/squeeze-water-filtration-system/","maker_hp":"https://sawyer.com","contact":"info@sawyer.com"},
    {"no":1175,"category":"アウトドア・スポーツ・旅行","name":"Black Diamond Spot 400-R Rechargeable Headlamp","maker":"Black Diamond","ec_site":"Black Diamond公式サイト","prod_url":"https://www.blackdiamondequipment.com/en_US/headlamps","maker_hp":"https://www.blackdiamondequipment.com","contact":"info@bdel.com"},
    # ── ペット用品 ─────────────────────────────────────
    {"no":1176,"category":"ペット用品","name":"Paws & Claws Automatic Ball Launcher","maker":"iFetch","ec_site":"iFetch公式サイト","prod_url":"https://www.goifetch.com/products","maker_hp":"https://www.goifetch.com","contact":"help@goifetch.com"},
    {"no":1177,"category":"ペット用品","name":"Wickedbone Smart Bone Auto Interactive Dog Toy","maker":"Wickedbone","ec_site":"Wickedbone公式サイト","prod_url":"https://www.wickedbone.com/products/wicked-bone","maker_hp":"https://www.wickedbone.com","contact":"support@wickedbone.com"},
    {"no":1178,"category":"ペット用品","name":"Petsafe Drinkwell Pagoda Ceramic Pet Fountain","maker":"PetSafe","ec_site":"PetSafe公式サイト","prod_url":"https://www.petsafe.net/drinkwell-pagoda-ceramic-pet-fountain","maker_hp":"https://www.petsafe.net","contact":"csc@petsafe.net"},
    {"no":1179,"category":"ペット用品","name":"Catit Pixi Smart 8-Meal Feeder","maker":"Catit","ec_site":"Catit公式サイト","prod_url":"https://catit.com/products/catit-pixi-smart-feeder","maker_hp":"https://catit.com","contact":"info@catit.com"},
    {"no":1180,"category":"ペット用品","name":"ScoopFree Smart Self-Cleaning Cat Litter Box","maker":"ScoopFree","ec_site":"ScoopFree公式サイト","prod_url":"https://www.scoopfree.com/products/scoopfree-smart-self-cleaning-cat-litter-box","maker_hp":"https://www.scoopfree.com","contact":"support@scoopfree.com"},
    {"no":1181,"category":"ペット用品","name":"Wopet Automatic Cat Feeder with Camera 6L","maker":"WOpet","ec_site":"WOpet公式サイト","prod_url":"https://www.wopet.com/collections/pet-feeder","maker_hp":"https://www.wopet.com","contact":"support@wopet.com"},
    {"no":1182,"category":"ペット用品","name":"PETKIT EVERSWEET 3 Pro Smart Pet Fountain","maker":"PETKIT","ec_site":"PETKIT公式サイト","prod_url":"https://www.petkit.com/collections/pet-fountain","maker_hp":"https://www.petkit.com","contact":"service@petkit.com"},
    {"no":1183,"category":"ペット用品","name":"Rollwalk eRW3 Motorized Dog Harness Leash","maker":"Rollwalk","ec_site":"Rollwalk公式サイト","prod_url":"https://rollwalk.com/products/erw3","maker_hp":"https://rollwalk.com","contact":"hello@rollwalk.com"},
    {"no":1184,"category":"ペット用品","name":"DoggyRide Mini20 Stroller for Pets","maker":"DoggyRide","ec_site":"DoggyRide公式サイト","prod_url":"https://www.doggyride.com/collections/strollers","maker_hp":"https://www.doggyride.com","contact":"info@doggyride.com"},
    {"no":1185,"category":"ペット用品","name":"Hepper Pod Bed Washable Cat Bed","maker":"Hepper","ec_site":"Hepper公式サイト","prod_url":"https://www.hepper.com/products/pod-bed","maker_hp":"https://www.hepper.com","contact":"hello@hepper.com"},
    {"no":1186,"category":"ペット用品","name":"Fluentpet HexTiles Dog Communication Buttons","maker":"FluentPet","ec_site":"FluentPet公式サイト","prod_url":"https://fluent.pet/products/hextile-starter-kit","maker_hp":"https://fluent.pet","contact":"support@fluent.pet"},
    {"no":1187,"category":"ペット用品","name":"Dogtap Digital Dog ID Tag NFC/QR","maker":"Dogtap","ec_site":"Dogtap公式サイト","prod_url":"https://www.dogtap.com/en/products","maker_hp":"https://www.dogtap.com","contact":"info@dogtap.com"},
    {"no":1188,"category":"ペット用品","name":"Pidan Cloud Cat Litter Box Enclosed","maker":"Pidan","ec_site":"Pidan公式サイト","prod_url":"https://www.pidan.com/collections/cat-litter-box","maker_hp":"https://www.pidan.com","contact":"service@pidan.com"},
    {"no":1189,"category":"ペット用品","name":"Hyper Pet Kannon K9 Fetch Ball Launcher","maker":"Hyper Pet","ec_site":"Hyper Pet公式サイト","prod_url":"https://hyperpet.com/collections/fetch","maker_hp":"https://hyperpet.com","contact":"support@hyperpet.com"},
    {"no":1190,"category":"ペット用品","name":"Barkio Dog Monitor App + Smart Camera Bundle","maker":"Barkio","ec_site":"Barkio公式サイト","prod_url":"https://barkio.com/products","maker_hp":"https://barkio.com","contact":"hello@barkio.com"},
    # ── テクノロジー・ガジェット ────────────────────────
    {"no":1191,"category":"テクノロジー・ガジェット","name":"Brilliant Smart Home Control 2-Switch Panel","maker":"Brilliant","ec_site":"Brilliant公式サイト","prod_url":"https://www.brilliant.tech/products/2-switch-panel","maker_hp":"https://www.brilliant.tech","contact":"hello@brilliant.tech"},
    {"no":1192,"category":"テクノロジー・ガジェット","name":"Clicks Creator Keyboard Case for iPhone 16 Pro","maker":"Clicks Technology","ec_site":"Clicks公式サイト","prod_url":"https://www.clicks.tech/products/clicks-for-iphone-16-pro","maker_hp":"https://www.clicks.tech","contact":"support@clicks.tech"},
    {"no":1193,"category":"テクノロジー・ガジェット","name":"Teenage Engineering OP-XY Synthesizer","maker":"Teenage Engineering","ec_site":"Teenage Engineering公式サイト","prod_url":"https://teenage.engineering/products/op-xy","maker_hp":"https://teenage.engineering","contact":"support@teenage.engineering"},
    {"no":1194,"category":"テクノロジー・ガジェット","name":"reMarkable 2 E-Paper Writing Tablet","maker":"reMarkable","ec_site":"reMarkable公式サイト","prod_url":"https://remarkable.com/store/remarkable-2","maker_hp":"https://remarkable.com","contact":"support@remarkable.com"},
    {"no":1195,"category":"テクノロジー・ガジェット","name":"Astro Robot Home Assistant by Amazon","maker":"Amazon Astro","ec_site":"Amazon公式サイト","prod_url":"https://www.amazon.com/astro","maker_hp":"https://www.amazon.com","contact":"astro-feedback@amazon.com"},
    {"no":1196,"category":"テクノロジー・ガジェット","name":"Panic Playdate Handheld Gaming Console","maker":"Panic Inc","ec_site":"Panic公式サイト","prod_url":"https://play.date","maker_hp":"https://panic.com","contact":"playdate@panic.com"},
    {"no":1197,"category":"テクノロジー・ガジェット","name":"Analogue Pocket Handheld Game System","maker":"Analogue","ec_site":"Analogue公式サイト","prod_url":"https://www.analogue.co/pocket","maker_hp":"https://www.analogue.co","contact":"support@analogue.co"},
    {"no":1198,"category":"テクノロジー・ガジェット","name":"Humane AI Pin Screenless AI Wearable","maker":"Humane","ec_site":"Humane公式サイト","prod_url":"https://hu.ma.ne/aipin","maker_hp":"https://hu.ma.ne","contact":"support@hu.ma.ne"},
    {"no":1199,"category":"テクノロジー・ガジェット","name":"NuraTrue Pro True Wireless Earbuds Personalized Sound","maker":"Nura","ec_site":"Nura公式サイト","prod_url":"https://www.nuraphone.com/products/nuratrue-pro","maker_hp":"https://www.nuraphone.com","contact":"support@nuraphone.com"},
    {"no":1200,"category":"テクノロジー・ガジェット","name":"Manta Sleep Mask Bluetooth Zero-Pressure Eye Mask","maker":"Manta Sleep","ec_site":"Manta公式サイト","prod_url":"https://mantasleep.com/products/manta-sleep-mask-bluetooth","maker_hp":"https://mantasleep.com","contact":"hello@mantasleep.com"},
    {"no":1201,"category":"テクノロジー・ガジェット","name":"OWC Thunderbolt Go Dock 11-Port Travel Hub","maker":"OWC","ec_site":"OWC公式サイト","prod_url":"https://www.owc.com/solutions/thunderbolt-go-dock","maker_hp":"https://www.owc.com","contact":"support@owc.com"},
    {"no":1202,"category":"テクノロジー・ガジェット","name":"Creality Halot-Mage Pro 8K Resin 3D Printer","maker":"Creality","ec_site":"Creality公式サイト","prod_url":"https://www.creality.com/products/halot-mage-pro-8k-resin-printer","maker_hp":"https://www.creality.com","contact":"support@creality.com"},
    {"no":1203,"category":"テクノロジー・ガジェット","name":"Flipper Zero Portable Multi-Tool Device","maker":"Flipper Devices","ec_site":"Flipper公式サイト","prod_url":"https://flipperzero.one/products","maker_hp":"https://flipperzero.one","contact":"support@flipper.net"},
    {"no":1204,"category":"テクノロジー・ガジェット","name":"Light Phone 3 Minimal Smartphone","maker":"Light","ec_site":"Light公式サイト","prod_url":"https://www.thelightphone.com/lightphone3","maker_hp":"https://www.thelightphone.com","contact":"hello@thelightphone.com"},
    {"no":1205,"category":"テクノロジー・ガジェット","name":"Tile Ultra Bluetooth Item Finder","maker":"Tile","ec_site":"Tile公式サイト","prod_url":"https://www.thetileapp.com/en-us/products/ultra","maker_hp":"https://www.thetileapp.com","contact":"support@thetileapp.com"},
    # ── 美容・スキンケア ───────────────────────────────
    {"no":1206,"category":"美容・スキンケア","name":"Lyma Laser Home Skincare Device","maker":"Lyma Life","ec_site":"Lyma公式サイト","prod_url":"https://lymalife.com/products/lyma-laser","maker_hp":"https://lymalife.com","contact":"hello@lymalife.com"},
    {"no":1207,"category":"美容・スキンケア","name":"Jovs Venus Pro II Laser Hair Removal","maker":"JOVS","ec_site":"JOVS公式サイト","prod_url":"https://www.jovs.com/products/venus-pro-ii","maker_hp":"https://www.jovs.com","contact":"support@jovs.com"},
    {"no":1208,"category":"美容・スキンケア","name":"Therabody TheraFace Depuffer Cooling Globes","maker":"Therabody","ec_site":"Therabody公式サイト","prod_url":"https://www.therabody.com/us/en-us/theraface-depuffer.html","maker_hp":"https://www.therabody.com","contact":"support@therabody.com"},
    {"no":1209,"category":"美容・スキンケア","name":"NEWA Lift Plus RF Anti-Aging Device","maker":"NEWA","ec_site":"NEWA公式サイト","prod_url":"https://www.newaskin.com/products/newa-lift-plus","maker_hp":"https://www.newaskin.com","contact":"support@newaskin.com"},
    {"no":1210,"category":"美容・スキンケア","name":"Omnilux Contour Face LED Mask","maker":"Omnilux","ec_site":"Omnilux公式サイト","prod_url":"https://omniluxled.com/products/omnilux-contour-face","maker_hp":"https://omniluxled.com","contact":"support@omniluxled.com"},
    {"no":1211,"category":"美容・スキンケア","name":"Cellreturn LED Platinum Band Neck Device","maker":"Cellreturn","ec_site":"Cellreturn公式サイト","prod_url":"https://www.cellreturn.com/en/products","maker_hp":"https://www.cellreturn.com","contact":"info@cellreturn.com"},
    {"no":1212,"category":"美容・スキンケア","name":"HigherDOSE Infrared Sauna Blanket V3","maker":"HigherDOSE","ec_site":"HigherDOSE公式サイト","prod_url":"https://higherdose.com/products/infrared-sauna-blanket","maker_hp":"https://higherdose.com","contact":"hi@higherdose.com"},
    {"no":1213,"category":"美容・スキンケア","name":"Droplette Micro-Infuser Skincare Device","maker":"Droplette","ec_site":"Droplette公式サイト","prod_url":"https://www.droplette.io/products/droplette-device","maker_hp":"https://www.droplette.io","contact":"hello@droplette.io"},
    {"no":1214,"category":"美容・スキンケア","name":"Revive Light Therapy DPL Nuve Neck Chest","maker":"reVive Light Therapy","ec_site":"reVive公式サイト","prod_url":"https://www.revivelighttherapy.com/products","maker_hp":"https://www.revivelighttherapy.com","contact":"support@revivelighttherapy.com"},
    {"no":1215,"category":"美容・スキンケア","name":"Clé de Peau Beauté Synactif Soap Bar","maker":"Clé de Peau Beauté","ec_site":"CdP公式サイト","prod_url":"https://www.cledepeaubeaute.com/en-us/body/synactif/soap","maker_hp":"https://www.cledepeaubeaute.com","contact":"customer.service@cledepeau-beaute.com"},
    {"no":1216,"category":"美容・スキンケア","name":"Tria Beauty Hair Removal Laser 4X","maker":"Tria Beauty","ec_site":"Tria公式サイト","prod_url":"https://www.triabeauty.com/products/laser-hair-removal-4x","maker_hp":"https://www.triabeauty.com","contact":"support@triabeauty.com"},
    {"no":1217,"category":"美容・スキンケア","name":"Baby Quasar PLUS Red Light Therapy Device","maker":"Baby Quasar","ec_site":"Baby Quasar公式サイト","prod_url":"https://babyquasar.com/products/baby-quasar-plus","maker_hp":"https://babyquasar.com","contact":"info@babyquasar.com"},
    {"no":1218,"category":"美容・スキンケア","name":"Refa CARAT FACE Facial Roller Massager","maker":"MTG","ec_site":"ReFa公式サイト","prod_url":"https://www.refa.net/en/product/refa_carat_face/","maker_hp":"https://www.refa.net","contact":"global@mtg.gr.jp"},
    {"no":1219,"category":"美容・スキンケア","name":"Inito Fertility Monitor at-Home Tracker","maker":"Inito","ec_site":"Inito公式サイト","prod_url":"https://inito.com/products/inito-fertility-monitor","maker_hp":"https://inito.com","contact":"support@inito.com"},
    {"no":1220,"category":"美容・スキンケア","name":"ORA Electric Gua Sha Face Sculptor","maker":"ORA","ec_site":"ORA公式サイト","prod_url":"https://oraorganics.com/products/electric-gua-sha","maker_hp":"https://oraorganics.com","contact":"hello@oraorganics.com"},
    # ── 子供・教育 ─────────────────────────────────────
    {"no":1221,"category":"子供・教育","name":"Moonbird Handheld Breathing Coach Device","maker":"Moonbird","ec_site":"Moonbird公式サイト","prod_url":"https://www.moonbird.life/products/moonbird","maker_hp":"https://www.moonbird.life","contact":"hello@moonbird.life"},
    {"no":1222,"category":"子供・教育","name":"Pictionary Air Smart Drawing Game","maker":"Mattel","ec_site":"Mattel公式サイト","prod_url":"https://www.mattel.com/products/pictionary-air","maker_hp":"https://www.mattel.com","contact":"consumer.affairs@mattel.com"},
    {"no":1223,"category":"子供・教育","name":"Merge Cube AR/VR Educational Cube","maker":"Merge EDU","ec_site":"Merge公式サイト","prod_url":"https://mergeedu.com/cube","maker_hp":"https://mergeedu.com","contact":"support@mergeedu.com"},
    {"no":1224,"category":"子供・教育","name":"Thames & Kosmos Gravity Maze Logic Game","maker":"Thames & Kosmos","ec_site":"Thames公式サイト","prod_url":"https://www.thamesandkosmos.com/products/gravity-maze","maker_hp":"https://www.thamesandkosmos.com","contact":"science@thamesandkosmos.com"},
    {"no":1225,"category":"子供・教育","name":"Roominate Smart House Building Kit STEM","maker":"Roominate","ec_site":"Roominate公式サイト","prod_url":"https://roominatetoy.com/products","maker_hp":"https://roominatetoy.com","contact":"hello@roominatetoy.com"},
    {"no":1226,"category":"子供・教育","name":"SmartMax My First Safari Animals Magnetic","maker":"SmartMax","ec_site":"SmartMax公式サイト","prod_url":"https://www.smartmax.com/en/products/","maker_hp":"https://www.smartmax.com","contact":"info@smartmax.com"},
    {"no":1227,"category":"子供・教育","name":"Kinetic Sand Sandisfying Set","maker":"Spin Master","ec_site":"Spin Master公式サイト","prod_url":"https://www.spinmaster.com/en-US/brands/kinetic-sand","maker_hp":"https://www.spinmaster.com","contact":"consumerservices@spinmaster.com"},
    {"no":1228,"category":"子供・教育","name":"Gravity Wells Physics Activity Set","maker":"University Games","ec_site":"University Games公式サイト","prod_url":"https://www.ugames.com/collections/science","maker_hp":"https://www.ugames.com","contact":"info@ugames.com"},
    {"no":1229,"category":"子供・教育","name":"Jimu Robot Mythical Series AstroBot Kit","maker":"UBTECH Robotics","ec_site":"UBTECH公式サイト","prod_url":"https://www.ubtrobot.com/collections/jimu-robot","maker_hp":"https://www.ubtrobot.com","contact":"support@ubtrobot.com"},
    {"no":1230,"category":"子供・教育","name":"Picasso Tiles 60-Piece Magnetic Building Blocks","maker":"PicassoTiles","ec_site":"PicassoTiles公式サイト","prod_url":"https://picassotiles.com/collections/sets","maker_hp":"https://picassotiles.com","contact":"support@picassotiles.com"},
    {"no":1231,"category":"子供・教育","name":"Gravity Maze Marble Run Brain Game","maker":"ThinkFun","ec_site":"ThinkFun公式サイト","prod_url":"https://www.thinkfun.com/products/gravity-maze","maker_hp":"https://www.thinkfun.com","contact":"info@thinkfun.com"},
    {"no":1232,"category":"子供・教育","name":"LittleCodr Coding Card Game for Kids","maker":"Firehose Games","ec_site":"Amazon","prod_url":"https://www.amazon.com/s?k=LittleCodr+coding+game","maker_hp":"https://firehosegames.com","contact":"info@firehosegames.com"},
    {"no":1233,"category":"子供・教育","name":"Anki Cozmo Robot Toy AI","maker":"Digital Dream Labs","ec_site":"DDL公式サイト","prod_url":"https://www.digitaldreamlabs.com/collections/cozmo","maker_hp":"https://www.digitaldreamlabs.com","contact":"hello@digitaldreamlabs.com"},
    {"no":1234,"category":"子供・教育","name":"Hackaball Programmable Smart Ball","maker":"Made by Many","ec_site":"Amazon","prod_url":"https://www.amazon.com/s?k=Hackaball+programmable+ball","maker_hp":"https://www.madebymany.com","contact":"hello@madebymany.com"},
    {"no":1235,"category":"子供・教育","name":"GraviTrax Pro Starter Set Ball Track System","maker":"Ravensburger","ec_site":"Ravensburger公式サイト","prod_url":"https://www.ravensburger.us/products/gravitrax/","maker_hp":"https://www.ravensburger.us","contact":"customerservice@ravensburger.com"},
    # ── ファッション・アクセサリー ──────────────────────
    {"no":1236,"category":"ファッション・アクセサリー","name":"Aviator Nation Fleece Sweatshirt","maker":"Aviator Nation","ec_site":"Aviator Nation公式サイト","prod_url":"https://aviatornation.com/collections/hoodies","maker_hp":"https://aviatornation.com","contact":"support@aviatornation.com"},
    {"no":1237,"category":"ファッション・アクセサリー","name":"Rains Waterproof Backpack Mini","maker":"Rains","ec_site":"Rains公式サイト","prod_url":"https://www.rains.com/collections/backpacks","maker_hp":"https://www.rains.com","contact":"info@rains.com"},
    {"no":1238,"category":"ファッション・アクセサリー","name":"Paravel Fold-Up Bag Carry-On","maker":"Paravel","ec_site":"Paravel公式サイト","prod_url":"https://tourparavel.com/collections/luggage","maker_hp":"https://tourparavel.com","contact":"hello@tourparavel.com"},
    {"no":1239,"category":"ファッション・アクセサリー","name":"Dagne Dover Medium Landon Carryall Bag","maker":"Dagne Dover","ec_site":"Dagne Dover公式サイト","prod_url":"https://www.dagnedover.com/collections/landon-carryall","maker_hp":"https://www.dagnedover.com","contact":"hello@dagnedover.com"},
    {"no":1240,"category":"ファッション・アクセサリー","name":"Lo & Sons Catalina Deluxe Travel Tote","maker":"Lo & Sons","ec_site":"Lo & Sons公式サイト","prod_url":"https://www.loandsons.com/collections/totes","maker_hp":"https://www.loandsons.com","contact":"support@loandsons.com"},
    {"no":1241,"category":"ファッション・アクセサリー","name":"Nomatic Travel Pack 20L","maker":"Nomatic","ec_site":"Nomatic公式サイト","prod_url":"https://www.nomatic.com/collections/travel-packs","maker_hp":"https://www.nomatic.com","contact":"support@nomatic.com"},
    {"no":1242,"category":"ファッション・アクセサリー","name":"Brevite Jumper Photo Backpack 24L","maker":"Brevite","ec_site":"Brevite公式サイト","prod_url":"https://brevite.co/products/the-jumper","maker_hp":"https://brevite.co","contact":"hello@brevite.co"},
    {"no":1243,"category":"ファッション・アクセサリー","name":"Topo Designs Global Travel Bag 40L","maker":"Topo Designs","ec_site":"Topo Designs公式サイト","prod_url":"https://topodesigns.com/collections/bags","maker_hp":"https://topodesigns.com","contact":"hello@topodesigns.com"},
    {"no":1244,"category":"ファッション・アクセサリー","name":"Alpaka Elements Tech Sling Bag","maker":"Alpaka","ec_site":"Alpaka公式サイト","prod_url":"https://alpakagear.com/collections/sling-bags","maker_hp":"https://alpakagear.com","contact":"support@alpakagear.com"},
    {"no":1245,"category":"ファッション・アクセサリー","name":"Waterfield Designs Air Porter Laptop Tote","maker":"WaterField Designs","ec_site":"WaterField公式サイト","prod_url":"https://www.sfbags.com/collections/laptop-bags","maker_hp":"https://www.sfbags.com","contact":"info@sfbags.com"},
    {"no":1246,"category":"ファッション・アクセサリー","name":"Mission Workshop Vandal Rolltop Backpack","maker":"Mission Workshop","ec_site":"Mission Workshop公式サイト","prod_url":"https://missionworkshop.com/collections/bags","maker_hp":"https://missionworkshop.com","contact":"support@missionworkshop.com"},
    {"no":1247,"category":"ファッション・アクセサリー","name":"Pitaka MagEZ Case 4 for iPhone 16 Pro","maker":"Pitaka","ec_site":"Pitaka公式サイト","prod_url":"https://www.pitaka.com/collections/iphone-cases","maker_hp":"https://www.pitaka.com","contact":"support@pitaka.com"},
    {"no":1248,"category":"ファッション・アクセサリー","name":"Moment Travelline Everyday Sling 10L","maker":"Moment","ec_site":"Moment公式サイト","prod_url":"https://www.shopmoment.com/collections/bags","maker_hp":"https://www.shopmoment.com","contact":"support@shopmoment.com"},
    {"no":1249,"category":"ファッション・アクセサリー","name":"United by Blue Moraine 25L Backpack","maker":"United By Blue","ec_site":"UBB公式サイト","prod_url":"https://unitedbyblue.com/collections/backpacks","maker_hp":"https://unitedbyblue.com","contact":"hello@unitedbyblue.com"},
    {"no":1250,"category":"ファッション・アクセサリー","name":"Hedgren Commute Backpack Anti-Theft RFID","maker":"Hedgren","ec_site":"Hedgren公式サイト","prod_url":"https://www.hedgren.com/us/collections/backpacks","maker_hp":"https://www.hedgren.com","contact":"info@hedgren.com"},
    # ── クリーニング・収納・整理 ───────────────────────
    {"no":1251,"category":"クリーニング・収納・整理","name":"Narwal Freo X Ultra Self-Cleaning Robot Vacuum","maker":"Narwal","ec_site":"Narwal公式サイト","prod_url":"https://www.narwal.com/products/narwal-freo-x-ultra","maker_hp":"https://www.narwal.com","contact":"support@narwal.com"},
    {"no":1252,"category":"クリーニング・収納・整理","name":"Dreame L20 Ultra Complete Robot Vacuum Mop","maker":"Dreame Technology","ec_site":"Dreame公式サイト","prod_url":"https://www.dreametech.com/products/dreame-l20-ultra-complete","maker_hp":"https://www.dreametech.com","contact":"support@dreametech.com"},
    {"no":1253,"category":"クリーニング・収納・整理","name":"Roborock S8 MaxV Ultra Robot Vacuum Cleaner","maker":"Roborock","ec_site":"Roborock公式サイト","prod_url":"https://us.roborock.com/pages/roborock-s8-maxv-ultra","maker_hp":"https://us.roborock.com","contact":"support@roborock.com"},
    {"no":1254,"category":"クリーニング・収納・整理","name":"Ecovacs Deebot X2 Omni Robot Vacuum","maker":"Ecovacs","ec_site":"Ecovacs公式サイト","prod_url":"https://www.ecovacs.com/us/deebot-robotic-vacuum-cleaner/DEEBOT-X2-OMNI","maker_hp":"https://www.ecovacs.com","contact":"support@ecovacs.com"},
    {"no":1255,"category":"クリーニング・収納・整理","name":"Dyson Gen5detect Absolute Cordless Vacuum","maker":"Dyson","ec_site":"Dyson公式サイト","prod_url":"https://www.dyson.com/en/vacuum-cleaners/cordless/gen5detect","maker_hp":"https://www.dyson.com","contact":"support@dyson.com"},
    {"no":1256,"category":"クリーニング・収納・整理","name":"iRobot Roomba Combo j9+ Robot Vacuum Mop","maker":"iRobot","ec_site":"iRobot公式サイト","prod_url":"https://www.irobot.com/en_US/roomba-combo-j9-plus/4703075.html","maker_hp":"https://www.irobot.com","contact":"support@irobot.com"},
    {"no":1257,"category":"クリーニング・収納・整理","name":"Tineco Floor One S7 Pro Smart Wet Dry Vacuum","maker":"Tineco","ec_site":"Tineco公式サイト","prod_url":"https://www.tineco.com/products/floor-one-s7-pro","maker_hp":"https://www.tineco.com","contact":"support@tineco.com"},
    {"no":1258,"category":"クリーニング・収納・整理","name":"Hoover ONEPWR Evolve Pet Cordless Upright","maker":"Hoover","ec_site":"Hoover公式サイト","prod_url":"https://www.hoover.com/cordless-vacuums/","maker_hp":"https://www.hoover.com","contact":"consumer@techtronic.com"},
    {"no":1259,"category":"クリーニング・収納・整理","name":"Brava Pure 6-Element Light Oven Self-Cleaning","maker":"Brava","ec_site":"Brava公式サイト","prod_url":"https://www.brava.com/pages/the-oven","maker_hp":"https://www.brava.com","contact":"support@brava.com"},
    {"no":1260,"category":"クリーニング・収納・整理","name":"Gladwell Gecko Robot Window Cleaner","maker":"Gladwell","ec_site":"Gladwell公式サイト","prod_url":"https://gladwell.us/products/gecko-window-cleaner","maker_hp":"https://gladwell.us","contact":"support@gladwell.us"},
    {"no":1261,"category":"クリーニング・収納・整理","name":"Leravan Smart Massage Pillow with Heat","maker":"Leravan","ec_site":"Leravan公式サイト","prod_url":"https://www.leravan.com/collections/massagers","maker_hp":"https://www.leravan.com","contact":"support@leravan.com"},
    {"no":1262,"category":"クリーニング・収納・整理","name":"IRIS USA Weathertight Storage Box 62qt Clear","maker":"IRIS USA","ec_site":"IRIS公式サイト","prod_url":"https://www.irisusainc.com/collections/storage-boxes","maker_hp":"https://www.irisusainc.com","contact":"customercare@irisusainc.com"},
    {"no":1263,"category":"クリーニング・収納・整理","name":"Elfa Starter Solution Closet System","maker":"Elfa","ec_site":"Container Store公式サイト","prod_url":"https://www.containerstore.com/elfa/","maker_hp":"https://www.elfa.com","contact":"info@elfa.com"},
    {"no":1264,"category":"クリーニング・収納・整理","name":"Laundry Jet Built-In Laundry Vacuum System","maker":"Laundry Jet","ec_site":"Laundry Jet公式サイト","prod_url":"https://laundryjet.com/products","maker_hp":"https://laundryjet.com","contact":"info@laundryjet.com"},
    {"no":1265,"category":"クリーニング・収納・整理","name":"Kärcher SC 4 EasyFix Steam Cleaner","maker":"Kärcher","ec_site":"Kärcher公式サイト","prod_url":"https://www.karcher.com/us/products/steam-cleaners.html","maker_hp":"https://www.karcher.com","contact":"info@karcher.com"},
    # ── キッチン続き ──────────────────────────────────
    {"no":1266,"category":"キッチン・調理器具","name":"Cinder Sensing Cooker Precision Grill","maker":"Cinder","ec_site":"Cinder公式サイト","prod_url":"https://www.cindersense.com/products/cinder-sensing-cooker","maker_hp":"https://www.cindersense.com","contact":"hello@cindersense.com"},
    {"no":1267,"category":"キッチン・調理器具","name":"Halo Elite Smart Oven 13-in-1 Air Fryer","maker":"Halo Products","ec_site":"Halo公式サイト","prod_url":"https://haloproducts.com/collections/ovens","maker_hp":"https://haloproducts.com","contact":"support@haloproducts.com"},
    {"no":1268,"category":"キッチン・調理器具","name":"GreenLife Soft Grip Ceramic 16-Piece Set","maker":"GreenLife","ec_site":"GreenLife公式サイト","prod_url":"https://www.greenlife.com/collections/cookware-sets","maker_hp":"https://www.greenlife.com","contact":"info@greenlife.com"},
    {"no":1269,"category":"キッチン・調理器具","name":"COSORI Pro Gen 2 Air Fryer 5.8 Qt","maker":"COSORI","ec_site":"COSORI公式サイト","prod_url":"https://cosori.com/collections/air-fryers","maker_hp":"https://cosori.com","contact":"support@cosori.com"},
    {"no":1270,"category":"キッチン・調理器具","name":"Ninja Thirsti Drink System Customizable","maker":"Ninja Kitchen","ec_site":"Ninja公式サイト","prod_url":"https://www.ninjakitchen.com/products/ninja-thirsti-drink-system","maker_hp":"https://www.ninjakitchen.com","contact":"support@ninjakitchen.com"},
    # ── スマートホーム続き ─────────────────────────────
    {"no":1271,"category":"スマートホーム・インテリア・照明","name":"Soma Smart Shades 3 Motorized Blinds Kit","maker":"Soma Smart Home","ec_site":"Soma公式サイト","prod_url":"https://soma-home.com/collections/shades","maker_hp":"https://soma-home.com","contact":"support@soma-home.com"},
    {"no":1272,"category":"スマートホーム・インテリア・照明","name":"ConnectSense Smart Outlet 2 Energy Monitor","maker":"ConnectSense","ec_site":"ConnectSense公式サイト","prod_url":"https://connectsense.com/products/smart-outlet-2","maker_hp":"https://connectsense.com","contact":"support@connectsense.com"},
    {"no":1273,"category":"スマートホーム・インテリア・照明","name":"Serene House Ultrasonic Aroma Diffuser Oslo","maker":"Serene House","ec_site":"Serene House公式サイト","prod_url":"https://www.serenehouse.com/collections/all","maker_hp":"https://www.serenehouse.com","contact":"info@serenehouse.com"},
    {"no":1274,"category":"スマートホーム・インテリア・照明","name":"ClearGrass Qingping Air Monitor Lite+","maker":"Qingping Technology","ec_site":"Qingping公式サイト","prod_url":"https://www.qingping.co/air-monitor-lite","maker_hp":"https://www.qingping.co","contact":"support@cleargrass.com"},
    {"no":1275,"category":"スマートホーム・インテリア・照明","name":"Airtame 2 Wireless HDMI Streaming Dongle","maker":"Airtame","ec_site":"Airtame公式サイト","prod_url":"https://airtame.com/products/airtame-2","maker_hp":"https://airtame.com","contact":"support@airtame.com"},
    # ── ウェアラブル続き ───────────────────────────────
    {"no":1276,"category":"ウェアラブル・ヘルス・フィットネス","name":"Withings BPM Connect Smart Blood Pressure Monitor","maker":"Withings","ec_site":"Withings公式サイト","prod_url":"https://www.withings.com/us/en/bpm-connect","maker_hp":"https://www.withings.com","contact":"support@withings.com"},
    {"no":1277,"category":"ウェアラブル・ヘルス・フィットネス","name":"Tempdrop Fertility Thermometer Wearable","maker":"Tempdrop","ec_site":"Tempdrop公式サイト","prod_url":"https://tempdrop.com/products/tempdrop-sensor","maker_hp":"https://tempdrop.com","contact":"support@tempdrop.com"},
    {"no":1278,"category":"ウェアラブル・ヘルス・フィットネス","name":"Myzone MZ-Switch Interchangeable Tracker","maker":"Myzone","ec_site":"Myzone公式サイト","prod_url":"https://us.myzone.org/pages/products","maker_hp":"https://us.myzone.org","contact":"support@myzone.org"},
    {"no":1279,"category":"ウェアラブル・ヘルス・フィットネス","name":"Spire Stone Mindfulness and Activity Tracker","maker":"Spire Health","ec_site":"Spire公式サイト","prod_url":"https://spirehealth.com/products/spire-stone","maker_hp":"https://spirehealth.com","contact":"support@spirehealth.com"},
    {"no":1280,"category":"ウェアラブル・ヘルス・フィットネス","name":"Skulpt Chisel Muscle Quality Scanner","maker":"Skulpt","ec_site":"Skulpt公式サイト","prod_url":"https://www.skulpt.me/products/chisel","maker_hp":"https://www.skulpt.me","contact":"support@skulpt.me"},
    # ── アウトドア続き ─────────────────────────────────
    {"no":1281,"category":"アウトドア・スポーツ・旅行","name":"Tern GSD S10 LX Electric Cargo Bike","maker":"Tern Bicycles","ec_site":"Tern公式サイト","prod_url":"https://www.ternbicycles.com/us/bikes/gsd","maker_hp":"https://www.ternbicycles.com","contact":"info@ternbicycles.com"},
    {"no":1282,"category":"アウトドア・スポーツ・旅行","name":"Hiplok Gold Wearable Bike Lock Chain","maker":"Hiplok","ec_site":"Hiplok公式サイト","prod_url":"https://hiplok.com/product/hiplok-gold/","maker_hp":"https://hiplok.com","contact":"support@hiplok.com"},
    {"no":1283,"category":"アウトドア・スポーツ・旅行","name":"LOON Outdoors Ultralight Trekking Poles","maker":"LOON Outdoors","ec_site":"LOON公式サイト","prod_url":"https://loonoutdoors.com/collections/trekking-poles","maker_hp":"https://loonoutdoors.com","contact":"hello@loonoutdoors.com"},
    {"no":1284,"category":"アウトドア・スポーツ・旅行","name":"Freefly Alta X Drone Heavy Lift Cinema","maker":"Freefly Systems","ec_site":"Freefly公式サイト","prod_url":"https://freeflysystems.com/alta-x","maker_hp":"https://freeflysystems.com","contact":"support@freeflysystems.com"},
    {"no":1285,"category":"アウトドア・スポーツ・旅行","name":"Shokz OpenRun Pro 2 Bone Conduction Headphones","maker":"Shokz","ec_site":"Shokz公式サイト","prod_url":"https://shokz.com/products/openrunpro2","maker_hp":"https://shokz.com","contact":"support@shokz.com"},
    # ── ペット続き ─────────────────────────────────────
    {"no":1286,"category":"ペット用品","name":"Leo's Paw Automatic Laser Cat Toy","maker":"Leo's Paw","ec_site":"Leo's Paw公式サイト","prod_url":"https://leospaw.com/products/automatic-laser-toy","maker_hp":"https://leospaw.com","contact":"support@leospaw.com"},
    {"no":1287,"category":"ペット用品","name":"Petcube Play 2 Laser Cat Camera Toy","maker":"Petcube","ec_site":"Petcube公式サイト","prod_url":"https://petcube.com/play2.html","maker_hp":"https://petcube.com","contact":"support@petcube.com"},
    {"no":1288,"category":"ペット用品","name":"Wagz Smart Dog Door with GPS Fence","maker":"Wagz","ec_site":"Wagz公式サイト","prod_url":"https://wagz.com/products/smart-door","maker_hp":"https://wagz.com","contact":"hello@wagz.com"},
    {"no":1289,"category":"ペット用品","name":"Furbo Dog Camera Nanny Alert 360 Degree","maker":"Furbo","ec_site":"Furbo公式サイト","prod_url":"https://www.furbo.com/products/furbo-nanny-cam","maker_hp":"https://www.furbo.com","contact":"support@furbo.com"},
    {"no":1290,"category":"ペット用品","name":"PawsInMotion Interactive Feather Wand Cat Toy","maker":"PawsInMotion","ec_site":"PawsInMotion公式サイト","prod_url":"https://pawsinmotion.com/products/interactive-feather","maker_hp":"https://pawsinmotion.com","contact":"info@pawsinmotion.com"},
    # ── テクノロジー続き ──────────────────────────────
    {"no":1291,"category":"テクノロジー・ガジェット","name":"Teenage Engineering Pocket Operator Series","maker":"Teenage Engineering","ec_site":"TE公式サイト","prod_url":"https://teenage.engineering/store/pocket-operator","maker_hp":"https://teenage.engineering","contact":"support@teenage.engineering"},
    {"no":1292,"category":"テクノロジー・ガジェット","name":"Opal C1 4K Webcam for Streaming","maker":"Opal Camera","ec_site":"Opal公式サイト","prod_url":"https://opalcamera.com/products/opal-c1","maker_hp":"https://opalcamera.com","contact":"support@opalcamera.com"},
    {"no":1293,"category":"テクノロジー・ガジェット","name":"Magewell USB Capture HDMI 4K Plus","maker":"Magewell","ec_site":"Magewell公式サイト","prod_url":"https://www.magewell.com/products/usb-capture-hdmi-4k-plus","maker_hp":"https://www.magewell.com","contact":"sales@magewell.com"},
    {"no":1294,"category":"テクノロジー・ガジェット","name":"Pimoroni Badger 2350 E-Ink Badge","maker":"Pimoroni","ec_site":"Pimoroni公式サイト","prod_url":"https://shop.pimoroni.com/products/badger-2350","maker_hp":"https://pimoroni.com","contact":"hello@pimoroni.com"},
    {"no":1295,"category":"テクノロジー・ガジェット","name":"Framework Laptop 16 Modular DIY Edition","maker":"Framework Computer","ec_site":"Framework公式サイト","prod_url":"https://frame.work/products/laptop-diy-16-gen1","maker_hp":"https://frame.work","contact":"support@frame.work"},
    # ── 美容続き ─────────────────────────────────────
    {"no":1296,"category":"美容・スキンケア","name":"Currentbody Skin Tone Device EMS","maker":"CurrentBody","ec_site":"CurrentBody公式サイト","prod_url":"https://www.currentbody.com/products/currentbody-skin-tone-device","maker_hp":"https://www.currentbody.com","contact":"info@currentbody.com"},
    {"no":1297,"category":"美容・スキンケア","name":"Nira Pro Anti-Aging Laser Device","maker":"Nira Skincare","ec_site":"Nira公式サイト","prod_url":"https://niraskincare.com/products/nira-pro","maker_hp":"https://niraskincare.com","contact":"support@niraskincare.com"},
    {"no":1298,"category":"美容・スキンケア","name":"Refa Active Massage Roller Face","maker":"MTG","ec_site":"ReFa公式サイト","prod_url":"https://www.refa.net/en/product/refa_active/","maker_hp":"https://www.refa.net","contact":"global@mtg.gr.jp"},
    {"no":1299,"category":"美容・スキンケア","name":"Geske SmartAppGuided MicroNeedle Face Roller","maker":"GESKE","ec_site":"GESKE公式サイト","prod_url":"https://www.geske.com/products/microneedle-face-roller","maker_hp":"https://www.geske.com","contact":"support@geske.com"},
    {"no":1300,"category":"美容・スキンケア","name":"InFace RF Beauty Instrument Skin Tightening","maker":"InFace","ec_site":"InFace公式サイト","prod_url":"https://inface.com/collections/beauty-devices","maker_hp":"https://inface.com","contact":"support@inface.com"},
    # ── 子供続き ─────────────────────────────────────
    {"no":1301,"category":"子供・教育","name":"Lego Education Spike Prime STEM Set","maker":"Lego Education","ec_site":"LEGO Education公式サイト","prod_url":"https://education.lego.com/en-us/products/lego-education-spike-prime","maker_hp":"https://education.lego.com","contact":"service@lego.com"},
    {"no":1302,"category":"子供・教育","name":"Matatalab Coding Robot Set for Kids","maker":"Matatalab","ec_site":"Matatalab公式サイト","prod_url":"https://matatalab.com/products/matatalab-coding-set","maker_hp":"https://matatalab.com","contact":"support@matatalab.com"},
    # ── ファッション続き ───────────────────────────────
    {"no":1303,"category":"ファッション・アクセサリー","name":"Evergoods MPL30 Mountain Panel Loader Pack","maker":"EVERGOODS","ec_site":"EVERGOODS公式サイト","prod_url":"https://evergoods.us/collections/packs","maker_hp":"https://evergoods.us","contact":"hello@evergoods.us"},
    {"no":1304,"category":"ファッション・アクセサリー","name":"Able Carry Max Plus Backpack 38L","maker":"Able Carry","ec_site":"Able Carry公式サイト","prod_url":"https://ablecarry.com/products/max-plus-backpack","maker_hp":"https://ablecarry.com","contact":"support@ablecarry.com"},
    # ── クリーニング続き ──────────────────────────────
    {"no":1305,"category":"クリーニング・収納・整理","name":"Cosmo XT Cordless Stick Vacuum & Handheld 2-in-1","maker":"Dreame Technology","ec_site":"Dreame公式サイト","prod_url":"https://www.dreametech.com/products/cosmo-xt","maker_hp":"https://www.dreametech.com","contact":"support@dreametech.com"},
]

# ===================== Styles =====================
cat_fills = {
    "キッチン・調理器具":             PatternFill("solid", start_color="2E75B6"),
    "スマートホーム・インテリア・照明": PatternFill("solid", start_color="538135"),
    "ウェアラブル・ヘルス・フィットネス": PatternFill("solid", start_color="7030A0"),
    "アウトドア・スポーツ・旅行":       PatternFill("solid", start_color="C55A11"),
    "ペット用品":                      PatternFill("solid", start_color="843C0C"),
    "テクノロジー・ガジェット":         PatternFill("solid", start_color="244185"),
    "美容・スキンケア":                PatternFill("solid", start_color="C00000"),
    "子供・教育":                      PatternFill("solid", start_color="375623"),
    "ファッション・アクセサリー":       PatternFill("solid", start_color="7B3F00"),
    "クリーニング・収納・整理":         PatternFill("solid", start_color="404040"),
}
white_font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
normal_font = Font(name="Arial", size=10)
link_font   = Font(name="Arial", size=10, color="0563C1", underline="single")
thin        = Side(style="thin", color="CCCCCC")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)
alt_fill    = PatternFill("solid", start_color="EBF3FB")
center_va   = Alignment(vertical="center", wrap_text=True)

wb = load_workbook(r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸1105件.xlsx")
ws = wb.active

for p in new_products:
    row_num = p["no"] + 1
    bg = alt_fill if row_num % 2 == 0 else None
    cat = p["category"]
    cat_fill = cat_fills.get(cat)

    contact_val = str(p["contact"]).strip()
    has_email = "@" in contact_val and not contact_val.startswith("http") and "問い合わせ" not in contact_val

    values = [p["no"], cat, p["name"], p["maker"], p["ec_site"], p["prod_url"], p["maker_hp"], contact_val]

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.border = border
        cell.alignment = center_va

        if col_idx == 2:
            if cat_fill:
                cell.fill = cat_fill
                cell.font = white_font
            else:
                cell.font = normal_font
        elif col_idx in (6, 7) and val and str(val).startswith("http"):
            cell.hyperlink = str(val)
            cell.font = link_font
            if bg:
                cell.fill = bg
        elif col_idx == 8:
            if has_email:
                gmail_url = build_gmail(contact_val, p["maker"], p["name"])
                cell.hyperlink = gmail_url
                cell.font = link_font
            elif "問い合わせ" in contact_val:
                url_part = contact_val.replace("（問い合わせページ）", "").strip()
                if url_part.startswith("http"):
                    cell.hyperlink = url_part
                    cell.value = "問い合わせページ"
                    cell.font = link_font
            if bg:
                cell.fill = bg
        else:
            cell.font = normal_font
            if bg and col_idx != 2:
                cell.fill = bg

output_path = r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸1305件.xlsx"
wb.save(output_path)
print(f"完了: {output_path}")
print(f"追加件数: {len(new_products)} 件 (No.1106〜1305)")
email_count = sum(1 for p in new_products if "@" in str(p["contact"]) and not str(p["contact"]).startswith("http"))
print(f"メールアドレスあり: {email_count} 件 ({email_count*100//len(new_products)}%)")
