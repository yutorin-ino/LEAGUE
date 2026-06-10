from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb = load_workbook(r"C:\Users\hiro\HP0524\ZECZEC_ガジェット日本未発売リスト.xlsx")
ws = wb.active

link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
link_font_bold = Font(name="Arial", size=10, color="0563C1", underline="single", bold=False)

# Column D = ZECZEC URL (col 4), Column E = メーカーHP (col 5)
for row in range(2, ws.max_row + 1):
    for col in [4, 5]:
        cell = ws.cell(row=row, column=col)
        val = cell.value
        if val and isinstance(val, str) and val.startswith("http"):
            cell.hyperlink = val
            cell.font = link_font

wb.save(r"C:\Users\hiro\HP0524\ZECZEC_ガジェット日本未発売リスト.xlsx")
print("ハイパーリンク設定完了")
