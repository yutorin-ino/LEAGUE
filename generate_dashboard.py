import json
from openpyxl import load_workbook
from urllib.parse import quote

wb = load_workbook(r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸1505件.xlsx")
ws = wb.active

def build_gmail_url(email, maker_name, product_name):
    subject = "Potential Distribution Partnership for Japan"
    body = f"""Dear {maker_name} Team,

My name is Hiroyuki Inoguchi from
Sumai pluS Co., Ltd. in Japan

I hope this message finds you all well.

Our company focuses on promoting products that enrich people's daily lives. We are currently seeking unique international brands to introduce to the Japanese market and support their growth.

I recently had the opportunity to review your product ({product_name}) and was very impressed by its potential in the Japanese market.

I am confident that your product will strongly appeal to Japanese customers, and I would like to explore the possibility of building a partnership with your company to help ensure its success.

Furthermore, our company has a proven track record of promoting overseas products through our proprietary sales network and Japanese distribution channels.

Would it be possible to schedule a brief online meeting sometime next week to discuss this matter?

I look forward to hearing from you.


Sincerely,

Hiroyuki Inoguchi
Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )

================================================

Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )
Hiroyuki Inoguchi

E-mail:yutorin.ino@gmail.com

Address: 49-5 Kitazakuno, Higashi-Itsushiro, Ichinomiya City, Aichi Prefecture, Japan

================================================="""
    sender = quote("yutorin.ino@gmail.com", safe='')
    return f"https://mail.google.com/mail/?authuser={sender}&view=cm&fs=1&to={quote(email,safe='')}&su={quote(subject,safe='')}&body={quote(body,safe='')}"

products = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    no       = row[0]
    category = row[1] or ""
    name     = row[2] or ""
    maker    = row[3] or ""
    ec_site  = row[4] or ""
    prod_url = row[5] or ""
    maker_hp = row[6] or ""
    contact  = row[7] or ""

    contact_str = str(contact).strip()
    has_email = "@" in contact_str and not contact_str.startswith("http") and "問い合わせ" not in contact_str
    gmail_url = build_gmail_url(contact_str, maker, name) if has_email else ""

    products.append({
        "no": no,
        "category": category,
        "name": name,
        "maker": maker,
        "ec_site": ec_site,
        "prod_url": prod_url,
        "maker_hp": maker_hp,
        "contact": contact_str,
        "has_email": has_email,
        "gmail_url": gmail_url,
    })

categories = sorted(set(p["category"] for p in products if p["category"]))

data_json = json.dumps(products, ensure_ascii=False)

html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>海外便利グッズ 一括メール送信ダッシュボード</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "Segoe UI", "Yu Gothic", sans-serif; background: #f0f4f8; color: #222; }}
  header {{ background: #1a3a6b; color: white; padding: 18px 24px; display: flex; align-items: center; gap: 16px; }}
  header h1 {{ font-size: 1.25rem; font-weight: 700; }}
  header .count {{ background: #f0a500; color: #1a3a6b; border-radius: 20px; padding: 3px 12px; font-weight: 700; font-size: 0.9rem; }}
  .toolbar {{ background: white; padding: 14px 24px; display: flex; flex-wrap: wrap; gap: 12px; align-items: center; border-bottom: 1px solid #ddd; position: sticky; top: 0; z-index: 100; box-shadow: 0 2px 6px rgba(0,0,0,0.08); }}
  .toolbar input[type=text] {{ border: 1.5px solid #bbb; border-radius: 8px; padding: 7px 12px; font-size: 0.95rem; width: 260px; outline: none; }}
  .toolbar input[type=text]:focus {{ border-color: #1a3a6b; }}
  .toolbar select {{ border: 1.5px solid #bbb; border-radius: 8px; padding: 7px 10px; font-size: 0.95rem; outline: none; }}
  .toolbar select:focus {{ border-color: #1a3a6b; }}
  .btn {{ padding: 8px 18px; border-radius: 8px; border: none; cursor: pointer; font-size: 0.92rem; font-weight: 600; transition: background 0.15s; }}
  .btn-primary {{ background: #1a3a6b; color: white; }}
  .btn-primary:hover {{ background: #2a5299; }}
  .btn-primary:disabled {{ background: #aaa; cursor: not-allowed; }}
  .btn-success {{ background: #28a745; color: white; }}
  .btn-success:hover {{ background: #218838; }}
  .btn-danger {{ background: #dc3545; color: white; }}
  .btn-danger:hover {{ background: #c82333; }}
  .btn-outline {{ background: white; color: #1a3a6b; border: 1.5px solid #1a3a6b; }}
  .btn-outline:hover {{ background: #eef2ff; }}
  .sel-count {{ font-size: 0.9rem; color: #555; white-space: nowrap; }}
  .sel-count span {{ font-weight: 700; color: #1a3a6b; }}
  .filter-info {{ font-size: 0.85rem; color: #777; }}

  .table-wrap {{ padding: 16px 24px; overflow-x: auto; }}
  table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,0.1); font-size: 0.88rem; }}
  thead {{ background: #1a3a6b; color: white; }}
  thead th {{ padding: 11px 10px; text-align: left; font-weight: 600; white-space: nowrap; }}
  thead th:first-child {{ width: 40px; text-align: center; }}
  tbody tr {{ border-bottom: 1px solid #eee; }}
  tbody tr:hover {{ background: #f5f8ff; }}
  tbody tr.selected {{ background: #dceeff; }}
  tbody tr.no-email {{ opacity: 0.55; }}
  tbody td {{ padding: 9px 10px; vertical-align: middle; }}
  tbody td:first-child {{ text-align: center; }}
  .no-col {{ color: #888; font-size: 0.8rem; width: 40px; text-align: center; }}
  .cat-badge {{ display: inline-block; padding: 2px 8px; border-radius: 12px; font-size: 0.75rem; font-weight: 600; white-space: nowrap; }}
  .name-col {{ max-width: 220px; font-weight: 500; }}
  .maker-col {{ max-width: 140px; color: #444; }}
  a.link {{ color: #1a6bc5; text-decoration: none; font-size: 0.82rem; }}
  a.link:hover {{ text-decoration: underline; }}
  .email-col {{ max-width: 200px; font-size: 0.82rem; color: #555; word-break: break-all; }}
  .no-email-tag {{ color: #aaa; font-size: 0.8rem; }}
  input[type=checkbox] {{ width: 16px; height: 16px; cursor: pointer; accent-color: #1a3a6b; }}

  .modal {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,0.5); z-index: 200; align-items: center; justify-content: center; }}
  .modal.show {{ display: flex; }}
  .modal-box {{ background: white; border-radius: 12px; padding: 28px; max-width: 480px; width: 90%; box-shadow: 0 8px 32px rgba(0,0,0,0.2); }}
  .modal-box h2 {{ font-size: 1.1rem; margin-bottom: 12px; color: #1a3a6b; }}
  .modal-box p {{ font-size: 0.92rem; color: #444; margin-bottom: 8px; line-height: 1.6; }}
  .modal-box .warn {{ background: #fff3cd; border: 1px solid #f0ad4e; border-radius: 6px; padding: 10px 14px; font-size: 0.85rem; color: #856404; margin: 14px 0; }}
  .modal-box .btns {{ display: flex; gap: 10px; justify-content: flex-end; margin-top: 18px; }}
  .progress {{ display: none; margin-top: 14px; }}
  .progress-bar {{ background: #eee; border-radius: 8px; height: 10px; overflow: hidden; }}
  .progress-fill {{ background: #28a745; height: 100%; transition: width 0.2s; }}
  .progress-text {{ font-size: 0.82rem; color: #555; margin-top: 6px; text-align: center; }}

  .cat-キッチン・調理器具 {{ background: #dbeafe; color: #1e40af; }}
  .cat-スマートホーム・インテリア・照明 {{ background: #d1fae5; color: #065f46; }}
  .cat-ウェアラブル・ヘルス・フィットネス {{ background: #ede9fe; color: #4c1d95; }}
  .cat-アウトドア・スポーツ・旅行 {{ background: #fee2e2; color: #991b1b; }}
  .cat-ペット用品 {{ background: #fef3c7; color: #78350f; }}
  .cat-テクノロジー・ガジェット {{ background: #e0f2fe; color: #0c4a6e; }}
  .cat-美容・スキンケア {{ background: #fce7f3; color: #831843; }}
  .cat-子供・教育 {{ background: #d1fae5; color: #14532d; }}
  .cat-ファッション・アクセサリー {{ background: #fdf2f8; color: #701a75; }}
  .cat-クリーニング・収納・整理 {{ background: #f3f4f6; color: #374151; }}
</style>
</head>
<body>

<header>
  <h1>🌐 海外便利グッズ 一括メール送信ダッシュボード</h1>
  <span class="count" id="totalCount"></span>
</header>

<div class="toolbar">
  <input type="text" id="searchBox" placeholder="🔍 商品名・メーカー名で検索..." oninput="applyFilter()">
  <select id="catFilter" onchange="applyFilter()">
    <option value="">すべてのカテゴリ</option>
    {''.join(f'<option value="{c}">{c}</option>' for c in categories)}
  </select>
  <select id="emailFilter" onchange="applyFilter()">
    <option value="">メール有無：すべて</option>
    <option value="yes">メールあり</option>
    <option value="no">メールなし</option>
  </select>
  <button class="btn btn-outline" onclick="selectAllVisible()">表示中を全選択</button>
  <button class="btn btn-outline" onclick="clearSelection()">選択解除</button>
  <span class="sel-count">選択中: <span id="selCount">0</span> 件</span>
  <span class="filter-info" id="filterInfo"></span>
  <button class="btn btn-success" id="sendBtn" onclick="confirmSend()" disabled>📧 選択した相手に一括送信</button>
</div>

<div class="table-wrap">
  <table id="mainTable">
    <thead>
      <tr>
        <th><input type="checkbox" id="headerCheck" onchange="toggleHeaderCheck(this)"></th>
        <th>No.</th>
        <th>カテゴリ</th>
        <th>商品名</th>
        <th>メーカー名</th>
        <th>販売ECサイト</th>
        <th>商品URL</th>
        <th>メーカーHP</th>
        <th>連絡先メール</th>
      </tr>
    </thead>
    <tbody id="tableBody"></tbody>
  </table>
</div>

<div class="modal" id="confirmModal">
  <div class="modal-box">
    <h2>📧 一括メール送信の確認</h2>
    <p id="confirmText"></p>
    <div class="warn">
      ⚠️ ブラウザのポップアップブロックを許可してください。<br>
      複数のGmail作成画面が新しいタブで開きます。<br>
      一度に大量送信する場合は、数十件ずつ分けることを推奨します。
    </div>
    <div class="progress" id="progressArea">
      <div class="progress-bar"><div class="progress-fill" id="progressFill"></div></div>
      <div class="progress-text" id="progressText"></div>
    </div>
    <div class="btns">
      <button class="btn btn-outline" onclick="closeModal()">キャンセル</button>
      <button class="btn btn-success" id="execBtn" onclick="executeSend()">送信開始</button>
    </div>
  </div>
</div>

<script>
const products = {data_json};

let filtered = [...products];
let selected = new Set();

const catColors = {{}};

function init() {{
  document.getElementById('totalCount').textContent = products.length + ' 件';
  renderTable(filtered);
}}

function applyFilter() {{
  const q = document.getElementById('searchBox').value.toLowerCase();
  const cat = document.getElementById('catFilter').value;
  const em = document.getElementById('emailFilter').value;
  filtered = products.filter(p => {{
    const matchQ = !q || p.name.toLowerCase().includes(q) || p.maker.toLowerCase().includes(q);
    const matchCat = !cat || p.category === cat;
    const matchEm = !em || (em === 'yes' ? p.has_email : !p.has_email);
    return matchQ && matchCat && matchEm;
  }});
  document.getElementById('filterInfo').textContent = `表示中: ${{filtered.length}} 件`;
  renderTable(filtered);
  updateSendBtn();
}}

function renderTable(rows) {{
  const tbody = document.getElementById('tableBody');
  tbody.innerHTML = '';
  rows.forEach(p => {{
    const tr = document.createElement('tr');
    if (selected.has(p.no)) tr.classList.add('selected');
    if (!p.has_email) tr.classList.add('no-email');
    tr.innerHTML = `
      <td><input type="checkbox" ${{!p.has_email ? 'disabled title="メールアドレスなし"' : ''}} ${{selected.has(p.no) ? 'checked' : ''}} onchange="toggleRow(${{p.no}}, this)"></td>
      <td class="no-col">${{p.no}}</td>
      <td><span class="cat-badge cat-${{p.category}}">${{p.category}}</span></td>
      <td class="name-col" title="${{p.name}}">${{truncate(p.name, 35)}}</td>
      <td class="maker-col" title="${{p.maker}}">${{truncate(p.maker, 22)}}</td>
      <td style="font-size:0.8rem;color:#555">${{p.ec_site}}</td>
      <td>${{p.prod_url ? `<a class="link" href="${{p.prod_url}}" target="_blank">🔗 商品ページ</a>` : ''}}</td>
      <td>${{p.maker_hp ? `<a class="link" href="${{p.maker_hp}}" target="_blank">🏠 HP</a>` : ''}}</td>
      <td class="email-col">${{p.has_email ? `<a class="link" href="${{p.gmail_url}}" target="_blank">✉️ ${{p.contact}}</a>` : `<span class="no-email-tag">メールなし</span>`}}</td>
    `;
    tbody.appendChild(tr);
  }});
  document.getElementById('headerCheck').checked = false;
}}

function truncate(str, n) {{
  return str.length > n ? str.slice(0, n) + '…' : str;
}}

function toggleRow(no, cb) {{
  if (cb.checked) selected.add(no);
  else selected.delete(no);
  const tr = cb.closest('tr');
  tr.classList.toggle('selected', cb.checked);
  updateSendBtn();
  document.getElementById('selCount').textContent = selected.size;
}}

function selectAllVisible() {{
  filtered.forEach(p => {{ if (p.has_email) selected.add(p.no); }});
  renderTable(filtered);
  updateSendBtn();
  document.getElementById('selCount').textContent = selected.size;
}}

function clearSelection() {{
  selected.clear();
  renderTable(filtered);
  updateSendBtn();
  document.getElementById('selCount').textContent = 0;
}}

function toggleHeaderCheck(cb) {{
  if (cb.checked) selectAllVisible();
  else clearSelection();
}}

function updateSendBtn() {{
  const btn = document.getElementById('sendBtn');
  btn.disabled = selected.size === 0;
  document.getElementById('selCount').textContent = selected.size;
}}

function confirmSend() {{
  const sel = products.filter(p => selected.has(p.no) && p.has_email);
  document.getElementById('confirmText').textContent =
    `選択した ${{sel.length}} 件の相手にGmailの作成画面を開きます。よろしいですか？`;
  document.getElementById('progressArea').style.display = 'none';
  document.getElementById('execBtn').disabled = false;
  document.getElementById('confirmModal').classList.add('show');
}}

function closeModal() {{
  document.getElementById('confirmModal').classList.remove('show');
}}

async function executeSend() {{
  const sel = products.filter(p => selected.has(p.no) && p.has_email);
  document.getElementById('execBtn').disabled = true;
  const prog = document.getElementById('progressArea');
  prog.style.display = 'block';
  const fill = document.getElementById('progressFill');
  const text = document.getElementById('progressText');

  for (let i = 0; i < sel.length; i++) {{
    const p = sel[i];
    window.open(p.gmail_url, '_blank');
    const pct = Math.round(((i + 1) / sel.length) * 100);
    fill.style.width = pct + '%';
    text.textContent = `${{i + 1}} / ${{sel.length}} 件 開きました...`;
    await new Promise(r => setTimeout(r, 400));
  }}

  text.textContent = `✅ ${{sel.length}} 件のGmail作成画面を開きました！`;
  setTimeout(() => closeModal(), 2000);
}}

init();
</script>
</body>
</html>
"""

output_path = r"C:\Users\hiro\HP0524\mail_dashboard.html"
with open(output_path, "w", encoding="utf-8") as f:
    f.write(html)

print(f"完了: {output_path}")
print(f"商品数: {len(products)} 件")
email_count = sum(1 for p in products if p["has_email"])
print(f"メールアドレスあり: {email_count} 件")
