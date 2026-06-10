from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "海外便利グッズリスト"

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="1F4E79")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

cat_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
cat_fills = {
    "キッチン・調理器具": PatternFill("solid", start_color="2E75B6"),
    "スマートホーム・インテリア・照明": PatternFill("solid", start_color="538135"),
    "ウェアラブル・ヘルス・フィットネス": PatternFill("solid", start_color="7030A0"),
    "アウトドア・スポーツ・旅行": PatternFill("solid", start_color="C55A11"),
    "ペット用品": PatternFill("solid", start_color="843C0C"),
    "テクノロジー・ガジェット": PatternFill("solid", start_color="244185"),
}
default_cat_fill = PatternFill("solid", start_color="404040")

data_font = Font(name="Arial", size=10)
data_align = Alignment(vertical="center", wrap_text=True)
alt_fill = PatternFill("solid", start_color="EBF3FB")

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["No.", "カテゴリ", "商品名", "メーカー名", "販売ECサイト名", "商品URL", "メーカーHP", "連絡先メール"]
ws.append(headers)
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
ws.row_dimensions[1].height = 30

# No., カテゴリ, 商品名, メーカー名, 販売ECサイト名, 商品URL, メーカーHP, 連絡先メール
data = [
    (1, "キッチン・調理器具", "ThermoWorks Gravitas Precision Kitchen Scale", "ThermoWorks", "ThermoWorks公式サイト", "https://www.thermoworks.com/products/gravitas", "https://www.thermoworks.com", "https://www.thermoworks.com/pages/contact-us（問い合わせページ）"),
    (2, "キッチン・調理器具", "LISSOME R1 Compact Smart Portable Dishwasher", "LISSOME", "LISSOME公式サイト", "https://lissomesmart.com/", "https://lissomesmart.com", "eakxu19495120@gmail.com"),
    (3, "キッチン・調理器具", "Cozytime Lumo AI Infrared Indoor Grill", "Cozytime", "Cozytime公式サイト", "https://cozytime-tech.com/pages/lumo", "https://cozytime-tech.com", "customers@cozytime-tech.com"),
    (4, "キッチン・調理器具", "Coolwill Ice Pop Maker", "Coolwill", "Coolwill公式サイト", "https://campaign.coolwill-tech.com/", "http://coolwill-tech.com", "support@coolwill-tech.com"),
    (5, "キッチン・調理器具", "Nosh Autonomous Cooking Robot", "Euphotic Labs", "Nosh公式サイト", "https://www.letsnosh.io/", "https://www.letsnosh.io", "support@letsnosh.io"),
    (6, "キッチン・調理器具", "Nama J3 Cold Press Juicer", "Nama", "Nama公式サイト", "https://namawell.com/products/nama-j3-cold-press-juicer", "https://namawell.com", "support@namawell.com"),
    (7, "キッチン・調理器具", "Karbo Grill X1 Portable Gas Grill", "Karbo Grill", "Karbo Grill公式サイト", "https://karbogrill.com/", "https://karbogrill.com", "https://karbogrill.com/pages/contact（問い合わせページ）"),
    (8, "キッチン・調理器具", "Typhur Sync Gold Wireless Meat Thermometer", "Typhur", "Typhur公式サイト", "https://www.typhur.com/sync-gold-series", "https://www.typhur.com", "support@typhur.com"),
    (9, "キッチン・調理器具", "Seattle Ultrasonics C-200 Ultrasonic Chef's Knife", "Seattle Ultrasonics", "Seattle Ultrasonics公式サイト", "https://seattleultrasonics.com/products/c-200-ultrasonic-8-chefs-knife", "https://seattleultrasonics.com", "hello@seattleultrasonics.com"),
    (10, "キッチン・調理器具", "Oulin AI Smart Sink", "Oulin", "Oulin公式サイト", "https://www.oulinworld.com/", "https://www.oulinworld.com", "https://www.oulinworld.com（問い合わせページ）"),
    (11, "キッチン・調理器具", "Fellow Espresso Series 1 Home Espresso Machine", "Fellow", "Fellow公式サイト", "https://fellowproducts.com/products/espresso-series-1", "https://fellowproducts.com", "hello@fellowproducts.com"),
    (12, "キッチン・調理器具", "Brisk It Neoma AI Countertop Oven", "Brisk It", "Brisk It公式サイト", "https://briskitgrills.com/", "https://briskitgrills.com", "support@briskitgrills.com"),
    (13, "キッチン・調理器具", "XBrew Lab EverNitro Nitro Drinks Machine", "XBrew Lab", "XBrew Lab公式サイト", "https://xbrewlab.com/", "https://xbrewlab.com", "https://xbrewlab.com/pages/contact（問い合わせページ）"),
    (14, "キッチン・調理器具", "GoveeLife Smart Nugget Ice Maker Pro", "GoveeLife", "GoveeLife公式サイト（US）", "https://us.goveelife.com/", "https://us.goveelife.com", "support@govee.com"),
    (15, "キッチン・調理器具", "VSSL Java G25 Hand Coffee Grinder", "VSSL", "VSSL公式サイト", "https://www.vsslgear.com/products/java-g25-grinder", "https://www.vsslgear.com", "info@vsslgear.com"),
    (16, "キッチン・調理器具", "Our Place Rice Cooker", "Our Place", "Our Place公式サイト", "https://fromourplace.com/products/rice-cooker", "https://fromourplace.com", "hello@fromourplace.com"),
    (17, "キッチン・調理器具", "Aiden Precision Coffee Maker", "Fellow", "Fellow公式サイト", "https://fellowproducts.com/products/aiden-precision-coffee-maker", "https://fellowproducts.com", "hello@fellowproducts.com"),
    (18, "キッチン・調理器具", "Kara Pure 2 Air-to-Water Dispenser", "Kara Water", "Kara Water公式サイト", "https://www.karawater.com/products/kara-pure-2", "https://www.karawater.com", "info@karawater.com"),
    (19, "キッチン・調理器具", "Waterdrop X12-PRO Reverse Osmosis Filter", "Waterdrop", "Waterdrop公式サイト", "https://www.waterdropfilter.com/", "https://www.waterdropfilter.com", "service@waterdropfilter.com"),
    (20, "キッチン・調理器具", "nutribullet Chill Ice Cream Maker", "nutribullet", "Gadget Flow", "https://thegadgetflow.com/product/nutribullet-chill-ice-cream-maker/", "https://www.nutribullet.com", "https://www.nutribullet.com/pages/contact（問い合わせページ）"),
    (21, "キッチン・調理器具", "Baratza Encore ESP Pro Coffee Grinder", "Baratza", "Gadget Flow", "https://thegadgetflow.com/product/baratza-encore-esp-pro-coffee-grinder/", "https://www.baratza.com", "https://www.baratza.com/pages/contact（問い合わせページ）"),
    (22, "キッチン・調理器具", "Thaan Grill Tabletop Charcoal Grill", "Thaan", "Gadget Flow", "https://thegadgetflow.com/product/thaan-grill-tabletop-charcoal-grill/", "https://www.thaangrill.com", "https://www.thaangrill.com（問い合わせページ）"),
    (23, "キッチン・調理器具", "Root Under-Sink Composting System", "Root", "Gadget Flow", "https://thegadgetflow.com/product/root-under-sink-composting-system/", "https://www.rootcompost.com", "https://www.rootcompost.com（問い合わせページ）"),
    (24, "キッチン・調理器具", "HiCOZY F3 Slushie Machine", "HiCOZY", "Gadget Flow", "https://thegadgetflow.com/product/hicozy-f3-slushie-machine-soft-serve-ice-cream-machine-slushie-machine/", "https://www.hicozy.com", "https://www.hicozy.com（問い合わせページ）"),
    (25, "キッチン・調理器具", "ESPRO P7 French Press Coffee Maker", "ESPRO", "Gadget Flow", "https://thegadgetflow.com/product/espro-p7-french-press-coffee-maker/", "https://www.espro.com", "https://www.espro.com/pages/contact（問い合わせページ）"),
    (26, "スマートホーム・インテリア・照明", "Poplight Tool-Free Rechargeable Wall Light", "Poplight", "Poplight公式サイト", "https://thepoplight.com/", "https://thepoplight.com", "hello@thepoplight.com"),
    (27, "スマートホーム・インテリア・照明", "SANDSARA Rectangular Kinetic Sand Art Coffee Table", "SANDSARA", "SANDSARA公式サイト", "https://www.sandsara.io/store", "https://www.sandsara.io", "hello@sandsara.io"),
    (28, "スマートホーム・インテリア・照明", "Gravity Universe Time 720° Levitating Clock", "Gravity", "Gadget Flow", "https://thegadgetflow.com/product/gravity-universe-time-levitating-clock/", "https://gravitylev.com", "https://gravitylev.com（問い合わせページ）"),
    (29, "スマートホーム・インテリア・照明", "LiberNovo Omni Smart Ergonomic Chair", "LiberNovo", "LiberNovo公式サイト", "https://libernovo.com/products/libernovo-omni-dynamic-ergonomic-chair", "https://libernovo.com", "support@libernovo.com"),
    (30, "スマートホーム・インテリア・照明", "LiberNovo Maxis Large-Body Ergonomic Chair", "LiberNovo", "Gadget Flow", "https://thegadgetflow.com/product/libernovo-maxis-ergonomic-chair/", "https://libernovo.com", "support@libernovo.com"),
    (31, "スマートホーム・インテリア・照明", "Orbitkey Grid Modular Desk Organizer", "Orbitkey", "Kickstarter / Orbitkey公式サイト", "https://www.orbitkey.com/blogs/newsroom/desk-organiser-kickstarter", "https://www.orbitkey.com", "hello@orbitkey.com"),
    (32, "スマートホーム・インテリア・照明", "Numi B1 Self-Changing Smart Bin", "Numi", "Gadget Flow", "https://thegadgetflow.com/product/numi-b1-self-changing-smart-bin/", "https://www.numi.ai", "https://www.numi.ai（問い合わせページ）"),
    (33, "スマートホーム・インテリア・照明", "LitONES EaseGlow H02 Dual-Direction Desk Lamp", "LitONES", "Gadget Flow", "https://thegadgetflow.com/product/litones-easeglow-h02-dual-direction-lighting-desk-lamp-for-home-offices/", "https://www.litones.com", "https://www.litones.com（問い合わせページ）"),
    (34, "スマートホーム・インテリア・照明", "MyRainbowClock LED Color Therapy Wall Clock", "Synergy Clock", "Gadget Flow", "https://thegadgetflow.com/product/synergy-clock-company-creative-clocks/", "https://myrainbowclock.com", "https://myrainbowclock.com（問い合わせページ）"),
    (35, "スマートホーム・インテリア・照明", "Laundry Chair（ランドリーレール付きチェア）", "Laundry Chair", "Gadget Flow", "https://thegadgetflow.com/product/laundry-chair-bedroom-storage-chair/", "https://laundrychairofficial.com", "https://laundrychairofficial.com（問い合わせページ）"),
    (36, "スマートホーム・インテリア・照明", "Sleepal AI Lamp Contactless Sleep Tracker", "Sleepal", "Sleepal公式サイト", "https://sleepal.ai/", "https://sleepal.ai", "support@sleepal.ai"),
    (37, "スマートホーム・インテリア・照明", "Mirakuru Self-Adjusting Air Pillow", "TVC (Mirakuru)", "Kickstarter", "https://www.kickstarter.com/projects/1735022062/mirakuru-pillow-better-mornings-better-days", "https://mirakurupillow.com", "https://mirakurupillow.com（問い合わせページ）"),
    (38, "スマートホーム・インテリア・照明", "Monar Hi-Fi Canvas Speaker", "Monar", "Gadget Flow", "https://thegadgetflow.com/product/monar-hi-fi-canvas-speaker-wireless-speaker/", "https://www.monarspeakers.com", "https://www.monarspeakers.com（問い合わせページ）"),
    (39, "スマートホーム・インテリア・照明", "TRETTITRE LP3/DP3/CP3 Wall-Mounted Audio System", "TRETTITRE", "Gadget Flow", "https://thegadgetflow.com/product/trettitre-wall-mounted-audio-system/", "https://www.trettitre.com", "https://www.trettitre.com（問い合わせページ）"),
    (40, "スマートホーム・インテリア・照明", "ZESTUM Scentbar Real-Time Scent Gaming Controller", "ZESTUM", "Gadget Flow", "https://thegadgetflow.com/product/the-scentbar-smell-your-games-scent-controller/", "https://www.zestum.com", "https://www.zestum.com（問い合わせページ）"),
    (41, "ウェアラブル・ヘルス・フィットネス", "Lumia 2 Smart Earrings Health Tracker", "Lumia Health", "Lumia Health公式サイト", "https://lumia2.lumiahealth.com/", "https://lumiahealth.com", "support@lumiahealth.com"),
    (42, "ウェアラブル・ヘルス・フィットネス", "VITA RING Proactive AI Smart Ring", "VITA Technology", "VITA RING公式サイト", "https://shop.vitaring.tech/", "https://vitaring.tech", "jingen@vitaring.tech"),
    (43, "ウェアラブル・ヘルス・フィットネス", "CUDIS 002 Sporty Smart Ring with Wearable Coach", "CUDIS", "CUDIS公式サイト", "https://www.cudis.xyz/products/cudis-002-classic-ring", "https://www.cudis.xyz", "support@cudis.xyz"),
    (44, "ウェアラブル・ヘルス・フィットネス", "PIN Pulse Smart Ring 16 Health Metrics", "PIN", "Gadget Flow", "https://thegadgetflow.com/product/pin-pulse-smart-ring-health-tracker/", "https://pinpulse.com", "https://pinpulse.com（問い合わせページ）"),
    (45, "ウェアラブル・ヘルス・フィットネス", "Omnilux Contour Face LED Therapy Mask", "Omnilux", "Gadget Flow", "https://thegadgetflow.com/product/omnilux-contour-face-led-therapy/", "https://omniluxled.com", "https://omniluxled.com/pages/contact（問い合わせページ）"),
    (46, "ウェアラブル・ヘルス・フィットネス", "Pebblebee Halo with Alert Live Personal Safety Device", "Pebblebee", "Pebblebee公式サイト", "https://pebblebee.com/products/halo", "https://pebblebee.com", "support@pebblebee.com"),
    (47, "ウェアラブル・ヘルス・フィットネス", "iRESTORE Elite Hair Restoration Helmet", "iRESTORE", "iRESTORE公式サイト", "https://www.irestorelaser.com/", "https://www.irestorelaser.com", "https://www.irestorelaser.com/pages/contact（問い合わせページ）"),
    (48, "ウェアラブル・ヘルス・フィットネス", "Circular Ring 2 Smart Ring（AFib検知搭載）", "Circular", "Circular公式サイト", "https://circular.app/", "https://circular.app", "https://circular.app/pages/contact（問い合わせページ）"),
    (49, "ウェアラブル・ヘルス・フィットネス", "Solly 280W Solar Power Bank with Wall Charger", "Solly", "Gadget Flow", "https://thegadgetflow.com/product/solly-280w-solar-power-bank/", "https://sollyenergy.com", "https://sollyenergy.com（問い合わせページ）"),
    (50, "ウェアラブル・ヘルス・フィットネス", "ANYPIN Note Pod P1 AI Wearable Voice Recorder", "ANYPIN", "Gadget Flow", "https://thegadgetflow.com/product/anypin-note-pod-p1-wi-fi-voice-recorder/", "https://anypin.io", "https://anypin.io（問い合わせページ）"),
    (51, "ウェアラブル・ヘルス・フィットネス", "Vocci Ring AI Note-Taking Ring", "Vocci", "Gadget Flow", "https://thegadgetflow.com/product/vocci-ring-ai-note-taking-ring/", "https://vocciring.com", "https://vocciring.com（問い合わせページ）"),
    (52, "ウェアラブル・ヘルス・フィットネス", "Looki L1 AI Wearable Clip-On Life Camera", "Looki", "Gadget Flow", "https://thegadgetflow.com/product/looki-l1-ai-wearable/", "https://looki.ai", "https://looki.ai（問い合わせページ）"),
    (53, "ウェアラブル・ヘルス・フィットネス", "Acer Aspire Badge AD Smart Wearable for Kids", "Acer", "Gadget Flow", "https://thegadgetflow.com/product/acer-aspire-badge-ad-smart-wearable-for-kids/", "https://www.acer.com", "https://www.acer.com/us-en/support（問い合わせページ）"),
    (54, "ウェアラブル・ヘルス・フィットネス", "HUAWEI WATCH FIT 5 Pro Smartwatch", "HUAWEI", "Gadget Flow", "https://thegadgetflow.com/product/huawei-watch-fit-5-pro-smartwatch/", "https://www.huawei.com", "https://consumer.huawei.com/us/support（問い合わせページ）"),
    (55, "ウェアラブル・ヘルス・フィットネス", "BodyPark ATOM AI Fitness Companion", "BodyPark", "BodyPark公式サイト", "https://bodypark.io", "https://bodypark.io", "https://bodypark.io（問い合わせページ）"),
    (56, "ウェアラブル・ヘルス・フィットネス", "Aeke K1 Smart Home Gym", "Aeke", "Aeke公式サイト", "https://www.aeke.com/", "https://www.aeke.com", "https://www.aeke.com/pages/contact（問い合わせページ）"),
    (57, "ウェアラブル・ヘルス・フィットネス", "Solos Wellness Smart Glasses（AIコーチ内蔵）", "Solos", "Solos公式サイト", "https://www.solosglasses.com/", "https://www.solosglasses.com", "https://www.solosglasses.com/pages/contact（問い合わせページ）"),
    (58, "ウェアラブル・ヘルス・フィットネス", "Momease PetalSoft Dual-Material Breast Pump Flange", "Momease", "Gadget Flow", "https://thegadgetflow.com/product/momease-petalsoft-flange-breast-pump/", "https://momease.com", "https://momease.com（問い合わせページ）"),
    (59, "ウェアラブル・ヘルス・フィットネス", "Hotsales Neck Fan Wearable Cooling", "Hotsales", "Gadget Flow", "https://thegadgetflow.com/product/hotsales-neck-fan-wearable-cooling/", "https://hotsales.com", "https://hotsales.com（問い合わせページ）"),
    (60, "ウェアラブル・ヘルス・フィットネス", "écoute TH2 Portable Vacuum Tube Hi-Fi Headphones", "écoute", "Gadget Flow", "https://thegadgetflow.com/product/ecoute-th2-wearable-hifi-headphones/", "https://ecouteaudio.com", "https://ecouteaudio.com（問い合わせページ）"),
    (61, "アウトドア・スポーツ・旅行", "HOVERAir AQUA Waterproof Self-Flying Drone", "HOVERAir", "HOVERAir公式サイト", "https://hoverair.com/pages/aqua", "https://hoverair.com", "support@hoverair.com"),
    (62, "アウトドア・スポーツ・旅行", "CYPLORE World's Lightest E-Assist Bike Kit", "CYPLORE", "CYPLORE公式サイト", "https://www.cyplore.com/", "https://www.cyplore.com", "https://www.cyplore.com（問い合わせページ）"),
    (63, "アウトドア・スポーツ・旅行", "Roam Ducky Pocket-Sized Universal Travel Adapter", "MOGICS", "Kickstarter", "https://www.kickstarter.com/projects/mogics/roam-ducky-all-in-one-travel-adapter-fits-in-your-pocket", "https://www.mogics.com", "info@mogics.com"),
    (64, "アウトドア・スポーツ・旅行", "Peak Design Roller Pro（モジュラー式キャリーオン）", "Peak Design", "Peak Design公式サイト", "https://www.peakdesign.com/", "https://www.peakdesign.com", "https://www.peakdesign.com/pages/contact（問い合わせページ）"),
    (65, "アウトドア・スポーツ・旅行", "VIPERPOD Full-Size Hanging Hook Tripod System", "VIPERPOD", "Gadget Flow", "https://thegadgetflow.com/product/viperpod-hanging-tripod/", "https://viperpod.com", "https://viperpod.com（問い合わせページ）"),
    (66, "アウトドア・スポーツ・旅行", "Reeflex Ultra Telephoto Lens for Smartphone", "Reeflex", "Gadget Flow", "https://thegadgetflow.com/product/reeflex-ultra-telephoto-smartphone-lens/", "https://reeflexlens.com", "https://reeflexlens.com（問い合わせページ）"),
    (67, "アウトドア・スポーツ・旅行", "BougeRV Portable Coffee Maker with 144Wh Battery", "BougeRV", "Gadget Flow", "https://thegadgetflow.com/product/bougerv-portable-coffee-maker-with-144wh-battery-camping-coffeemaker/", "https://www.bougerv.com", "https://www.bougerv.com/pages/contact（問い合わせページ）"),
    (68, "アウトドア・スポーツ・旅行", "CamperKit 11-in-1 Outdoor Solution", "Aecooly", "Kickstarter", "https://www.kickstarter.com/projects/aecooly/camperkit-11-in-1-outdoor-solution-for-every-adventure", "https://aecooly.com", "https://aecooly.com（問い合わせページ）"),
    (69, "アウトドア・スポーツ・旅行", "I'm Back Roll APS-C Digital Film Roll for 35mm Cameras", "I'm Back", "Gadget Flow", "https://thegadgetflow.com/product/im-back-roll-digital-film-converter/", "https://www.imback.eu", "https://www.imback.eu（問い合わせページ）"),
    (70, "アウトドア・スポーツ・旅行", "Collapsy Folding Travel Hanger", "Collapsy", "Gadget Flow", "https://thegadgetflow.com/product/collapsy-folding-hanger/", "https://collapsy.com", "https://collapsy.com（問い合わせページ）"),
    (71, "アウトドア・スポーツ・旅行", "Dometic Recon 360 Faucet（アウトドア用多機能蛇口）", "Dometic", "Gadget Flow", "https://thegadgetflow.com/product/dometic-recon-360-faucet/", "https://www.dometic.com", "https://www.dometic.com/en-us/contact（問い合わせページ）"),
    (72, "アウトドア・スポーツ・旅行", "Anker SOLIX S2000 Portable Power Station", "Anker", "Gadget Flow / Amazon.com", "https://thegadgetflow.com/product/anker-solix-s2000-portable-power-station/", "https://www.anker.com", "https://www.anker.com/pages/contact（問い合わせページ）"),
    (73, "アウトドア・スポーツ・旅行", "Nayo Smart Herman Pro Half Roll-Top Backpack", "Nayo Smart", "Gadget Flow", "https://thegadgetflow.com/product/nayo-smart-herman-pro-half-roll-top-backpack/", "https://nayosmart.com", "https://nayosmart.com（問い合わせページ）"),
    (74, "アウトドア・スポーツ・旅行", "Aceii One AI Tennis Training Robot", "Aceii", "Gadget Flow", "https://thegadgetflow.com/product/aceii-one-ai-tennis-robot/", "https://aceii.ai", "https://aceii.ai（問い合わせページ）"),
    (75, "アウトドア・スポーツ・旅行", "Pinned Golf SoundStick PRO Golf Speaker", "Pinned", "Gadget Flow", "https://thegadgetflow.com/product/pinned-golfsound-stick-pro-golf-speaker/", "https://pinnedgolf.com", "https://pinnedgolf.com（問い合わせページ）"),
    (76, "ペット用品", "Fetchly Instant-Lock Dog Leash（磁気ワンタッチリード）", "Fetchly", "Fetchly公式サイト", "https://fetchly.eu/", "https://fetchly.eu", "https://fetchly.eu（問い合わせページ）"),
    (77, "ペット用品", "PettiChat Real-Time AI Pet Translator", "PettiChat", "PettiChat公式サイト", "https://pettichat.com/", "https://pettichat.com", "pettichatofficial@gmail.com"),
    (78, "ペット用品", "Petscast Smart AI Pet Monitoring Camera", "Petscast", "Kickstarter", "https://www.kickstarter.com/projects/petscast-petcamera/petscast-smart-ai-pet-monitoring-camera", "https://petscast.app", "https://petscast.app（問い合わせページ）"),
    (79, "ペット用品", "Groomatic Cat Brush（自動起動ネコ用ブラシ）", "Groomatic", "Groomatic公式サイト", "https://groomatic.com/", "https://groomatic.com", "https://groomatic.com（問い合わせページ）"),
    (80, "ペット用品", "Cheerble Wicked Ball Interactive Pet Toy", "Cheerble", "Cheerble公式サイト", "https://www.cheerble.com/", "https://www.cheerble.com", "https://www.cheerble.com/pages/contact（問い合わせページ）"),
    (81, "ペット用品", "Easy Lock Magnetic Pet Harness & Collar System", "Easy Lock", "Easy Lock公式サイト", "https://www.easylocksmart.com/", "https://www.easylocksmart.com", "https://www.easylocksmart.com（問い合わせページ）"),
    (82, "ペット用品", "Sleeptracker AI Pet Sleep Monitor", "Sleeptracker", "Sleeptracker公式サイト", "https://sleeptracker.com/", "https://sleeptracker.com", "https://sleeptracker.com/en/support（問い合わせページ）"),
    (83, "テクノロジー・ガジェット", "Shargeek 300 Transparent 300W Power Bank", "SHARGE", "SHARGE公式サイト", "https://sharge.com/products/shargeek-300", "https://sharge.com", "info@sharge.com"),
    (84, "テクノロジー・ガジェット", "Revopoint POP 4 Hybrid Blue Laser 3D Scanner", "Revopoint 3D", "Revopoint公式サイト", "https://www.revopoint3d.com/", "https://www.revopoint3d.com", "customer@revopoint3d.com"),
    (85, "テクノロジー・ガジェット", "Makera Z1 Desktop CNC Machine for Metal/Wood", "Makera", "Makera公式サイト", "https://www.makera.com/pages/makera-z1", "https://www.makera.com", "info@makera.com"),
    (86, "テクノロジー・ガジェット", "CODEE Screen-Free AI Coding Robot Lab for Kids", "CODEE", "Gadget Flow", "https://thegadgetflow.com/product/codee-ai-powered-coding-lab/", "https://codee.ai", "https://codee.ai（問い合わせページ）"),
    (87, "テクノロジー・ガジェット", "ZIMO1 27-inch Glasses-Free 3D Monitor", "Zondision", "Gadget Flow", "https://thegadgetflow.com/product/zimo1-27inch-3d-display/", "https://zondision.com", "https://zondision.com（問い合わせページ）"),
    (88, "テクノロジー・ガジェット", "Vastnaut One 4x4 Exoskeleton for Rugged Terrain", "Vastnaut", "Gadget Flow", "https://thegadgetflow.com/product/vastnaut-one-4x4-exoskeleton-for-rugged-terrain/", "https://vastnaut.com", "https://vastnaut.com（問い合わせページ）"),
    (89, "テクノロジー・ガジェット", "LumioClaw AI Projector Learning Buddy for Kids", "LumioClaw", "Gadget Flow", "https://thegadgetflow.com/product/lumioclaw-ai-projector-learning-buddy/", "https://lumioclaw.com", "https://lumioclaw.com（問い合わせページ）"),
    (90, "テクノロジー・ガジェット", "Tembo Magnetic Drum Machine and Sampler", "Tembo", "Gadget Flow", "https://thegadgetflow.com/product/tembo-a-new-musical-instrument-for-playful-music-making-magnetic-drum-machine/", "https://tembomusic.com", "https://tembomusic.com（問い合わせページ）"),
    (91, "テクノロジー・ガジェット", "Melo-D AI Guitar Anyone Can Play Instantly", "Melo-D", "Gadget Flow", "https://thegadgetflow.com/product/melo-d-ai-guitar/", "https://melod.io", "https://melod.io（問い合わせページ）"),
    (92, "テクノロジー・ガジェット", "Pixboom Spark 4K High-Speed Global Shutter Camera", "Pixboom", "Gadget Flow", "https://thegadgetflow.com/product/pixboom-spark-high-speed-camera/", "https://pixboom.io", "https://pixboom.io（問い合わせページ）"),
    (93, "テクノロジー・ガジェット", "MarkYin Artifact AI Card for Photos/Video/Audio", "MarkYin", "Gadget Flow", "https://thegadgetflow.com/product/markyin-artifact-ai-card/", "https://markyin.ai", "https://markyin.ai（問い合わせページ）"),
    (94, "テクノロジー・ガジェット", "open_slate 2-in-1 Tablet with Linux and Android", "open_slate", "Gadget Flow", "https://thegadgetflow.com/product/open-slate-private-2-in-1-tablet/", "https://openslate.tech", "https://openslate.tech（問い合わせページ）"),
    (95, "テクノロジー・ガジェット", "R-Go Waver Ambidextrous Vertical Ergonomic Mouse", "R-Go Tools", "Gadget Flow", "https://thegadgetflow.com/product/r-go-waver-ergonomic-mouse/", "https://www.r-go-tools.com", "https://www.r-go-tools.com/en/contact（問い合わせページ）"),
    (96, "テクノロジー・ガジェット", "GAMR Active Gaming Play Pad 360° Sensor", "GAMR", "Gadget Flow", "https://thegadgetflow.com/product/gamr-active-gaming-play-pad-active-gaming-pad/", "https://gamrplay.com", "https://gamrplay.com（問い合わせページ）"),
    (97, "テクノロジー・ガジェット", "Vellvii DOX Luxury Docking Storage Charging Vault", "Vellvii", "Gadget Flow", "https://thegadgetflow.com/product/vellvii-dox-luxury-docking-system/", "https://vellvii.com", "https://vellvii.com（問い合わせページ）"),
    (98, "テクノロジー・ガジェット", "BB-777 Retro Boombox with Modern Audio Features", "BB-777", "Gadget Flow", "https://thegadgetflow.com/product/bb-777-iconic-boombox-retro-speaker/", "https://bb777.audio", "https://bb777.audio（問い合わせページ）"),
    (99, "テクノロジー・ガジェット", "Skywave X100 Dual 9.2.6 THX Wireless Home Theater", "Ultimea", "Gadget Flow", "https://thegadgetflow.com/product/skywave-x100-dual-wireless-soundbar/", "https://ultimea.com", "https://ultimea.com/pages/contact（問い合わせページ）"),
    (100, "テクノロジー・ガジェット", "Pixelpaw Labs Phase Modular Gaming Mouse Controller", "Pixelpaw Labs", "Gadget Flow", "https://thegadgetflow.com/product/phase-modular-gaming-controller/", "https://pixelpawlabs.com", "https://pixelpawlabs.com（問い合わせページ）"),
    (101, "テクノロジー・ガジェット", "Sleepal AI Lamp（接触なし睡眠トラッカー）", "Sleepal", "Sleepal公式サイト", "https://sleepal.ai", "https://sleepal.ai", "support@sleepal.ai"),
    (102, "アウトドア・スポーツ・旅行", "HOVERAir AQUA（防水自律飛行カメラドローン）", "HOVERAir", "HOVERAir公式サイト（グローバル）", "https://global.hoverair.com/products/hoverair-aqua", "https://hoverair.com", "support@hoverair.com"),
    (103, "アウトドア・スポーツ・旅行", "VIPERPOD Hanging Tripod System", "VIPERPOD", "Gadget Flow", "https://thegadgetflow.com/product/viperpod-hanging-tripod/", "https://viperpod.com", "https://viperpod.com（問い合わせページ）"),
    (104, "ウェアラブル・ヘルス・フィットネス", "écoute TH2 Portable Vacuum Tube Headphones", "écoute", "Gadget Flow", "https://thegadgetflow.com/product/ecoute-th2-wearable-hifi-headphones/", "https://ecouteaudio.com", "https://ecouteaudio.com（問い合わせページ）"),
    (105, "ウェアラブル・ヘルス・フィットネス", "Lumia 2 Smart Earrings Health Tracker（追加モデル）", "Lumia Health", "Lumia Health公式サイト", "https://lumiahealth.com/", "https://lumiahealth.com", "support@lumiahealth.com"),
]

for i, row in enumerate(data, 2):
    for col, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.font = data_font
        cell.border = border
        cell.alignment = data_align
        if i % 2 == 0:
            cell.fill = alt_fill
    cat = row[1]
    cat_cell = ws.cell(row=i, column=2)
    cat_cell.font = cat_font
    cat_cell.fill = cat_fills.get(cat, default_cat_fill)
    ws.row_dimensions[i].height = 28

col_widths = [6, 22, 55, 22, 28, 65, 35, 50]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{len(data)+1}"

# Sheet 2: 優先推奨TOP10
ws2 = wb.create_sheet("優先推奨TOP10")
top10_headers = ["順位", "商品名", "メーカー名", "販売ECサイト名", "日本向け推奨理由", "商品URL", "連絡先メール"]
ws2.append(top10_headers)
for col, h in enumerate(top10_headers, 1):
    cell = ws2.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
ws2.row_dimensions[1].height = 30

top10 = [
    (1, "Cozytime Lumo AI Infrared Grill", "Cozytime", "Cozytime公式サイト", "煙なし室内BBQ。日本の集合住宅事情に最適。CES 2026話題作", "https://cozytime-tech.com/pages/lumo", "customers@cozytime-tech.com"),
    (2, "LISSOME R1 Compact Smart Portable Dishwasher", "LISSOME", "LISSOME公式サイト", "一人暮らし・狭小キッチン向けポータブル食洗機、日本の需要大", "https://lissomesmart.com/", "eakxu19495120@gmail.com"),
    (3, "Sleepal AI Lamp Contactless Sleep Tracker", "Sleepal", "Sleepal公式サイト", "睡眠意識の高い日本市場に最適。ウェアラブル不要で差別化", "https://sleepal.ai/", "support@sleepal.ai"),
    (4, "Seattle Ultrasonics C-200 Ultrasonic Chef's Knife", "Seattle Ultrasonics", "Seattle Ultrasonics公式サイト", "超音波技術×料理ガジェット。日本の料理文化・プロ需要に直結", "https://seattleultrasonics.com/products/c-200-ultrasonic-8-chefs-knife", "hello@seattleultrasonics.com"),
    (5, "Lumia 2 Smart Earrings Health Tracker", "Lumia Health", "Lumia Health公式サイト", "世界初の健康追跡スマートイヤリング。女性市場に強く訴求", "https://lumia2.lumiahealth.com/", "support@lumiahealth.com"),
    (6, "PettiChat Real-Time AI Pet Translator", "PettiChat", "PettiChat公式サイト", "ペット大国日本で高い需要見込み。話題性・SNS拡散力大", "https://pettichat.com/", "pettichatofficial@gmail.com"),
    (7, "Nosh Autonomous Cooking Robot", "Euphotic Labs", "Nosh公式サイト", "AI自律調理ロボット。共働き家庭・時短需要に対応", "https://www.letsnosh.io/", "support@letsnosh.io"),
    (8, "CYPLORE World's Lightest E-Assist Bike Kit", "CYPLORE", "CYPLORE公式サイト", "自転車文化の強い日本に最適な電動アシストキット", "https://www.cyplore.com/", "https://www.cyplore.com（問い合わせページ）"),
    (9, "Orbitkey Grid Modular Desk Organizer", "Orbitkey", "Kickstarter / Orbitkey公式サイト", "テレワーク普及で需要拡大中のモジュラーデスク収納", "https://www.orbitkey.com/blogs/newsroom/desk-organiser-kickstarter", "hello@orbitkey.com"),
    (10, "Shargeek 300 Transparent 300W Power Bank", "SHARGE", "SHARGE公式サイト", "ガジェット好きに刺さる透明デザイン×高出力300W", "https://sharge.com/products/shargeek-300", "info@sharge.com"),
]
for i, row in enumerate(top10, 2):
    for col, val in enumerate(row, 1):
        cell = ws2.cell(row=i, column=col, value=val)
        cell.font = data_font
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if i % 2 == 0:
            cell.fill = alt_fill
    ws2.row_dimensions[i].height = 40

top10_widths = [8, 50, 22, 28, 55, 65, 45]
for col, width in enumerate(top10_widths, 1):
    ws2.column_dimensions[get_column_letter(col)].width = width
ws2.freeze_panes = "A2"

wb.save(r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸105件.xlsx")
print("Done")
