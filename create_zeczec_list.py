from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "ZECZEC日本未発売ガジェット"

headers = ["カテゴリ", "商品名", "メーカー名", "ZECZEC商品ページURL", "メーカーHP", "コンタクトメール", "日本市場スコア", "日本市場コメント"]

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

products = [
    # ガジェット・充電
    ("ガジェット・充電", "Acer Gadget 碁速充 Qi2 五合一 35W 固態行動電源", "Acer Gadget", "https://www.zeczec.com/projects/acer-gadget-jisuchong", "https://www.acer.com/tw", "要調査", "3/5", "Qi2対応の多機能モバイルバッテリー。Acerブランドの信頼性はあるが日本でのQi2充電器市場は競争激しい"),
    ("ガジェット・充電", "Mag Slim Qi2 磁吸行動電源 世界最薄7.3mm", "Mag Slim", "https://www.zeczec.com/projects/mag-slim-qi2-2025-10-16", "要調査", "要調査", "4/5", "世界最薄7.3mmのQi2磁気モバイルバッテリー。iPhone×MagSafeユーザーが多い日本市場に高い訴求"),
    ("ガジェット・充電", "Allite 氮化鎵快充 史上最小65W雙孔", "Allite", "https://www.zeczec.com/projects/allite", "https://www.allite.com.tw", "要調査", "4/5", "GaN採用の超小型65W充電器。デザイン性が高くAnkerとの差別化が図れる"),
    ("ガジェット・充電", "Allite WP1 Qi2 磁吸行動電源 追劇支架", "Allite", "https://www.zeczec.com/projects/wp1", "https://www.allite.com.tw", "要調査", "4/5", "Qi2公式認証取得・スタンド機能付き磁気モバイルバッテリー。iPhoneユーザーの多い日本に強い訴求"),
    ("ガジェット・充電", "Encore™ Wireless 無線座充 三機一体充電", "Encore", "https://www.zeczec.com/projects/encore-wireless", "要調査", "要調査", "3/5", "スマホ・イヤホン・スマートウォッチを同時ワイヤレス充電するデスクステーション"),
    ("ガジェット・充電", "Evo 超輕小旅行充電器 世界対応", "Evo", "https://www.zeczec.com/projects/evo", "要調査", "要調査", "3/5", "折り畳み式コンパクト旅行充電器。日本人旅行者に適しているが競合多数"),
    ("ガジェット・充電", "e-Fusion 行動電源×充電器×無線充電 三合一", "e-Fusion", "https://www.zeczec.com/projects/e-fusion", "要調査", "要調査", "3/5", "モバイルバッテリー+充電器+ワイヤレス充電の三合一デバイス。出張・旅行ユーザー向け"),
    ("ガジェット・充電", "LilNob 3C1A 四孔充電器 最高100W", "LilNob", "https://www.zeczec.com/projects/lilnob-3c1a", "要調査", "要調査", "3/5", "4ポートGaN 100W充電器。デザイン性と機能性のバランスで日本市場参入可能"),
    ("ガジェット・充電", "TITANSHIELD 固態行動電源 防爆五合一", "GrantClassic", "https://www.zeczec.com/projects/grantclassic", "要調査", "要調査", "3/5", "固体電池採用の安全設計モバイルバッテリー。爆発しない安全訴求は日本のシニア層に刺さる"),
    ("ガジェット・充電", "DEZCTOP 七合一 USB-C 集線器", "DEZCTOP", "https://www.zeczec.com/projects/dezctop-7in1-HUB", "要調査", "要調査", "3/5", "7-in-1 USB-CハブでMacBookユーザー多い日本のリモートワーカーに需要あり"),
    ("ガジェット・充電", "Roommi P02 多功能行動電源供應器", "Roommi", "https://www.zeczec.com/projects/roommi-P02", "要調査", "要調査", "3/5", "多機能モバイル電源。アウトドア・キャンプブームの日本市場に合う可能性"),
    ("ガジェット・充電", "CASA Hub O7 集線×無線充電デスクハブ", "CASA", "https://www.zeczec.com/projects/casahubo7", "https://www.casa.com.tw", "要調査", "3/5", "ハブとワイヤレス充電を統合したデスクアクセサリー。MacBookユーザーが多い日本に需要あり"),
    ("ガジェット・充電", "SuperCharger2.0 三合一 超級充電器", "SuperCharger", "https://www.zeczec.com/projects/supercharger2-0", "要調査", "要調査", "3/5", "充電アダプター+モバイルバッテリー+カードリーダー三合一。旅行・出張ユーザー向け"),
    ("ガジェット・充電", "EnergieMAX 60W PD+QC3.0 USB-C三孔旅充器", "EnergieMAX", "https://www.zeczec.com/projects/otca400pd", "要調査", "要調査", "3/5", "60W PD対応マルチポート旅行充電器。海外旅行需要が回復した日本市場で需要あり"),
    ("ガジェット・充電", "POLYBATT CAT 22W磁吸式自帯線行動電源", "POLYBATT", "https://www.zeczec.com/projects/POLYBATT_CAT", "要調査", "要調査", "3/5", "MagSafe対応ケーブル内蔵22Wモバイルバッテリー。カメラ傷防止設計が独自性あり"),
    ("ガジェット・充電", "GENKI Nintendo Switch 専用無線ユニット", "Human Things (GENKI)", "https://www.zeczec.com/projects/genki", "https://genkithings.com", "support@genkithings.com", "4/5", "Nintendo Switch専用Bluetooth音声送信機。日本のSwitch愛好家に刺さる製品"),

    # AR/VR・ウェアラブル
    ("AR/VR・ウェアラブル", "Rokid AR Spatial 三視窗智慧眼鏡 300吋プライベートスペース", "Rokid", "https://www.zeczec.com/projects/rokid-ar-spatial-300", "https://www.rokid.com", "要調査", "4/5", "300インチ仮想画面3ウィンドウARグラス。日本のAR眼鏡市場拡大中。エンタメ・ビジネス両用で高需要"),
    ("AR/VR・ウェアラブル", "Rokid Glasses 樂奇AI智慧眼鏡 未来から見える", "Rokid", "https://www.zeczec.com/projects/rokid-glasses", "https://www.rokid.com", "要調査", "4/5", "AI機能搭載スマートグラス。Meta Ray-Ban対抗製品として日本の先進ガジェット層に訴求"),
    ("AR/VR・ウェアラブル", "ASUS AirVision M1 100吋穿戴式顯示器", "ASUS", "https://www.zeczec.com/projects/airvision-m1", "https://www.asus.com/jp", "要調査", "3/5", "100インチウェアラブルディスプレイ87g超軽量。ASUSは日本展開済みだがこのモデルは台湾先行"),
    ("AR/VR・ウェアラブル", "Dream Glass 4K AR 攜帶式智慧眼鏡", "Dream Glass", "https://www.zeczec.com/projects/dreamglass4kAR", "要調査", "要調査", "3/5", "4K対応携帯型ARグラス。日本でのAR体験需要は高いが認知度構築が課題"),
    ("AR/VR・ウェアラブル", "SmartEcho AI智慧眼鏡 全台首款AR同步翻譯", "SmartEcho", "https://www.zeczec.com/projects/smartecho-glasses", "要調査", "要調査", "4/5", "リアルタイム翻訳ARグラス。インバウンド対応・海外旅行需要がある日本市場にドンピシャ"),
    ("AR/VR・ウェアラブル", "ROG XREAL R1 電競AR眼鏡 171吋 240Hz", "ASUS ROG × XREAL", "https://www.zeczec.com/projects/rog-xreal-r1-ar-glasses", "https://rog.asus.com", "要調査", "4/5", "世界初240Hz Micro OLED搭載ゲーミングARグラス。日本の熱烈なゲーマー層に高い訴求力"),
    ("AR/VR・ウェアラブル", "大朋E4 VR眼鏡 4K一体機 超軽量", "DPVR", "https://www.zeczec.com/projects/dpvr", "https://www.dpvrvr.com", "要調査", "3/5", "4K・120Hz VRヘッドセット。Meta Questの競合。日本市場参入には認知度向上が必要"),
    ("AR/VR・ウェアラブル", "RingConn Gen 2 AI 智慧戒指 睡眠健康", "RingConn", "https://www.zeczec.com/projects/cheertoktw-2025-06-13", "https://www.ringconn.com", "要調査", "4/5", "AI搭載スマートリング。Oura Ring対抗として日本の健康意識の高いユーザーに高い需要"),
    ("AR/VR・ウェアラブル", "LARMI 2025 AI 智能手錶 AI翻訳×GPS×5ATM", "LARMI", "https://www.zeczec.com/projects/Ai6", "要調査", "要調査", "3/5", "AIリアルタイム翻訳・GPS・5ATM防水スマートウォッチ。多機能×価格競争力で日本でも戦える"),
    ("AR/VR・ウェアラブル", "Wearbuds™ Hi-Fi 真無線耳機×智能腕錶一体型", "Wearbuds", "https://www.zeczec.com/projects/wearbuds", "要調査", "要調査", "3/5", "イヤホン×スマートウォッチ一体型デバイス。ユニークなコンセプトで日本の新しもの好きに訴求"),
    ("AR/VR・ウェアラブル", "PHILIPS TAT2630 タッチスクリーン搭載Bluetoothイヤホン", "Philips", "https://www.zeczec.com/projects/philips-tat2630", "https://www.philips.com/ja-jp", "要調査", "3/5", "1.45インチタッチスクリーン搭載TWS。Bluetooth 6.0対応でペット探し・タイマー機能付き"),

    # PC周辺・入力機器
    ("PC周辺・入力機器", "Vision Keyboard 10吋液晶螢幕内蔵キーボード AI対応", "Vision Keyboard", "https://www.zeczec.com/projects/visionkeyboard", "要調査", "要調査", "4/5", "10インチタッチスクリーン内蔵キーボード。ChatGPT連携機能付きで在宅ワーカーに高需要"),
    ("PC周辺・入力機器", "TRACPOINT 実体+仮想 二合一マウス×プレゼンペン", "TRACPOINT", "https://www.zeczec.com/projects/tracpoint", "要調査", "要調査", "4/5", "物理+仮想一体のプレゼンテーションペン/マウス。日本のビジネスマン・教育者に強い訴求力"),
    ("PC周辺・入力機器", "Winmaxle B1 迷你體感鍵盤 掌中BluetoothIME", "Winmaxle", "https://www.zeczec.com/projects/winmaxle-b1", "要調査", "要調査", "4/5", "ジャイロセンサー搭載84gの手のひらサイズBTキーボード。XR/VR用途やプレゼンに使えるユニーク製品"),
    ("PC周辺・入力機器", "DINOTAENG QUOKKA 無線キーボードマウスセット", "Norns", "https://www.zeczec.com/projects/norns_DINOTAENG", "要調査", "要調査", "3/5", "韓国キャラクターDINOTAENG×デザイン周辺機器。韓流ファン層の多い日本で訴求できる"),
    ("PC周辺・入力機器", "MOZTECH カカ螢幕燈 折疊携帯+無線モニターライト", "MOZTECH", "https://www.zeczec.com/projects/kakalight", "https://www.moztech.com.tw", "要調査", "4/5", "折り畳み式ワイヤレスモニターライト。在宅ワーク×コンパクト設計で日本のデスク環境に最適"),
    ("PC周辺・入力機器", "GOOVIS LITE 世界最鮮明 Micro-OLED 3D HMD", "GOOVIS", "https://www.zeczec.com/projects/goovis-lite", "https://www.goovis.com", "要調査", "3/5", "Micro-OLED搭載高解像度HMD。4K映像愛好家・シネマ体験需要がある日本でニッチ層に刺さる"),
    ("PC周辺・入力機器", "Photontree PRO 3D HMD 965吋 デュアル2K", "Photontree", "https://www.zeczec.com/projects/PT-PRO", "要調査", "要調査", "3/5", "デュアル2K IPS搭載3D HMD。映画・ゲーム愛好者向けプレミアムHMD"),
    ("PC周辺・入力機器", "ldeao hub 24吋 タッチ集線ディスプレイ エルゴノミクス", "Ideao", "https://www.zeczec.com/projects/FTI24JECTOR", "要調査", "要調査", "3/5", "24インチタッチ機能付きUSB-C集線ディスプレイ。デザイナー・クリエイター向けに日本でも需要あり"),
    ("PC周辺・入力機器", "AIFA i-Ctrl Pro Plus 家電遠端遙控×智慧防霉", "AIFA (艾電家)", "https://www.zeczec.com/projects/AIFASmart-iCtrl-Pro-Plus", "https://www.aifa.com.tw", "service@aifa.com.tw", "3/5", "全家電赤外線リモコン×AIカビ防止機能IoTデバイス。古い家電が多い日本でスマートホーム化に貢献"),
    ("PC周辺・入力機器", "JCH422 iPad/iPhone-Windows対傳視訊分享器", "JCH", "https://www.zeczec.com/projects/JCH422", "要調査", "要調査", "4/5", "iPad/iPhoneとWindows間クロスプラットフォーム映像共有デバイス。Apple×Windows混在環境の多い日本企業向け"),
    ("PC周辺・入力機器", "LG StanbyME 無線移式タッチスクリーン 27型", "LG", "https://www.zeczec.com/projects/lgstanbyme", "https://www.lg.com/jp", "jp.lgservice@lge.com", "3/5", "ワイヤレス・移動式27型タッチスクリーン。LGは日本展開済みだがStanbyMeの台湾特化モデルの可能性"),
    ("PC周辺・入力機器", "SAMSUNG Smart Monitor M7 43吋 4K UHD", "Samsung", "https://www.zeczec.com/projects/SAMSUNG-M7", "https://www.samsung.com/jp", "要調査", "2/5", "Samsung 43型4K UHDスマートモニター。Samsungは日本市場展開済みのため独自性は低い"),
    ("PC周辺・入力機器", "Muigic Tu1 小沐智能タッチ制御FHDプロジェクター", "Muigic", "https://www.zeczec.com/projects/MuigicTu1", "要調査", "要調査", "4/5", "タッチ操作対応FHD高画質プロジェクター。ホームシアター化が進む日本市場に高需要"),

    # オーディオ・録音
    ("オーディオ・録音", "HiDock P1 AI録音器 会話整理AI秘書", "HiDock", "https://www.zeczec.com/projects/hidock_p1", "要調査", "要調査", "5/5", "AI文字起こし・要約機能搭載Bluetooth録音機。議事録自動化ニーズが高い日本ビジネス市場に超高需要"),
    ("オーディオ・録音", "Slimca 超薄録音カード 随録随傳", "Slimca", "https://www.zeczec.com/projects/slimca", "要調査", "要調査", "4/5", "超薄型ICレコーダーカード。名刺サイズの携帯性で日本のビジネスパーソンに刺さる"),
    ("オーディオ・録音", "TileRec 隠形録音片 全球最小智慧録音設備", "TileRec", "https://www.zeczec.com/projects/tilerec", "要調査", "要調査", "3/5", "世界最小級スマート録音デバイス。超小型・長時間録音で日本のビジネスユーザーに需要あり"),
    ("オーディオ・録音", "Kando 藍牙喇叭麥克風組 一鍵消原音KTV", "Kando", "https://www.zeczec.com/projects/kando-ktv", "要調査", "要調査", "3/5", "Bluetoothスピーカー+デュアルマイク一体型。日本のカラオケ文化との親和性高い"),
    ("オーディオ・録音", "PHILIPS 小福麥 360°環繞Bluetoothスピーカー×マイク", "Philips", "https://www.zeczec.com/projects/philips-fumai", "https://www.philips.com/ja-jp", "要調査", "3/5", "360度サラウンドBTスピーカー×AIマイク一体型。Philipsは日本展開済みだがこのモデルは台湾先行"),
    ("オーディオ・録音", "CELIA & PERAH 即興自組音響 モジュラーカスタム", "CELIA & PERAH", "https://www.zeczec.com/projects/Freestyle-Moment", "要調査", "要調査", "3/5", "カスタマイズ可能なモジュラースピーカー。DIY・個性重視の日本オーディオ愛好家に訴求"),
    ("オーディオ・録音", "ZMI PurPods 真無線耳機 雙動圈設計 高音質", "ZMI (紫米)", "https://www.zeczec.com/projects/zmi_purpods_TW101", "https://www.zmi.com", "要調査", "3/5", "デュアルドライバー採用コスパ優秀TWS。AirPods対抗として日本の学生・若年層に訴求"),
    ("オーディオ・録音", "M6 擴增音響 自由搭配 聴覚拡張", "M6 Audio", "https://www.zeczec.com/projects/CnPm6", "要調査", "要調査", "3/5", "モジュール式拡張対応スピーカーシステム。カスタム志向の日本のオーディオ市場にニッチ需要"),

    # スマートロック・セキュリティ
    ("スマートロック・セキュリティ", "小島快鎖 TX 雙鏡頭監視器×掌靜脈×人臉辨識", "小島快鎖", "https://www.zeczec.com/projects/TXsmartlock", "要調査", "要調査", "5/5", "顔認証+掌静脈認証+デュアルカメラ搭載スマートロック。防犯意識の高い日本で最高水準の需要"),
    ("スマートロック・セキュリティ", "Acer Gadget 碁智鎖 五合一智能電子門鎖", "Acer Gadget", "https://www.zeczec.com/projects/acer", "https://www.acer.com/tw", "要調査", "4/5", "指紋・カード・パスワード等5種解錠対応スマートロック。Acerブランドで信頼性あり日本市場参入余地大"),
    ("スマートロック・セキュリティ", "FAMMIX SAFER-F4 防水人臉+ナノ波指紋 十合一電子鎖", "FAMMIX", "https://www.zeczec.com/projects/fammix-lock-safer-f4", "要調査", "要調査", "4/5", "完全防水・顔+ナノ波指紋認証の屋外対応スマートロック。日本の戸建て住宅向けに高需要"),
    ("スマートロック・セキュリティ", "Arpha M2 Ultra 3D人臉辨識靜音智慧電子鎖", "Arpha", "https://www.zeczec.com/projects/arpha-m2-ultra", "要調査", "要調査", "4/5", "3D顔認証×静音設計スマートロック。集合住宅の多い日本の「静か」文化に合致"),
    ("スマートロック・セキュリティ", "Philips Alpha VP 6合1智慧視訊電子鎖 遠端通話", "Philips", "https://www.zeczec.com/projects/alpha-vp", "https://www.philips.com/ja-jp", "要調査", "4/5", "ビデオ通話対応スマートロック6合一。Philipsブランドで日本展開可能性高い"),
    ("スマートロック・セキュリティ", "mimax 掌門人視訊電子鎖 掌靜脈×人臉辨識×180°夜視", "mimax", "https://www.zeczec.com/projects/mimax-Master-E-Lock", "要調査", "要調査", "4/5", "180度夜間視野+ビデオ通話+掌静脈認証スマートロック。日本の高齢者セキュリティ需要に合致"),
    ("スマートロック・セキュリティ", "3E 小島快鎖 顔認証 3E智慧鎖 国際特許", "3E", "https://www.zeczec.com/projects/EEElock", "要調査", "要調査", "4/5", "8種解錠方法対応の顔認証スマートロック。日本の住宅構造への適合性確認が必要だが需要は高い"),
    ("スマートロック・セキュリティ", "Clever Sky 無線智慧監視セキュリティシステム", "Clever Sky", "https://www.zeczec.com/projects/clever-sky", "要調査", "要調査", "3/5", "ワイヤレスセキュリティカメラシステム。日本の防犯意識の高まりで需要あり"),
    ("スマートロック・セキュリティ", "SGUDA U-LOCK 聲控+無線遠端智慧鎖", "SGUDA", "https://www.zeczec.com/projects/SGUDA", "要調査", "要調査", "3/5", "音声認識+リモート制御スマートロック。ハンズフリー操作ニーズに応える"),
    ("スマートロック・セキュリティ", "Philips EASYKEY 智能補助門鎖503 穴開け不要", "Philips", "https://www.zeczec.com/projects/EasyKey503", "https://www.philips.com/ja-jp", "要調査", "4/5", "既存錠に追加するだけの補助スマートロック。賃貸が多い日本に最適な「穴を開けない」取付"),

    # スマートホーム・IoT・ロボット
    ("スマートホーム・IoT・ロボット", "Eilik 桌面型陪伴ロボット 1000以上表情 感情表現", "Energy Robotics", "https://www.zeczec.com/projects/eilik-2025-11-14", "https://www.energyroboticstech.com", "要調査", "5/5", "1000以上の表情を持つ感情表現デスクロボット。日本のロボット・キャラクター好き文化に完全合致"),
    ("スマートホーム・IoT・ロボット", "Yeelight LED 智慧彩光吸頂燈 3分鐘安装1600万色", "Yeelight", "https://www.zeczec.com/projects/yeelightled", "https://www.yeelight.com", "support@yeelight.com", "3/5", "3分で取付できる1600万色スマートシーリングライト。Mi Home/HomeKit対応で日本でも展開可能"),
    ("スマートホーム・IoT・ロボット", "Lytmi Neo LED 智慧氛圍燈帯 HDMI同期沈浸体験", "Lytmi", "https://www.zeczec.com/projects/lytmi-neo-led", "https://www.lytmi.com", "要調査", "4/5", "HDMI同期バックライトLEDストリップ。ゲーマー・シアタールーム需要の高い日本に刺さる"),
    ("スマートホーム・IoT・ロボット", "Smart光 LED全光譜螢幕掛燈 護眼+RGB自動感光", "Smart光", "https://www.zeczec.com/projects/smartlight", "要調査", "要調査", "4/5", "自動調光+RGB雰囲気照明モニターライト。在宅ワーク×目の保護意識の高い日本に最適"),
    ("スマートホーム・IoT・ロボット", "MOZTECH 変変燈 百変支架 LED無線センサーライト", "MOZTECH", "https://www.zeczec.com/projects/MOZTECH-BANG-LIGHT", "https://www.moztech.com.tw", "要調査", "4/5", "取付場所自由・センサー式ワイヤレスLEDライト。賃貸が多い日本の「工事不要」ニーズにぴったり"),
    ("スマートホーム・IoT・ロボット", "彩虹光 全彩LED智慧檯燈 全球唯一全彩", "彩虹光", "https://www.zeczec.com/projects/al-spectrum-led", "要調査", "要調査", "3/5", "世界唯一フルカラーLEDスマートデスクライト。子供の学習環境や在宅ワークで需要あり"),
    ("スマートホーム・IoT・ロボット", "Ka Ka Light 卡卡燈 輕鬆拆解 自由組合全方位テーブルランプ", "Ka Ka", "https://www.zeczec.com/projects/ka-ka-light", "要調査", "要調査", "3/5", "分解・組み立て自由のモジュール式デスクライト。DIY志向の日本インテリア市場にユニーク"),
    ("スマートホーム・IoT・ロボット", "LIGHTSHIELD LED照明 抗菌×抗空汙 次世代照明", "LIGHTSHIELD", "https://www.zeczec.com/projects/LIGHTSHIELD", "要調査", "要調査", "3/5", "抗菌・抗汚染機能搭載LED照明。コロナ以降の衛生意識が高い日本市場で一定の需要"),

    # 空気清浄・除湿機
    ("空気清浄・除湿機", "CS101 PLUS 克立淨超能雷神空氣清淨機 除濕旗艦", "CleanStation (克立淨)", "https://www.zeczec.com/projects/cleanstation-cs101", "要調査", "要調査", "3/5", "空気清浄+除湿のフラッグシップモデル。日本でもダイキン等競合が強いがコスパで差別化可能"),
    ("空気清浄・除湿機", "Smini 零耗材空氣清淨機 永久使用 離子技術", "Smini", "https://www.zeczec.com/projects/smini", "要調査", "要調査", "4/5", "フィルター消耗品不要のイオン技術空気清浄機。「ランニングコスト0」訴求で日本のコスト意識層に響く"),
    ("空気清浄・除湿機", "FLAT 壁掛畫空淨機 外銷突破20國 アートパネル型", "FLAT", "https://www.zeczec.com/projects/flat", "要調査", "要調査", "4/5", "アートパネル型壁掛け空気清浄機。インテリアを損なわない設計で日本のスタイリッシュ志向層に高需要"),
    ("空気清浄・除湿機", "AO-600 真無線空氣清淨機 四大強效コードレス", "AO", "https://www.zeczec.com/projects/AO-600", "要調査", "要調査", "3/5", "コードレス・持ち運び可能な空気清浄機。部屋移動しながら使える柔軟性は日本の1K住宅に合う"),
    ("空気清浄・除湿機", "DIKE 淨化者EVO HEPA14 五年換濾網不要 医療グレード", "DIKE", "https://www.zeczec.com/projects/HCF630", "要調査", "要調査", "4/5", "医療グレードHEPA14フィルター×5年交換不要。ランニングコストの低さで日本のシニア層にも訴求"),
    ("空気清浄・除湿機", "Ionix Flow 攜帯式空氣清淨器 航空機グレード離子", "Ionix", "https://www.zeczec.com/projects/ionixflow", "要調査", "要調査", "3/5", "航空機グレードのポータブル空気清浄機。通勤・移動中の花粉対策需要がある日本市場に合致"),
    ("空気清浄・除湿機", "DIKE 寵愛抗敏空氣清淨機 HCF610 去病毒臭味", "DIKE", "https://www.zeczec.com/projects/DIKE_HCF610", "要調査", "要調査", "3/5", "ペットアレルギー対応空気清浄機。日本のペット飼育率向上と合わせて需要拡大中"),
    ("空気清浄・除湿機", "LG 12L UVC 雙變頻除濕機 一級能效×紫外線", "LG", "https://www.zeczec.com/projects/lg-12l-dehumidifier-2025-07-11", "https://www.lg.com/jp", "jp.lgservice@lge.com", "2/5", "LGブランド除湿機。日本でもLG製品は販売されているため独自性は低い"),
    ("空気清浄・除湿機", "KOIZUMI 16.9L 智能除濕機 18坪対応 一級能效", "KOIZUMI", "https://www.zeczec.com/projects/koizumi169", "要調査", "要調査", "2/5", "KOIZUMIブランド除湿機。日本のKOIZUMIとの関係要確認。仮に台湾独自展開なら参入余地あり"),

    # ペットテック
    ("ペットテック", "LG 喵太座 キャットタワー×空清×健康監測 All in One", "LG", "https://www.zeczec.com/projects/lg-aerocat-tower", "https://www.lg.com/jp", "jp.lgservice@lge.com", "5/5", "キャットタワー+空気清浄+健康監視オールインワン。日本の猫ブームと健康意識の高さで超高需要"),
    ("ペットテック", "PETUS 寵物智能項圈 5G GPS 無距離制限", "PETUS", "https://www.zeczec.com/projects/petus", "要調査", "要調査", "4/5", "5G IoT対応GPSペットスマートカラー。迷子防止・位置確認ニーズが高い日本のペット飼育者に需要"),
    ("ペットテック", "amicoipet 寵物居家生理監測スマート項圈", "amicoi", "https://www.zeczec.com/projects/amicoipet", "要調査", "要調査", "4/5", "バイタル監視機能付きペットスマートカラー。ペットの健康管理に熱心な日本市場に高い訴求力"),
    ("ペットテック", "Pawbby 智慧寵物零食機 APP操作 リモート給餌", "Pawbby", "https://www.zeczec.com/projects/pawbby", "要調査", "要調査", "4/5", "スマホ操作の自動おやつ投与機。日本のペットカメラ・自動給餌器市場は拡大中"),
    ("ペットテック", "ALNPET 寵物智能偵測殺菌飲水機 99.9%抗菌", "ALNPET", "https://www.zeczec.com/projects/alnpet-powered-by-pettofund", "要調査", "要調査", "3/5", "殺菌・水質検知機能付きペット自動給水機。ペットの健康管理意識が高い日本で需要あり"),
    ("ペットテック", "ebot 24時間寵物管家 AI見守りシステム", "ebot", "https://www.zeczec.com/projects/ebotaiwan", "要調査", "要調査", "4/5", "AI搭載24時間ペット監視管理システム。共働き世帯が多い日本でペット留守番管理に高需要"),

    # 健康・フィットネス・睡眠
    ("健康・フィットネス・睡眠", "Cooler Master Hybrid M 電競工學按摩椅 三願一次", "Cooler Master", "https://www.zeczec.com/projects/coolermaster-hybrid-m", "https://www.coolermaster.com/jp", "要調査", "3/5", "マッサージ機能付きゲーミングチェア。日本ゲーマー×健康意識層に訴求。Cooler Masterは既に日本展開"),
    ("健康・フィットネス・睡眠", "RheoFit 全自動AI按摩滾輪 運動リカバリー", "RheoFit", "https://www.zeczec.com/projects/rheofit", "要調査", "要調査", "4/5", "AI制御全自動マッサージローラー。運動後のリカバリー製品として日本のスポーツ愛好家に高需要"),
    ("健康・フィットネス・睡眠", "OYO NOVA 攜帶型全方位居家健身神器 NASA使用", "OYO Fitness", "https://www.zeczec.com/projects/oyo-nova", "要調査", "要調査", "4/5", "NASA採用技術の持ち運びコンパクトフィットネス器具。在宅ワーク×健康ブームの日本市場に刺さる"),
    ("健康・フィットネス・睡眠", "BGYM BQQ 迷你健步機 座りながらマッサージ×運動", "BGYM", "https://www.zeczec.com/projects/bgym-bqq", "要調査", "要調査", "4/5", "座ったまま使えるマッサージ付きミニステッパー。デスクワーカーが多い日本のオフィス需要に合致"),
    ("健康・フィットネス・睡眠", "MAGICON 體感遊戲健身環 スマホゲーム連動", "MAGICON", "https://www.zeczec.com/projects/EL-FR002", "要調査", "要調査", "3/5", "スマホ連動ゲーミングフィットネスリング。Nintendo Switch Ringとの差別化明確化が必要"),
    ("健康・フィットネス・睡眠", "O5PRO 8D 瞬眠枕 -5℃瞬涼雙層テクノロジー枕", "O5PRO", "https://www.zeczec.com/projects/o5pro-pillow", "要調査", "要調査", "3/5", "冷感8Dメモリーフォーム枕。日本でも機能性枕市場は拡大中だが競合が強い"),
    ("健康・フィットネス・睡眠", "歐歐睏助眠燈 台湾工科大学院チーム開発助眠ライト", "OuOu Sleep", "https://www.zeczec.com/projects/osleepy", "要調査", "要調査", "4/5", "科学的設計の助眠ライト。日本の睡眠問題意識の高さで高需要期待。7年間の研究開発品"),
    ("健康・フィットネス・睡眠", "Qi Ease 撓定 エネルギーモバイル充電 ストレスゼロ", "撓定", "https://www.zeczec.com/projects/qi-ease-1", "要調査", "要調査", "2/5", "エネルギー充電・ストレス軽減デバイス（気功研究ベース）。日本では科学的根拠の提示が重要"),
    ("健康・フィットネス・睡眠", "Heal 好眠貼 三層設計 鼻呼吸促進助眠", "Heal", "https://www.zeczec.com/projects/heal", "要調査", "要調査", "3/5", "医療グレード三層構造の口閉じ睡眠促進テープ。口呼吸改善・睡眠改善ニーズで日本でも需要あり"),
    ("健康・フィットネス・睡眠", "白噪音安眠枕 5分鐘入睡 自然音周波数内蔵枕", "Future Legend", "https://www.zeczec.com/projects/Future-Legend", "要調査", "要調査", "3/5", "ホワイトノイズ・自然音内蔵の安眠枕。日本の睡眠改善市場で独自性あるアプローチ"),
    ("健康・フィットネス・睡眠", "T-Sox 飆動自行車訓練台 Zwift/Rouvy対応 MIT", "T-Sox", "https://www.zeczec.com/projects/t-sox-mit", "要調査", "要調査", "3/5", "Zwift/Rouvy対応スマートバイクトレーナー。日本のインドアサイクリスト向けに需要あり"),

    # 電動モビリティ・昇降
    ("電動モビリティ・昇降", "UDITER 像素騎士 LED智能電動スケートボード 世界初", "UDITER", "https://www.zeczec.com/projects/uditer-pixel-rider-led", "要調査", "要調査", "3/5", "LED搭載世界初スマート電動スケートボード。日本でも電動移動体市場は拡大中"),
    ("電動モビリティ・昇降", "Scooty RIDE 電動滑板車 折りたたみ携帯型", "Scooty", "https://www.zeczec.com/projects/scooty-ride", "要調査", "要調査", "3/5", "折り畳み携帯型電動スケートボード。日本の電動モビリティ規制緩和後に参入余地あり"),
    ("電動モビリティ・昇降", "Scorpion™ スケートボード×電動板 50cm以下コンパクト", "Scorpion", "https://www.zeczec.com/projects/scorpion", "要調査", "要調査", "2/5", "超コンパクト電動スケートボード。日本では公道使用に法規制があるため用途が限定される"),
    ("電動モビリティ・昇降", "自行車電動裝置 掛載式 電動アシスト変換キット", "バイクパワー", "https://www.zeczec.com/projects/bike-power", "要調査", "要調査", "3/5", "既存自転車を電動アシスト化するキット。日本の電動アシスト自転車市場の拡大に乗れる可能性"),
    ("電動モビリティ・昇降", "E-ach UP 桌上型電動升降スタンドデスク", "AKA", "https://www.zeczec.com/projects/aka-eachup", "要調査", "要調査", "4/5", "デスクトップ電動昇降スタンドデスク。日本の在宅ワーク×健康意識で電動昇降デスク需要急増"),
    ("電動モビリティ・昇降", "Stair-Rover™ 八輪スケートボード 城市遨遊", "Stair-Rover", "https://www.zeczec.com/projects/allrover", "要調査", "要調査", "2/5", "8輪設計の段差対応スケートボード。日本では安全・法規制面で課題があるがユニーク製品"),
]

row_colors = ["F5F5F5", "FFFFFF"]
for row_idx, product in enumerate(products, 2):
    for col_idx, value in enumerate(product, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=row_colors[(row_idx) % 2])

column_widths = [20, 44, 24, 52, 38, 26, 14, 62]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[1].height = 35
for row in range(2, len(products) + 2):
    ws.row_dimensions[row].height = 70

ws.freeze_panes = "A2"

# Score color highlight
score_fill_high = PatternFill("solid", fgColor="C6EFCE")
score_fill_mid = PatternFill("solid", fgColor="FFEB9C")
score_fill_low = PatternFill("solid", fgColor="FFC7CE")

for row in range(2, len(products) + 2):
    score_cell = ws.cell(row=row, column=7)
    score = score_cell.value or ""
    if score.startswith("5") or score.startswith("4"):
        score_cell.fill = score_fill_high
    elif score.startswith("3"):
        score_cell.fill = score_fill_mid
    else:
        score_cell.fill = score_fill_low
    score_cell.font = Font(name="Arial", size=10, bold=True)
    score_cell.alignment = Alignment(horizontal="center", vertical="top")

output_path = r"C:\Users\hiro\HP0524\ZECZEC_ガジェット日本未発売リスト.xlsx"
wb.save(output_path)
print(f"保存完了: {output_path}")
print(f"商品数: {len(products)}")
