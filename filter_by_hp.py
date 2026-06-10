from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

src = load_workbook(r"C:\Users\hiro\HP0524\ZECZEC_ガジェット日本未発売リスト.xlsx")
ws_src = src.active

wb = Workbook()
ws = wb.active
ws.title = "メーカーHP有り"

# ヘッダーコピー
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col in range(1, ws_src.max_column + 1):
    cell = ws.cell(row=1, column=col, value=ws_src.cell(row=1, column=col).value)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# メーカーHP列(E=5)がhttpで始まる行のみ抽出
link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
row_colors = ["F5F5F5", "FFFFFF"]
new_row = 2

for row in range(2, ws_src.max_row + 1):
    hp_val = ws_src.cell(row=row, column=5).value
    if hp_val and isinstance(hp_val, str) and hp_val.startswith("http"):
        for col in range(1, ws_src.max_column + 1):
            src_cell = ws_src.cell(row=row, column=col)
            new_cell = ws.cell(row=new_row, column=col, value=src_cell.value)
            new_cell.font = Font(name="Arial", size=10)
            new_cell.alignment = Alignment(vertical="top", wrap_text=True)
            new_cell.fill = PatternFill("solid", fgColor=row_colors[new_row % 2])

            # URLはハイパーリンクに
            if col in [4, 5] and src_cell.value and str(src_cell.value).startswith("http"):
                new_cell.hyperlink = src_cell.value
                new_cell.font = link_font

            # スコアセルの色付け
            if col == 7:
                score = src_cell.value or ""
                if score.startswith("5") or score.startswith("4"):
                    new_cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif score.startswith("3"):
                    new_cell.fill = PatternFill("solid", fgColor="FFEB9C")
                else:
                    new_cell.fill = PatternFill("solid", fgColor="FFC7CE")
                new_cell.font = Font(name="Arial", size=10, bold=True)
                new_cell.alignment = Alignment(horizontal="center", vertical="top")

        new_row += 1

# 列幅・行高
column_widths = [20, 44, 24, 52, 38, 26, 14, 62]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[1].height = 35
for row in range(2, new_row):
    ws.row_dimensions[row].height = 70

ws.freeze_panes = "A2"

out_path = r"C:\Users\hiro\HP0524\ZECZEC_メーカーHP有りリスト.xlsx"
wb.save(out_path)
print(f"保存完了: {out_path}")
print(f"商品数: {new_row - 2}件")
