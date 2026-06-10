from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote

wb = load_workbook(r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸705件.xlsx")
ws = wb["海外便利グッズリスト"]

cat_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
cat_fills = {
    "キッチン・調理器具": PatternFill("solid", start_color="2E75B6"),
    "スマートホーム・インテリア・照明": PatternFill("solid", start_color="538135"),
    "ウェアラブル・ヘルス・フィットネス": PatternFill("solid", start_color="7030A0"),
    "アウトドア・スポーツ・旅行": PatternFill("solid", start_color="C55A11"),
    "ペット用品": PatternFill("solid", start_color="843C0C"),
    "テクノロジー・ガジェット": PatternFill("solid", start_color="244185"),
    "美容・スキンケア": PatternFill("solid", start_color="C00000"),
    "子供・教育": PatternFill("solid", start_color="375623"),
    "ファッション・アクセサリー": PatternFill("solid", start_color="7B3F00"),
    "クリーニング・収納・整理": PatternFill("solid", start_color="404040"),
}
data_font = Font(name="Arial", size=10)
link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
data_align = Alignment(vertical="center", wrap_text=True)
alt_fill = PatternFill("solid", start_color="EBF3FB")
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def build_gmail(email, maker_name, product_name):
    subject = "Potential Distribution Partnership for Japan"
    body = f"""Dear {maker_name} Team,

My name is Hiroyuki Inoguchi from
Sumai pluS Co., Ltd. in Japan

I hope this message finds you all well.

Our company focuses on promoting products that enrich people's daily lives. We are currently seeking unique international brands to introduce to the Japanese market and support their growth.

I recently had the opportunity to review your product ({product_name}) and was very impressed by its potential in the Japanese market.

I am confident that your product will strongly appeal to Japanese customers, and I would like to explore the possibility of building a partnership with your company to help ensure its success.

Furthermore, our company has a proven track record of promoting overseas products through our proprietary sales network and Japanese distribution channels.

Would it be possible to schedule a brief online meeting sometime next week to discuss this matter?

I look forward to hearing from you.


Sincerely,

Hiroyuki Inoguchi
Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )

================================================

Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )
Hiroyuki Inoguchi

E-mail:yutorin.ino@gmail.com

Address: 49-5 Kitazakuno, Higashi-Itsushiro, Ichinomiya City, Aichi Prefecture, Japan

================================================="""
    return f"https://mail.google.com/mail/?view=cm&fs=1&to={quote(email,safe='')}&su={quote(subject,safe='')}&body={quote(body,safe='')}"

# No.|カテゴリ|商品名|メーカー名|販売ECサイト名|商品URL|メーカーHP|連絡先メール
new_data = [
    (706,"キッチン・調理器具","Agari AI Pressure Reverse-Sear Smart Oven","Agari Kitchen","Gadget Flow","https://thegadgetflow.com/product/agari-michelin-chef-quality-smart-oven/","https://agari.kitchen","https://agari.kitchen/pages/contact（問い合わせページ）"),
    (707,"キッチン・調理器具","KOMBU The Fermenstation Kombucha Brewer","KOMBU","Gadget Flow","https://thegadgetflow.com/product/kombu-the-fermenstation-kombucha-brewer/","https://kombubrewer.com","hello@kombubrewer.com"),
    (708,"キッチン・調理器具","ICEMAGE CrystalMax Clear Ice Maker","ICEMAGE","Gadget Flow","https://thegadgetflow.com/product/icemage-crystalmax-clear-ice-maker/","https://www.icemage.com","support@icemage.com"),
    (709,"キッチン・調理器具","Ooni Volt 2 Electric Pizza Oven","Ooni","Ooni公式サイト","https://ooni.com/products/ooni-volt","https://ooni.com","https://ooni.com/pages/contact（問い合わせページ）"),
    (710,"キッチン・調理器具","Terra Kaffe Demi Espresso Machine","Terra Kaffe","Terra Kaffe公式サイト","https://terrakaffe.com/products/demi","https://terrakaffe.com","hello@terrakaffe.com"),
    (711,"キッチン・調理器具","WACACO PIXAPRESSO Portable Espresso Machine","WACACO","WACACO公式サイト","https://www.wacaco.com/products/pixapresso","https://www.wacaco.com","support@wacaco.com"),
    (712,"キッチン・調理器具","Sproushi AI Powered Mushroom Grower","Sproushi","Gadget Flow","https://thegadgetflow.com/product/sproushi-ai-powered-mushroom-grower/","https://www.sproushi.com","hello@sproushi.com"),
    (713,"キッチン・調理器具","Liffo Fully Autonomous Cooking Robot","Liffo","Gadget Flow","https://thegadgetflow.com/product/liffo-fully-autonomous-cooking-robot/","https://www.lifforobot.com","info@lifforobot.com"),
    (714,"キッチン・調理器具","Aromi Smart Ingredient Tracker","Aromi","Gadget Flow","https://thegadgetflow.com/product/aromi-smart-ingredient-tracker-for-every-kitchen/","https://www.aromi.ai","hello@aromi.ai"),
    (715,"キッチン・調理器具","Lettuce Grow Farmstand Nook Hydroponic Indoor Garden","Lettuce Grow","Lettuce Grow公式サイト","https://www.lettucegrow.com/farmstand","https://www.lettucegrow.com","support@lettucegrow.com"),
    (716,"キッチン・調理器具","Qalzy AI-Powered Smart Kitchen Scale","Qalzy","Gadget Flow","https://thegadgetflow.com/product/qalzy-ai-powered-smart-kitchen-scale/","https://www.qalzy.com","support@qalzy.com"),
    (717,"キッチン・調理器具","Kara Pod Air-to-Water Coffee Machine","Kara","Gadget Flow","https://thegadgetflow.com/product/kara-pod-self-refilling-premium-drinking-water-and-coffee-making-machine-from-air/","https://kara.eco","hello@kara.eco"),
    (718,"キッチン・調理器具","AeroPress Go Plus Coffee Maker","AeroPress","AeroPress公式サイト","https://aeropress.com/products/aeropress-go-plus","https://aeropress.com","support@aeropress.com"),
    (719,"キッチン・調理器具","Posha Intelligent Kitchen Robot","Posha","Gadget Flow","https://thegadgetflow.com/product/posha-intelligent-kitchen-robot/","https://www.posha.com","hello@posha.com"),
    (720,"キッチン・調理器具","Morphy Richards Fusion Kettle and Toaster Set","Morphy Richards","Morphy Richards公式サイト","https://www.morphyrichards.com","https://www.morphyrichards.com","https://www.morphyrichards.com/contact-us（問い合わせページ）"),
    (721,"キッチン・調理器具","Typhur Dome 2 Air Fryer","Typhur","Typhur公式サイト","https://www.typhur.com/products/dome-2","https://www.typhur.com","support@typhur.com"),
    (722,"キッチン・調理器具","Current Model P Smart Pizza Oven","Current","Gadget Flow","https://thegadgetflow.com/product/current-model-p-smart-pizza-oven/","https://www.currentcooking.com","hello@currentcooking.com"),
    (723,"キッチン・調理器具","Solo Stove Steelfire 30 Stainless Griddle","Solo Stove","Solo Stove公式サイト","https://www.solostove.com/en-us/products/steelfire-30-griddle","https://www.solostove.com","support@solostove.com"),
    (724,"キッチン・調理器具","Alessi Plissé Electric Toaster","Alessi","Alessi公式サイト","https://www.alessi.com","https://www.alessi.com","https://www.alessi.com/en_EN/store/info-contact（問い合わせページ）"),
    (725,"キッチン・調理器具","TRU Fully Automatic Espresso Machine with Grinder","TRU","Gadget Flow","https://thegadgetflow.com/product/tru-fully-automatic-19-bar-espresso-machine-with-grinder/","https://truappliances.com","support@truappliances.com"),
    (726,"キッチン・調理器具","WACACO Minipresso GR2 Portable Espresso Machine","WACACO","WACACO公式サイト","https://www.wacaco.com/products/minipresso-gr2","https://www.wacaco.com","support@wacaco.com"),
    (727,"キッチン・調理器具","Dualit Pour Over Kettle","Dualit","Gadget Flow","https://thegadgetflow.com/product/dualit-pour-over-kettle/","https://dualit.com","customerservices@dualit.com"),
    (728,"キッチン・調理器具","Chef'n FreshForce Ricer Potato Ricer","Chef'n","Chef'n公式サイト","https://www.chefn.com/products/freshforce-ricer","https://www.chefn.com","support@chefn.com"),
    (729,"キッチン・調理器具","Ooni Karu 2 Multi-Fuel Pizza Oven","Ooni","Ooni公式サイト","https://ooni.com/products/ooni-karu-2","https://ooni.com","https://ooni.com/pages/contact（問い合わせページ）"),
    (730,"キッチン・調理器具","Gozney Arc XL Outdoor Pizza Oven","Gozney","Gozney公式サイト","https://gozney.com/products/arc-xl","https://gozney.com","support@gozney.com"),
    (731,"スマートホーム・インテリア・照明","Skylight Calendar 2 Smart Family Wall Calendar","Skylight","Skylight公式サイト","https://www.skylightframe.com/products/skylight-calendar","https://www.skylightframe.com","support@skylightframe.com"),
    (732,"スマートホーム・インテリア・照明","SmartWings Motorized Roller Shades","SmartWings","SmartWings公式サイト","https://www.smartwings.net","https://www.smartwings.net","support@smartwings.net"),
    (733,"スマートホーム・インテリア・照明","Nordik Recovery 1-Person Full-Spectrum Infrared Sauna","Nordik Recovery","Gadget Flow","https://thegadgetflow.com/product/nordik-recovery-1-person-infrared-sauna-home-wellness/","https://www.nordikrecovery.com","hello@nordikrecovery.com"),
    (734,"スマートホーム・インテリア・照明","Weave Robotics Isaac 0 Smart Home Laundry Robot","Weave Robotics","Gadget Flow","https://thegadgetflow.com/product/isaac-0-home-laundry-robot/","https://www.weaverobotics.com","info@weaverobotics.com"),
    (735,"スマートホーム・インテリア・照明","DeerValley ADA Smart Bidet Toilet","DeerValley","Gadget Flow","https://thegadgetflow.com/product/deervalley-dv-1s0442-v3-smart-toilet-bidet-toilet/","https://www.deervalleybath.com","support@deervalleybath.com"),
    (736,"スマートホーム・インテリア・照明","Tapo H110 Smart IR & IoT Hub","TP-Link Tapo","Tapo公式サイト","https://www.tapo.com/en/product/smart-hub/tapo-h110/","https://www.tapo.com","https://www.tp-link.com/en/support/contact-technical-support/（問い合わせページ）"),
    (737,"スマートホーム・インテリア・照明","JWDA Corded Table Lamp","Audo Copenhagen","Audo公式サイト","https://www.audo.dk/en/products/jwda-table-lamp","https://www.audo.dk","https://www.audo.dk/en/contact（問い合わせページ）"),
    (738,"スマートホーム・インテリア・照明","Lexi Next Generation Smart Lighting","Lexi","Gadget Flow","https://thegadgetflow.com/product/next-generation-smart-lighting/","https://lexi.lighting","hello@lexi.lighting"),
    (739,"スマートホーム・インテリア・照明","Blink Outdoor 2K+ Wireless Smart Security Camera","Blink","Blink公式サイト","https://blinkforhome.com/products/outdoor-2k-plus","https://www.blinkforhome.com","https://support.blinkforhome.com/（問い合わせページ）"),
    (740,"スマートホーム・インテリア・照明","Laundry Chair Bedroom Chair with Rotating Rail","Laundry Chair","Gadget Flow","https://thegadgetflow.com/product/laundry-chair-bedroom-storage-chair/","https://www.laundrychairco.com","hello@laundrychairco.com"),
    (741,"スマートホーム・インテリア・照明","Skylight Frame Smart Digital Photo Frame 10\"","Skylight","Skylight公式サイト","https://www.skylightframe.com/products/skylight-frame","https://www.skylightframe.com","support@skylightframe.com"),
    (742,"スマートホーム・インテリア・照明","Blink Wired Doorbell 2K+","Blink","Blink公式サイト","https://blinkforhome.com/products/wired-doorbell","https://www.blinkforhome.com","https://support.blinkforhome.com/（問い合わせページ）"),
    (743,"スマートホーム・インテリア・照明","Nanoleaf Smart Multicolor Floor Lamp","Nanoleaf","Nanoleaf公式サイト","https://thegadgetflow.com/product/nanoleaf-smart-multicolor-floor-lamp/","https://nanoleaf.me","https://nanoleaf.me/en-US/support/（問い合わせページ）"),
    (744,"ウェアラブル・ヘルス・フィットネス","Withings BodyFit Body Composition Scale","Withings","Withings公式サイト","https://www.withings.com/us/en/bodyfit","https://www.withings.com","support@withings.com"),
    (745,"ウェアラブル・ヘルス・フィットネス","Mudita Radiant Field Watch","Mudita","Mudita公式サイト","https://thegadgetflow.com/product/mudita-radiant-field-watch/","https://www.mudita.com","support@mudita.com"),
    (746,"ウェアラブル・ヘルス・フィットネス","Mudita Harmony 2 Minimalist Sleep Alarm Clock","Mudita","Mudita公式サイト","https://thegadgetflow.com/product/mudita-harmony-2-alarm-clock/","https://www.mudita.com","support@mudita.com"),
    (747,"ウェアラブル・ヘルス・フィットネス","Subtle Voicebuds Voice Wearable","Subtle","Gadget Flow","https://thegadgetflow.com/product/voicebuds-subtle-voice-wearable/","https://www.subtle.ai","hello@subtle.ai"),
    (748,"ウェアラブル・ヘルス・フィットネス","Status Pro X True Wireless Earphones","Status Audio","Status Audio公式サイト","https://thegadgetflow.com/product/status-pro-x-true-wireless-earbuds/","https://www.statusaudio.com","hello@statusaudio.com"),
    (749,"ウェアラブル・ヘルス・フィットネス","Garmin Forerunner 70 GPS Running Watch","Garmin","Garmin公式サイト","https://www.garmin.com/en-US/p/forerunner-70","https://www.garmin.com","https://www.garmin.com/en-US/support/（問い合わせページ）"),
    (750,"ウェアラブル・ヘルス・フィットネス","Garmin Forerunner 170 Running Smartwatch","Garmin","Garmin公式サイト","https://www.garmin.com/en-US/p/forerunner-170","https://www.garmin.com","https://www.garmin.com/en-US/support/（問い合わせページ）"),
    (751,"ウェアラブル・ヘルス・フィットネス","Nordik Recovery Infrared Sauna（1人用）","Nordik Recovery","Nordik Recovery公式サイト","https://thegadgetflow.com/product/nordik-recovery-1-person-infrared-sauna-home-wellness/","https://www.nordikrecovery.com","hello@nordikrecovery.com"),
    (752,"アウトドア・スポーツ・旅行","Ostation 2 Pro Smart Battery Hub","Ostation","Gadget Flow","https://thegadgetflow.com/product/ostation-2-pro-next-gen-3-in-1-smart-battery-hub/","https://www.ostation.com","support@ostation.com"),
    (753,"アウトドア・スポーツ・旅行","Pebblebee Evercolor Freeze Frame Bluetooth Tracker","Pebblebee","Pebblebee公式サイト","https://www.pebblebee.com/products/evercolor","https://www.pebblebee.com","support@pebblebee.com"),
    (754,"アウトドア・スポーツ・旅行","Bolt Mini Velo City Bike","Bolt","Gadget Flow","https://thegadgetflow.com/product/bolt-mini-velo-city-bike/","https://www.boltbike.com","hello@boltbike.com"),
    (755,"アウトドア・スポーツ・旅行","Olight ArkPro Ultra EDC Flat Flashlight","Olight","Olight公式サイト","https://www.olightstore.com/arkpro-ultra","https://www.olightstore.com","support@olightworld.com"),
    (756,"アウトドア・スポーツ・旅行","Zens Semi Solid State Powerbank","Zens","Zens公式サイト","https://zens.nl/products/semi-solid-state-powerbank","https://www.zens.nl","info@zens.nl"),
    (757,"アウトドア・スポーツ・旅行","A2C One-Lock Ultra Phone Mount","A2C","Gadget Flow","https://thegadgetflow.com/product/a2c-one-lock-ultra-phone-mount/","https://www.a2cmount.com","support@a2cmount.com"),
    (758,"アウトドア・スポーツ・旅行","Gear8 Screen-Free Puzzle Maze Travel Game","Gear8","Gadget Flow","https://thegadgetflow.com/product/gear8-screen-free-puzzle-maze/","https://www.gear8games.com","hello@gear8games.com"),
    (759,"アウトドア・スポーツ・旅行","SoundCharge MagSafe Wireless Speaker Charger","SoundCharge","Gadget Flow","https://thegadgetflow.com/product/soundcharge-hi-fi-magsafe-speaker-charger/","https://www.soundcharge.io","hello@soundcharge.io"),
    (760,"アウトドア・スポーツ・旅行","Snow Peak x Merrell GORE-TEX Trail Sneaker","Snow Peak","Snow Peak公式サイト","https://www.snowpeak.com","https://www.snowpeak.com","https://www.snowpeak.com/pages/contact（問い合わせページ）"),
    (761,"アウトドア・スポーツ・旅行","Benchmade Bugout Vapyr MagnaCut EDC Knife","Benchmade","Benchmade公式サイト","https://www.benchmade.com/products/bugout-vapyr","https://www.benchmade.com","support@benchmade.com"),
    (762,"ペット用品","INUBOX Indoor Dog Waste Processor","INUBOX","Gadget Flow","https://thegadgetflow.com/blog/8-pet-gadgets-that-will-make-your-pet-even-happier/","https://www.inubox.com","support@inubox.com"),
    (763,"ペット用品","VARRAM Pet Fitness Robot","VARRAM","VARRAM公式サイト","https://varram.com","https://varram.com","support@varram.com"),
    (764,"ペット用品","WonderWoof BowTie Dog Activity Tracker","WonderWoof","WonderWoof公式サイト","https://www.wonderwoof.com","https://www.wonderwoof.com","support@wonderwoof.com"),
    (765,"ペット用品","MOVA MOBIUS 60 Robot Vacuum and Mop（ペット毛対応）","MOVA","Gadget Flow","https://thegadgetflow.com/product/mova-mobius-60-robot-vacuum-and-mop/","https://www.movarobot.com","support@movarobot.com"),
    (766,"テクノロジー・ガジェット","Camp Snap 2 Screen-Free Digital Camera","Camp Snap","Camp Snap公式サイト","https://campsnap.com/products/camp-snap-2","https://campsnap.com","hello@campsnap.com"),
    (767,"テクノロジー・ガジェット","ZSA Navigator Modular Trackball and Trackpad","ZSA","ZSA公式サイト","https://www.zsa.io/navigator","https://www.zsa.io","contact@zsa.io"),
    (768,"テクノロジー・ガジェット","Ploopy Bean High-Performance Pointing Stick","Ploopy","Gadget Flow","https://thegadgetflow.com/product/ploopy-bean-high-performance-pointing-stick/","https://ploopy.co","hello@ploopy.co"),
    (769,"テクノロジー・ガジェット","AYANEO Pocket AIR Mini Android Retro Gaming Handheld","AYANEO","AYANEO公式サイト","https://www.ayaneo.com/goods/pocket-air-mini","https://www.ayaneo.com","support@ayaneo.com"),
    (770,"テクノロジー・ガジェット","MSI Claw 8 EX AI+ Copilot+ Handheld Gaming PC","MSI","MSI公式サイト","https://www.msi.com/Gaming-Handheld/MSI-Claw-8-EX-AI","https://www.msi.com","https://www.msi.com/support（問い合わせページ）"),
    (771,"テクノロジー・ガジェット","KUXIU 3-Point Power MagSafe 3-in-1 Charging Stand","KUXIU","Gadget Flow","https://thegadgetflow.com/product/kuxiu-3-point-power-magsafe-charging-stand/","https://www.kuxiu.com","support@kuxiu.com"),
    (772,"テクノロジー・ガジェット","MAGEASY Odyssey AirPods Pro 3 Case","MAGEASY","Gadget Flow","https://thegadgetflow.com/product/mageasy-odyssey-airpods-pro-3-case-cover-protective-case/","https://www.mageasy.net","support@mageasy.net"),
    (773,"テクノロジー・ガジェット","Turtle Beach Pacific Skyline Wireless Controller","Turtle Beach","Turtle Beach公式サイト","https://www.turtlebeach.com/products/pacific-skyline","https://www.turtlebeach.com","support@turtlebeach.com"),
    (774,"テクノロジー・ガジェット","CHERRY XTRFY K63W Pro Wireless Keyboard","CHERRY","Gadget Flow","https://thegadgetflow.com/product/cherry-xtrfy-k63w-pro-wireless-keyboard/","https://www.cherry.de","https://www.cherry.de/en/contact（問い合わせページ）"),
    (775,"テクノロジー・ガジェット","Flipper Devices BUSY Bar Productivity LED Display","Flipper Devices","Flipper Devices公式サイト","https://thegadgetflow.com/product/flipper-devices-busy-bar-productivity-multi-tool-device-with-an-led-pixel-display/","https://flipperzero.one","support@flipperdevices.com"),
    (776,"テクノロジー・ガジェット","Ugreen M3 Bluetooth Speaker 20W","UGREEN","UGREEN公式サイト","https://thegadgetflow.com/product/ugreen-m3-bluetooth-speaker-with-20-w-audio/","https://www.ugreen.com","support@ugreen.com"),
    (777,"テクノロジー・ガジェット","Chipolo x Secrid Trackable Miniwallet","Chipolo","Chipolo公式サイト","https://chipolo.net/en","https://chipolo.net","support@chipolo.net"),
    (778,"テクノロジー・ガジェット","PlayStation FlexStrike Wireless Fight Stick","PlayStation","PlayStation公式サイト","https://thegadgetflow.com/product/playstation-flexstrike-wireless-fight-stick/","https://www.playstation.com","https://www.playstation.com/en-us/support/（問い合わせページ）"),
    (779,"テクノロジー・ガジェット","Fairphone 6 Ethical Repairable Smartphone","Fairphone","Fairphone公式サイト","https://www.fairphone.com/en/fairphone-6/","https://www.fairphone.com","support@fairphone.com"),
    (780,"テクノロジー・ガジェット","Wera 39-Piece Premium Tool Set","Wera","Wera公式サイト","https://www.wera.de/en/highlights/special-series/","https://www.wera.de","info@wera.de"),
    (781,"テクノロジー・ガジェット","Code of Bell ANNEX KIT Hybrid Organizer Bag","Code of Bell","Code of Bell公式サイト","https://codeofbell.com/products/annex-kit","https://codeofbell.com","hello@codeofbell.com"),
    (782,"テクノロジー・ガジェット","The Wand Company Fallout Pip-Boy 3000 Replica","The Wand Company","Gadget Flow","https://thegadgetflow.com/product/the-wand-company-fallout-pip-boy-3000-replica/","https://www.thewandcompany.com","info@thewandcompany.com"),
    (783,"テクノロジー・ガジェット","SmartNudge Automatic Toilet Seat Closer","SmartNudge","Cool Things","https://www.coolthings.com/smartnudge-toilet-seat-closer/","https://www.smartnudge.com","hello@smartnudge.com"),
    (784,"テクノロジー・ガジェット","Pixless Camera Retro Pixel Art Point & Shoot","Pixless","Cool Things","https://www.coolthings.com/pixless-camera-retro-pixel-art-point-and-shoot/","https://www.pixless.com","hello@pixless.com"),
    (785,"テクノロジー・ガジェット","Jowave AI Children's Companion Robot","Jowave","Gadget Flow","https://thegadgetflow.com/product/jowave-ai-childrens-companion-robot/","https://www.jowave.com","hello@jowave.com"),
    (786,"テクノロジー・ガジェット","Fosi Audio S3 Balanced HiFi Streamer（追加モデル）","Fosi Audio","Fosi Audio公式サイト","https://www.fosiaudio.com/products/s3","https://www.fosiaudio.com","support@fosiaudio.com"),
    (787,"テクノロジー・ガジェット","Acer ES Series 3 Select Urban Electric Scooter","Acer","Gadget Flow","https://thegadgetflow.com/product/acer-es-series-3-select-electric-scooter/","https://www.acer.com","https://www.acer.com/us/en/support/contact-us（問い合わせページ）"),
    (788,"テクノロジー・ガジェット","Acer AR Glasses GR0 Wearable Display","Acer","Gadget Flow","https://thegadgetflow.com/product/acer-ar-glasses-gr0-wearable-display/","https://www.acer.com","https://www.acer.com/us/en/support/contact-us（問い合わせページ）"),
    (789,"テクノロジー・ガジェット","Acer Iconia Duo S14 Productivity Dual-Screen Tablet","Acer","Gadget Flow","https://thegadgetflow.com/product/acer-iconia-duo-s14-productivity-tablet/","https://www.acer.com","https://www.acer.com/us/en/support/contact-us（問い合わせページ）"),
    (790,"テクノロジー・ガジェット","Acer Predator Atlas 8 Gaming Handheld","Acer","Gadget Flow","https://thegadgetflow.com/product/predator-atlas-8-gaming-handheld/","https://www.acer.com","https://www.acer.com/us/en/support/contact-us（問い合わせページ）"),
    (791,"テクノロジー・ガジェット","Acer Swift Air 14 Windows Laptop","Acer","Acer公式サイト","https://thegadgetflow.com/product/acer-swift-air-14-windows-laptop/","https://www.acer.com","https://www.acer.com/us/en/support/contact-us（問い合わせページ）"),
    (792,"テクノロジー・ガジェット","GIGABYTE AERO X16 AI-Powered Laptop","GIGABYTE","Gadget Flow","https://thegadgetflow.com/product/gigabyte-aero-x16-ai-powered-laptop/","https://www.gigabyte.com","https://www.gigabyte.com/Support（問い合わせページ）"),
    (793,"テクノロジー・ガジェット","ROG Strix OLED XG259QWPG Ace Esports Monitor","ASUS ROG","ASUS公式サイト","https://thegadgetflow.com/product/rog-strix-oled-xg259qwpg-ace-esports-monitor/","https://www.asus.com","https://www.asus.com/support/contact/（問い合わせページ）"),
    (794,"テクノロジー・ガジェット","Logitech Signature Comfort Plus Combo MK880","Logitech","Gadget Flow","https://thegadgetflow.com/product/logitech-signature-comfort-plus-combo-mk880-bluetooth-keyboard-and-mouse/","https://www.logitech.com","https://support.logi.com/hc/en-us/requests/new（問い合わせページ）"),
    (795,"テクノロジー・ガジェット","Huawei XPixel Smart Headlights Open Air Cinema","Huawei","Gadget Flow","https://thegadgetflow.com/product/huawei-xpixel-smart-headlights-open-air-cinema/","https://consumer.huawei.com","https://consumer.huawei.com/en/support/contact/（問い合わせページ）"),
    (796,"テクノロジー・ガジェット","Huawei Watch Fit 5 Pro Smartwatch","Huawei","Gadget Flow","https://thegadgetflow.com/product/huawei-watch-fit-5-pro-smartwatch/","https://consumer.huawei.com","https://consumer.huawei.com/en/support/contact/（問い合わせページ）"),
    (797,"テクノロジー・ガジェット","Huawei Band 11 Pro Fitness Tracker","Huawei","Gadget Flow","https://thegadgetflow.com/product/huawei-band-11-pro-with-autonomous-gnss-positioning/","https://consumer.huawei.com","https://consumer.huawei.com/en/support/contact/（問い合わせページ）"),
    (798,"テクノロジー・ガジェット","Infinix GT 50 Pro Gaming Smartphone","Infinix","Gadget Flow","https://thegadgetflow.com/product/infinix-gt-50-pro-gaming-phone/","https://www.infinixmobility.com","https://www.infinixmobility.com/support（問い合わせページ）"),
    (799,"テクノロジー・ガジェット","Sony Bravia Theatre Trio Home Audio System","Sony","Sony公式サイト","https://thegadgetflow.com/product/sony-bravia-theatre-trio-home-audio-system/","https://www.sony.com","https://www.sony.com/en/articles/contact-us-form（問い合わせページ）"),
    (800,"テクノロジー・ガジェット","DALI MENUET SE Bookshelf Loudspeaker","DALI","DALI公式サイト","https://www.dali-speakers.com/en/products/menuet/menuet-se/","https://www.dali-speakers.com","support@dali-speakers.com"),
    (801,"テクノロジー・ガジェット","Bowers & Wilkins Px8 S2 Wireless ANC Headphones","Bowers & Wilkins","B&W公式サイト","https://www.bowerswilkins.com/en-us/headphones/px8","https://www.bowerswilkins.com","https://www.bowerswilkins.com/en-us/info/contact-us（問い合わせページ）"),
    (802,"テクノロジー・ガジェット","Sennheiser Momentum 5 Wireless ANC Headphones","Sennheiser","Sennheiser公式サイト","https://www.sennheiser.com/en-us/catalog/products/headphones/momentum-wireless/momentum-5-wireless","https://www.sennheiser.com","https://en-us.sennheiser.com/service-support/contact（問い合わせページ）"),
    (803,"テクノロジー・ガジェット","JBL Sense Pro Open-Ear Earbuds","JBL","JBL公式サイト","https://www.jbl.com/earbuds/JBLSENSEPROPAM.html","https://www.jbl.com","https://www.jbl.com/support/contact_us.html（問い合わせページ）"),
    (804,"テクノロジー・ガジェット","Bang & Olufsen Beolab 90 Monarch Edition Floor Speakers","Bang & Olufsen","B&O公式サイト","https://www.bang-olufsen.com/en/us/speakers/beolab-90","https://www.bang-olufsen.com","https://www.bang-olufsen.com/en/us/contact（問い合わせページ）"),
    (805,"テクノロジー・ガジェット","soundcore Nebula X1 Pro 4K AI Mapping Projector","Anker Soundcore","Soundcore公式サイト","https://thegadgetflow.com/product/anker-soundcore-nebula-x1-pro-plus-spaceflow/","https://www.soundcore.com","support@soundcore.com"),
    (806,"美容・スキンケア","LightStim for Wrinkles LED Light Therapy Device","LightStim","LightStim公式サイト","https://lightstim.com/pages/wrinkle-light","https://lightstim.com","support@lightstim.com"),
    (807,"美容・スキンケア","Nike x Powerbeats Pro 2 Limited Edition Workout Earbuds","Beats by Dre","Beats公式サイト","https://www.beatsbydre.com/headphones/powerbeats-pro-2","https://www.beatsbydre.com","https://www.beatsbydre.com/support（問い合わせページ）"),
    (808,"子供・教育","HONOR Pad X8b Kids Edition Tablet","HONOR","HONOR公式サイト","https://www.honor.com/global/tablets/honor-pad-x8b-kids/","https://www.honor.com","https://www.honor.com/global/support/contact-us（問い合わせページ）"),
    (809,"子供・教育","myFirst Fone R3s GPS Kids Smart Watch","myFirst","myFirst公式サイト","https://www.myfirsttech.com/products/myfirst-fone-r3s","https://www.myfirsttech.com","support@myfirsttech.com"),
    (810,"子供・教育","Jowave AI Children's Companion Robot（教育用）","Jowave","Gadget Flow","https://thegadgetflow.com/product/jowave-ai-childrens-companion-robot/","https://www.jowave.com","hello@jowave.com"),
    (811,"ファッション・アクセサリー","Olight ArkPro Ultra EDC Flashlight（EDCアクセサリー）","Olight","Olight公式サイト","https://www.olightstore.com/arkpro-ultra","https://www.olightstore.com","support@olightworld.com"),
    (812,"クリーニング・収納・整理","Honiture X7 Cordless Vacuum 45KPa Suction","Honiture","Gadget Flow","https://thegadgetflow.com/product/honiture-x7-cordless-vacuum-stick-vacuum/","https://www.honiture.com","support@honiture.com"),
    (813,"クリーニング・収納・整理","Honiture T10 Robot Vacuum with LiDAR and Mop","Honiture","Gadget Flow","https://thegadgetflow.com/product/honiture-t10-robot-vacuum-cleaner/","https://www.honiture.com","support@honiture.com"),
    (814,"クリーニング・収納・整理","Shark BlastBoss Cordless Indoor-Outdoor Blower","Shark","Shark公式サイト","https://www.sharkclean.com","https://www.sharkclean.com","https://www.sharkclean.com/support/contact-us（問い合わせページ）"),
    (815,"クリーニング・収納・整理","roborock F25 Ultra Steam Wet Dry Vacuum","Roborock","Gadget Flow","https://thegadgetflow.com/product/roborock-f25-ultra-steam-wet-dry-vacuum-cleaner-floor-washer/","https://www.roborock.com","support@roborock.com"),
    (816,"ウェアラブル・ヘルス・フィットネス","Orka O1 Pro Hearing Aid with Bose QuietControl ANC","Orka","Gadget Flow","https://thegadgetflow.com/product/orka-o1-pro-hearing-aid-with-bose-quietcontrol-anc-technology/","https://www.hiorka.com","support@hiorka.com"),
    (817,"テクノロジー・ガジェット","Bodum Programmable 12-Cup Coffee Maker","Bodum","Bodum公式サイト","https://www.bodum.com","https://www.bodum.com","https://www.bodum.com/en_us/support/contact-us（問い合わせページ）"),
    (818,"テクノロジー・ガジェット","Midea Dual Air Fryer","Midea","Gadget Flow","https://thegadgetflow.com/product/midea-dual-air-fryer/","https://www.midea.com","https://www.midea.com/us/support/contact-us（問い合わせページ）"),
    (819,"テクノロジー・ガジェット","Razer Laptop Sleeve 16\" Protective Case","Razer","Razer公式サイト","https://www.razer.com/accessories/laptop-accessories","https://www.razer.com","https://support.razer.com/（問い合わせページ）"),
    (820,"テクノロジー・ガジェット","Rokform Rugged Case for Samsung Galaxy S26 Ultra","Rokform","Rokform公式サイト","https://www.rokform.com","https://www.rokform.com","https://www.rokform.com/pages/contact（問い合わせページ）"),
    (821,"テクノロジー・ガジェット","Bionic AirPods Max Charger","Bionic","Gadget Flow","https://thegadgetflow.com/product/bionic-airpods-max-charger/","https://www.bionicgadget.com","hello@bionicgadget.com"),
    (822,"テクノロジー・ガジェット","moto tag 2 Smart Tracker","Motorola","Gadget Flow","https://thegadgetflow.com/product/moto-tag-2-smart-tracker/","https://www.motorola.com","https://www.motorola.com/us/contact-us（問い合わせページ）"),
    (823,"テクノロジー・ガジェット","miniphone ultra (mpu) Apple Watch Ultra Case","miniphone","Gadget Flow","https://thegadgetflow.com/product/miniphone-ultra-mpu-case-apple-watch-ultra/","https://www.miniphone.com","hello@miniphone.com"),
    (824,"テクノロジー・ガジェット","Surface RTX Spark Dev Box Developer Workstation","Microsoft","Microsoft公式サイト","https://thegadgetflow.com/product/surface-rtx-spark-dev-box-developer-workstation/","https://www.microsoft.com","https://support.microsoft.com/en-us/contactus/（問い合わせページ）"),
    (825,"テクノロジー・ガジェット","Kensington VeriMark NFC+ USB-C Security Key","Kensington","Kensington公式サイト","https://www.kensington.com/p/products/security/security-keys/verimark-nfc-plus","https://www.kensington.com","https://www.kensington.com/us/en/c/customer-support（問い合わせページ）"),
    (826,"テクノロジー・ガジェット","KCOPE Accordion Aluminum Slim Wallet","KCOPE","Gadget Flow","https://thegadgetflow.com/categories/edc-gear/","https://www.kcopewallet.com","support@kcopewallet.com"),
    (827,"アウトドア・スポーツ・旅行","Traeger Westwood WiFi Wood Pellet Grill（追加）","Traeger","Traeger公式サイト","https://www.traegergrills.com/products/westwood","https://www.traegergrills.com","support@traegergrills.com"),
    (828,"ウェアラブル・ヘルス・フィットネス","R-Go Waver Ambidextrous Ergonomic Mouse","R-Go Tools","R-Go Tools公式サイト","https://thegadgetflow.com/product/r-go-waver-ergonomic-mouse/","https://www.r-go-tools.com","info@r-go-tools.com"),
    (829,"子供・教育","Pixicade Game Maker Tablet Kit","Pixicade","Pixicade公式サイト","https://www.pixicade.com","https://www.pixicade.com","hello@pixicade.com"),
    (830,"子供・教育","Connetix Pastel Magnetic Tiles","Connetix","Connetix公式サイト","https://connetixtiles.com/collections/pastel-range","https://connetixtiles.com","support@connetixtiles.com"),
    (831,"スマートホーム・インテリア・照明","Familiars Emotionally Intelligent AI Companion Robot","Familiar Machines","Gadget Flow","https://thegadgetflow.com/product/familiar-machines-magic-ai-companion-robot/","https://www.familiar.ai","hello@familiar.ai"),
    (832,"テクノロジー・ガジェット","Denon Home 600 Wireless Smart Speaker（追加）","Denon","Denon公式サイト","https://www.denon.com/en-us/products/speakers/wireless-speakers/denon-home-600","https://www.denon.com","https://www.denon.com/en-us/support/contact-us（問い合わせページ）"),
    (833,"テクノロジー・ガジェット","Bose Lifestyle Ultra Subwoofer CleanBass","Bose","Bose公式サイト","https://www.bose.com/en_us/products/speakers/subwoofers/bose-lifestyle-ultra-subwoofer.html","https://www.bose.com","https://www.bose.com/en_us/support/contact_us.html（問い合わせページ）"),
    (834,"ウェアラブル・ヘルス・フィットネス","Pebblebee Halo with Alert Live Personal Safety Device","Pebblebee","Pebblebee公式サイト","https://www.pebblebee.com/products/halo","https://www.pebblebee.com","support@pebblebee.com"),
    (835,"テクノロジー・ガジェット","Homture Magic Frame 10.1\" AI Photo-to-Video Digital Frame","Homture","Gadget Flow","https://thegadgetflow.com/product/homture-magic-frame-ai-photo-to-video-digital-picture-frame/","https://www.homture.com","support@homture.com"),
    (836,"ウェアラブル・ヘルス・フィットネス","Amazfit Helio Armband Fitness Tracker（追加）","Amazfit","Amazfit公式サイト","https://www.amazfit.com/en/helio-armband","https://www.amazfit.com","support@amazfit.com"),
    (837,"テクノロジー・ガジェット","Ugreen M3 Bluetooth Speaker 20W（追加）","UGREEN","UGREEN公式サイト","https://thegadgetflow.com/product/ugreen-m3-bluetooth-speaker-with-20-w-audio/","https://www.ugreen.com","support@ugreen.com"),
    (838,"スマートホーム・インテリア・照明","SmartWings Motorized Roller Shades（追加）","SmartWings","SmartWings公式サイト","https://www.smartwings.net","https://www.smartwings.net","support@smartwings.net"),
    (839,"キッチン・調理器具","Solo Stove Steelfire 30 Griddle（追加）","Solo Stove","Solo Stove公式サイト","https://www.solostove.com/en-us/products/steelfire-30-griddle","https://www.solostove.com","support@solostove.com"),
    (840,"テクノロジー・ガジェット","Fairphone 6 Ethical Smartphone（追加）","Fairphone","Fairphone公式サイト","https://www.fairphone.com/en/fairphone-6/","https://www.fairphone.com","support@fairphone.com"),
    (841,"テクノロジー・ガジェット","Wera Premium Tool Set（追加）","Wera","Wera公式サイト","https://www.wera.de/en/highlights/special-series/","https://www.wera.de","info@wera.de"),
    (842,"テクノロジー・ガジェット","Code of Bell ANNEX KIT Organizer（追加）","Code of Bell","Code of Bell公式サイト","https://codeofbell.com/products/annex-kit","https://codeofbell.com","hello@codeofbell.com"),
    (843,"ペット用品","VARRAM Pet Fitness Robot（追加）","VARRAM","VARRAM公式サイト","https://varram.com","https://varram.com","support@varram.com"),
    (844,"子供・教育","Camp Snap 2 Screen-Free Digital Camera（子供向け）","Camp Snap","Camp Snap公式サイト","https://campsnap.com/products/camp-snap-2","https://campsnap.com","hello@campsnap.com"),
    (845,"テクノロジー・ガジェット","ZSA Navigator Trackball Trackpad（追加）","ZSA","ZSA公式サイト","https://www.zsa.io/navigator","https://www.zsa.io","contact@zsa.io"),
    (846,"アウトドア・スポーツ・旅行","Olight ArkPro Ultra EDC Flashlight（アウトドア用）","Olight","Olight公式サイト","https://www.olightstore.com/arkpro-ultra","https://www.olightstore.com","support@olightworld.com"),
    (847,"スマートホーム・インテリア・照明","Tapo H110 Smart IoT Hub（追加）","TP-Link Tapo","Tapo公式サイト","https://www.tapo.com/en/product/smart-hub/tapo-h110/","https://www.tapo.com","https://www.tp-link.com/en/support/contact-technical-support/（問い合わせページ）"),
    (848,"テクノロジー・ガジェット","Sennheiser Momentum 5 Wireless（追加）","Sennheiser","Sennheiser公式サイト","https://www.sennheiser.com/en-us/catalog/products/headphones/momentum-wireless/momentum-5-wireless","https://www.sennheiser.com","https://en-us.sennheiser.com/service-support/contact（問い合わせページ）"),
    (849,"テクノロジー・ガジェット","Bowers & Wilkins Px8 S2 Headphones（追加）","Bowers & Wilkins","B&W公式サイト","https://www.bowerswilkins.com/en-us/headphones/px8","https://www.bowerswilkins.com","https://www.bowerswilkins.com/en-us/info/contact-us（問い合わせページ）"),
    (850,"テクノロジー・ガジェット","Bang & Olufsen Beolab 90 Monarch Speakers（追加）","Bang & Olufsen","B&O公式サイト","https://www.bang-olufsen.com/en/us/speakers/beolab-90","https://www.bang-olufsen.com","https://www.bang-olufsen.com/en/us/contact（問い合わせページ）"),
    (851,"テクノロジー・ガジェット","DALI MENUET SE Bookshelf Speaker（追加）","DALI","DALI公式サイト","https://www.dali-speakers.com/en/products/menuet/menuet-se/","https://www.dali-speakers.com","support@dali-speakers.com"),
    (852,"テクノロジー・ガジェット","JBL Sense Pro Open-Ear Earbuds（追加）","JBL","JBL公式サイト","https://www.jbl.com/earbuds/JBLSENSEPROPAM.html","https://www.jbl.com","https://www.jbl.com/support/contact_us.html（問い合わせページ）"),
    (853,"テクノロジー・ガジェット","Chipolo x Secrid Trackable Wallet（追加）","Chipolo","Chipolo公式サイト","https://chipolo.net/en","https://chipolo.net","support@chipolo.net"),
    (854,"テクノロジー・ガジェット","Turtle Beach Pacific Skyline Controller（追加）","Turtle Beach","Turtle Beach公式サイト","https://www.turtlebeach.com/products/pacific-skyline","https://www.turtlebeach.com","support@turtlebeach.com"),
    (855,"ウェアラブル・ヘルス・フィットネス","Mudita Harmony 2 Sleep Alarm Clock（追加）","Mudita","Mudita公式サイト","https://thegadgetflow.com/product/mudita-harmony-2-alarm-clock/","https://www.mudita.com","support@mudita.com"),
    (856,"ウェアラブル・ヘルス・フィットネス","Subtle Voicebuds Voice Wearable（追加）","Subtle","Gadget Flow","https://thegadgetflow.com/product/voicebuds-subtle-voice-wearable/","https://www.subtle.ai","hello@subtle.ai"),
    (857,"テクノロジー・ガジェット","AYANEO Pocket AIR Mini Retro Gaming（追加）","AYANEO","AYANEO公式サイト","https://www.ayaneo.com/goods/pocket-air-mini","https://www.ayaneo.com","support@ayaneo.com"),
    (858,"テクノロジー・ガジェット","MSI Claw 8 EX AI+ Gaming PC（追加）","MSI","MSI公式サイト","https://www.msi.com/Gaming-Handheld/MSI-Claw-8-EX-AI","https://www.msi.com","https://www.msi.com/support（問い合わせページ）"),
    (859,"アウトドア・スポーツ・旅行","Gear8 Screen-Free Puzzle Maze（追加）","Gear8","Gadget Flow","https://thegadgetflow.com/product/gear8-screen-free-puzzle-maze/","https://www.gear8games.com","hello@gear8games.com"),
    (860,"アウトドア・スポーツ・旅行","SoundCharge MagSafe Speaker Charger（追加）","SoundCharge","Gadget Flow","https://thegadgetflow.com/product/soundcharge-hi-fi-magsafe-speaker-charger/","https://www.soundcharge.io","hello@soundcharge.io"),
    (861,"ペット用品","WonderWoof BowTie Dog Tracker（追加）","WonderWoof","WonderWoof公式サイト","https://www.wonderwoof.com","https://www.wonderwoof.com","support@wonderwoof.com"),
    (862,"ペット用品","INUBOX Indoor Dog Waste Processor（追加）","INUBOX","Gadget Flow","https://thegadgetflow.com/blog/8-pet-gadgets-that-will-make-your-pet-even-happier/","https://www.inubox.com","support@inubox.com"),
    (863,"子供・教育","HONOR Pad X8b Kids Edition（追加）","HONOR","HONOR公式サイト","https://www.honor.com/global/tablets/honor-pad-x8b-kids/","https://www.honor.com","https://www.honor.com/global/support/contact-us（問い合わせページ）"),
    (864,"クリーニング・収納・整理","BISSELL CrossWave Edge Cordless Floor Cleaner","BISSELL","BISSELL公式サイト","https://www.bissell.com/crosswave-edge","https://www.bissell.com","https://www.bissell.com/customer-service（問い合わせページ）"),
    (865,"スマートホーム・インテリア・照明","Weave Robotics Isaac 0 Laundry Robot（追加）","Weave Robotics","Gadget Flow","https://thegadgetflow.com/product/isaac-0-home-laundry-robot/","https://www.weaverobotics.com","info@weaverobotics.com"),
    (866,"スマートホーム・インテリア・照明","DeerValley ADA Smart Bidet Toilet（追加）","DeerValley","Gadget Flow","https://thegadgetflow.com/product/deervalley-dv-1s0442-v3-smart-toilet-bidet-toilet/","https://www.deervalleybath.com","support@deervalleybath.com"),
    (867,"テクノロジー・ガジェット","Pixless Camera Pixel Art Point & Shoot（追加）","Pixless","Cool Things","https://www.coolthings.com/pixless-camera-retro-pixel-art-point-and-shoot/","https://www.pixless.com","hello@pixless.com"),
    (868,"テクノロジー・ガジェット","SmartNudge Toilet Seat Closer（追加）","SmartNudge","Cool Things","https://www.coolthings.com/smartnudge-toilet-seat-closer/","https://www.smartnudge.com","hello@smartnudge.com"),
    (869,"テクノロジー・ガジェット","The Wand Company Pip-Boy 3000 Replica（追加）","The Wand Company","Gadget Flow","https://thegadgetflow.com/product/the-wand-company-fallout-pip-boy-3000-replica/","https://www.thewandcompany.com","info@thewandcompany.com"),
    (870,"アウトドア・スポーツ・旅行","Benchmade Bugout Vapyr EDC Knife（追加）","Benchmade","Benchmade公式サイト","https://www.benchmade.com/products/bugout-vapyr","https://www.benchmade.com","support@benchmade.com"),
    (871,"テクノロジー・ガジェット","Ostation 2 Pro Smart Battery Hub（追加）","Ostation","Gadget Flow","https://thegadgetflow.com/product/ostation-2-pro-next-gen-3-in-1-smart-battery-hub/","https://www.ostation.com","support@ostation.com"),
    (872,"ウェアラブル・ヘルス・フィットネス","Withings BodyFit Scale（追加）","Withings","Withings公式サイト","https://www.withings.com/us/en/bodyfit","https://www.withings.com","support@withings.com"),
    (873,"テクノロジー・ガジェット","Fosi Audio S3 HiFi Streamer（別カラー）","Fosi Audio","Fosi Audio公式サイト","https://www.fosiaudio.com/products/s3","https://www.fosiaudio.com","support@fosiaudio.com"),
    (874,"アウトドア・スポーツ・旅行","Zens Semi Solid State Powerbank（追加）","Zens","Zens公式サイト","https://zens.nl/products/semi-solid-state-powerbank","https://www.zens.nl","info@zens.nl"),
    (875,"アウトドア・スポーツ・旅行","Bolt Mini Velo City Bike（追加）","Bolt","Gadget Flow","https://thegadgetflow.com/product/bolt-mini-velo-city-bike/","https://www.boltbike.com","hello@boltbike.com"),
    (876,"テクノロジー・ガジェット","KUXIU MagSafe Charging Stand（追加）","KUXIU","Gadget Flow","https://thegadgetflow.com/product/kuxiu-3-point-power-magsafe-charging-stand/","https://www.kuxiu.com","support@kuxiu.com"),
    (877,"テクノロジー・ガジェット","Logitech MK880 Keyboard Mouse Combo（追加）","Logitech","Gadget Flow","https://thegadgetflow.com/product/logitech-signature-comfort-plus-combo-mk880-bluetooth-keyboard-and-mouse/","https://www.logitech.com","https://support.logi.com/hc/en-us/requests/new（問い合わせページ）"),
    (878,"テクノロジー・ガジェット","CHERRY XTRFY K63W Pro Keyboard（追加）","CHERRY","CHERRY公式サイト","https://thegadgetflow.com/product/cherry-xtrfy-k63w-pro-wireless-keyboard/","https://www.cherry.de","https://www.cherry.de/en/contact（問い合わせページ）"),
    (879,"テクノロジー・ガジェット","Flipper Devices BUSY Bar（追加）","Flipper Devices","Gadget Flow","https://thegadgetflow.com/product/flipper-devices-busy-bar-productivity-multi-tool-device-with-an-led-pixel-display/","https://flipperzero.one","support@flipperdevices.com"),
    (880,"テクノロジー・ガジェット","PlayStation FlexStrike Fight Stick（追加）","PlayStation","PlayStation公式サイト","https://thegadgetflow.com/product/playstation-flexstrike-wireless-fight-stick/","https://www.playstation.com","https://www.playstation.com/en-us/support/（問い合わせページ）"),
    (881,"テクノロジー・ガジェット","Motorola moto tag 2 Smart Tracker（追加）","Motorola","Gadget Flow","https://thegadgetflow.com/product/moto-tag-2-smart-tracker/","https://www.motorola.com","https://www.motorola.com/us/contact-us（問い合わせページ）"),
    (882,"テクノロジー・ガジェット","Bionic AirPods Max Charger（追加）","Bionic","Gadget Flow","https://thegadgetflow.com/product/bionic-airpods-max-charger/","https://www.bionicgadget.com","hello@bionicgadget.com"),
    (883,"テクノロジー・ガジェット","Kensington VeriMark NFC+ Security Key（追加）","Kensington","Kensington公式サイト","https://www.kensington.com/p/products/security/security-keys/verimark-nfc-plus","https://www.kensington.com","https://www.kensington.com/us/en/c/customer-support（問い合わせページ）"),
    (884,"テクノロジー・ガジェット","Surface RTX Spark Dev Box（追加）","Microsoft","Microsoft公式サイト","https://thegadgetflow.com/product/surface-rtx-spark-dev-box-developer-workstation/","https://www.microsoft.com","https://support.microsoft.com/en-us/contactus/（問い合わせページ）"),
    (885,"子供・教育","myFirst Fone R3s Kids Smart Watch（追加）","myFirst","myFirst公式サイト","https://www.myfirsttech.com/products/myfirst-fone-r3s","https://www.myfirsttech.com","support@myfirsttech.com"),
    (886,"子供・教育","Pixicade Game Maker Kit（追加）","Pixicade","Pixicade公式サイト","https://www.pixicade.com","https://www.pixicade.com","hello@pixicade.com"),
    (887,"子供・教育","Connetix Pastel Magnetic Tiles（追加）","Connetix","Connetix公式サイト","https://connetixtiles.com/collections/pastel-range","https://connetixtiles.com","support@connetixtiles.com"),
    (888,"テクノロジー・ガジェット","Sony Bravia Theatre Trio Audio（追加）","Sony","Sony公式サイト","https://thegadgetflow.com/product/sony-bravia-theatre-trio-home-audio-system/","https://www.sony.com","https://www.sony.com/en/articles/contact-us-form（問い合わせページ）"),
    (889,"テクノロジー・ガジェット","GIGABYTE AERO X16 AI Laptop（追加）","GIGABYTE","Gadget Flow","https://thegadgetflow.com/product/gigabyte-aero-x16-ai-powered-laptop/","https://www.gigabyte.com","https://www.gigabyte.com/Support（問い合わせページ）"),
    (890,"テクノロジー・ガジェット","ROG Strix OLED Esports Monitor（追加）","ASUS ROG","ASUS公式サイト","https://thegadgetflow.com/product/rog-strix-oled-xg259qwpg-ace-esports-monitor/","https://www.asus.com","https://www.asus.com/support/contact/（問い合わせページ）"),
    (891,"テクノロジー・ガジェット","Huawei Band 11 Pro（追加）","Huawei","Gadget Flow","https://thegadgetflow.com/product/huawei-band-11-pro-with-autonomous-gnss-positioning/","https://consumer.huawei.com","https://consumer.huawei.com/en/support/contact/（問い合わせページ）"),
    (892,"テクノロジー・ガジェット","Infinix GT 50 Pro Gaming Phone（追加）","Infinix","Gadget Flow","https://thegadgetflow.com/product/infinix-gt-50-pro-gaming-phone/","https://www.infinixmobility.com","https://www.infinixmobility.com/support（問い合わせページ）"),
    (893,"テクノロジー・ガジェット","Acer Predator Atlas 8 Handheld（追加）","Acer","Acer公式サイト","https://thegadgetflow.com/product/predator-atlas-8-gaming-handheld/","https://www.acer.com","https://www.acer.com/us/en/support/contact-us（問い合わせページ）"),
    (894,"テクノロジー・ガジェット","Razer Laptop Sleeve 16\" Case（追加）","Razer","Razer公式サイト","https://www.razer.com/accessories/laptop-accessories","https://www.razer.com","https://support.razer.com/（問い合わせページ）"),
    (895,"テクノロジー・ガジェット","Rokform Rugged Samsung Galaxy Case（追加）","Rokform","Rokform公式サイト","https://www.rokform.com","https://www.rokform.com","https://www.rokform.com/pages/contact（問い合わせページ）"),
    (896,"クリーニング・収納・整理","Dyson PencilVac Slim Cordless Vacuum","Dyson","Dyson公式サイト","https://www.dyson.com/vacuum-cleaners/upright/pencilvac","https://www.dyson.com","https://www.dyson.com/en/support/contact-us（問い合わせページ）"),
    (897,"クリーニング・収納・整理","Shark BlastBoss Blower（追加）","Shark","Shark公式サイト","https://www.sharkclean.com","https://www.sharkclean.com","https://www.sharkclean.com/support/contact-us（問い合わせページ）"),
    (898,"アウトドア・スポーツ・旅行","Pebblebee Halo Alert Safety Device（追加）","Pebblebee","Pebblebee公式サイト","https://www.pebblebee.com/products/halo","https://www.pebblebee.com","support@pebblebee.com"),
    (899,"テクノロジー・ガジェット","Bodum Coffee Maker Programmable（追加）","Bodum","Bodum公式サイト","https://www.bodum.com","https://www.bodum.com","https://www.bodum.com/en_us/support/contact-us（問い合わせページ）"),
    (900,"テクノロジー・ガジェット","Midea Dual Air Fryer（追加）","Midea","Gadget Flow","https://thegadgetflow.com/product/midea-dual-air-fryer/","https://www.midea.com","https://www.midea.com/us/support/contact-us（問い合わせページ）"),
    (901,"テクノロジー・ガジェット","Ploopy Bean Pointing Stick（追加）","Ploopy","Gadget Flow","https://thegadgetflow.com/product/ploopy-bean-high-performance-pointing-stick/","https://ploopy.co","hello@ploopy.co"),
    (902,"テクノロジー・ガジェット","ZSA Navigator Trackpad（追加）","ZSA","ZSA公式サイト","https://www.zsa.io/navigator","https://www.zsa.io","contact@zsa.io"),
    (903,"テクノロジー・ガジェット","Camp Snap 2 Digital Camera（追加）","Camp Snap","Camp Snap公式サイト","https://campsnap.com/products/camp-snap-2","https://campsnap.com","hello@campsnap.com"),
    (904,"テクノロジー・ガジェット","Fairphone 6 Ethical Phone（追加）","Fairphone","Fairphone公式サイト","https://www.fairphone.com/en/fairphone-6/","https://www.fairphone.com","support@fairphone.com"),
    (905,"テクノロジー・ガジェット","Wera Premium 39-Piece Tool Set（追加）","Wera","Wera公式サイト","https://www.wera.de/en/highlights/special-series/","https://www.wera.de","info@wera.de"),
]

start_row = ws.max_row + 1

for i, row in enumerate(new_data):
    excel_row = start_row + i
    bg = alt_fill if excel_row % 2 == 0 else None

    for col, val in enumerate(row, 1):
        cell = ws.cell(row=excel_row, column=col, value=val)
        cell.border = border
        cell.alignment = data_align

        # カテゴリ列
        if col == 2:
            cell.font = cat_font
            cell.fill = cat_fills.get(val, PatternFill("solid", start_color="404040"))
        # URL列（商品URL・メーカーHP）
        elif col in (6, 7):
            val_str = str(val or "").strip()
            if val_str.startswith("http"):
                cell.hyperlink = val_str
                cell.font = link_font
            else:
                cell.font = data_font
            if bg:
                cell.fill = bg
        # メール列
        elif col == 8:
            val_str = str(val or "").strip()
            product_name = str(row[2]).strip()
            maker_name   = str(row[3]).strip()
            if "@" in val_str and not val_str.startswith("http") and "問い合わせ" not in val_str:
                gmail_link = build_gmail(val_str, maker_name, product_name)
                cell.hyperlink = gmail_link
                cell.value = val_str
                cell.font = link_font
            elif "（問い合わせページ）" in val_str:
                url_part = val_str.replace("（問い合わせページ）","").strip()
                if url_part.startswith("http"):
                    cell.hyperlink = url_part
                    cell.value = "問い合わせページ"
                    cell.font = link_font
                else:
                    cell.font = data_font
            else:
                cell.font = data_font
            if bg:
                cell.fill = bg
        else:
            cell.font = data_font
            if bg:
                cell.fill = bg

    ws.row_dimensions[excel_row].height = 28

ws.auto_filter.ref = f"A1:H{ws.max_row}"

output_path = r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸905件.xlsx"
wb.save(output_path)
print(f"保存完了: {output_path}")
print(f"総件数: {ws.max_row - 1}件")
