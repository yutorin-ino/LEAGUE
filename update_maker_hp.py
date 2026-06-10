from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# 調査済みメーカーHP辞書 {商品名キーワード: (メーカーHP, メール)}
MAKER_DATA = {
    # ZECZEC
    "Allite": ("https://www.allite.com.tw", ""),
    "CASA Hub": ("https://www.casa.com.tw", ""),
    "GENKI": ("https://genkithings.com", "support@genkithings.com"),
    "Rokid": ("https://www.rokid.com", ""),
    "ASUS AirVision": ("https://www.asus.com/jp", ""),
    "ROG XREAL": ("https://rog.asus.com", ""),
    "DPVR": ("https://www.dpvrvr.com", ""),
    "RingConn": ("https://www.ringconn.com", ""),
    "MOZTECH": ("https://www.moztech.com.tw", ""),
    "GOOVIS": ("https://www.goovis.com", ""),
    "AIFA": ("https://www.aifa.com.tw", "service@aifa.com.tw"),
    "Yeelight": ("https://www.yeelight.com", "support@yeelight.com"),
    "Lytmi": ("https://www.lytmi.com", ""),
    "Philips": ("https://www.philips.com/ja-jp", ""),
    "LG": ("https://www.lg.com/jp", "jp.lgservice@lge.com"),
    "Energy Robotics": ("https://www.energyroboticstech.com", ""),
    "HiDock": ("https://www.hidock.com", "support@hidock.com"),
    "ZMI": ("https://www.zmi.com", ""),
    "Acer": ("https://www.acer.com/tw", ""),
    "Samsung": ("https://www.samsung.com/jp", ""),
    "ECOVACS": ("https://www.ecovacs.com/jp", "jp@ecovacs.com"),
    "Swiftpoint": ("https://www.swiftpoint.com", ""),
    "TileRec": ("https://www.atto-digital.com", ""),
    "Slimca": ("https://www.atto-digital.com", ""),
    "Smini": ("https://www.brise.com.tw", ""),
    "FLAT": ("https://www.sauberair.com", ""),
    "RheoFit": ("https://rheofit.com", ""),
    "OYO": ("https://www.oyogym.com", ""),
    "FAMMIX": ("https://www.yomix.com.tw", ""),
    "Cooler Master": ("https://www.coolermaster.com/jp", ""),
    # Indiegogo
    "Looktech": ("https://www.looktech.ai", ""),
    "NarTick": ("https://nartick.com", "cs@nartick.com"),
    "Majestic Devices": ("https://www.majesticdevices.com", ""),
    "DASUNG": ("https://shop.dasung.com", "contact@dasung.com"),
    "Sequent": ("https://www.sequentag.com", ""),
    "VIITA": ("https://www.viitawatches.com", ""),
    # FlyingV
    "BEZALEL": ("https://bezalel.co", ""),
    "POIEMA": ("https://www.poiema.com.tw", "custcare@poiema.com.tw"),
    "Bloomin8": ("https://bloomin8.com", ""),
    "COFO": ("https://cofoglobal.com", ""),
    "RayNeo": ("https://www.rayneo.com", ""),
    "MAD Gaze": ("https://www.madgaze.com", ""),
    "Gravastar": ("https://www.gravastar.com", ""),
    "Libratone": ("https://www.libratone.com", ""),
    "Cricut": ("https://www.cricut.com", ""),
    "NECCHI": ("https://www.necchi.com", ""),
    "智米": ("https://www.smartmi.com", ""),
    "Smartmi": ("https://www.smartmi.com", ""),
    "Dream Glass": ("https://dreamglasslab.com", ""),
}

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")

link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
normal_font = Font(name="Arial", size=10)

updated = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        product_name = str(ws.cell(row=row, column=2).value or "")
        maker_name   = str(ws.cell(row=row, column=3).value or "")
        hp_cell      = ws.cell(row=row, column=5)
        mail_cell    = ws.cell(row=row, column=6)

        if hp_cell.value and str(hp_cell.value).startswith("http"):
            continue  # 既に入力済み

        # キーワードマッチ
        for keyword, (hp, mail) in MAKER_DATA.items():
            if keyword.lower() in product_name.lower() or keyword.lower() in maker_name.lower():
                hp_cell.value = hp
                hp_cell.hyperlink = hp
                hp_cell.font = link_font
                if mail and (not mail_cell.value or mail_cell.value == "要調査"):
                    mail_cell.value = mail
                    mail_cell.font = normal_font
                updated += 1
                break

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"更新完了: {updated}件のメーカーHPを追加")

# 未入力件数の確認
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    remain = sum(1 for r in range(2, ws.max_row + 1)
                 if not str(ws.cell(row=r, column=5).value or "").startswith("http"))
    print(f"  [{sheet_name}] 要調査残り: {remain}件")
