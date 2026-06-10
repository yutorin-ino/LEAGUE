from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 第3弾 新規追加製品リスト（22件）
# ZECZEC 12件 / FlyingV 4件 / Indiegogo 6件
# 日本CF（Makuake / Campfire / Greenfunding）未掲載確認済み
# ============================================================
#
# 列構成: カテゴリ(1) | 商品名(2) | メーカー名(3) | CFページURL(4)
#         メーカーHP(5) | コンタクトメール(6) | スコア(7) | コメント(8)
#
NEW_PRODUCTS = {
    "ZECZEC": [
        # (カテゴリ, 商品名, メーカー名, CFページURL, メーカーHP, メール, スコア, コメント)
        (
            "オーディオ・録音",
            "RingStar Air｜耳夾式Hi-Res耳機 Z型不壓迫 千元內",
            "LinkLike / RingStar",
            "https://www.zeczec.com/projects/linklike-ringstar-air",
            "要調査", "要調査", 4,
            "耳掛け式Hi-Resイヤホン 圧迫感ゼロ設計"
        ),
        (
            "ガジェット・充電",
            "FREEZE-X 智冷充｜Qi2.2 25W 磁吸無線充電器 半導體冷卻",
            "FREEZE-X",
            "https://www.zeczec.com/projects/freeze-x",
            "要調査", "要調査", 4,
            "冷却機能付きQi2.2ワイヤレス充電スタンド 世界初"
        ),
        (
            "ガジェット・充電",
            "TITANSHIELD 穩如鈦山 Pro｜Qi2.2 固態行動電源 防爆",
            "GrantClassic",
            "https://www.zeczec.com/projects/titanshield-pro",
            "要調査", "要調査", 4,
            "Qi2.2対応 防爆固態バッテリー 2026年新型"
        ),
        (
            "ペットテック",
            "PetPivot AS11 開放式自動貓砂機｜全球首創深度清潔",
            "PetPivot",
            "https://www.zeczec.com/projects/petpivot-as11",
            "要調査", "要調査", 4,
            "世界初 深度クリーニング 汚れ貼り付き問題を解決 WiFi不要"
        ),
        (
            "ペットテック",
            "Nareca 喵淨界｜智慧開放貓砂機 首創階梯式落砂",
            "Nareca",
            "https://www.zeczec.com/projects/nareca",
            "要調査", "要調査", 3,
            "初の階段式砂落ち防止設計 5重安全 大容量"
        ),
        (
            "映像・プロジェクター",
            "SHOBOX 全能PC投影機｜全球首創五合一 1000ANSI 100吋",
            "SHOBOX",
            "https://www.zeczec.com/projects/SHOBOX",
            "要調査", "要調査", 4,
            "TV+PC+プロジェクター+ゲーム機+スピーカー 世界初5合1"
        ),
        (
            "オーディオ・録音",
            "SANSUI 聲空艙 SS2｜5.1.2 Soundbar 360°天空聲道 600W 杜比全景聲",
            "SANSUI (山水電氣)",
            "https://www.zeczec.com/projects/SANSUI-SS2",
            "要調査", "要調査", 4,
            "Dolby Atmos 5.1.2ch 360°天空音道 600W"
        ),
        (
            "PC周辺・入力機器",
            "AI ONE 智慧語音集線器｜十孔擴充×語音轉寫×15種AI機能",
            "AI ONE",
            "https://www.zeczec.com/projects/AIONEHUB",
            "要調査", "要調査", 4,
            "USBハブ×AI音声文字起こし×15種AI機能 一体型"
        ),
        (
            "E-Ink・ディスプレイ",
            "ememo 電子便條紙｜世界首創天然竹料電子顯示器 磁吸設計",
            "ememo",
            "https://www.zeczec.com/projects/ememo",
            "要調査", "要調査", 4,
            "世界初 天然竹製E-Inkメモボード マグネット着脱"
        ),
        (
            "E-Ink・ディスプレイ",
            "ULANI 永恆曆 ∞｜世界首款彩色電子紙桌曆",
            "ULANI",
            "https://www.zeczec.com/projects/ulani",
            "要調査", "要調査", 4,
            "世界初 カラーE-Ink卓上カレンダー"
        ),
        (
            "スマートホーム・IoT・ロボット",
            "PhotoTag 永恆足｜全球首款彩色電子紙定位防丟器 照片＋名片",
            "PhotoTag",
            "https://www.zeczec.com/projects/phototag",
            "要調査", "要調査", 4,
            "世界初 カラーE-Ink GPS紛失防止タグ 写真+名刺+追跡"
        ),
        (
            "スマートホーム・IoT・ロボット",
            "Govee FunCube AI 光影玩咖｜AI影音燈光同步盒＋智慧燈帶",
            "Govee",
            "https://www.zeczec.com/projects/govee-funcube-ai",
            "https://www.govee.com",
            "要調査", 3,
            "AI画面同期アンビエントライトボックス スマートLEDストリップ付"
        ),
    ],

    "FlyingV": [
        (
            "ペットテック",
            "CATLINK Open-X 翻滾喵｜開放式智慧貓砂盆 全天候自由呼吸",
            "CATLINK",
            "https://www.flyingv.cc/projects/36921",
            "https://www.catlink.net",
            "要調査", 4,
            "360°オープン型スマート猫トイレ 達成率141,098%"
        ),
        (
            "ペットテック",
            "WANTECH G5 開放式智慧貓砂盆｜360°零壓迫視野 為貓咪而生",
            "WANTECH",
            "https://www.flyingv.cc/projects/36536",
            "要調査", "要調査", 3,
            "360°完全視野 猫ストレスゼロ設計 スマート猫トイレ"
        ),
        (
            "家電・キッチン",
            "GE 智慧瞬熱烤箱｜260℃碳纖光波×±5℃數位精準溫控 鷗翼瞬熱",
            "GE (General Electric)",
            "https://www.flyingv.cc/projects/37328",
            "https://www.geappliances.com",
            "要調査", 3,
            "カーボンファイバー光波×±5℃精密温度制御 瞬間加熱スマートオーブン"
        ),
        (
            "オーディオ・録音",
            "PHILIPS 小飛筆 Pro｜最聰明的AI人生副手 語音轉寫×摘要×翻譯",
            "PHILIPS",
            "https://www.flyingv.cc/projects/36816",
            "https://www.philips.com.tw",
            "要調査", 4,
            "AIペン型録音筆 音声文字起こし・要約・翻訳 達成率10,124%"
        ),
    ],

    "Indiegogo": [
        (
            "スマートリング・ウォッチ",
            "LUNA Ring: 70 Health Metrics. 1 Smart Ring｜No Subscription",
            "Noise",
            "https://www.indiegogo.com/projects/luna-ring-70-health-metrics-1-smart-ring",
            "https://www.noise.com",
            "要調査", 4,
            "70種健康指標 月額不要 睡眠・活動・健康スコア"
        ),
        (
            "スマートリング・ウォッチ",
            "MOOV Ring: Revolutionary BP Monitoring Smart Ring｜AI-Powered",
            "MOOV",
            "https://www.indiegogo.com/projects/moov-ring-revolutionary-bp-monitoring-smart-ring",
            "要調査", "要調査", 4,
            "AIで血圧・心拍・睡眠・感情など14+指標を監視"
        ),
        (
            "PC周辺・入力機器",
            "E-inkey Dynamic Keyboard｜ePaper Keys Adapt to Any App",
            "E-inkey",
            "https://www.indiegogo.com/projects/e-inkey-dynamic-keyboard",
            "要調査", "要調査", 4,
            "電子ペーパーで表示変わる動的キーボード アプリに合わせ自動切替"
        ),
        (
            "E-Ink・ディスプレイ",
            "Paperlike 253: The First 25.3 inch E-ink Monitor｜3K HD",
            "Paperlike",
            "https://www.indiegogo.com/projects/paperlike-253-the-first-25-3-inch-e-ink-monitor",
            "https://paperlike.com",
            "要調査", 5,
            "世界初 25.3型3K E-Inkモニター 目に優しい大型電子ペーパー"
        ),
        (
            "映像・プロジェクター",
            "Lighty: World's First Robotic HD Laser Projector｜Auto-Follow",
            "Lighty",
            "https://www.indiegogo.com/projects/lighty-world-s-first-robotic-hd-laser-projector",
            "要調査", "要調査", 4,
            "世界初 自動追従ロボット型レーザープロジェクター"
        ),
        (
            "オーディオ・録音",
            "Infinix AI Buds: Hear the World Through AI｜162-Language Real-Time Translation",
            "Infinix",
            "https://www.indiegogo.com/projects/infinix-hear-the-world-through-ai",
            "https://www.infinixmobile.com",
            "要調査", 4,
            "162言語リアルタイムAI翻訳イヤホン アクティブNR搭載"
        ),
    ],
}

# ── スタイル定義 ──
normal_font   = Font(name="Arial", size=10)
center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

sheet_fills = {
    "ZECZEC":    PatternFill(fill_type="solid", fgColor="FFF2CC"),   # 薄黄
    "Indiegogo": PatternFill(fill_type="solid", fgColor="FCE4D6"),   # 薄橙
    "FlyingV":   PatternFill(fill_type="solid", fgColor="E2EFDA"),   # 薄緑
}

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total_added = 0

print("=== 第3弾 新規製品追加 ===")
for sheet_name, products in NEW_PRODUCTS.items():
    ws = wb[sheet_name]
    fill = sheet_fills[sheet_name]
    start_row = ws.max_row + 1

    print(f"\n[{sheet_name}] {len(products)}件追加 (行{start_row}〜)")
    for i, (cat, name, maker, cf_url, hp, mail, score, comment) in enumerate(products):
        r = start_row + i
        values = [cat, name, maker, cf_url, hp, mail, score, comment]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font   = normal_font
            c.fill   = fill
            c.border = border
            # col1(カテゴリ)=中央揃え、それ以外=左揃え
            c.alignment = center_align if col == 1 else left_align
        ws.row_dimensions[r].height = 30
        total_added += 1
        print(f"  + {name[:40]}")

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"\n✓ 保存完了: 合計 {total_added}件追加")

# ── 追加後の件数確認 ──
print("\n=== 各シート件数（追加後） ===")
wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
grand_total = 0
for sn in wb2.sheetnames:
    if sn == "削除済みリスト":
        continue
    ws2 = wb2[sn]
    # 空行を除いた実データ行数カウント
    cnt = sum(1 for r in range(2, ws2.max_row + 1) if ws2.cell(r, 2).value)
    print(f"  [{sn}] {cnt}件")
    grand_total += cnt
print(f"  合計: {grand_total}件")
