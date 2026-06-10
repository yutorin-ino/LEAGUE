from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = ["カテゴリ", "商品名", "メーカー名", "クラウドファンディングページURL", "メーカーHP", "コンタクトメール", "日本市場スコア", "日本市場コメント"]
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
ROW_COLORS = ["F5F5F5", "FFFFFF"]
SCORE_HIGH = PatternFill("solid", fgColor="C6EFCE")
SCORE_MID  = PatternFill("solid", fgColor="FFEB9C")
SCORE_LOW  = PatternFill("solid", fgColor="FFC7CE")
COL_WIDTHS = [20, 44, 24, 54, 38, 26, 14, 62]

def write_sheet(ws, products, title_color="1F4E79"):
    hfill = PatternFill("solid", fgColor=title_color)
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT; c.fill = hfill; c.alignment = HEADER_ALIGN
    for ri, prod in enumerate(products, 2):
        for ci, val in enumerate(prod, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=ROW_COLORS[ri % 2])
            if ci in [4, 5] and val and str(val).startswith("http"):
                cell.hyperlink = val; cell.font = LINK_FONT
            if ci == 7:
                s = val or ""
                cell.fill = SCORE_HIGH if s.startswith(("5","4")) else (SCORE_MID if s.startswith("3") else SCORE_LOW)
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 35
    for r in range(2, len(products) + 2):
        ws.row_dimensions[r].height = 70
    ws.freeze_panes = "A2"

# ========== ZECZEC データ ==========
zeczec_products = [
    ("ガジェット・充電", "Acer Gadget 碁速充 Qi2 五合一 35W 固態行動電源", "Acer Gadget", "https://www.zeczec.com/projects/acer-gadget-jisuchong", "https://www.acer.com/tw", "要調査", "3/5", "Qi2対応の多機能モバイルバッテリー。Acerブランドの信頼性はあるが日本でのQi2充電器市場は競争激しい"),
    ("ガジェット・充電", "Mag Slim Qi2 磁吸行動電源 世界最薄7.3mm", "Mag Slim", "https://www.zeczec.com/projects/mag-slim-qi2-2025-10-16", "要調査", "要調査", "4/5", "世界最薄7.3mmのQi2磁気モバイルバッテリー。iPhoneユーザーが多い日本市場に高い訴求"),
    ("ガジェット・充電", "Allite 氮化鎵快充 史上最小65W雙孔", "Allite", "https://www.zeczec.com/projects/allite", "https://www.allite.com.tw", "要調査", "4/5", "GaN採用の超小型65W充電器。デザイン性が高くAnkerとの差別化が図れる"),
    ("ガジェット・充電", "Allite WP1 Qi2 磁吸行動電源 追劇支架", "Allite", "https://www.zeczec.com/projects/wp1", "https://www.allite.com.tw", "要調査", "4/5", "Qi2公式認証取得・スタンド機能付き磁気モバイルバッテリー。iPhoneユーザーの多い日本に強い訴求"),
    ("ガジェット・充電", "Encore™ Wireless 無線座充 三機一体充電", "Encore", "https://www.zeczec.com/projects/encore-wireless", "要調査", "要調査", "3/5", "スマホ・イヤホン・スマートウォッチを同時ワイヤレス充電するデスクステーション"),
    ("ガジェット・充電", "Evo 超輕小旅行充電器 世界対応", "Evo", "https://www.zeczec.com/projects/evo", "要調査", "要調査", "3/5", "折り畳み式コンパクト旅行充電器。日本人旅行者に適しているが競合多数"),
    ("ガジェット・充電", "e-Fusion 行動電源×充電器×無線充電 三合一", "e-Fusion", "https://www.zeczec.com/projects/e-fusion", "要調査", "要調査", "3/5", "モバイルバッテリー+充電器+ワイヤレス充電の三合一デバイス。出張・旅行ユーザー向け"),
    ("ガジェット・充電", "LilNob 3C1A 四孔充電器 最高100W", "LilNob", "https://www.zeczec.com/projects/lilnob-3c1a", "要調査", "要調査", "3/5", "4ポートGaN 100W充電器。デザイン性と機能性のバランスで日本市場参入可能"),
    ("ガジェット・充電", "TITANSHIELD 固態行動電源 防爆五合一", "GrantClassic", "https://www.zeczec.com/projects/grantclassic", "要調査", "要調査", "3/5", "固体電池採用の安全設計モバイルバッテリー。爆発しない安全訴求は日本のシニア層に刺さる"),
    ("ガジェット・充電", "DEZCTOP 七合一 USB-C 集線器", "DEZCTOP", "https://www.zeczec.com/projects/dezctop-7in1-HUB", "要調査", "要調査", "3/5", "7-in-1 USB-CハブでMacBookユーザー多い日本のリモートワーカーに需要あり"),
    ("ガジェット・充電", "CASA Hub O7 集線×無線充電デスクハブ", "CASA", "https://www.zeczec.com/projects/casahubo7", "https://www.casa.com.tw", "要調査", "3/5", "ハブとワイヤレス充電を統合したデスクアクセサリー。MacBookユーザーが多い日本に需要あり"),
    ("ガジェット・充電", "POLYBATT CAT 22W磁吸式自帯線行動電源", "POLYBATT", "https://www.zeczec.com/projects/POLYBATT_CAT", "要調査", "要調査", "3/5", "MagSafe対応ケーブル内蔵22Wモバイルバッテリー。カメラ傷防止設計が独自性あり"),
    ("ガジェット・充電", "GENKI Nintendo Switch 専用無線ユニット", "Human Things (GENKI)", "https://www.zeczec.com/projects/genki", "https://genkithings.com", "support@genkithings.com", "4/5", "Nintendo Switch専用Bluetooth音声送信機。日本のSwitch愛好家に刺さる製品"),
    ("AR/VR・ウェアラブル", "Rokid AR Spatial 三視窗智慧眼鏡 300吋", "Rokid", "https://www.zeczec.com/projects/rokid-ar-spatial-300", "https://www.rokid.com", "要調査", "4/5", "300インチ仮想画面3ウィンドウARグラス。日本のAR眼鏡市場拡大中。エンタメ・ビジネス両用"),
    ("AR/VR・ウェアラブル", "Rokid Glasses 樂奇AI智慧眼鏡", "Rokid", "https://www.zeczec.com/projects/rokid-glasses", "https://www.rokid.com", "要調査", "4/5", "AI機能搭載スマートグラス。Meta Ray-Ban対抗製品として日本の先進ガジェット層に訴求"),
    ("AR/VR・ウェアラブル", "ASUS AirVision M1 100吋穿戴式顯示器", "ASUS", "https://www.zeczec.com/projects/airvision-m1", "https://www.asus.com/jp", "要調査", "3/5", "100インチウェアラブルディスプレイ87g超軽量。ASUSは日本展開済みだがこのモデルは台湾先行"),
    ("AR/VR・ウェアラブル", "SmartEcho AI智慧眼鏡 AR同步翻譯", "SmartEcho", "https://www.zeczec.com/projects/smartecho-glasses", "要調査", "要調査", "4/5", "リアルタイム翻訳ARグラス。インバウンド対応・海外旅行需要がある日本市場にドンピシャ"),
    ("AR/VR・ウェアラブル", "ROG XREAL R1 電競AR眼鏡 171吋 240Hz", "ASUS ROG × XREAL", "https://www.zeczec.com/projects/rog-xreal-r1-ar-glasses", "https://rog.asus.com", "要調査", "4/5", "世界初240Hz Micro OLED搭載ゲーミングARグラス。日本の熱烈なゲーマー層に高い訴求力"),
    ("AR/VR・ウェアラブル", "大朋E4 VR眼鏡 4K一体機 超軽量", "DPVR", "https://www.zeczec.com/projects/dpvr", "https://www.dpvrvr.com", "要調査", "3/5", "4K・120Hz VRヘッドセット。Meta Questの競合。日本市場参入には認知度向上が必要"),
    ("AR/VR・ウェアラブル", "RingConn Gen 2 AI 智慧戒指 睡眠健康", "RingConn", "https://www.zeczec.com/projects/cheertoktw-2025-06-13", "https://www.ringconn.com", "要調査", "4/5", "AI搭載スマートリング。Oura Ring対抗として日本の健康意識の高いユーザーに高い需要"),
    ("AR/VR・ウェアラブル", "LARMI 2025 AI 智能手錶 AI翻訳×GPS×5ATM", "LARMI", "https://www.zeczec.com/projects/Ai6", "要調査", "要調査", "3/5", "AIリアルタイム翻訳・GPS・5ATM防水スマートウォッチ。多機能×価格競争力で日本でも戦える"),
    ("PC周辺・入力機器", "Vision Keyboard 10吋液晶螢幕内蔵キーボード AI対応", "Vision Keyboard", "https://www.zeczec.com/projects/visionkeyboard", "要調査", "要調査", "4/5", "10インチタッチスクリーン内蔵キーボード。ChatGPT連携機能付きで在宅ワーカーに高需要"),
    ("PC周辺・入力機器", "TRACPOINT 実体+仮想 二合一マウス×プレゼンペン", "TRACPOINT", "https://www.zeczec.com/projects/tracpoint", "要調査", "要調査", "4/5", "物理+仮想一体のプレゼンテーションペン/マウス。日本のビジネスマン・教育者に強い訴求力"),
    ("PC周辺・入力機器", "MOZTECH カカ螢幕燈 折疊携帯+無線モニターライト", "MOZTECH", "https://www.zeczec.com/projects/kakalight", "https://www.moztech.com.tw", "要調査", "4/5", "折り畳み式ワイヤレスモニターライト。在宅ワーク×コンパクト設計で日本のデスク環境に最適"),
    ("PC周辺・入力機器", "GOOVIS LITE 世界最鮮明 Micro-OLED 3D HMD", "GOOVIS", "https://www.zeczec.com/projects/goovis-lite", "https://www.goovis.com", "要調査", "3/5", "Micro-OLED搭載高解像度HMD。4K映像愛好家・シネマ体験需要がある日本でニッチ層に刺さる"),
    ("PC周辺・入力機器", "AIFA i-Ctrl Pro Plus 家電遠端遙控×智慧防霉", "AIFA (艾電家)", "https://www.zeczec.com/projects/AIFASmart-iCtrl-Pro-Plus", "https://www.aifa.com.tw", "service@aifa.com.tw", "3/5", "全家電赤外線リモコン×AIカビ防止機能IoTデバイス。古い家電が多い日本でスマートホーム化に貢献"),
    ("PC周辺・入力機器", "JCH422 iPad/iPhone-Windows対傳視訊分享器", "JCH", "https://www.zeczec.com/projects/JCH422", "要調査", "要調査", "4/5", "iPad/iPhoneとWindows間クロスプラットフォーム映像共有デバイス。Apple×Windows混在環境の多い日本企業向け"),
    ("オーディオ・録音", "HiDock P1 AI録音器 会話整理AI秘書", "HiDock", "https://www.zeczec.com/projects/hidock_p1", "要調査", "要調査", "5/5", "AI文字起こし・要約機能搭載Bluetooth録音機。議事録自動化ニーズが高い日本ビジネス市場に超高需要"),
    ("オーディオ・録音", "Slimca 超薄録音カード 随録随傳", "Slimca", "https://www.zeczec.com/projects/slimca", "要調査", "要調査", "4/5", "超薄型ICレコーダーカード。名刺サイズの携帯性で日本のビジネスパーソンに刺さる"),
    ("オーディオ・録音", "TileRec 隠形録音片 全球最小智慧録音設備", "TileRec", "https://www.zeczec.com/projects/tilerec", "要調査", "要調査", "3/5", "世界最小級スマート録音デバイス。超小型・長時間録音で日本のビジネスユーザーに需要あり"),
    ("オーディオ・録音", "ZMI PurPods 真無線耳機 雙動圈設計 高音質", "ZMI (紫米)", "https://www.zeczec.com/projects/zmi_purpods_TW101", "https://www.zmi.com", "要調査", "3/5", "デュアルドライバー採用コスパ優秀TWS。AirPods対抗として日本の学生・若年層に訴求"),
    ("スマートロック・セキュリティ", "小島快鎖 TX 雙鏡頭監視器×掌靜脈×人臉辨識", "小島快鎖", "https://www.zeczec.com/projects/TXsmartlock", "要調査", "要調査", "5/5", "顔認証+掌静脈認証+デュアルカメラ搭載スマートロック。防犯意識の高い日本で最高水準の需要"),
    ("スマートロック・セキュリティ", "Acer Gadget 碁智鎖 五合一智能電子門鎖", "Acer Gadget", "https://www.zeczec.com/projects/acer", "https://www.acer.com/tw", "要調査", "4/5", "指紋・カード・パスワード等5種解錠対応スマートロック。Acerブランドで信頼性あり日本市場参入余地大"),
    ("スマートロック・セキュリティ", "FAMMIX SAFER-F4 防水人臉+ナノ波指紋 十合一電子鎖", "FAMMIX", "https://www.zeczec.com/projects/fammix-lock-safer-f4", "要調査", "要調査", "4/5", "完全防水・顔+ナノ波指紋認証の屋外対応スマートロック。日本の戸建て住宅向けに高需要"),
    ("スマートロック・セキュリティ", "Philips Alpha VP 6合1智慧視訊電子鎖 遠端通話", "Philips", "https://www.zeczec.com/projects/alpha-vp", "https://www.philips.com/ja-jp", "要調査", "4/5", "ビデオ通話対応スマートロック6合一。Philipsブランドで日本展開可能性高い"),
    ("スマートホーム・IoT・ロボット", "Eilik 桌面型陪伴ロボット 1000以上表情 感情表現", "Energy Robotics", "https://www.zeczec.com/projects/eilik-2025-11-14", "https://www.energyroboticstech.com", "要調査", "5/5", "1000以上の表情を持つ感情表現デスクロボット。日本のロボット・キャラクター好き文化に完全合致"),
    ("スマートホーム・IoT・ロボット", "Lytmi Neo LED 智慧氛圍燈帯 HDMI同期沈浸体験", "Lytmi", "https://www.zeczec.com/projects/lytmi-neo-led", "https://www.lytmi.com", "要調査", "4/5", "HDMI同期バックライトLEDストリップ。ゲーマー・シアタールーム需要の高い日本に刺さる"),
    ("スマートホーム・IoT・ロボット", "MOZTECH 変変燈 百変支架 LED無線センサーライト", "MOZTECH", "https://www.zeczec.com/projects/MOZTECH-BANG-LIGHT", "https://www.moztech.com.tw", "要調査", "4/5", "取付場所自由・センサー式ワイヤレスLEDライト。賃貸が多い日本の「工事不要」ニーズにぴったり"),
    ("スマートホーム・IoT・ロボット", "Yeelight LED 智慧彩光吸頂燈 3分鐘安装", "Yeelight", "https://www.zeczec.com/projects/yeelightled", "https://www.yeelight.com", "support@yeelight.com", "3/5", "3分で取付できる1600万色スマートシーリングライト。Mi Home/HomeKit対応で日本でも展開可能"),
    ("空気清浄・除湿機", "Smini 零耗材空氣清淨機 永久使用 離子技術", "Smini", "https://www.zeczec.com/projects/smini", "要調査", "要調査", "4/5", "フィルター消耗品不要のイオン技術空気清浄機。「ランニングコスト0」訴求で日本のコスト意識層に響く"),
    ("空気清浄・除湿機", "FLAT 壁掛畫空淨機 外銷突破20國 アートパネル型", "FLAT", "https://www.zeczec.com/projects/flat", "要調査", "要調査", "4/5", "アートパネル型壁掛け空気清浄機。インテリアを損なわない設計で日本のスタイリッシュ志向層に高需要"),
    ("空気清浄・除湿機", "DIKE 淨化者EVO HEPA14 五年換濾網不要 医療グレード", "DIKE", "https://www.zeczec.com/projects/HCF630", "要調査", "要調査", "4/5", "医療グレードHEPA14フィルター×5年交換不要。ランニングコストの低さで日本のシニア層にも訴求"),
    ("ペットテック", "LG 喵太座 キャットタワー×空清×健康監測 All in One", "LG", "https://www.zeczec.com/projects/lg-aerocat-tower", "https://www.lg.com/jp", "jp.lgservice@lge.com", "5/5", "キャットタワー+空気清浄+健康監視オールインワン。日本の猫ブームと健康意識の高さで超高需要"),
    ("ペットテック", "PETUS 寵物智能項圈 5G GPS 無距離制限", "PETUS", "https://www.zeczec.com/projects/petus", "要調査", "要調査", "4/5", "5G IoT対応GPSペットスマートカラー。迷子防止・位置確認ニーズが高い日本のペット飼育者に需要"),
    ("ペットテック", "ebot 24時間寵物管家 AI見守りシステム", "ebot", "https://www.zeczec.com/projects/ebotaiwan", "要調査", "要調査", "4/5", "AI搭載24時間ペット監視管理システム。共働き世帯が多い日本でペット留守番管理に高需要"),
    ("健康・フィットネス・睡眠", "HiDock P1 AI録音器 会話整理AI秘書", "HiDock", "https://www.zeczec.com/projects/hidock_p1", "要調査", "要調査", "5/5", "AI文字起こし・要約機能搭載録音機"),
    ("健康・フィットネス・睡眠", "RheoFit 全自動AI按摩滾輪 運動リカバリー", "RheoFit", "https://www.zeczec.com/projects/rheofit", "要調査", "要調査", "4/5", "AI制御全自動マッサージローラー。運動後のリカバリー製品として日本のスポーツ愛好家に高需要"),
    ("健康・フィットネス・睡眠", "OYO NOVA 攜帶型全方位居家健身神器 NASA仕様", "OYO Fitness", "https://www.zeczec.com/projects/oyo-nova", "要調査", "要調査", "4/5", "NASA採用技術の持ち運びコンパクトフィットネス器具。在宅ワーク×健康ブームの日本市場に刺さる"),
    ("健康・フィットネス・睡眠", "歐歐睏助眠燈 台湾工科大学院チーム開発助眠ライト", "OuOu Sleep", "https://www.zeczec.com/projects/osleepy", "要調査", "要調査", "4/5", "科学的設計の助眠ライト。日本の睡眠問題意識の高さで高需要期待。7年間の研究開発品"),
    ("電動モビリティ・昇降", "E-ach UP 桌上型電動升降スタンドデスク", "AKA", "https://www.zeczec.com/projects/aka-eachup", "要調査", "要調査", "4/5", "デスクトップ電動昇降スタンドデスク。日本の在宅ワーク×健康意識で電動昇降デスク需要急増"),
]

# ========== Indiegogo データ ==========
indiegogo_products = [
    ("AR・スマートグラス", "RayNeo X2 World's 1st AI-Powered True AR Glasses", "RayNeo", "https://www.indiegogo.com/projects/rayneo-x2-world-s-1st-ai-powered-true-ar-glasses", "https://www.rayneo.com", "要調査", "4/5", "フルカラーMicroLED搭載AIAARグラス。日本のAR市場拡大中でガジェット先進層に高訴求"),
    ("AR・スマートグラス", "Looktech AI: The Smart Glasses That Truly Know You", "Looktech AI", "https://www.indiegogo.com/projects/looktech-ai-the-smart-glasses-that-truly-know-you--2", "要調査", "要調査", "4/5", "GPT-4o搭載・13MPカメラ・2K動画のパーソナライズドAIグラス。日本の先進ユーザー層に刺さる"),
    ("AR・スマートグラス", "Rokid Air AR Glasses with Voice Control AI", "Rokid", "https://www.indiegogo.com/projects/rokid-air-ar-glasses-with-voice-control-ai", "https://www.rokid.com", "要調査", "4/5", "音声AI対応ARグラス。全デバイス対応で日本のビジネス・エンタメ両用に期待"),
    ("スマートリング・ウォッチ", "Y-RING: Your Smart Future", "Y-RING", "https://www.indiegogo.com/projects/y-ring-your-smart-future", "要調査", "要調査", "4/5", "NFC・チタン製・10ATM防水・24h健康トラッキングスマートリング。日本の健康意識層に高訴求"),
    ("スマートリング・ウォッチ", "Ring One: The Most Advanced Smart Ring", "Ring One", "https://www.indiegogo.com/projects/ring-one-the-most-advanced-smart-ring-for-you", "要調査", "要調査", "4/5", "AI搭載・NFC決済対応の高機能スマートリング。Oura Ring対抗として日本市場参入余地大"),
    ("スマートリング・ウォッチ", "AI-Powered Next Gen Health and Sport Smart Ring", "TRENDLOX", "https://www.indiegogo.com/projects/ai-powered-next-gen-health-and-sport-smart-ring", "要調査", "要調査", "3/5", "AI搭載スポーツ・健康管理スマートリング。サブスクなしで24h健康トラッキング"),
    ("スマートリング・ウォッチ", "SEQUENT: The World's First Self-Charging Smartwatch", "Sequent AG", "https://www.indiegogo.com/projects/sequent-the-world-s-first-self-charging-smartwatch", "https://www.sequentag.com", "要調査", "3/5", "動作エネルギーで自己充電するスイス製スマートウォッチ。電池交換不要という独自訴求"),
    ("スマートリング・ウォッチ", "Oru Watch - World's First Dual Display Smart Watch", "Oru Watch", "https://www.indiegogo.com/projects/oru-watch-world-s-first-dual-display-smart-watch", "要調査", "要調査", "3/5", "E-paper+TFT デュアルディスプレイスマートウォッチ。スマホ充電機能付きという独自性あり"),
    ("スマートリング・ウォッチ", "Wearbuds Watch: Smartwatch that Houses Its Earbuds", "Aipower", "https://www.indiegogo.com/projects/wearbuds-watch-smartwatch-that-houses-its-earbuds", "要調査", "要調査", "3/5", "スマートウォッチにイヤホンを収納する一体型デバイス。独自コンセプトで日本の新しもの好きに訴求"),
    ("スマートリング・ウォッチ", "VIITA Smartwatch - AI Wearable Fitness Coach", "VIITA Watches", "https://www.indiegogo.com/projects/viita-smartwatch-ai-wearable-fitness-coach", "https://www.viitawatches.com", "要調査", "3/5", "4週間バッテリーのAIフィットネスコーチスマートウォッチ。ストレス・脱水症状管理機能付き"),
    ("スマートリング・ウォッチ", "Aether X Powered by TikTik Designer Smartwatch", "TikTik", "https://www.indiegogo.com/projects/aether-x-powered-by-tiktik-designer-smartwatch", "要調査", "要調査", "3/5", "デザイン性重視のスマートウォッチ。ファッション感覚で日本の若年層に訴求可能"),
    ("充電・モバイル", "oCharger: The Next-Generation Smart Charger", "oCharger", "https://www.indiegogo.com/projects/ocharger-the-next-generation-smart-charger", "要調査", "要調査", "3/5", "AppleWatch・Android・Qi全対応オールインワンスマート充電器。バッテリー+ケーブル+無線充電一体"),
    ("充電・モバイル", "20W Multi-Device MagSafe Wireless Power Bank Stand", "Tmobi", "https://www.indiegogo.com/projects/20w-multi-device-magsafe-wireless-power-bank-stand", "要調査", "要調査", "3/5", "20W PD×10W Qi×スタンド機能付きMagSafe対応10000mAhモバイルバッテリー"),
    ("充電・モバイル", "MagInk: AI Customizable Smart E-Ink Power Bank", "Tmobi", "https://www.indiegogo.com/projects/magink-ai-customizable-smart-e-ink-power-bank", "要調査", "要調査", "4/5", "4色E-inkスクリーン搭載AIカスタマイズ可能モバイルバッテリー。デザイン×機能で日本市場に差別化"),
    ("キーボード・入力機器", "Glyph AI Smart Keyboard for Infinite Task Flow", "Majestic Devices", "https://www.indiegogo.com/projects/glyph-ai-smart-keyboard-for-infinite-task-flow", "要調査", "要調査", "4/5", "OLEDタッチスクリーン×ホットスワップ対応AIスマートキーボード。日本のメカキーボード愛好家に刺さる"),
    ("キーボード・入力機器", "The Minimal Phone: First E-Ink QWERTY Phone", "Minimal Company", "https://www.indiegogo.com/projects/the-minimal-phone-first-e-ink-qwerty-phone", "要調査", "要調査", "4/5", "E-inkディスプレイ+物理QWERTYキーボードスマートフォン。デジタルデトックス需要の高い日本に刺さる"),
    ("E-Ink・ディスプレイ", "NarTick AI E-Ink Calendar Ultimate AI Scheduler", "NarTick", "https://www.indiegogo.com/projects/nartick-e-ink-calendar-ultimate-ai-scheduler--2", "要調査", "要調査", "4/5", "世界初AI音声対応E-inkカレンダー。2週間バッテリー×スケジュール自動管理で日本のビジネス層に訴求"),
    ("E-Ink・ディスプレイ", "DASUNG Link: The World First E-Ink Phone Monitor", "DASUNG", "https://www.indiegogo.com/projects/dasung-link-the-world-first-e-ink-phone-monitor", "https://www.dasung.com", "要調査", "4/5", "6.7型E-inkスマートフォンモニター。目の疲れ対策ニーズが高い日本で需要大"),
    ("E-Ink・ディスプレイ", "Blotch: The World's Best E-Ink Smart Frame", "Blotch", "https://www.indiegogo.com/projects/blotch-the-world-s-best-e-ink-smart-frame", "要調査", "要調査", "3/5", "13.3型・100以上ウィジェット対応E-inkスマートフレーム。ミニマルインテリア好きな日本市場向け"),
    ("ロボット・AI機器", "LOOI: Turn Your Smartphone into a Desktop Robot!", "Tangible Future", "https://www.indiegogo.com/projects/looi-turn-your-smartphone-into-a-desktop-robot", "要調査", "要調査", "4/5", "スマートフォンをデスクロボットに変えるガジェット。日本のロボット・キャラクター文化に合致"),
    ("スポーツ・フィットネス", "The Smartest Cycling Computer in the World", "Absolute Cycling", "https://www.indiegogo.com/projects/the-smartest-cycling-computer-in-the-world--2", "要調査", "要調査", "3/5", "高精度ナビ搭載の最スマートサイクリングコンピューター。日本のサイクリングブームに乗れる可能性"),
    ("スポーツ・フィットネス", "Siimple Projector", "Siimple", "https://www.indiegogo.com/projects/siimple-projector", "要調査", "要調査", "3/5", "シンプルで高品質なポータブルプロジェクター。ホームシアター需要の高い日本でコンパクト性が強み"),
    ("スポーツ・フィットネス", "A.I Watch Smartwatch: The New Intelligence Life", "AI Watch", "https://www.indiegogo.com/projects/a-i-watch-smartwatch-the-new-intelligence-life-is-here", "要調査", "要調査", "3/5", "AIアシスタント搭載マルチ機能スマートウォッチ。スマートフォン・フィットネス・ガイド機能統合"),
]

# ========== FlyingV データ ==========
flyingv_products = [
    ("充電・モバイル", "QV Safe 類固態行動電源", "Quantum Volt", "https://www.flyingv.cc/projects/36739", "要調査", "要調査", "3/5", "固体電池採用の安全設計モバイルバッテリー。爆発リスクなしの安全訴求は日本市場に適合"),
    ("充電・モバイル", "Prelude Z MagSafe 固態全能充", "BEZALEL (倍加能)", "https://www.flyingv.cc/projects/37184", "https://www.bezalel.com", "要調査", "4/5", "MagSafe対応固体電池全能充電器。BEZALEL は世界展開ブランドで日本市場参入余地あり"),
    ("充電・モバイル", "Prelude S MagSafe 磁吸無線行動電源 世界最薄", "BEZALEL (倍加能)", "https://www.flyingv.cc/projects/35539", "https://www.bezalel.com", "要調査", "4/5", "世界最薄7.5mmQi2対応磁気モバイルバッテリー。iPhoneユーザーが多い日本に高い訴求"),
    ("充電・モバイル", "THINKTHING 雙向磁吸充電組 全球首創充裝置×充行動電源", "THINKTHING", "https://www.flyingv.cc/projects/35661", "要調査", "要調査", "3/5", "Apple Watch充電スタンド付き双方向磁気充電セット。iPhoneユーザーの多い日本向けに刺さる"),
    ("AR・ディスプレイ", "RayNeo Air4 Pro 口袋電視 ARグラス", "RayNeo Taiwan", "https://www.flyingv.cc/projects/36980", "https://www.rayneo.com", "要調査", "4/5", "ポケットサイズの没入型ARグラス。RayNeoはIndiegogoでも展開中。日本のAR市場参入に好機"),
    ("AR・ディスプレイ", "MAD Gaze GLOW 智慧MR眼鏡 身歷其境", "MAD Gaze", "https://www.flyingv.cc/projects/25217", "https://www.madgaze.com", "要調査", "3/5", "MR対応スマートグラス。MAD GazeはHKスタートアップで日本展開すれば先進ユーザー層に訴求"),
    ("AR・ディスプレイ", "Dream Glass Flow 遊戲AR眼鏡 ゲーム体験拡張", "Dream Glass", "https://www.flyingv.cc/projects/34445", "要調査", "要調査", "3/5", "ゲーム向けARグラス。日本のゲーマー向けに展開可能だがコンシューマー認知度構築が課題"),
    ("スマートロック", "Acer Gadget 碁智鎖 Pro 五合一智能電子門鎖", "旺德電通 / Acer Gadget", "https://www.flyingv.cc/projects/36332", "https://www.acer.com/tw", "要調査", "4/5", "ZECZECにも掲載されるAcer Gadgetスマートロック上位モデル。日本の防犯ニーズに合致"),
    ("家電・清淨機", "POIEMA 新氣几 零耗材空氣清淨機 水洗フィルター", "POIEMA", "https://www.flyingv.cc/projects/37111", "https://www.poiema.com", "要調査", "4/5", "水洗いでフィルター交換不要の空気清浄機。ランニングコスト0の訴求で日本のコスト意識層に響く"),
    ("家電・清淨機", "MyComfyTech Ultra AIR 隨身空氣清淨機 携帯型", "MyComfyTech", "https://www.flyingv.cc/projects/37298", "要調査", "要調査", "3/5", "携帯型空気清浄機。通勤・移動中の花粉・PM2.5対策需要がある日本市場に合致"),
    ("家電・清淨機", "SoleusAir 吸毛大師 旗艦除毛抗敏寵物清淨機", "SOLEUS AIR", "https://www.flyingv.cc/projects/36469", "要調査", "要調査", "4/5", "ペットの毛吸引×抗アレルギー機能付きペット専用空気清浄機。日本のペット飼育者に高需要"),
    ("家電・清淨機", "智米 空氣清淨機 400m³/h CADR VOC監測", "智米 (Zhimi)", "https://www.flyingv.cc/projects/27506", "https://www.smartmi.com", "要調査", "3/5", "CADR400・VOCセンサー搭載空気清浄機。Xiaomiエコシステムとの連携で日本のMiユーザーに訴求"),
    ("ロボット・掃除機", "ECOVACS X8 PRO OMNI 全自動水洗掃地機器人", "ECOVACS", "https://www.flyingv.cc/projects/36044", "https://www.ecovacs.com/jp", "jp@ecovacs.com", "4/5", "全自動水洗床掃除ロボット。ECOVACSは日本でも展開中。この特定モデルの日本先行確認要"),
    ("PC周辺・キーボード", "Mercury K1 Pro 異星機械鍵盤 2024最帥キーボード", "Gravastar / Mercury", "https://www.flyingv.cc/projects/35827", "https://www.gravastar.com", "要調査", "4/5", "SF的デザインのメカニカルキーボード。個性重視の日本のゲーマー・クリエイター層に刺さる"),
    ("PC周辺・キーボード", "DEMON ART 金屬經典鍵盤 有線×Bluetooth", "DEMON ART", "https://www.flyingv.cc/projects/2394", "要調査", "要調査", "3/5", "金属製クラシックメカニカルキーボード。有線・Bluetooth両対応で日本のキーボードマニアに訴求"),
    ("オーディオ", "AERO 真無線藍牙耳機 零感延遅×臨場環繞", "AERO", "https://www.flyingv.cc/projects/26447", "要調査", "要調査", "3/5", "50ms超低遅延・XROUNDサラウンド音響TWS。ゲーミング用途で日本のeスポーツ層に刺さる"),
    ("オーディオ", "quandio ID 本我 真無線耳機 バランス三周波", "quandio", "https://www.flyingv.cc/projects/24729", "要調査", "要調査", "3/5", "バランス型三周波CVC8.0ノイズキャンセリングTWS。コスパ重視の日本の若年層に訴求"),
    ("オーディオ", "Libratone Track Air+ 真無線耳機 アクティブNC", "Libratone", "https://www.flyingv.cc/projects/26454", "https://www.libratone.com", "要調査", "3/5", "デンマーク発の高音質ANC対応TWS。Libratoneは欧州ブランドで日本展開余地あり"),
    ("家具・チェア", "COFO 浮雲工學椅 2.0 日本人体工学", "COFO Taiwan", "https://www.flyingv.cc/projects/37345", "https://www.cofo.com", "要調査", "4/5", "日本人体工学に基づいたエルゴノミックチェア。在宅ワーク需要が続く日本でプレミアム椅子市場に参入"),
    ("家具・チェア", "COFO 良木升降桌 x 浮雲工學椅 日本人体工学", "COFO Taiwan", "https://www.flyingv.cc/projects/34509", "https://www.cofo.com", "要調査", "4/5", "電動昇降デスク×エルゴノミックチェアセット。日本の在宅ワーク環境改善ニーズに直接応える"),
    ("E-Ink・デジタルフレーム", "BLOOMIN8 E-Ink 無線電子畫框 首創無線設計", "Bloomin8", "https://www.flyingv.cc/projects/37264", "要調査", "要調査", "4/5", "ワイヤレス給電E-inkデジタルフォトフレーム。インテリアとしても優秀で日本のミニマル志向層に刺さる"),
    ("クラフト・ミシン", "NECCHI NE210 智慧刺繡縫紉兩用機 イタリア製", "NECCHI / 立家手創館", "https://www.flyingv.cc/projects/37342", "https://www.necchi.com", "要調査", "3/5", "イタリア発刺繍×ミシン両用スマートミシン。日本のハンドメイドブーム×DIY市場に一定需要"),
    ("クラフト・ミシン", "Cricut™ 多功能智慧裁切機 クリカット", "Cricut / igogosport", "https://www.flyingv.cc/projects/37369", "https://www.cricut.com", "要調査", "4/5", "クラフト用多機能スマートカッティングマシン。日本のハンドメイド・DIYコミュニティで高需要"),
    ("ペットテック", "eterno 智慧恆溫寵物墊 温度管理スマートペットマット", "eterno", "https://www.flyingv.cc/projects/15052", "要調査", "要調査", "3/5", "温度自動管理スマートペットマット。日本の高齢ペット・小型犬猫向けに需要あり"),
]

wb = Workbook()

# シート1: ZECZEC (青)
ws1 = wb.active
ws1.title = "ZECZEC"
write_sheet(ws1, zeczec_products, "1F4E79")

# シート2: Indiegogo (オレンジ)
ws2 = wb.create_sheet("Indiegogo")
write_sheet(ws2, indiegogo_products, "C55A11")

# シート3: FlyingV (緑)
ws3 = wb.create_sheet("FlyingV")
write_sheet(ws3, flyingv_products, "375623")

# タブの色設定
ws1.sheet_properties.tabColor = "1F4E79"
ws2.sheet_properties.tabColor = "C55A11"
ws3.sheet_properties.tabColor = "375623"

out = r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx"
wb.save(out)
print(f"保存完了: {out}")
print(f"ZECZEC: {len(zeczec_products)}件 / Indiegogo: {len(indiegogo_products)}件 / FlyingV: {len(flyingv_products)}件")
