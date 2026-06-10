from openpyxl import load_workbook
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 日本CFサイト（Makuake / Campfire / Greenfunding）で
# 先行販売済みが確認された製品リスト
# ============================================================
# 形式: {シート名: {行番号: (製品名, 確認サイト), ...}}
JAPAN_CF_ROWS = {
    "ZECZEC": {
        14: ("GENKI Nintendo Switch 専用無線ユニット",    "Campfire"),
        15: ("Rokid AR Spatial 三視窗智慧眼鏡",         "Makuake"),
        16: ("Rokid Glasses 樂奇AI智慧眼鏡",            "Makuake"),
        17: ("ASUS AirVision M1",                       "Greenfunding"),
        21: ("RingConn Gen 2 AI 智慧戒指",               "Makuake"),
        26: ("GOOVIS LITE Micro-OLED 3D HMD",           "Makuake / Campfire"),
        29: ("HiDock P1 AI録音器（オーディオ）",          "Makuake"),
        37: ("Eilik 桌面型陪伴ロボット",                  "Campfire"),
        41: ("Smini 零耗材空氣清淨機",                   "Makuake / Campfire"),
        44: ("LG 喵太座 キャットタワー×空清",              "Greenfunding"),
        46: ("ebot 24時間寵物管家 AI見守りシステム",       "Makuake / Greenfunding"),
        47: ("HiDock P1 AI録音器（フィットネス）",         "Makuake"),
        49: ("OYO NOVA 攜帶型全方位居家健身神器",          "Makuake"),
    },
    "Indiegogo": {
         4: ("Rokid Air AR Glasses",                    "Makuake"),
        10: ("Wearbuds Watch Smartwatch+Earbuds",        "Makuake / Campfire / Greenfunding"),
        18: ("NarTick AI E-Ink Calendar",                "Greenfunding"),
        21: ("LOOI Desktop Robot",                       "Makuake"),
    },
    "FlyingV": {
         3: ("BEZALEL Prelude Z MagSafe 固態全能充",     "Makuake / Greenfunding"),
         4: ("BEZALEL Prelude S 磁吸無線行動電源",        "Makuake / Greenfunding"),
         7: ("MAD Gaze GLOW 智慧MR眼鏡",                "Campfire"),
         8: ("Dream Glass Flow 遊戲AR眼鏡",              "Campfire"),
        17: ("AERO 真無線藍牙耳機",                      "Makuake / Greenfunding"),
        19: ("Libratone Track Air+ 真無線耳機",           "Makuake / Greenfunding"),
        20: ("COFO 浮雲工學椅 2.0",                      "Makuake"),
        21: ("COFO 良木升降桌 × 浮雲工學椅",              "Makuake"),
        24: ("Cricut 多功能智慧裁切機",                   "Campfire"),
    },
}

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")

# ─── プレビュー表示 ───
total_delete = 0
for sn in wb.sheetnames:
    rows = JAPAN_CF_ROWS.get(sn, {})
    if rows:
        print(f"\n=== [{sn}] 削除予定 {len(rows)}件 ===")
        for r in sorted(rows):
            name, site = rows[r]
            print(f"  行{r:2d}: {name}  ← {site}")
        total_delete += len(rows)

print(f"\n合計削除予定: {total_delete}件")
print("─" * 50)

# ─── 実際の削除（行番号が大きい順に削除してズレを防ぐ） ───
deleted = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = JAPAN_CF_ROWS.get(sn, {})
    for r in sorted(rows.keys(), reverse=True):
        ws.delete_rows(r)
        deleted += 1

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"\n削除完了: {deleted}件")

# ─── 削除後の件数確認 ───
wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total = 0
for sn in wb2.sheetnames:
    ws = wb2[sn]
    cnt = ws.max_row - 1
    print(f"  [{sn}] 残り {cnt}件")
    total += cnt
print(f"  合計残り: {total}件")
