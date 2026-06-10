from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 削除した26件の記録
DELETED = [
    # (元シート, カテゴリ, 商品名, メーカー名, 確認サイト)
    ("ZECZEC",    "ガジェット・充電",          "GENKI Nintendo Switch 専用無線ユニット",          "Human Things (GENKI)",    "Campfire"),
    ("ZECZEC",    "AR/VR・ウェアラブル",        "Rokid AR Spatial 三視窗智慧眼鏡 300吋",          "Rokid",                   "Makuake"),
    ("ZECZEC",    "AR/VR・ウェアラブル",        "Rokid Glasses 樂奇AI智慧眼鏡",                  "Rokid",                   "Makuake"),
    ("ZECZEC",    "AR/VR・ウェアラブル",        "ASUS AirVision M1 100吋穿戴式顯示器",            "ASUS",                    "Greenfunding"),
    ("ZECZEC",    "AR/VR・ウェアラブル",        "RingConn Gen 2 AI 智慧戒指",                    "RingConn",                "Makuake（歴代1位）"),
    ("ZECZEC",    "PC周辺・入力機器",           "GOOVIS LITE Micro-OLED 3D HMD",                "GOOVIS",                  "Makuake / Campfire"),
    ("ZECZEC",    "オーディオ・録音",            "HiDock P1 AI録音器 会話整理AI秘書",              "HiDock",                  "Makuake"),
    ("ZECZEC",    "スマートホーム・IoT・ロボット","Eilik 桌面型陪伴ロボット 1000以上表情",           "Energy Robotics",         "Campfire"),
    ("ZECZEC",    "空気清浄・除湿機",            "Smini 零耗材空氣清淨機 永久使用 離子技術",         "Smini",                   "Makuake / Campfire"),
    ("ZECZEC",    "ペットテック",               "LG 喵太座 キャットタワー×空清×健康監測",           "LG",                      "Greenfunding（達成率4378%）"),
    ("ZECZEC",    "ペットテック",               "ebot 24時間寵物管家 AI見守りシステム",             "ebot (Enabot)",           "Makuake / Greenfunding"),
    ("ZECZEC",    "健康・フィットネス・睡眠",      "HiDock P1 AI録音器（重複エントリ）",              "HiDock",                  "Makuake"),
    ("ZECZEC",    "健康・フィットネス・睡眠",      "OYO NOVA 攜帶型全方位居家健身神器",               "OYO Fitness",             "Makuake"),
    ("Indiegogo", "AR・スマートグラス",          "Rokid Air AR Glasses with Voice Control AI",   "Rokid",                   "Makuake"),
    ("Indiegogo", "スマートリング・ウォッチ",      "Wearbuds Watch: Smartwatch that Houses Earbuds","Aipower",                 "Makuake / Campfire / Greenfunding"),
    ("Indiegogo", "E-Ink・ディスプレイ",         "NarTick AI E-Ink Calendar Ultimate AI Scheduler","NarTick",                "Greenfunding"),
    ("Indiegogo", "ロボット・AI機器",             "LOOI: Turn Your Smartphone into a Desktop Robot","Tangible Future",         "Makuake"),
    ("FlyingV",   "充電・モバイル",              "Prelude Z MagSafe 固態全能充",                  "BEZALEL (倍加能)",         "Makuake / Greenfunding"),
    ("FlyingV",   "充電・モバイル",              "Prelude S MagSafe 磁吸無線行動電源 世界最薄",     "BEZALEL (倍加能)",         "Makuake / Greenfunding"),
    ("FlyingV",   "AR・ディスプレイ",            "MAD Gaze GLOW 智慧MR眼鏡 身歷其境",             "MAD Gaze",                "Campfire"),
    ("FlyingV",   "AR・ディスプレイ",            "Dream Glass Flow 遊戲AR眼鏡 ゲーム体験拡張",     "Dream Glass",             "Campfire"),
    ("FlyingV",   "オーディオ",                 "AERO 真無線藍牙耳機 零感延遅×臨場環繞",           "XROUND / AERO",           "Makuake / Greenfunding"),
    ("FlyingV",   "オーディオ",                 "Libratone Track Air+ 真無線耳機 アクティブNC",   "Libratone",               "Makuake / Greenfunding"),
    ("FlyingV",   "家具・チェア",               "COFO 浮雲工學椅 2.0 日本人体工学",               "COFO Taiwan",             "Makuake"),
    ("FlyingV",   "家具・チェア",               "COFO 良木升降桌 x 浮雲工學椅 日本人体工学",       "COFO Taiwan",             "Makuake"),
    ("FlyingV",   "クラフト・ミシン",             "Cricut™ 多功能智慧裁切機 クリカット",             "Cricut / igogosport",     "Campfire"),
]

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")

# 既存の削除済みシートがあれば削除
if "削除済みリスト" in wb.sheetnames:
    del wb["削除済みリスト"]

ws = wb.create_sheet("削除済みリスト")

# ── スタイル定義 ──
header_fill   = PatternFill(fill_type="solid", fgColor="C00000")   # 濃い赤
header_font   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
title_fill    = PatternFill(fill_type="solid", fgColor="FF0000")    # 赤
title_font    = Font(name="Arial", size=13, bold=True, color="FFFFFF")
normal_font   = Font(name="Arial", size=10)
center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── タイトル行 ──
ws.merge_cells("A1:E1")
ws["A1"] = "日本クラウドファンディングサイト掲載確認済み製品リスト（削除済み）"
ws["A1"].font  = title_font
ws["A1"].fill  = title_fill
ws["A1"].alignment = center_align
ws.row_dimensions[1].height = 30

# ── ヘッダー行 ──
headers = ["元シート", "カテゴリ", "商品名", "メーカー名", "確認サイト（Makuake / Campfire / Greenfunding）"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font      = header_font
    c.fill      = header_fill
    c.alignment = center_align
    c.border    = border
ws.row_dimensions[2].height = 22

# ── データ行 ──
# シートごとに行の色を変える
sheet_fills = {
    "ZECZEC":    PatternFill(fill_type="solid", fgColor="FFF2CC"),  # 薄黄
    "Indiegogo": PatternFill(fill_type="solid", fgColor="FCE4D6"),  # 薄橙
    "FlyingV":   PatternFill(fill_type="solid", fgColor="E2EFDA"),  # 薄緑
}

for i, (sheet, cat, prod, maker, site) in enumerate(DELETED, start=3):
    fill = sheet_fills.get(sheet, PatternFill())
    values = [sheet, cat, prod, maker, site]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.font      = normal_font
        c.fill      = fill
        c.border    = border
        c.alignment = left_align if col >= 3 else center_align
    ws.row_dimensions[i].height = 30

# ── 列幅 ──
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 42
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 32

# ── 合計行 ──
last_row = 3 + len(DELETED)
ws.merge_cells(f"A{last_row}:D{last_row}")
ws[f"A{last_row}"] = f"合計  {len(DELETED)}件削除（残り71件 / 元97件）"
ws[f"A{last_row}"].font = Font(name="Arial", size=10, bold=True)
ws[f"A{last_row}"].fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
ws[f"A{last_row}"].alignment = center_align
ws[f"A{last_row}"].border = border
ws.row_dimensions[last_row].height = 20

# シートをリストの先頭に移動
wb.move_sheet("削除済みリスト", offset=-wb.sheetnames.index("削除済みリスト"))

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print("削除済みリスト シートを作成しました")
print(f"  削除件数: {len(DELETED)}件")
print(f"  残り件数: 71件 (元97件)")
