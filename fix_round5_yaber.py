from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 第5弾 修正スクリプト
# Yaber T1 Pro 小影霸 → Greenfunding掲載済みが判明 → 削除
# 代替: Wanbo TOGO PRO 勁量行動投影機 を追加
# ============================================================

YABER_URL = "https://www.zeczec.com/projects/yaber-smart-projector"

REPLACEMENT = (
    "映像・プロジェクター",
    "Wanbo TOGO PRO 勁量行動投影機｜99Wh大電量 × 500lm × 360°雲台 × GoogleTV",
    "Wanbo (萬播)",
    "https://www.zeczec.com/projects/togo-pro-battery",
    "要調査", "要調査", 4,
    "99Wh大容量内蔵バッテリー 500lm 360°首振り Google TV搭載 屋内外対応"
)

# スタイル
normal_font  = Font(name="Arial", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
fill   = PatternFill(fill_type="solid", fgColor="FFF2CC")   # ZECZEC=薄黄

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
ws = wb["ZECZEC"]

# ── 1. Yaber行を検索・削除 ──
yaber_row = None
for r in range(2, ws.max_row + 1):
    if ws.cell(r, 4).value == YABER_URL:
        yaber_row = r
        break

if yaber_row:
    removed_name = ws.cell(yaber_row, 2).value
    ws.delete_rows(yaber_row)
    print(f"✓ 削除: 行{yaber_row} → {removed_name[:50]}")
    print("  理由: Greenfunding (Kibidango) projects/9143 に掲載済み")
else:
    print("⚠ Yaber行が見つかりませんでした（既に削除済み？）")

# ── 2. 代替製品を末尾に追加 ──
cat, name, maker, cf_url, hp, mail, score, comment = REPLACEMENT
r_new = ws.max_row + 1
values = [cat, name, maker, cf_url, hp, mail, score, comment]
for col, val in enumerate(values, 1):
    c = ws.cell(row=r_new, column=col, value=val)
    c.font      = normal_font
    c.fill      = fill
    c.border    = border
    c.alignment = center_align if col == 1 else left_align
ws.row_dimensions[r_new].height = 30
print(f"✓ 代替追加: {name[:55]}")

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print("\n✓ 保存完了")

# ── 件数確認 ──
print("\n=== 各シート件数（修正後） ===")
wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
grand_total = 0
for sn in wb2.sheetnames:
    if sn == "削除済みリスト":
        continue
    ws2 = wb2[sn]
    cnt = sum(1 for r in range(2, ws2.max_row + 1) if ws2.cell(r, 2).value)
    print(f"  [{sn}] {cnt}件")
    grand_total += cnt
print(f"  合計: {grand_total}件")
