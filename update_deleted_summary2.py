from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 第2弾 削除確認済み11件
NEW_DELETED = [
    # (元シート, カテゴリ, 商品名, メーカー名, 確認サイト)
    ("ZECZEC",   "ガジェット・充電",      "Allite 氮化鎵快充 史上最小65W雙孔",                "Allite",           "Greenfunding"),
    ("ZECZEC",   "ガジェット・充電",      "LilNob 3C1A 四孔充電器 最高100W",                  "LilNob / CIO",     "Makuake / Greenfunding"),
    ("ZECZEC",   "ガジェット・充電",      "TITANSHIELD 固態行動電源 防爆五合一",               "GrantClassic",     "Makuake"),
    ("ZECZEC",   "PC周辺・入力機器",      "TRACPOINT 実体+仮想 二合一マウス×プレゼンペン",     "TRACPOINT",        "Campfire"),
    ("ZECZEC",   "PC周辺・入力機器",      "AIFA i-Ctrl Pro Plus 家電遠端遙控×智慧防霉",        "AIFA (艾電家)",    "Makuake"),
    ("ZECZEC",   "オーディオ・録音",      "Slimca 超薄録音カード 随録随傳",                    "Slimca",           "Makuake"),
    ("ZECZEC",   "オーディオ・録音",      "TileRec 隠形録音片 全球最小智慧録音設備",            "TileRec",          "Makuake / Campfire / Greenfunding"),
    ("ZECZEC",   "空気清浄・除湿機",      "FLAT 壁掛畫空淨機 外銷突破20國 アートパネル型",      "FLAT",             "Makuake (Sauberairとして)"),
    ("Indiegogo","AR・スマートグラス",    "RayNeo X2 World's 1st AI-Powered True AR Glasses", "RayNeo",           "Greenfunding"),
    ("Indiegogo","AR・スマートグラス",    "Looktech AI: The Smart Glasses That Truly Know You","Looktech AI",      "Makuake"),
    ("FlyingV",  "家電・清淨機",          "智米 空氣清淨機 400m³/h CADR VOC監測",              "智米 (Zhimi)",     "Greenfunding"),
]

# ── スタイル定義 ──
header_fill   = PatternFill(fill_type="solid", fgColor="C00000")
header_font   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
normal_font   = Font(name="Arial", size=10)
bold_font     = Font(name="Arial", size=10, bold=True)
center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

sheet_fills = {
    "ZECZEC":    PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "Indiegogo": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "FlyingV":   PatternFill(fill_type="solid", fgColor="E2EFDA"),
}

# 第2弾を示す薄めの色（少し濃くして区別）
new_fills = {
    "ZECZEC":    PatternFill(fill_type="solid", fgColor="FFE699"),
    "Indiegogo": PatternFill(fill_type="solid", fgColor="F4B183"),
    "FlyingV":   PatternFill(fill_type="solid", fgColor="C6E0B4"),
}

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
ws = wb["削除済みリスト"]

# ─── 既存の合計行（最終行）を削除 ───
last_row = ws.max_row
# 合計行を特定（D列が空、A列に「合計」が含まれる行）
while last_row > 2 and ws.cell(last_row, 2).value is None:
    ws.delete_rows(last_row)
    last_row = ws.max_row

# 合計行（最終の非空行が合計行のはず）を削除
if "合計" in str(ws.cell(last_row, 1).value or ""):
    ws.delete_rows(last_row)
    last_row = ws.max_row

# ─── 区切り行を追加（第2弾であることを示す） ───
separator_row = last_row + 1
ws.merge_cells(f"A{separator_row}:E{separator_row}")
ws[f"A{separator_row}"] = "▼ 第2弾 追加削除確認分（2026年5月調査）"
ws[f"A{separator_row}"].font = Font(name="Arial", size=10, bold=True, color="595959")
ws[f"A{separator_row}"].fill = PatternFill(fill_type="solid", fgColor="EDEDED")
ws[f"A{separator_row}"].alignment = center_align
ws[f"A{separator_row}"].border = border
ws.row_dimensions[separator_row].height = 18

# ─── 新しいデータ行を追加 ───
start_row = separator_row + 1
for i, (sheet, cat, prod, maker, site) in enumerate(NEW_DELETED):
    r = start_row + i
    fill = new_fills.get(sheet, PatternFill())
    values = [sheet, cat, prod, maker, site]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font      = normal_font
        c.fill      = fill
        c.border    = border
        c.alignment = left_align if col >= 3 else center_align
    ws.row_dimensions[r].height = 30

# ─── 新しい合計行 ───
TOTAL_DELETED = 26 + len(NEW_DELETED)   # 26 (第1弾) + 11 (第2弾) = 37
REMAINING     = 97 - TOTAL_DELETED       # 97 - 37 = 60
new_last_row  = start_row + len(NEW_DELETED)

ws.merge_cells(f"A{new_last_row}:D{new_last_row}")
ws[f"A{new_last_row}"] = (
    f"合計  {TOTAL_DELETED}件削除"
    f"（第1弾26件 ＋ 第2弾{len(NEW_DELETED)}件）"
    f"　残り{REMAINING}件 / 元97件"
)
ws[f"A{new_last_row}"].font = Font(name="Arial", size=10, bold=True)
ws[f"A{new_last_row}"].fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
ws[f"A{new_last_row}"].alignment = center_align
ws[f"A{new_last_row}"].border = border
ws.row_dimensions[new_last_row].height = 20

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"削除済みリスト を更新しました")
print(f"  追加件数: {len(NEW_DELETED)}件")
print(f"  総削除件数: {TOTAL_DELETED}件（第1弾26 + 第2弾{len(NEW_DELETED)}）")
print(f"  残り件数: {REMAINING}件 / 元97件")
