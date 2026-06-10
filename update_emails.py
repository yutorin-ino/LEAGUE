from openpyxl import load_workbook
from openpyxl.styles import Font

# HP別サイト調査で確認したコンタクトメールアドレス辞書
# {キーワード: メールアドレス}
EMAIL_DATA = {
    # 充電・モバイル
    "CASA":              "casa@casa.com.tw",
    "GrantClassic":      "support@grantclassic.com",
    "TITANSHIELD":       "support@grantclassic.com",
    "LilNob":            "info@connectinternationalone.co.jp",
    "CIO":               "info@connectinternationalone.co.jp",
    "POLYBATT":          "sales03@poweraegis.com",
    "ZMI":               "support@zmi.com",
    "紫米":               "support@zmi.com",
    "MOZTECH":           "contact@moztech.cc",
    "DEZCTOP":           "support@dezctop.com",
    "TileRec":           "support@atto-digital.com",
    "Slimca":            "support@atto-digital.com",
    # AR/VR・ウェアラブル
    "Rokid":             "Glasses.Global@rokid.com",
    "RayNeo":            "service@rayneo.com",
    "DPVR":              "support@dpvr.com",
    "GOOVIS":            "goovishw@nedoptics.com.cn",
    "RingConn":          "cs@ringconn.com",
    "VIITA":             "hello@viita-watches.com",
    "Sequent":           "care@sequent.ch",
    "Aipower":           "support@myaipower.com",
    "Wearbuds":          "support@myaipower.com",
    "TikTik":            "writetous@tiktikglobal.com",
    "Aether X":          "writetous@tiktikglobal.com",
    "Muse Wearables":    "hello@musewearables.com",
    "Ring One":          "hello@musewearables.com",
    # スマホ・PC
    "Minimal Company":   "press@tryminimal.com",
    "Minimal Phone":     "press@tryminimal.com",
    "Majestic Devices":  "support@majesticdevices.com",
    "Glyph":             "support@majesticdevices.com",
    "Absolute Cycling":  "service@absolutecycling.com",
    # IoT・ロボット
    "LOOI":              "service@looirobot.com",
    "Tangible Future":   "service@looirobot.com",
    "Blotch":            "contact@blotch.app",
    "Eilik":             "info@enabot.com",
    "ebot":              "info@enabot.com",
    "Enabot":            "info@enabot.com",
    # セキュリティ
    "FAMMIX":            "service@vichys.com.tw",
    "小島快鎖":           "3elocks@gmail.com",
    "3E":                "3elocks@gmail.com",
    # オーディオ
    "XROUND":            "sales01@xroundaudio.com",
    "AERO":              "sales01@xroundaudio.com",
    "Libratone":         "info@libratone.com",
    "Gravastar":         "service@gravastar.com",
    "Mercury K1":        "service@gravastar.com",
    # 空気清浄
    "MyComfyTech":       "hello@mycomfytech.com",
    "Sauberair":         "biz@sauberair.com",
    "FLAT":              "biz@sauberair.com",
    "Smini":             "support@aethersi.com",
    # スマートホーム・充電
    "THINKTHING":        "info@thinkthingstudio.com",
    "BEZALEL":           "support@bezalel.co",
    "倍加能":             "support@bezalel.co",
    # フィットネス
    "RheoFit":           "service@rheofit.com",
    # ウォレット・IoT
    "COFO":              "support@cofoglobal.com",
    # E-ink
    "Tmobi":             "support@tmobi.info",
    "MagInk":            "support@tmobi.info",
    "DASUNG":            "contact@dasung.com",
    "Lytmi":             "support@lytmi.co",
}

normal_font = Font(name="Arial", size=10)

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
updated = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        hp_cell   = ws.cell(row=row, column=5)
        mail_cell = ws.cell(row=row, column=6)

        # HPが設定済みでメールが未入力の行のみ対象
        if not str(hp_cell.value or "").startswith("http"):
            continue
        if str(mail_cell.value or "") not in ("要調査", "", "None"):
            continue

        product = str(ws.cell(row=row, column=2).value or "")
        maker   = str(ws.cell(row=row, column=3).value or "")
        combined = product + " " + maker

        for kw, mail in EMAIL_DATA.items():
            if kw.lower() in combined.lower():
                mail_cell.value = mail
                mail_cell.font = normal_font
                updated += 1
                break

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"メール更新件数: {updated}件")

# 結果サマリー
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total_hp = total_mail = 0
for sn in wb2.sheetnames:
    ws = wb2[sn]
    s_hp   = sum(1 for r in range(2, ws.max_row + 1)
                 if str(ws.cell(row=r, column=5).value or "").startswith("http"))
    s_mail = sum(1 for r in range(2, ws.max_row + 1)
                 if str(ws.cell(row=r, column=6).value or "") not in ("要調査", "", "None")
                 and str(ws.cell(row=r, column=5).value or "").startswith("http"))
    print(f"  [{sn}] HP入力済: {s_hp}件 / メール入力済: {s_mail}件 / メール未入力: {s_hp-s_mail}件")
    total_hp += s_hp; total_mail += s_mail
print(f"  合計: HP入力済{total_hp}件 / メール入力済{total_mail}件({round(total_mail/total_hp*100)}%) / 未入力{total_hp-total_mail}件")
