from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = load_workbook(r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸705件.xlsx")

link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
link_font_cat_row = Font(name="Arial", size=10, color="0563C1", underline="single")

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

cat_fills = {
    "キッチン・調理器具": PatternFill("solid", start_color="2E75B6"),
    "スマートホーム・インテリア・照明": PatternFill("solid", start_color="538135"),
    "ウェアラブル・ヘルス・フィットネス": PatternFill("solid", start_color="7030A0"),
    "アウトドア・スポーツ・旅行": PatternFill("solid", start_color="C55A11"),
    "ペット用品": PatternFill("solid", start_color="843C0C"),
    "テクノロジー・ガジェット": PatternFill("solid", start_color="244185"),
    "美容・スキンケア": PatternFill("solid", start_color="C00000"),
    "子供・教育": PatternFill("solid", start_color="375623"),
    "ファッション・アクセサリー": PatternFill("solid", start_color="7B3F00"),
    "クリーニング・収納・整理": PatternFill("solid", start_color="404040"),
}
alt_fill = PatternFill("solid", start_color="EBF3FB")

# URL列: F列(6)=商品URL, G列(7)=メーカーHP
# メール列: H列(8)=連絡先メール（「（問い合わせページ）」を含む場合はURLとして処理）

for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=2):
        row_num = row[0].row
        bg = alt_fill if row_num % 2 == 0 else None

        for cell in row:
            col = cell.column
            val = cell.value
            if not val:
                continue

            val_str = str(val).strip()

            # 商品URL列 (col 6) とメーカーHP列 (col 7)
            if col in (6, 7):
                if val_str.startswith("http"):
                    cell.hyperlink = val_str
                    cell.value = val_str
                    cell.font = link_font
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = border
                    if bg:
                        cell.fill = bg

            # 連絡先メール列 (col 8)
            elif col == 8:
                if "（問い合わせページ）" in val_str:
                    # URLを抽出してハイパーリンクに
                    url_part = val_str.replace("（問い合わせページ）", "").strip()
                    if url_part.startswith("http"):
                        cell.hyperlink = url_part
                        cell.value = "問い合わせページ"
                        cell.font = link_font
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
                        cell.border = border
                        if bg:
                            cell.fill = bg
                elif "@" in val_str and not val_str.startswith("http"):
                    # メールアドレス
                    cell.hyperlink = f"mailto:{val_str}"
                    cell.value = val_str
                    cell.font = link_font
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = border
                    if bg:
                        cell.fill = bg

output_path = r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸705件.xlsx"
wb.save(output_path)
print("完了: ハイパーリンク設定済み")
print(f"保存先: {output_path}")
