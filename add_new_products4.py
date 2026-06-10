from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 第4弾 新規追加製品リスト（12件）
# ZECZEC 8件 / FlyingV 1件 / Indiegogo 3件
# 日本CF（Makuake / Campfire / Greenfunding）未掲載確認済み
# ============================================================
#
# 列構成: カテゴリ(1) | 商品名(2) | メーカー名(3) | CFページURL(4)
#         メーカーHP(5) | コンタクトメール(6) | スコア(7) | コメント(8)

NEW_PRODUCTS = {
    "ZECZEC": [
        (
            "ガジェット・充電",
            "FluxGo Pro 固態行動電源｜容量大2倍 20,000mAh×65W防爆快充",
            "FluxGo",
            "https://www.zeczec.com/projects/fluxgo-pro",
            "要調査", "要調査", 4,
            "一般比2倍の20,000mAh 防爆固態電池 65W急速充電"
        ),
        (
            "ガジェット・充電",
            "SOLA Pro 萬能充｜10,000mAh 固態電芯 × 自帶雙線 × 旅遊通勤",
            "SOLA",
            "https://www.zeczec.com/projects/sola-pro",
            "要調査", "要調査", 4,
            "10,000mAh 固態電池 Lightning+USB-C内蔵ケーブル一体型"
        ),
        (
            "オーディオ・録音",
            "OpenRock Link20｜AI「翻譯×錄音×降噪」磁吸麥克風 開放式無線耳機",
            "OpenRock / 1More",
            "https://www.zeczec.com/projects/link20-ai",
            "https://openrock.audio",
            "要調査", 4,
            "AI翻訳×録音×ノイキャン搭載 Bluetooth 6.0 マグネット式マイク付き"
        ),
        (
            "PC周辺・入力機器",
            "ZERO｜地表最小 附著式翻譯機 40+語言 88口音 超薄貼附型",
            "ZERO",
            "https://www.zeczec.com/projects/zero-trans",
            "要調査", "要調査", 4,
            "世界最小クラス 40言語88アクセント対応 貼り付け型翻訳機"
        ),
        (
            "スマートホーム・IoT・ロボット",
            "INXNI 雙拖布螺旋掃拖機器人 X3・自動集塵掃拖機器人 X1",
            "INXNI",
            "https://www.zeczec.com/projects/INXNI",
            "要調査", "要調査", 3,
            "デュアル回転モップ 自動集塵 螺旋クリーニングシステム"
        ),
        (
            "ガジェット・モビリティ",
            "Bikniks® 超電力野餐行旅車｜移動電源 × 城市電輔 × 行動野餐",
            "Bikniks",
            "https://www.zeczec.com/projects/bikniks",
            "要調査", "要調査", 4,
            "世界初 電動アシスト×ピクニックバスケット×モバイルバッテリー一体型自転車"
        ),
        (
            "ペットテック",
            "Acerpure PRO 毛管家｜AI恆控寵物空淨機 360°吸毛 99%抗菌",
            "Acerpure (Acer)",
            "https://www.zeczec.com/projects/acerpure-p5-airpurifier",
            "https://www.acerpure.com",
            "要調査", 4,
            "AI制御 ペット特化空気清浄機 360°全方位吸毛 2026台湾優良デザイン賞"
        ),
        (
            "ガジェット・カメラ",
            "AXV 360 行車記錄器｜1抵5鏡頭 × 360°再擴充 × 一鍵輸出免編輯",
            "AXV",
            "https://www.zeczec.com/projects/axv-360",
            "要調査", "要調査", 3,
            "1台5カメラ相当 360°ドライブレコーダー ワンクリック出力"
        ),
    ],

    "FlyingV": [
        (
            "家電・キッチン",
            "GE 廚餘博士｜櫥下廚餘機 急速粉碎 把惱人廚餘通通沖走",
            "GE (General Electric)",
            "https://www.flyingv.cc/projects/36949",
            "https://www.geappliances.com",
            "要調査", 3,
            "シンク下設置型 キッチン生ごみ処理機 GE製 高速粉砕"
        ),
    ],

    "Indiegogo": [
        (
            "PC周辺・入力機器",
            "XNote: AI-Powered Smart Writing Set｜Handwriting + Digital Sync",
            "Kodion Innovation",
            "https://www.indiegogo.com/projects/xnote-ai-powered-smart-writing-set",
            "要調査", "要調査", 4,
            "スマートペン+ノート AIで手書きをデジタル同期 AIチャット・検索連携"
        ),
        (
            "PC周辺・入力機器",
            "NEWYES Scan Reader Pen 3 Pro｜AI Translator + OCR + Recorder",
            "NEWYES",
            "https://www.indiegogo.com/projects/newyes-scan-reader-pen-3-pro-ai-translator",
            "https://www.newyes.com",
            "要調査", 4,
            "スキャン翻訳×録音×OCR一体型AIペン 多言語対応"
        ),
        (
            "AR・スマートグラス",
            "StarV Air - MYVU AR Smart Glasses｜Real-Time Translation 13 Languages",
            "StarV",
            "https://www.indiegogo.com/projects/starv-air-myvu-ar-smart-glasses--2",
            "要調査", "要調査", 4,
            "13言語リアルタイム翻訳 ARスマートグラス テキスト文字起こし対応"
        ),
    ],
}

# ── スタイル定義 ──
normal_font  = Font(name="Arial", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

sheet_fills = {
    "ZECZEC":    PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "Indiegogo": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "FlyingV":   PatternFill(fill_type="solid", fgColor="E2EFDA"),
}

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total_added = 0

print("=== 第4弾 新規製品追加 ===")
for sheet_name, products in NEW_PRODUCTS.items():
    ws = wb[sheet_name]
    fill = sheet_fills[sheet_name]
    start_row = ws.max_row + 1

    print(f"\n[{sheet_name}] {len(products)}件追加 (行{start_row}〜)")
    for i, (cat, name, maker, cf_url, hp, mail, score, comment) in enumerate(products):
        r = start_row + i
        values = [cat, name, maker, cf_url, hp, mail, score, comment]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = normal_font
            c.fill      = fill
            c.border    = border
            c.alignment = center_align if col == 1 else left_align
        ws.row_dimensions[r].height = 30
        total_added += 1
        print(f"  + {name[:45]}")

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"\n✓ 保存完了: 合計 {total_added}件追加")

# ── 追加後の件数確認 ──
print("\n=== 各シート件数（追加後） ===")
wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
grand_total = 0
for sn in wb2.sheetnames:
    if sn == "削除済みリスト":
        continue
    ws2 = wb2[sn]
    cnt = sum(1 for r in range(2, ws2.max_row + 1) if ws2.cell(r, 2).value)
    print(f"  [{sn}] {cnt}件")
    grand_total += cnt
print(f"  合計: {grand_total}件")
