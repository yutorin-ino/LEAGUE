from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 第9弾 新規追加製品リスト（8件）
# ZECZEC 3件 / FlyingV 2件 / Indiegogo 3件
# 日本CF（Makuake / Campfire / Greenfunding）未掲載確認済み
# ============================================================
# 精査除外済み:
#   HiDock P1                → Makuake掲載済み（5600万円超達成）
#   CATLINK X3 ProVision     → 同弾CATLINK Open-Xと重複回避
#   Tairona R-02             → 2023年製品・ZECZEC R-01既に掲載済み
#   COFO Chair Pro 2         → Makuake掲載済み（2億円超）
#   Tineco S9 Artist Pro     → Makuake掲載済み
#   Olight Ostation 2        → Greenfunding掲載済み
#   Wanbo Mozart 1 Pro       → Amazon Japan一般販売済み
#   INFINITE AIR ポータブル風力発電 → Makuake+Greenfunding+CAMPFIRE掲載済み
#   FREEZE-X 智冷充           → 類似製品がCAMPFIRE掲載済み
#   PLAUD NotePin            → Amazon Japan/日本公式展開済み
#
# 列構成: カテゴリ(1) | 商品名(2) | メーカー名(3) | CFページURL(4)
#         メーカーHP(5) | コンタクトメール(6) | スコア(7) | コメント(8)

NEW_PRODUCTS = {
    "ZECZEC": [
        (
            "スマートグラス・AR",
            "【ROG XREAL R1】電競AR眼鏡｜世界初240Hz Micro-OLED 171吋沉浸式 0.01ms 91g",
            "ASUS ROG × XREAL",
            "https://www.zeczec.com/projects/rog-xreal-r1-ar-glasses",
            "https://rog.asus.com", "要調査", 5,
            "世界初240Hz Micro-OLED ARゲーミングメガネ 171インチ仮想画面 0.01ms応答 Boseオーディオ 電気変色レンズ 91g"
        ),
        (
            "ガジェット・充電",
            "OmniCell+ 萬變磁型行動電源｜全球首款「充電座×飛航旅架×車充」究極三合一 Qi2",
            "ZERO 零式創作",
            "https://www.zeczec.com/projects/omnicell",
            "要調査", "要調査", 4,
            "世界初 充電スタンド+旅行スタンド+車載充電器 三合一 10000mAh Qi2 15W 4台同時 UL/PSE認証 2M NTD達成"
        ),
        (
            "PC周辺・入力機器",
            "DFM80 輕量化滑鼠｜CP值霸主 52g極輕盈 無線藍牙三模 100時間以上電量持続",
            "DFM",
            "https://www.zeczec.com/projects/dfm80-52g",
            "要調査", "要調査", 3,
            "52g超軽量ワイヤレスマウス 三モード(BT/2.4G/有線) 2H充電→100H以上 BSMI/NCC認証 韓国フロステッドデザイン"
        ),
    ],

    "FlyingV": [
        (
            "スマートホーム・IoT・ロボット",
            "CATLINK Open-X 翻滾喵｜全天候開放式智慧貓砂盆 360°環景 AI辨識 20貓管理",
            "CATLINK",
            "https://www.flyingv.cc/projects/36921",
            "https://www.catlink.com", "要調査", 3,
            "全天候開放型スマート猫用トイレ 360°パノラマビュー 12L大容量 AI行動認識 最大20匹管理 15kgまで対応 オゾン脱臭"
        ),
        (
            "スマートホーム・IoT・ロボット",
            "GE 智慧瞬熱烤箱｜廚房最強「鷗翼瞬熱」260℃碳纖光波×±5℃數位精準溫控",
            "GE Appliances",
            "https://www.flyingv.cc/projects/37328",
            "https://www.geappliances.com", "要調査", 3,
            "GEブランド鷗翼開放ドア設計 炭素繊維光波ヒーター260℃ ±5℃精密デジタル温度制御 家庭用オーブントースター"
        ),
    ],

    "Indiegogo": [
        (
            "E-Ink・ディスプレイ",
            "Paperlike 13K: World's First 13.3\" Color E-Ink Monitor｜300PPI × 37Hz × Touch",
            "DASUNG",
            "https://www.indiegogo.com/en/projects/dasung/paperlike-13k-world-first-13-3-color-eink-monitor",
            "https://shop.dasung.com", "要調査", 4,
            "世界初13.3型カラーE-Inkモニター 3200×2400/300PPI 37Hz高速リフレッシュ タッチ対応 Type-C+HDMI 目に優しい"
        ),
        (
            "スマートリング・ウォッチ",
            "Fulcrum X1: The Smart Companion for Every Watch｜Transforms ANY Watch into Smart",
            "Fulcrum Wearables",
            "https://www.indiegogo.com/projects/fulcrum-x1-the-smart-companion-for-every-watch",
            "https://fulcrumwearables.com", "要調査", 3,
            "あらゆる腕時計をスマートウォッチ化 マイクロサクション吸着 血中酸素+心拍+睡眠 25時間バッテリー 28mm径"
        ),
        (
            "PC周辺・入力機器",
            "DASUNG Paperlike 103: World's First 60Hz E-ink Monitor｜10.3\" Fast Refresh",
            "DASUNG",
            "https://www.indiegogo.com/en/projects/dasung/paperlike-103-world-s-first-60hz-e-ink-monitor",
            "https://shop.dasung.com", "要調査", 4,
            "世界初60Hz高速リフレッシュE-Inkモニター 10.3型 LCD並みの滑らかさ 目疲れゼロ 超軽量ポータブル"
        ),
    ],
}

# ── スタイル定義 ──
normal_font  = Font(name="Arial", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

sheet_fills = {
    "ZECZEC":    PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "Indiegogo": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "FlyingV":   PatternFill(fill_type="solid", fgColor="E2EFDA"),
}

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total_added = 0

print("=== 第9弾 新規製品追加 ===")
for sheet_name, products in NEW_PRODUCTS.items():
    ws = wb[sheet_name]
    fill = sheet_fills[sheet_name]
    start_row = ws.max_row + 1

    print(f"\n[{sheet_name}] {len(products)}件追加 (行{start_row}〜)")
    for i, (cat, name, maker, cf_url, hp, mail, score, comment) in enumerate(products):
        r = start_row + i
        values = [cat, name, maker, cf_url, hp, mail, score, comment]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = normal_font
            c.fill      = fill
            c.border    = border
            c.alignment = center_align if col == 1 else left_align
        ws.row_dimensions[r].height = 30
        total_added += 1
        print(f"  + {name[:50]}")

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"\n✓ 保存完了: 合計 {total_added}件追加")

# ── 追加後の件数確認 ──
print("\n=== 各シート件数（追加後） ===")
wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
grand_total = 0
for sn in wb2.sheetnames:
    if sn == "削除済みリスト":
        continue
    ws2 = wb2[sn]
    cnt = sum(1 for r in range(2, ws2.max_row + 1) if ws2.cell(r, 2).value)
    print(f"  [{sn}] {cnt}件")
    grand_total += cnt
print(f"  合計: {grand_total}件")
