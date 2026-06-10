from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 第6弾 新規追加製品リスト（13件）
# ZECZEC 6件 / FlyingV 2件 / Indiegogo 5件
# 日本CF（Makuake / Campfire / Greenfunding）未掲載確認済み
# ============================================================
# 精査除外済み:
#   BodyPedia     → Greenfunding掲載済み
#   Cubo AI       → CAMPFIRE掲載済み
#   KoKonna 墨想框 → Makuake掲載済み
#   Rokid Glasses → Makuake掲載済み
#   AINOTE 2      → Makuake掲載済み (iFLYTEK製)
#   ECOVACS W2S   → Makuake掲載済み (WINBOT W2 OMNI)
#
# 列構成: カテゴリ(1) | 商品名(2) | メーカー名(3) | CFページURL(4)
#         メーカーHP(5) | コンタクトメール(6) | スコア(7) | コメント(8)

NEW_PRODUCTS = {
    "ZECZEC": [
        (
            "スマートホーム・IoT・ロボット",
            "PhotoCube AI｜世界首款無線備份 一走進家門手機立刻自動開始備份",
            "PhotoCube",
            "https://www.zeczec.com/projects/photocubeai",
            "要調査", "要調査", 4,
            "世界初ワイヤレス自動写真バックアップ 帰宅即起動 ケーブル/月額不要"
        ),
        (
            "スマートホーム・IoT・ロボット",
            "Ionix Flow｜主動淨界 離子艙級空氣革命 攜帶式空氣清淨器",
            "Ionix",
            "https://www.zeczec.com/projects/ionixflow",
            "要調査", "要調査", 4,
            "航空機グレードイオン浄化技術 フィルター不要 オゾンゼロ ポケットサイズ空気清浄器"
        ),
        (
            "オーディオ・録音",
            "EDIFIER Lolli｜耳夾式耳機 健康檢測×主動降噪 開放式無感配戴 解放耳道",
            "EDIFIER",
            "https://www.zeczec.com/projects/lolli",
            "https://www.edifier.com", "要調査", 4,
            "耳掛け型オープンイヤー 心拍・血中酸素監視+ANC 39時間持続 IP56防水"
        ),
        (
            "スマートホーム・IoT・ロボット",
            "大同寶寶智慧音箱｜全球首創人偶造型 AI 智慧音箱 全向高音質 Google 語音助理",
            "大同 (Tatung)",
            "https://www.zeczec.com/projects/t-boy",
            "https://www.tatung.com", "要調査", 3,
            "世界初 人形型AIスマートスピーカー 360°全方位高音質 Google音声AI 台湾文化アイコン"
        ),
        (
            "ヘルスケア・フィットネス",
            "MotionAI 動智運動紀錄器｜用聽的運動紀錄器 不需藍牙Wi-Fi 全家都能用",
            "MotionAI",
            "https://www.zeczec.com/projects/motionai",
            "要調査", "要調査", 4,
            "音声認識方式運動記録 BT/WiFi不要 クリップ装着 家族全員で使える運動トラッカー"
        ),
        (
            "PC周辺・入力機器",
            "GetNew 熱點分享穩流機｜即插即連 × 輕巧便攜 × 超低功耗 不再斷線 熱銷80萬",
            "GetNew",
            "https://www.zeczec.com/projects/getnew",
            "要調査", "要調査", 3,
            "モバイルホットスポット安定化デバイス 即接続 超省電力 断線ゼロ 80万TWD超ヒット"
        ),
    ],

    "FlyingV": [
        (
            "E-Ink・ディスプレイ",
            "BLOOMIN8 E-Ink 無線電子畫框｜首創無線設計 Spectra 6色彩 真實紙感藝術畫框",
            "BLOOMIN8",
            "https://www.flyingv.cc/projects/37264",
            "要調査", "要調査", 4,
            "世界初コードレスE-Inkアートフレーム Spectra 6色 電源コード不要 紙のような質感"
        ),
        (
            "ガジェット・充電",
            "Prelude Z｜MagSafe 固態全能充 45W快充 × 自帶插頭 × 防爆 4台同時充電",
            "Prelude Z",
            "https://www.flyingv.cc/projects/37184",
            "要調査", "要調査", 4,
            "MagSafe対応 固態防爆電池 45W 内蔵プラグ ノート/スマホ/AirPods/Apple Watch同時充電"
        ),
    ],

    "Indiegogo": [
        (
            "PC周辺・入力機器",
            "Glyph: AI Smart Keyboard for Infinite Task Flow｜OLED Keys + Hot-Swap + Hub",
            "Majestic Devices",
            "https://www.indiegogo.com/projects/glyph-ai-smart-keyboard-for-infinite-task-flow",
            "要調査", "要調査", 4,
            "OLEDキー表示 ホットスワップ対応 AIタスクフロー USBハブ一体型スマートキーボード"
        ),
        (
            "映像・プロジェクター",
            "Wemax Go: World's Thinnest ALPD Smart Laser Projector｜1\" Thin × 4K HDR × Built-in Battery",
            "Wemax",
            "https://www.indiegogo.com/en/projects/wemaxofficial/wemax-go-thinnest-alpd-smart-laser-projector",
            "https://www.wemax.com", "要調査", 4,
            "世界最薄1インチ ALPDスマートレーザープロジェクター 4K HDR 内蔵バッテリー 120インチ対応"
        ),
        (
            "スマートリング・ウォッチ",
            "VIITA Smartwatch: AI Wearable Fitness Coach｜4-Week Battery × Stress × Hydration AI",
            "VIITA",
            "https://www.indiegogo.com/projects/viita-smartwatch-ai-wearable-fitness-coach",
            "要調査", "要調査", 4,
            "4週間バッテリー AIパーソナルフィットネスコーチ ストレス/水分補給管理 24時間健康追跡"
        ),
        (
            "ヘルスケア・フィットネス",
            "Doctoh: The Multidimensional Wearable Health System｜AI Multi-Sensor Monitoring",
            "Doctoh",
            "https://www.indiegogo.com/projects/doctoh-the-multidimensional-wearable-health-system",
            "要調査", "要調査", 3,
            "多次元センサー搭載ウェアラブル AI総合健康監視システム 運動・睡眠・バイタル管理"
        ),
        (
            "スマートリング・ウォッチ",
            "AI-Powered Next Gen Smart Ring by TRENDLOX｜24/7 Health + Sleep + Sport No Subscription",
            "TRENDLOX Ltd.",
            "https://www.indiegogo.com/projects/ai-powered-next-gen-health-and-sport-smart-ring",
            "要調査", "要調査", 3,
            "月額不要 24時間健康追跡 睡眠/女性周期/スポーツ対応 AIスマートリング"
        ),
    ],
}

# ── スタイル定義 ──
normal_font  = Font(name="Arial", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

sheet_fills = {
    "ZECZEC":    PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "Indiegogo": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "FlyingV":   PatternFill(fill_type="solid", fgColor="E2EFDA"),
}

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total_added = 0

print("=== 第6弾 新規製品追加 ===")
for sheet_name, products in NEW_PRODUCTS.items():
    ws = wb[sheet_name]
    fill = sheet_fills[sheet_name]
    start_row = ws.max_row + 1

    print(f"\n[{sheet_name}] {len(products)}件追加 (行{start_row}〜)")
    for i, (cat, name, maker, cf_url, hp, mail, score, comment) in enumerate(products):
        r = start_row + i
        values = [cat, name, maker, cf_url, hp, mail, score, comment]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = normal_font
            c.fill      = fill
            c.border    = border
            c.alignment = center_align if col == 1 else left_align
        ws.row_dimensions[r].height = 30
        total_added += 1
        print(f"  + {name[:50]}")

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"\n✓ 保存完了: 合計 {total_added}件追加")

# ── 追加後の件数確認 ──
print("\n=== 各シート件数（追加後） ===")
wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
grand_total = 0
for sn in wb2.sheetnames:
    if sn == "削除済みリスト":
        continue
    ws2 = wb2[sn]
    cnt = sum(1 for r in range(2, ws2.max_row + 1) if ws2.cell(r, 2).value)
    print(f"  [{sn}] {cnt}件")
    grand_total += cnt
print(f"  合計: {grand_total}件")
