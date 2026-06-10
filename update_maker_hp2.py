from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

# 商品名 or メーカー名に含まれるキーワード → (HP, メール)
MAKER_DATA = {
    # 充電・モバイル
    "Allite":          ("https://www.allite.com.tw", ""),
    "CASA Hub":        ("https://www.casa.com.tw", ""),
    "GENKI":           ("https://genkithings.com", "support@genkithings.com"),
    "Mag Slim":        ("https://www.magslim.com", ""),
    "BEZALEL":         ("https://bezalel.co", ""),
    "倍加能":           ("https://bezalel.co", ""),
    # AR/VR
    "Rokid":           ("https://www.rokid.com", ""),
    "RayNeo":          ("https://www.rayneo.com", ""),
    "MAD Gaze":        ("https://www.madgaze.com", ""),
    "Dream Glass":     ("https://dreamglasslab.com", ""),
    "DPVR":            ("https://www.dpvrvr.com", ""),
    "SmartEcho":       ("https://www.smartecho.ai", ""),
    "Looktech":        ("https://www.looktech.ai", ""),
    # PC周辺
    "ASUS":            ("https://www.asus.com/jp", ""),
    "ROG":             ("https://rog.asus.com", ""),
    "MOZTECH":         ("https://www.moztech.com.tw", ""),
    "GOOVIS":          ("https://www.goovis.com", ""),
    "AIFA":            ("https://www.aifa.com.tw", "service@aifa.com.tw"),
    "Swiftpoint":      ("https://www.swiftpoint.com", ""),
    "TRACPOINT":       ("https://www.swiftpoint.com", ""),
    "Gravastar":       ("https://www.gravastar.com", ""),
    "Mercury K1":      ("https://www.gravastar.com", ""),
    "DASUNG":          ("https://shop.dasung.com", "contact@dasung.com"),
    "Majestic Devices":("https://www.majesticdevices.com", ""),
    "Glyph":           ("https://www.majesticdevices.com", ""),
    # ウェアラブル
    "RingConn":        ("https://www.ringconn.com", ""),
    "Ring One":        ("https://join.musewearables.com", ""),
    "Muse Wearables":  ("https://join.musewearables.com", ""),
    "VIITA":           ("https://www.viitawatches.com", ""),
    "Sequent":         ("https://www.sequentag.com", ""),
    # オーディオ
    "HiDock":          ("https://www.hidock.com", "support@hidock.com"),
    "ZMI":             ("https://www.zmi.com", ""),
    "紫米":             ("https://www.zmi.com", ""),
    "TileRec":         ("https://www.atto-digital.com", ""),
    "Slimca":          ("https://www.atto-digital.com", ""),
    "Libratone":       ("https://www.libratone.com", ""),
    # セキュリティ
    "FAMMIX":          ("https://www.yomix.com.tw", ""),
    "Acer Gadget":     ("https://www.acer.com/tw", ""),
    "Acer":            ("https://www.acer.com/tw", ""),
    # スマートホーム
    "Yeelight":        ("https://www.yeelight.com", "support@yeelight.com"),
    "Lytmi":           ("https://www.lytmi.com", ""),
    "Energy Robotics": ("https://www.energyroboticstech.com", ""),
    "Eilik":           ("https://www.energyroboticstech.com", ""),
    "ECOVACS":         ("https://www.ecovacs.com/jp", "jp@ecovacs.com"),
    # 大手ブランド
    "Philips":         ("https://www.philips.com/ja-jp", ""),
    "LG":              ("https://www.lg.com/jp", "jp.lgservice@lge.com"),
    "Samsung":         ("https://www.samsung.com/jp", ""),
    "Cooler Master":   ("https://www.coolermaster.com/jp", ""),
    # 空気清浄
    "DIKE":            ("https://www.dike.com.tw", ""),
    "POIEMA":          ("https://www.poiema.com.tw", "custcare@poiema.com.tw"),
    "FLAT":            ("https://www.sauberair.com", ""),
    "Sauberair":       ("https://www.sauberair.com", ""),
    "Smini":           ("https://www.brise.com.tw", ""),
    "智米":             ("https://www.smartmi.com", ""),
    "Smartmi":         ("https://www.smartmi.com", ""),
    # フィットネス
    "RheoFit":         ("https://rheofit.com", ""),
    "OYO":             ("https://www.oyogym.com", ""),
    # ペット
    "ebot":            ("https://www.enabot.com", ""),
    "Enabot":          ("https://www.enabot.com", ""),
    # E-ink
    "NarTick":         ("https://nartick.com", "cs@nartick.com"),
    "Bloomin8":        ("https://bloomin8.com", ""),
    # 家具
    "COFO":            ("https://cofoglobal.com", ""),
    # クラフト
    "Cricut":          ("https://www.cricut.com", ""),
    "NECCHI":          ("https://www.necchi.com", ""),
}

link_font   = Font(name="Arial", size=10, color="0563C1", underline="single")
normal_font = Font(name="Arial", size=10)

wb = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
updated = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        product = str(ws.cell(row=row, column=2).value or "")
        maker   = str(ws.cell(row=row, column=3).value or "")
        hp_cell   = ws.cell(row=row, column=5)
        mail_cell = ws.cell(row=row, column=6)

        if str(hp_cell.value or "").startswith("http"):
            continue  # 既に入力済み

        combined = product + " " + maker
        for kw, (hp, mail) in MAKER_DATA.items():
            if kw.lower() in combined.lower():
                hp_cell.value = hp
                hp_cell.hyperlink = hp
                hp_cell.font = link_font
                hp_cell.alignment = Alignment(vertical="top", wrap_text=True)
                if mail and str(mail_cell.value or "") in ("要調査", ""):
                    mail_cell.value = mail
                    mail_cell.font = normal_font
                updated += 1
                break

wb.save(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
print(f"更新件数: {updated}件")

wb2 = load_workbook(r"C:\Users\hiro\HP0524\海外クラウドファンディング_日本未発売リスト.xlsx")
total = remain = 0
for sn in wb2.sheetnames:
    ws = wb2[sn]
    s_total = ws.max_row - 1
    s_remain = sum(1 for r in range(2, ws.max_row + 1)
                   if not str(ws.cell(row=r, column=5).value or "").startswith("http"))
    print(f"  [{sn}] 全{s_total}件 / HP未入力: {s_remain}件")
    total += s_total; remain += s_remain
print(f"  合計: {total}件 / HP未入力: {remain}件 / 入力済: {total-remain}件")
